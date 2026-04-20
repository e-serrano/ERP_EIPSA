from config.config_functions import config_database, config_sql_engine
import psycopg2
from windows.Excel_Export_Templates import material_order
import pandas as pd
from datetime import *
import PySide6.QtCore
from PySide6.QtWidgets import QFileDialog
import openpyxl
import re
from utils.Database_Manager import Database_Connection
from utils.Show_Message import MessageHelper
from fractions import Fraction


def flow_matorder(proxy, model, numorder, numorder_pedmat, variable, state):
    """
    Processes material raw orders for flow items by inserting new entries into the fabrication orders database.

    Args:
        proxy (QAbstractProxyModel): The proxy model containing the current data view.
        model (QAbstractItemModel): The model containing the main data.
        numorder (str): The order number to process.
        numorder_pedmat (str): The base order number for material orders.
        variable (str): A variable that determines the type of processing to be done. The specific usage
                        of this variable is not detailed in the function.

    Returns:
        None: This function does not return a value but modifies the database state.
    """
    id_list=[]
    orifice_flange_list = []
    line_flange_list = []
    gasket_list = []
    bolts_list = []
    plugs_list = []
    extractor_list = []
    plate_list = []
    nipple_list = []
    handle_list = []
    chring_list = []
    tube_list = []
    piece2_list = []
    bar_handle_list = []

    proxy_data = proxy.data
    proxy_index = proxy.index

    data = model.data
    index = model.index

    if numorder[0] == 'P':
        id_list = [
            proxy_data(proxy_index(row, 0))
            for row in range(proxy.rowCount())
            if proxy_data(proxy_index(row, 2)) == "PURCHASED" and proxy_data(proxy_index(row, 6)) != "ZZZ"
        ]
    elif numorder[0] == 'O':
        id_list = [
            proxy_data(proxy_index(row, 0))
            for row in range(proxy.rowCount())
            if proxy_data(proxy_index(row, 2)) == "QUOTED" and str(proxy_data(proxy_index(row, 5))) == ''
        ]

    commands_numot = ("""SELECT "ot_num"
                        FROM fabrication.fab_order
                        WHERE NOT "ot_num" LIKE '90%'
                        ORDER BY "ot_num" ASC
                        """)

    check_otpedmat = f"SELECT * FROM fabrication.fab_order WHERE id = '{numorder_pedmat + '-PEDMAT'}'"

    commands_otpedmat = ("""
                            INSERT INTO fabrication.fab_order (
                            "id","tag","element","qty_element",
                            "ot_num","qty_ot","start_date")
                            VALUES (%s,%s,%s,%s,%s,%s,%s)
                            """)

    if state == 'Offer':
        num_ot = '0'
    else:
        try:
            with Database_Connection(config_database()) as conn:
                with conn.cursor() as cur:
                    cur.execute(commands_numot)
                    results=cur.fetchall()
                    num_ot=results[-1][0]

            excel_file_path = r"\\ERP-EIPSA-DATOS\Comunes\EIPSA Sistemas de Gestion\MasterCTF\Bases\Contador.xlsm"
            workbook = openpyxl.load_workbook(excel_file_path, keep_vba=True)
            worksheet = workbook.active
            num_ot = worksheet['B2'].value
            with Database_Connection(config_database()) as conn:
                with conn.cursor() as cur:
                    cur.execute(check_otpedmat)
                    results=cur.fetchall()

            if len(results) == 0:
                data_numot=(numorder_pedmat + '-PEDMAT', numorder_pedmat, 'PEDIDO DE MATERIALES', 1, '{:06}'.format(int(num_ot) + 1), len(id_list), date.today().strftime("%d/%m/%Y"))
                with Database_Connection(config_database()) as conn:
                    with conn.cursor() as cur:
                        cur.execute(commands_otpedmat, data_numot)
                    conn.commit()

                worksheet['B2'].value = '{:06}'.format(int(num_ot) + 1)
                workbook.save(excel_file_path)

        except (Exception, psycopg2.DatabaseError) as error:
            MessageHelper.show_message("Ha ocurrido el siguiente error:\n"
                        + str(error), "critical")

    row_map = {
        data(index(row, 0)): row
        for row in range(model.rowCount())
    }

    for element in id_list:
        row = row_map.get(element)
        if row is None:
            continue

        flange_material = data(index(row, 13))
        sch = data(index(row, 12))
        design_flange = str(data(index(row, 61))).replace('.', ',') # pipe internal diameter
        size = f"{data(index(row,9))} {data(index(row,10))} {data(index(row,11))}"

        all_list_parts =[]

        # setting list for eache element [code_element, code_fab_element, trad_element, design_element, process_element, material_element, qty_element, code_purch_element]
        code_orifice_flange = data(index(row, 162))
        if code_orifice_flange:
            orifice_flange_list.append([
                code_orifice_flange,
                data(index(row, 174)), # code fab orifice flange
                data(index(row, 198)), # trad orifice flange
                sch,
                design_flange,
                "", #data(index(row, 37)),
                flange_material,
                int(data(index(row, 186))) * int(data(index(row, 35))), # quantity orifice flange per equipment * number of equipments
                data(index(row, 210)) # code purch orifice flange
                ])
            all_list_parts.append(orifice_flange_list)

        code_line_flange = data(index(row, 163))
        if code_line_flange:
            line_flange_list.append([
                code_line_flange,
                data(index(row, 175)), # code fab line flange
                data(index(row, 199)), # trad line flange
                sch,
                design_flange,
                "", #data(index(row, 37)),
                flange_material,
                int(data(index(row, 187))) * int(data(index(row, 35))), # quantity line flange per equipment * number of equipments
                data(index(row, 211)) # code purch line flange
                ])
            all_list_parts.append(line_flange_list)

        code_gasket = data(index(row, 164))
        if code_gasket:
            gasket_list.append([
                code_gasket,
                data(index(row, 176)), # code fab gasket
                data(index(row, 200)), # trad gasket
                size,
                '',
                '',
                '',
                int(data(index(row, 42))) * int(data(index(row, 35))), # quantity gasket per equipment * number of equipments
                data(index(row, 212))]) # code purch gasket
            all_list_parts.append(gasket_list)

        code_bolts = data(index(row, 165))
        if code_bolts:
            bolts_list.append([
                code_bolts,
                data(index(row, 177)), # code fab bolts
                data(index(row, 201)), # trad bolts
                size,
                ('esp. placa ' + data(index(row, 21))),
                '',
                data(index(row, 24)) + " / " + data(index(row, 25)),
                (int(data(index(row, 44))) if data(index(row, 44)) != '' else 0) * int(data(index(row, 35))), # quantity bolts per equipment * number of equipments
                data(index(row, 213))]) # code purch bolts
            all_list_parts.append(bolts_list)

        code_extractor = data(index(row, 167))
        if code_extractor:
            extractor_list.append([
                code_extractor,
                data(index(row, 179)), # code fab extractor
                data(index(row, 203)), # trad extractor
                size,
                ('esp. placa ' + data(index(row, 21))),
                '',
                data(index(row, 47)),
                int(data(index(row, 49))) * int(data(index(row, 35))), # quantity extractor per equipment * number of equipments
                data(index(row, 215)) # code purch extractor
                ])
            all_list_parts.append(extractor_list)

        code_plate = data(index(row, 168))
        if code_plate:
            plate_list.append([
                code_plate,
                data(index(row, 180)), # code fab plate
                data(index(row, 204)), # trad plate
                ('ESP ' + data(index(row, 21)) + 'mm'),
                data(index(row, 62)),
                'ARAMCO' if data(index(row, 22)) =='ARA' else '',
                data(index(row, 19)),
                int(data(index(row, 28)) if data(index(row, 8)) == "MULTISTAGE RO" else 1) * int(data(index(row, 35))), # quantity of plates per equipment * number of equipments
                data(index(row, 216)) # code purch plate
                ])
            all_list_parts.append(plate_list)

        code_nipple = data(index(row, 169))
        if code_nipple:
            nipple_list.append([
                code_nipple,
                data(index(row, 181)), # code fab nipple
                data(index(row, 205)), # trad nipple
                '',
                '',
                '',
                data(index(row, 13)),
                int(data(index(row, 193))) * int(data(index(row, 35))), # quantity nipple per equipment * quantity of equipments
                data(index(row, 217)) # code purch nipple
                ])
            all_list_parts.append(nipple_list)

        code_handle = data(index(row, 170))
        if code_handle and data(index(row, 21)) not in ['3', '1/8" (3)']:
            handle_list.append([
                code_handle,
                data(index(row, 182)), # code fab handle
                data(index(row, 206)), # trad handle
                '' if data(index(row, 11)) == 'RTJ' else (data(index(row, 64)) + "x" + data(index(row, 65)) + "x" + data(index(row, 66)) +' mm'),
                '' if data(index(row, 11)) == 'RTJ' else data(index(row, 22)),
                '',
                '316SS',
                1 * int(data(index(row, 35))), # quantity of handles per equipment * quantity of equipments
                data(index(row, 218)) # code purch handle
                ])
            all_list_parts.append(handle_list)

        if code_handle and data(index(row, 21)) not in ['3', '1/8" (3)'] and data(index(row, 11)) == 'RTJ':
            bar_handle_list.append([
                'Barra Mango RTJ',
                'Barra Mango RTJ',
                'BARRA MANGO',
                '',
                '',
                '',
                '316SS',
                ((int(float(data(index(row, 64)))) - 30) if 'datos' not in data(index(row, 64)) else 0) * int(data(index(row, 35))), # length of bar handle per equipment * quantity of equipments
                ''
                ])
            all_list_parts.append(bar_handle_list)

        code_chring = data(index(row, 171))
        if code_chring:
            chring_list.append([
                code_chring,
                data(index(row, 183)), # code fab chring
                data(index(row, 207)), # trad chring
                'ESP ' if data(index(row, 11)) == "RTJ" else 'ESP 38,5mm ACABADO',
                'ø' + str(data(index(row, 62))),
                '', #data(index(row, 37)),
                data(index(row, 19)),
                1 * int(data(index(row, 35))), # quantity chring per equipment * quantity of equipments
                data(index(row, 219)) # code purch chring
                ])
            all_list_parts.append(chring_list)

        code_plugs = data(index(row, 166))
        if code_plugs != '':
            plugs_list.append([
                code_plugs,
                data(index(row, 178)), # code fab plug
                data(index(row, 202)), # trad plug
                '',
                '',
                '',
                data(index(row, 45)), # material plug
                (int(data(index(row, 46))) if data(index(row, 46)) != '' else 0) * int(data(index(row, 35))), # quantity plugs per equipment * quantity of equipment
                data(index(row, 214))
                ])
            all_list_parts.append(plugs_list)

        code_tube = data(index(row, 172))
        if code_tube:
            tube_list.append([
                code_tube,
                data(index(row, 184)), # code fab tube
                data(index(row, 208)), # trad tube
                sch,
                design_flange,
                '',
                data(index(row,15)),
                float(data(index(row, 196))) * int(data(index(row, 35))), # quantity tube per equipment (length of tube) * quantity of equipments
                data(index(row, 220)) # code purch tube
                ])
            all_list_parts.append(tube_list)

        code_piece2 = data(index(row, 173))
        if code_piece2:
            commands_thk = ("""
                SELECT wall_thk
                FROM validation_data.pipe_diam
                WHERE (line_size = %s
                AND
                sch = %s)
                """)

            try:
                with Database_Connection(config_database()) as conn:
                    with conn.cursor() as cur:
                        cur.execute(commands_thk,(data(index(row, 9)), data(index(row, 12)),))
                        results=cur.fetchone()
                        thkmin=results[0]

            except (Exception, psycopg2.DatabaseError) as error:
                MessageHelper.show_message("Ha ocurrido el siguiente error:\n"
                            + str(error), "critical")

            commands_flangecode = ("""
                SELECT code
                FROM validation_data.flow_flange_material
                WHERE flange_material = %s
                """)

            commands_sheetmaterial = ("""
                SELECT sheet_material
                FROM validation_data.flow_sheet_material
                WHERE code = %s
                """)

            try:
                with Database_Connection(config_database()) as conn:
                    with conn.cursor() as cur:
                        cur.execute(commands_flangecode,(data(index(row, 13)),))
                        results=cur.fetchone()
                        code=results[0]

                        cur.execute(commands_sheetmaterial,(code,))
                        results=cur.fetchall()
                        materialpiece2 = results[0]

            except (Exception, psycopg2.DatabaseError) as error:
                MessageHelper.show_message("Ha ocurrido el siguiente error:\n"
                            + str(error), "critical")

            piece2_list.append([
                code_piece2,
                data(index(row, 185)), # code fab wedge
                data(index(row, 209)), # trad wedge
                ('Th mín ' + thkmin + 'mm'),
                '',
                '',
                materialpiece2,
                1 * int(data(index(row, 35))), # quantity of wedge parts per equipment * quantity of equipments
                data(index(row, 221)) # code purch wedge
                ])
            all_list_parts.append(piece2_list)

        if state == 'Order':
            columns_equipments = ["code_equipment", "code_fab_equipment", "translate_equipment", "section_type",
                                "f_orifice_flange", "qty_f_orifice_flange", "f_line_flange", "qty_f_line_flange",
                                "f_gasket", "qty_f_gasket", "f_bolts", "qty_f_bolts",
                                "f_plug", "qty_f_plug", "f_extractor", "qty_f_extractor",
                                "f_plate", "qty_f_plate", "f_nipple", "qty_f_nipple",
                                "f_handle", "qty_f_handle", "f_chring", "qty_f_chring",
                                "f_tube", "qty_f_tube", "f_piece2", "qty_f_piece2"]
            columns_parts = ["code_part", "code_fab_part", "code_element", "model", "design", "process", "material", "section_type"]
            columns_tags = ["code", "equipment", "num_order","order_material","contractual_date","inspection"]

            values_equipments = [data(index(row, 159)), data(index(row, 160)), data(index(row, 161)), "Q-CAUD",
                                data(index(row, 162)), data(index(row, 186)), data(index(row, 163)), data(index(row, 187)),
                                data(index(row, 164)), data(index(row, 188)), data(index(row, 165)), data(index(row, 189)),
                                data(index(row, 166)), data(index(row, 190)), data(index(row, 167)), data(index(row, 191)),
                                data(index(row, 168)), data(index(row, 192)), data(index(row, 169)), data(index(row, 193)),
                                data(index(row, 170)), data(index(row, 194)), data(index(row, 171)), data(index(row, 195)),
                                data(index(row, 172)), data(index(row, 196)), data(index(row, 173)), data(index(row, 197))]

            values_tags = [data(index(row, 4)) + "-" + data(index(row, 8)) + "-" + data(index(row, 1)), 
                            data(index(row, 159)), data(index(row, 4)), data(index(row, 94)),
                            data(index(row, 39)), data(index(row, 141))]

            columns_equipments  = ", ".join([f'"{column}"' for column in columns_equipments])
            values_equipments =  ", ".join(['NULL' if value == '' or value == 0 else (str(value) if isinstance(value, (int, float)) else f"'{str(value)}'") for value in values_equipments])

            columns_tags  = ", ".join([f'"{column}"' for column in columns_tags])
            values_tags =  ", ".join(['NULL' if value == '' or value == PySide6.QtCore.QDate() else (str(value) if isinstance(value, (int, float)) else (f"'{value.toString('yyyy-MM-dd')}'" if isinstance(value, PySide6.QtCore.QDate) else f"'{str(value)}'")) for value in values_tags])

            columns_parts = ", ".join([f'"{column}"' for column in columns_parts])

            commands_equipments = f"INSERT INTO fabrication.equipments ({columns_equipments}) VALUES ({values_equipments})"
            commands_tags = f"INSERT INTO fabrication.tags ({columns_tags}) VALUES ({values_tags})"

            check_equipments = f"SELECT * FROM fabrication.equipments WHERE code_equipment = '{data(index(row, 159))}'"

            try:
                with Database_Connection(config_database()) as conn:
                    with conn.cursor() as cur:
                        cur.execute(check_equipments)
                        results=cur.fetchall()

                if len(results) == 0:
                    with Database_Connection(config_database()) as conn:
                        with conn.cursor() as cur:
                            cur.execute(commands_equipments)
                        conn.commit()

                else:
                    set_clause = ", ".join([f"{column} = {value}" for column, value in zip(columns_equipments.split(", ")[1:], values_equipments.split(", ")[1:])])
                    update_equipments = f"UPDATE fabrication.equipments SET {set_clause} WHERE code_equipment = '{data(index(row, 159))}'"
                    with Database_Connection(config_database()) as conn:
                        with conn.cursor() as cur:
                            cur.execute(update_equipments)
                        conn.commit()

                for list_part in all_list_parts:
                    check_parts = f"SELECT * FROM fabrication.parts WHERE code_part = '{list_part[0][0]}'"
                    with Database_Connection(config_database()) as conn:
                        with conn.cursor() as cur:
                            cur.execute(check_parts)
                            results=cur.fetchall()

                    if len(results) == 0:
                        list_part_modified = list_part[0][:8].copy()
                        list_part_modified[-1] = 'Q-CAUD'
                        values_parts = ", ".join('NULL' if value == '' else (str(value) if isinstance(value, (int, float)) else f"'{str(value)}'") for value in list_part_modified)
                        commands_parts = f"INSERT INTO fabrication.parts ({columns_parts}) VALUES ({values_parts})"
                        with Database_Connection(config_database()) as conn:
                            with conn.cursor() as cur:
                                cur.execute(commands_parts)
                            conn.commit()

                    else:
                        list_part_modified = list_part[0][:8].copy()
                        list_part_modified[-1] = 'Q-CAUD'
                        values_parts = ", ".join('NULL' if value == '' else (str(value) if isinstance(value, (int, float)) else f"'{str(value)}'") for value in list_part_modified)
                        set_clause = ", ".join([f"{column} = {value}" for column, value in zip(columns_parts.split(", ")[1:], values_parts.split(", ")[1:])])
                        update_parts = f"UPDATE fabrication.parts SET {set_clause} WHERE code_part = '{list_part[0][0]}'"
                        with Database_Connection(config_database()) as conn:
                            with conn.cursor() as cur:
                                cur.execute(update_parts)
                            conn.commit()

                with Database_Connection(config_database()) as conn:
                    with conn.cursor() as cur:
                        cur.execute(commands_tags)
                    conn.commit()

            except (Exception, psycopg2.DatabaseError) as error:
                MessageHelper.show_message("Ha ocurrido el siguiente error:\n"
                            + str(error), "critical")

# Turn all lists in dataframe and grouped in order to sum same items
    data_lists = [
    (orifice_flange_list, "df_orifice_flange"),
    (line_flange_list, "df_line_flange"),
    (gasket_list, "df_gasket"),
    (bolts_list, "df_bolts"),
    (plugs_list, "df_plugs"),
    (extractor_list, "df_extractor"),
    (plate_list, "df_plate"),
    (nipple_list, "df_nipple"),
    (handle_list, "df_handle"),
    (chring_list, "df_chring"),
    (tube_list, "df_tube"),
    (piece2_list, "df_piece2"),
    (bar_handle_list, "df_bar_handle")]

    data_frames_with_data = []

    for data_list, df_name in data_lists:
        if data_list:
            sublists = [sublist[2:] for sublist in data_list]
            df = pd.DataFrame(sublists, columns=['descripción', 'modelo', 'diseño', 'proceso', 'material', 'cantidad', 'suministro'])
            df = df.groupby(['descripción', 'modelo', 'diseño', 'proceso', 'material', 'suministro'])['cantidad'].sum().reset_index()
            data_frames_with_data.append(df)

    if data_frames_with_data:
        df_final = pd.concat(data_frames_with_data, ignore_index=True)

        values_supplies = df_final['suministro'].dropna().unique().tolist()

        query = """
        SELECT reference AS suministro, physical_stock AS st_fisico, available_stock AS st_disponible, pending_stock as st_pend, virtual_stock AS st_virtual
        FROM purch_fact.supplies
        WHERE reference IN %(values)s
        """

        with Database_Connection(config_database()) as conn:
            df_supplies = pd.read_sql(query, config_sql_engine(), params={"values": tuple(values_supplies)})

        df_final = (df_final.merge(df_supplies, on='suministro', how='left')
                    [['descripción', 'modelo', 'diseño', 'proceso', 'material', 'cantidad', 'suministro',
                        'st_fisico', 'st_disponible', 'st_pend', 'st_virtual']])

        df_final['almacen_si'] = ''
        df_final['almacen_no'] = ''
        df_final['proveedor'] = ''
        df_final['fecha_pedido'] = ''
        df_final['fecha_prevista'] = ''

        df_final = df_final[
            ['descripción', 'modelo', 'diseño', 'proceso', 'material',
            'cantidad', 'almacen_si', 'almacen_no', 'suministro',
            'proveedor', 'fecha_pedido', 'fecha_prevista',
            'st_fisico', 'st_disponible', 'st_pend', 'st_virtual']
        ]

    commands_client_order = ("""
                SELECT orders."num_order",orders."num_offer",offers."client"
                FROM offers
                INNER JOIN orders ON (offers."num_offer"=orders."num_offer")
                WHERE UPPER(orders."num_order") LIKE UPPER('%%'||%s||'%%')
                """)

    commands_client_offer = ("""
                SELECT offers."num_offer",offers."client"
                FROM offers
                WHERE UPPER(offers."num_offer") LIKE UPPER('%%'||%s||'%%')
                """)

    try:
        with Database_Connection(config_database()) as conn:
            with conn.cursor() as cur:
                if state == 'Order':
                    cur.execute(commands_client_order,(numorder,))
                    results=cur.fetchone()
                    client=results[2]
                else:
                    cur.execute(commands_client_offer,(numorder,))
                    results=cur.fetchone()
                    client=results[1]

    except (Exception, psycopg2.DatabaseError) as error:
        MessageHelper.show_message("Ha ocurrido el siguiente error:\n"
                    + str(error), "critical")

    excel_mat_order = material_order(df_final, numorder_pedmat, client, variable, num_ot)
    excel_mat_order.save_excel()


def temp_matorder(proxy, model, numorder, numorder_pedmat, variable, state):
    """
    Processes material raw orders for temp items by inserting new entries into the fabrication orders database.

    Args:
        proxy (QAbstractProxyModel): The proxy model containing the current data view.
        model (QAbstractItemModel): The model containing the main data.
        numorder (str): The order number to process.
        numorder_pedmat (str): The base order number for material orders.
        variable (str): A variable that determines the type of processing to be done. The specific usage
                        of this variable is not detailed in the function.

    Returns:
        None: This function does not return a value but modifies the database state.
    """
    id_list=[]
    bar_list = []
    tube_list = []
    flange_list = []
    sensor_list = []
    head_list = []
    btb_list = []
    nipple_list = []
    spring_list = []
    plug_list = []
    puntal_list = []
    tw_list = []
    extcable_list = []

    proxy_data = proxy.data
    proxy_index = proxy.index

    data = model.data
    index = model.index

    if numorder[0] == 'P':
        id_list = [
            proxy_data(proxy_index(row, 0))
            for row in range(proxy.rowCount())
            if proxy_data(proxy_index(row, 2)) == "PURCHASED" and proxy_data(proxy_index(row, 6)) != "ZZZ"
        ]
    elif numorder[0] == 'O':
        id_list = [
            proxy_data(proxy_index(row, 0))
            for row in range(proxy.rowCount())
            if proxy_data(proxy_index(row, 2)) == "QUOTED" and str(proxy_data(proxy_index(row, 5))) == ''
        ]

    commands_numot = ("""SELECT "ot_num"
                        FROM fabrication.fab_order
                        WHERE NOT "ot_num" LIKE '90%'
                        ORDER BY "ot_num" ASC
                        """)

    check_otpedmat = f"SELECT * FROM fabrication.fab_order WHERE id = '{numorder_pedmat + '-PEDMAT'}'"

    commands_otpedmat = ("""
                            INSERT INTO fabrication.fab_order (
                            "id","tag","element","qty_element",
                            "ot_num","qty_ot","start_date")
                            VALUES (%s,%s,%s,%s,%s,%s,%s)
                            """)

    if state == 'Offer':
        num_ot = '0'
    else:
        try:
            with Database_Connection(config_database()) as conn:
                with conn.cursor() as cur:
                    cur.execute(commands_numot)
                    results=cur.fetchall()
                    num_ot=results[-1][0]

            excel_file_path = r"\\ERP-EIPSA-DATOS\Comunes\EIPSA Sistemas de Gestion\MasterCTF\Bases\Contador.xlsm"
            workbook = openpyxl.load_workbook(excel_file_path, keep_vba=True)
            worksheet = workbook.active
            num_ot = worksheet['B2'].value

            with Database_Connection(config_database()) as conn:
                with conn.cursor() as cur:
                    cur.execute(check_otpedmat)
                    results=cur.fetchall()

            if len(results) == 0:
                data_numot=(numorder_pedmat + '-PEDMAT', numorder_pedmat, 'PEDIDO DE MATERIALES', 1, '{:06}'.format(int(num_ot) + 1), len(id_list), date.today().strftime("%d/%m/%Y"))
                with Database_Connection(config_database()) as conn:
                    with conn.cursor() as cur:
                        cur.execute(commands_otpedmat, data_numot)
                    conn.commit()

                worksheet['B2'].value = '{:06}'.format(int(num_ot) + 1)
                workbook.save(excel_file_path)

        except (Exception, psycopg2.DatabaseError) as error:
            MessageHelper.show_message("Ha ocurrido el siguiente error:\n"
                        + str(error), "critical")

    row_map = {
        data(index(row, 0)): row
        for row in range(model.rowCount())
    }

    for element in id_list:
        row = row_map.get(element)
        if row is None:
            continue

        all_list_parts =[]

        tw_type = data(index(row, 9))

        # setting list for eache element [code_element, code_fab_element, trad_element, design_elemente, process_element, material_element, qty_element, code_purch_element]
        code_bar = data(index(row, 142))
        if code_bar:
            bar_list.append([
                code_bar,
                data(index(row, 154)),
                data(index(row, 178)) if 'Helical' not in tw_type else 'VAINA HELICOIDAL' + (' BRIDADA ' + data(index(row, 10)) + ' ' + data(index(row, 11)) + ' ' + data(index(row, 12)) if tw_type == 'Flanged Helical' else ''),
                ('U=' + data(index(row, 16)) + ' /L=' + data(index(row, 15)) if 'Stone' in tw_type or 'Helical' in tw_type
                        else 'Barra ø=' + (data(index(row, 50)))),
                ('RAÍZ ø=' + data(index(row, 17))) if tw_type == 'Van-Stone TW' else '',
                '',
                data(index(row, 14)),
                data(index(row, 166)) if 'Helical' not in tw_type else 1,
                data(index(row, 190))
                ])
            all_list_parts.append(bar_list)

        code_tube = data(index(row, 143))
        if code_tube:
            tube_list.append([
                code_tube,
                data(index(row, 155)),
                data(index(row, 179)),
                data(index(row, 38)),
                '',
                '',
                data(index(row, 14)),
                data(index(row, 167)),
                data(index(row, 191))
                ])
            all_list_parts.append(tube_list)

        code_flange = data(index(row, 144))
        if code_flange:
            tw_types_list = ['Buttweld TW','Forged Flanged TW','Threaded Helical','Van-Stone Helical','VORTICRACK']
            flange_list.append([
                code_flange,
                data(index(row, 156)),
                data(index(row, 180)) if tw_type not in tw_types_list else '',
                '',
                '',
                '',
                (data(index(row, 35)) if tw_type == 'Van-Stone TW' else (data(index(row, 14)) if tw_type not in tw_types_list else '')),
                1 if tw_type not in tw_types_list else '',
                data(index(row, 192))
                ])
            all_list_parts.append(flange_list)

        code_sensor = data(index(row, 145))
        if code_sensor:
            sensor_list.append([
                code_sensor,
                data(index(row, 157)),
                data(index(row, 181)),
                (data(index(row, 33)) + '-' + data(index(row, 32))) if code_sensor[:4] == 'Bime' else '',
                (data(index(row, 27)) + '-' + data(index(row, 28))) if code_sensor[:4] == 'Bime' else '',
                '',
                'PLATINO' if data(index(row, 181))[:5] == 'PT100' else ('AC. INOX.' if data(index(row, 24)) == 'St.Steel' else data(index(row, 24))),
                1 if (data(index(row, 181))[:5] == 'PT100' or code_sensor[:4] == 'Bime') else ((float(data(index(row, 56)))/1000) if data(index(row, 56)) != '' else ''),
                data(index(row, 193))
                ])
            all_list_parts.append(sensor_list)

        code_head = data(index(row, 146))
        if code_head:
            head_list.append([
                code_head,
                data(index(row, 158)),
                data(index(row, 182)),
                data(index(row, 31)),
                '',
                data(index(row, 33)),
                ('ALUMINIO' if data(index(row, 31))[-2:] == 'AL' 
                    else ('AC.CARBONO' if data(index(row, 31))[-2:] == 'CS' 
                    else ('AC.INOXIDABLE' if data(index(row, 31))[-2:] == 'SS' 
                    else 'MATERIAL CABEZA NO DEFINIDO'))),
                1,
                data(index(row, 194))
                ])
            all_list_parts.append(head_list)

        code_btb = data(index(row, 147))
        if code_btb:
            btb_list.append([
                code_btb,
                data(index(row, 159)),
                data(index(row, 183)),
                ("RANGO " + data(index(row, 27)) + '-' + data(index(row, 28))) if code_btb[:2] == 'BI' else '',
                '',
                '',
                data(index(row, 24)) if code_btb[:2] == 'BI' else ('CERÁMICO' if code_btb[:2] == 'CE' else ''),
                data(index(row, 171)),
                data(index(row, 195))
                ])
            all_list_parts.append(btb_list)

        code_nipple = data(index(row, 148))
        if code_nipple:
            nipple_list.append([
                code_nipple,
                data(index(row, 160)),
                data(index(row, 184)),
                ('' if data(index(row, 30)) == 'N/A' or data(index(row, 30))=='' else data(index(row, 30))),
                '',
                '',
                'A-105/A106' if data(index(row, 184))[data(index(row, 184)).find('('):data(index(row, 184)).find('(')+9] == '(CS)' else 'AISI-316',
                1,
                data(index(row, 196))
                ])
            all_list_parts.append(nipple_list)

        code_spring = data(index(row, 149))
        if code_spring:
            spring_list.append([
                code_spring,
                data(index(row, 161)),
                data(index(row, 185)),
                '',
                '',
                '',
                'AC.INOX',
                1,
                data(index(row, 197))
                ])
            all_list_parts.append(spring_list)

        code_plug = data(index(row, 151))
        if code_plug:
            plug_list.append([
                code_plug,
                data(index(row, 163)),
                data(index(row, 187)),
                '',
                '',
                '',
                data(index(row, 187))[data(index(row, 187)).find('('):data(index(row, 187)).find('(')+9],
                1,
                data(index(row, 199))
                ])
            all_list_parts.append(plug_list)

        code_puntal = data(index(row, 150))
        if code_puntal:
            puntal_list.append([
                code_puntal,
                data(index(row, 162)),
                data(index(row, 186)),
                '',
                '',
                '',
                data(index(row, 14)),
                float(code_puntal[1:8])/1000,
                data(index(row, 198))
                ])
            all_list_parts.append(puntal_list)

        code_tw = data(index(row, 152))
        if code_tw and ('Van-Stone TW' in tw_type or 'Forged' in tw_type):
            tw_list.append([
                code_tw,
                data(index(row, 164)),
                data(index(row, 188)),
                'U=' + data(index(row, 16)) + ' / L=' + data(index(row, 15)),
                '',
                '',
                data(index(row, 14)),
                data(index(row, 176)),
                data(index(row, 200))
                ])
            all_list_parts.append(tw_list)

        code_extcable = data(index(row, 153))
        if code_extcable != '':
            extcable_list.append([
                code_extcable,
                data(index(row, 165)),
                data(index(row, 189)),
                '',
                '',
                '',
                'AC. INOX.' if data(index(row, 24)) in ['AISI-304', 'AISI-310', 'AISI-316', 'AISI-321', 'St.Steel'] else data(index(row, 24)),
                float(data(index(row, 177))) if data(index(row, 177)) != '' else 0,
                data(index(row, 201))
                ])
            all_list_parts.append(extcable_list)

        columns_equipments = ["code_equipment", "code_fab_equipment", "translate_equipment", "section_type",
                                        "t_bar", "qty_t_bar", "t_tube", "qty_t_tube",
                                        "t_flange", "qty_t_flange", "t_sensor", "qty_t_sensor",
                                        "t_head", "qty_t_head", "t_btb", "qty_t_btb",
                                        "t_nippleextcomp", "qty_t_nippleextcomp", "t_spring", "qty_t_spring",
                                        "t_puntal", "qty_t_puntal", "t_plug", "qty_t_plug",
                                        "t_tw", "qty_t_tw", "t_extcable", "qty_t_extcable"]

        columns_parts = ["code_part", "code_fab_part", "code_element", "model", "design", "process", "material", "section_type"]
        columns_tags = ["code", "equipment", "num_order", "order_material", "contractual_date", "inspection"]

        values_equipments = [data(index(row, 139)), data(index(row, 140)), data(index(row, 141)), "T-TEMP",
                            data(index(row, 142)), data(index(row, 166)), data(index(row, 143)), data(index(row, 167)),
                            data(index(row, 144)), data(index(row, 168)), data(index(row, 145)), data(index(row, 169)),
                            data(index(row, 146)), data(index(row, 170)), data(index(row, 147)), data(index(row, 171)),
                            data(index(row, 148)), data(index(row, 172)), data(index(row, 149)), data(index(row, 173)),
                            data(index(row, 150)), data(index(row, 174)), data(index(row, 151)), data(index(row, 175)),
                            data(index(row, 152)), data(index(row, 176)), data(index(row, 153)), data(index(row, 177))]

        values_tags = [data(index(row, 4)) + "-" + data(index(row, 8)) + "-" + data(index(row, 1)), 
                        data(index(row, 139)), data(index(row, 4)), data(index(row, 66)),
                        data(index(row, 43)), data(index(row, 121))]

        columns_equipments  = ", ".join([f'"{column}"' for column in columns_equipments])
        values_equipments =  ", ".join(['NULL' if value == '' or value == 0 else (str(value) if isinstance(value, (int, float)) else f"'{str(value)}'") for value in values_equipments])

        columns_tags  = ", ".join([f'"{column}"' for column in columns_tags])
        values_tags =  ", ".join(['NULL' if value == '' or value == PySide6.QtCore.QDate() else (str(value) if isinstance(value, (int, float)) else (f"'{value.toString('yyyy-MM-dd')}'" if isinstance(value, PySide6.QtCore.QDate) else f"'{str(value)}'")) for value in values_tags])

        columns_parts = ", ".join([f'"{column}"' for column in columns_parts])

        commands_equipments = f"INSERT INTO fabrication.equipments ({columns_equipments}) VALUES ({values_equipments})"
        commands_tags = f"INSERT INTO fabrication.tags ({columns_tags}) VALUES ({values_tags})"

        check_equipments = f"SELECT * FROM fabrication.equipments WHERE code_equipment = '{data(index(row, 139))}'"

        if state == 'Order':
            try:
                with Database_Connection(config_database()) as conn:
                    with conn.cursor() as cur:
                        cur.execute(check_equipments)
                        results=cur.fetchall()

                if len(results) == 0:
                    with Database_Connection(config_database()) as conn:
                        with conn.cursor() as cur:
                            cur.execute(commands_equipments)
                        conn.commit()

                else:
                    set_clause = ", ".join([f"{column} = {value}" for column, value in zip(columns_equipments.split(", ")[1:], values_equipments.split(", ")[1:])])
                    update_equipments = f"UPDATE fabrication.equipments SET {set_clause} WHERE code_equipment = '{data(index(row, 139))}'"
                    with Database_Connection(config_database()) as conn:
                        with conn.cursor() as cur:
                            cur.execute(update_equipments)
                        conn.commit()

                for list_part in all_list_parts:
                    check_parts = f"SELECT * FROM fabrication.parts WHERE code_part = '{list_part[0][0]}'"
                    with Database_Connection(config_database()) as conn:
                        with conn.cursor() as cur:
                            cur.execute(check_parts)
                            results=cur.fetchall()

                    if len(results) == 0:
                        list_part_modified = list_part[0][:8].copy()
                        list_part_modified[-1] = 'T-TEMP'
                        values_parts = ", ".join('NULL' if value == '' else (str(value) if isinstance(value, (int, float)) else f"'{str(value)}'") for value in list_part_modified)
                        commands_parts = f"INSERT INTO fabrication.parts ({columns_parts}) VALUES ({values_parts})"
                        with Database_Connection(config_database()) as conn:
                            with conn.cursor() as cur:
                                cur.execute(commands_parts)
                            conn.commit()
                    else:
                        list_part_modified = list_part[0][:8].copy()
                        list_part_modified[-1] = 'T-TEMP'
                        values_parts = ", ".join('NULL' if value == '' else (str(value) if isinstance(value, (int, float)) else f"'{str(value)}'") for value in list_part_modified)
                        set_clause = ", ".join([f"{column} = {value}" for column, value in zip(columns_parts.split(", ")[1:], values_parts.split(", ")[1:])])
                        update_parts = f"UPDATE fabrication.parts SET {set_clause} WHERE code_part = '{list_part[0][0]}'"
                        with Database_Connection(config_database()) as conn:
                            with conn.cursor() as cur:
                                cur.execute(update_parts)
                            conn.commit()

                with Database_Connection(config_database()) as conn:
                    with conn.cursor() as cur:
                        cur.execute(commands_tags)
                    conn.commit()

            except (Exception, psycopg2.DatabaseError) as error:
                MessageHelper.show_message("Ha ocurrido el siguiente error:\n"
                            + str(error), "critical")

# Turn all lists in dataframe and grouped in order to sum same items
    data_lists = [
    (bar_list, "df_bar"),
    (tube_list, "df_tube"),
    (flange_list, "df_flange"),
    (sensor_list, "df_sensor"),
    (head_list, "df_head"),
    (btb_list, "df_btb"),
    (nipple_list, "df_nipple"),
    (spring_list, "df_spring"),
    (plug_list, "df_plug"),
    (puntal_list, "df_puntal"),
    (extcable_list, "df_extcable"),
    (tw_list, "df_tw")]

    data_frames_with_data = []

    for data_list, df_name in data_lists:
        if data_list:
            sublists = [sublist[2:] for sublist in data_list]
            df = pd.DataFrame(sublists, columns=['descripción', 'modelo', 'diseño', 'proceso', 'material', 'cantidad', 'suministro'])
            df = df.groupby(['descripción', 'modelo', 'diseño', 'proceso', 'material', 'suministro'])['cantidad'].sum().reset_index()
            data_frames_with_data.append(df)

    if data_frames_with_data:
        df_final = pd.concat(data_frames_with_data, ignore_index=True)

        values_supplies = df_final['suministro'].dropna().unique().tolist()

        query = """
        SELECT reference AS suministro, physical_stock AS st_fisico, available_stock AS st_disponible, pending_stock as st_pend, virtual_stock AS st_virtual
        FROM purch_fact.supplies
        WHERE reference IN %(values)s
        """

        with Database_Connection(config_database()) as conn:
            df_supplies = pd.read_sql(query, config_sql_engine(), params={"values": tuple(values_supplies)})

        df_final = (df_final.merge(df_supplies, on='suministro', how='left')
                    [['descripción', 'modelo', 'diseño', 'proceso', 'material', 'cantidad', 'suministro',
                        'st_fisico', 'st_disponible', 'st_pend', 'st_virtual']])

        df_final['almacen_si'] = ''
        df_final['almacen_no'] = ''
        df_final['proveedor'] = ''
        df_final['fecha_pedido'] = ''
        df_final['fecha_prevista'] = ''

        df_final = df_final[
            ['descripción', 'modelo', 'diseño', 'proceso', 'material',
            'cantidad', 'almacen_si', 'almacen_no', 'suministro',
            'proveedor', 'fecha_pedido', 'fecha_prevista',
            'st_fisico', 'st_disponible', 'st_pend', 'st_virtual']
        ]

    commands_client_order = ("""
                SELECT orders."num_order",orders."num_offer",offers."client"
                FROM offers
                INNER JOIN orders ON (offers."num_offer"=orders."num_offer")
                WHERE UPPER(orders."num_order") LIKE UPPER('%%'||%s||'%%')
                """)

    commands_client_offer = ("""
                SELECT offers."num_offer",offers."client"
                FROM offers
                WHERE UPPER(offers."num_offer") LIKE UPPER('%%'||%s||'%%')
                """)

    try:
        with Database_Connection(config_database()) as conn:
            with conn.cursor() as cur:
                if state == 'Order':
                    cur.execute(commands_client_order,(numorder,))
                    results=cur.fetchone()
                    client=results[2]
                else:
                    cur.execute(commands_client_offer,(numorder,))
                    results=cur.fetchone()
                    client=results[1]

    except (Exception, psycopg2.DatabaseError) as error:
        MessageHelper.show_message("Ha ocurrido el siguiente error:\n"
                    + str(error), "critical")

    excel_mat_order = material_order(df_final, numorder_pedmat, client, variable, num_ot)
    excel_mat_order.save_excel()


def level_matorder(proxy, model, numorder, numorder_pedmat, variable, state):
    """
    Processes material raw orders for level items by inserting new entries into the fabrication orders database.

    Args:
        proxy (QAbstractProxyModel): The proxy model containing the current data view.
        model (QAbstractItemModel): The model containing the main data.
        numorder (str): The order number to process.
        numorder_pedmat (str): The base order number for material orders.
        variable (str): A variable that determines the type of processing to be done. The specific usage
                        of this variable is not detailed in the function.

    Returns:
        None: This function does not return a value but modifies the database state.
    """
    def expand_illuminators_from_list(illuminator_list):
        """
        Expand combinations like '1 x 1-I 218-T + 2 x 1-I 228-T'
        mantaining fields and original structure.
        """
        rows = []
        for item in illuminator_list:
            code_ill, codefab_ill, tradcod, model, design, process, material, qty = item

            parts = [p.strip() for p in str(tradcod).split("+")]
            for p in parts:
                match = re.match(r"(\d+)\s*x\s*(1-I\s*\d+-T)", p)
                if match:
                    quantity = int(match.group(1)) * qty
                    illuminator = "ILUMINADOR " + match.group(2).replace(" ", "")
                    rows.append([
                        code_ill, codefab_ill, illuminator, model, design, process, material, quantity
                    ])
        return rows

    id_list = []
    body_list = []
    cover_list = []
    glass_list = []
    gasket_list = []
    mica_list = []
    bolts_list = []
    nipplehex_list = []
    valve_list = []
    flangevalve_list = []
    nippletube_list = []
    dv_list = []
    plug_list = []
    antifrost_list = []
    illuminator_list = []

    proxy_data = proxy.data
    proxy_index = proxy.index

    data = model.data
    index = model.index

    if numorder[0] == 'P':
        id_list = [
            proxy_data(proxy_index(row, 0))
            for row in range(proxy.rowCount())
            if proxy_data(proxy_index(row, 2)) == "PURCHASED" and proxy_data(proxy_index(row, 6)) != "ZZZ"
        ]
    elif numorder[0] == 'O':
        id_list = [
            proxy_data(proxy_index(row, 0))
            for row in range(proxy.rowCount())
            if proxy_data(proxy_index(row, 2)) == "QUOTED" and str(proxy_data(proxy_index(row, 5))) == ''
        ]

    commands_numot = ("""SELECT "ot_num"
                        FROM fabrication.fab_order
                        WHERE NOT "ot_num" LIKE '90%'
                        ORDER BY "ot_num" ASC
                        """)

    check_otpedmat = f"SELECT * FROM fabrication.fab_order WHERE id = '{numorder_pedmat + '-PEDMAT'}'"

    commands_otpedmat = ("""
                            INSERT INTO fabrication.fab_order (
                            "id","tag","element","qty_element",
                            "ot_num","qty_ot","start_date")
                            VALUES (%s,%s,%s,%s,%s,%s,%s)
                            """)

    if state == 'Offer':
        num_ot = '0'
    else:
        try:
            with Database_Connection(config_database()) as conn:
                with conn.cursor() as cur:
                    cur.execute(commands_numot)
                    results=cur.fetchall()
                    num_ot=results[-1][0]

            excel_file_path = r"\\ERP-EIPSA-DATOS\Comunes\EIPSA Sistemas de Gestion\MasterCTF\Bases\Contador.xlsm"
            workbook = openpyxl.load_workbook(excel_file_path, keep_vba=True)
            worksheet = workbook.active
            num_ot = worksheet['B2'].value

            with Database_Connection(config_database()) as conn:
                with conn.cursor() as cur:
                    cur.execute(check_otpedmat)
                    results=cur.fetchall()

            if len(results) == 0:
                data_numot=(numorder_pedmat + '-PEDMAT', numorder_pedmat, 'PEDIDO DE MATERIALES', 1, '{:06}'.format(int(num_ot) + 1), len(id_list), date.today().strftime("%d/%m/%Y"))
                with Database_Connection(config_database()) as conn:
                    with conn.cursor() as cur:
                        cur.execute(commands_otpedmat, data_numot)
                    conn.commit()

                worksheet['B2'].value = '{:06}'.format(int(num_ot) + 1)
                workbook.save(excel_file_path)

        except (Exception, psycopg2.DatabaseError) as error:
            MessageHelper.show_message("Ha ocurrido el siguiente error:\n"
                        + str(error), "critical")

    row_map = {
        data(index(row, 0)): row
        for row in range(model.rowCount())
    }

    for element in id_list:
        row = row_map.get(element)
        if row is None:
            continue

        code_scale = data(index(row, 90))
        code_fab_scale = data(index(row, 91))

        code_float = data(index(row, 102))
        codefab_float = data(index(row, 103))

        all_list_parts = []

        model_num = data(index(row, 9))
        model_value = model_num[:6] if model_num[2:4] !='HH' else model_num[:7]
        level_type = data(index(row, 8))
        conn_type = data(index(row, 15))
        nipplehexdim = data(index(row, 32))[:8]
        nippletubedim = data(index(row, 33))[:8]
        cc_length = int(data(index(row, 17)))

        # setting list for eache element [code_element, code_fab_element, trad_element, design_elemente, process_element, material_element, qty_element, code_purch_element]
        code_body = data(index(row, 69))
        if code_body:
            body_list.append([
                code_body,
                data(index(row, 70)),
                data(index(row, 121)),
                nipplehexdim,
                '40x40' if model_value[2:3] != 'H' else ('100x50'if model_value[2:4] != 'HH' else '80x40'),
                (nipplehexdim + '-M'),
                'A-105' if data(index(row, 10)) == 'Carbon Steel' else data(index(row, 10)),
                data(index(row, 71)),
                data(index(row, 181))
                ])
            all_list_parts.append(body_list)

        code_cover = data(index(row, 72))
        if code_cover:
            commands_coverdim = ("""
                SELECT *
                FROM validation_data.level_cover_dim
                WHERE cover = %s
                """)

            try:
                cover_num = model_value[2:6] if model_value[2:4] !='HH' else model_value[3:7]
                cover_num = cover_num[:2] + '1' + cover_num[3:]
                with Database_Connection(config_database()) as conn:
                    with conn.cursor() as cur:
                        cur.execute(commands_coverdim,(cover_num,))
                        results=cur.fetchone()
                        length=results[1]
                        bores=results[2]

            except (Exception, psycopg2.DatabaseError) as error:
                MessageHelper.show_message("Ha ocurrido el siguiente error:\n"
                            + str(error), "critical")

            cover_list.append([
                code_cover,
                data(index(row, 73)),
                data(index(row, 122)),
                ('L=' + str(length)),
                '80x30' if model_value[2:4] != 'HH' else '90x40',
                (str(bores) + ' taladros'),
                'A-105' if data(index(row, 27)) == 'Carbon Steel' else data(index(row, 27)),
                data(index(row, 74)),
                data(index(row, 182))
                ])
            all_list_parts.append(cover_list)

        code_glass = data(index(row, 99))
        if code_glass:
            glass_list.append([
                code_glass,
                data(index(row, 100)),
                data(index(row, 131)),
                'TRANSPARENCIA' if level_type == 'Transparent' else 'REFLEXIÓN',
                '',
                '',
                'BOROSILICATO',
                model.data(model.index(row, 101)),
                model.data(model.index(row, 191))
                ])
            all_list_parts.append(glass_list)

        code_gasket = data(index(row, 96))
        if code_gasket:
            gasket_list.append([
                code_gasket,
                data(index(row, 97)),
                data(index(row, 130)),
                'TRANSPARENCIA' if level_type == 'Transparent' else 'REFLEXIÓN',
                '',
                '',
                'GRAFOIL',
                data(index(row, 98)),
                data(index(row, 190))
                ])
            all_list_parts.append(gasket_list)

        code_mica = data(index(row, 105))
        if code_mica:
            mica_list.append([
                code_mica,
                data(index(row, 106)),
                data(index(row, 133)),
                'TRANSPARENCIA',
                '',
                '',
                'MICA',
                data(index(row, 107)),
                data(index(row, 193))
                ])
            all_list_parts.append(mica_list)

        code_bolts = data(index(row, 75))
        if code_bolts:
            bolts_list.append([
                code_bolts,
                data(index(row, 76)),
                data(index(row, 123)),
                'TRANSPARENCIA' if level_type == 'Transparent' else 'REFLEXIÓN',
                '' if model_value[2:4] == 'HH' else ('M10x132 mm' if level_type == 'Transparent' else ''),
                '' if model_value[2:4] == 'HH' else ('cabeza exag 17 e/c' if level_type == 'Transparent' else ''),
                'B7/2H' if level_type in ['Transparent','Reflex'] else data(index(row, 24)),
                data(index(row, 77)),
                data(index(row, 183))
                ])
            all_list_parts.append(bolts_list)

        code_nipplehex = data(index(row, 78))
        if code_nipplehex:
            nipplehex_list.append([
                code_nipplehex,
                data(index(row, 79)),
                data(index(row, 124)),
                (str((cc_length-int(get_number_before_mm(data(index(row, 121))))-72)/2+22) + ' mm'),
                '',
                '',
                'A-105' if data(index(row, 10)) == 'Carbon Steel' else data(index(row, 10)), 
                data(index(row, 80)),
                data(index(row, 184))
                ])
            all_list_parts.append(nipplehex_list)

        code_valve = data(index(row, 81))
        if code_valve:
            valve_list.append([
                code_valve,
                data(index(row, 82)),
                data(index(row, 125)),
                nipplehexdim[:4] + ' x ' + data(index(row, 20)),
                nipplehexdim[-3:] + '-H',
                '',
                'A-105' if data(index(row, 18))[-2:] == 'NB' else '316 SS',
                data(index(row, 83)),
                data(index(row, 185))
                ])
            all_list_parts.append(valve_list)

        code_flangevalve = data(index(row, 84))
        if code_flangevalve:
            flangevalve_list.append([
                code_flangevalve,
                data(index(row, 85)),
                data(index(row, 126)),
                '',
                '',
                '',
                'A-105' if data(index(row, 10)) == 'Carbon Steel' else data(index(row, 10)),
                data(index(row, 86)),
                data(index(row, 186))
                ])
            all_list_parts.append(flangevalve_list)

        code_dv = data(index(row, 87))
        if code_dv:
            dv_list.append([
                code_dv,
                data(index(row, 88)),
                data(index(row, 127)),
                '',
                '',
                '',
                'A-105' if data(index(row, 10)) == 'Carbon Steel' else data(index(row, 10)),
                data(index(row, 89)),
                data(index(row, 187))
                ])
            all_list_parts.append(dv_list)

            if data(index(row, 127))[:3] == 'VÁL':
                plug_list.append([
                    'TAPÓN NORMAL ' + data(index(row, 20)) + data(index(row, 21)),
                    '',
                    '',
                    '',
                    'A-105' if data(index(row, 10)) == 'Carbon Steel' else data(index(row, 10)),
                    2,
                    'TA ' + data(index(row, 20)) + data(index(row, 21))
                    ])

        code_nippletube = data(index(row, 114))
        if code_nippletube:
            nippletube_list.append([
                code_nippletube,
                data(index(row, 115)),
                data(index(row, 136)),
                '80 mm',
                '',
                '',
                'A-106' if data(index(row, 10)) in ['Carbon Steel','ASTM A350 LF2 CL2'] else data(index(row, 10)),
                data(index(row, 116)),
                data(index(row, 196))
                ])
            all_list_parts.append(nippletube_list)

        code_illuminator = data(index(row, 93))
        if code_illuminator:
            illuminator_list.append([
                code_illuminator,
                data(index(row, 94)),
                data(index(row, 129)),
                '',
                '',
                '',
                'HIERRO',
                data(index(row, 95)),
                data(index(row, 189))
                ])
            all_list_parts.append(illuminator_list)

        code_antifrost = data(index(row, 117))
        if code_antifrost:
            antifrost_list.append([
                code_antifrost,
                data(index(row, 118)),
                data(index(row, 137)),
                '',
                '',
                '',
                'METACRILATO',
                data(index(row, 119)),
                data(index(row, 197))
                ])
            all_list_parts.append(antifrost_list)

        columns_equipments = ["code_equipment", "code_fab_equipment", "translate_equipment", "section_type",
                                        "l_body", "qty_l_body", "l_cover", "qty_l_cover",
                                        "l_studs", "qty_l_studs", "l_nipplehex", "qty_l_nipplehex",
                                        "l_valve", "qty_l_valve", "l_flange", "qty_l_flange",
                                        "l_dv", "qty_l_dv", "l_scale", "qty_l_scale",
                                        "l_illuminator", "qty_l_illuminator", "l_gasketglass", "qty_l_gasketglass",
                                        "l_glass", "qty_l_glass", "l_float", "qty_l_float",
                                        "l_mica", "qty_l_mica", "l_flags", "qty_l_flags",
                                        "l_gasketflange", "qty_l_gasketflange", "l_nippletub", "qty_l_nippletub",
                                        "l_antifrost", "qty_l_antifrost"]

        columns_parts = ["code_part", "code_fab_part", "code_element", "model", "design", "process", "material", "section_type"]

        columns_tags = ["code", "equipment", "num_order","order_material","contractual_date","inspection"]

        values_equipments = [data(index(row, 66)), data(index(row, 67)), data(index(row, 68)), "N-Niveles",
                            data(index(row, 69)), data(index(row, 71)), data(index(row, 72)), data(index(row, 74)),
                            data(index(row, 75)), data(index(row, 77)), data(index(row, 78)), data(index(row, 80)),
                            data(index(row, 81)), data(index(row, 83)), data(index(row, 84)), data(index(row, 86)),
                            data(index(row, 87)), data(index(row, 89)), data(index(row, 90)), data(index(row, 92)),
                            data(index(row, 93)), data(index(row, 95)), data(index(row, 96)), data(index(row, 98)),
                            data(index(row, 99)), data(index(row, 101)), data(index(row, 102)), data(index(row, 104)),
                            data(index(row, 105)), data(index(row, 107)), data(index(row, 108)), data(index(row, 110)),
                            data(index(row, 111)), data(index(row, 113)), data(index(row, 114)), data(index(row, 116)),
                            data(index(row, 117)), data(index(row, 119))]

        values_tags = [data(index(row, 4)) + "-" + data(index(row, 8)) + "-" + data(index(row, 1)), 
                        data(index(row, 66)), data(index(row, 4)), data(index(row, 48)),
                        data(index(row, 39)), data(index(row, 62))]

        columns_equipments  = ", ".join([f'"{column}"' for column in columns_equipments])
        values_equipments =  ", ".join(['NULL' if value == '' or value == 0 else (str(value) if isinstance(value, (int, float)) else f"'{str(value)}'") for value in values_equipments])

        columns_tags  = ", ".join([f'"{column}"' for column in columns_tags])
        values_tags =  ", ".join(['NULL' if value == '' or value == PySide6.QtCore.QDate() else (str(value) if isinstance(value, (int, float)) else (f"'{value.toString('yyyy-MM-dd')}'" if isinstance(value, PySide6.QtCore.QDate) else f"'{str(value)}'")) for value in values_tags])

        columns_parts = ", ".join([f'"{column}"' for column in columns_parts])

        commands_equipments = f"INSERT INTO fabrication.equipments ({columns_equipments}) VALUES ({values_equipments})"
        commands_tags = f"INSERT INTO fabrication.tags ({columns_tags}) VALUES ({values_tags})"

        check_equipments = f"SELECT * FROM fabrication.equipments WHERE code_equipment = '{data(index(row, 66))}'"

        if state == 'Order':
            try:
                with Database_Connection(config_database()) as conn:
                    with conn.cursor() as cur:
                        cur.execute(check_equipments)
                        results=cur.fetchall()

                if len(results) == 0:
                    with Database_Connection(config_database()) as conn:
                        with conn.cursor() as cur:
                            cur.execute(commands_equipments)
                        conn.commit()

                else:
                    set_clause = ", ".join([f"{column} = {value}" for column, value in zip(columns_equipments.split(", ")[1:], values_equipments.split(", ")[1:])])
                    update_equipments = f"UPDATE fabrication.equipments SET {set_clause} WHERE code_equipment = '{data(index(row, 66))}'"
                    with Database_Connection(config_database()) as conn:
                        with conn.cursor() as cur:
                            cur.execute(update_equipments)
                        conn.cursor()

                for list_part in all_list_parts:
                    check_parts = f"SELECT * FROM fabrication.parts WHERE code_part = '{list_part[0][0]}'"
                    with Database_Connection(config_database()) as conn:
                        with conn.cursor() as cur:
                            cur.execute(check_parts)
                            results=cur.fetchall()

                    if len(results) == 0:
                        list_part_modified = list_part[0].copy()
                        list_part_modified[-1] = 'N-Niveles'
                        values_parts = ", ".join('NULL' if value == '' else (str(value) if isinstance(value, (int, float)) else f"'{str(value)}'") for value in list_part_modified)
                        commands_parts = f"INSERT INTO fabrication.parts ({columns_parts}) VALUES ({values_parts})"
                        with Database_Connection(config_database()) as conn:
                            with conn.cursor() as cur:
                                cur.execute(commands_parts)
                            conn.commit()

                    else:
                        list_part_modified = list_part[0].copy()
                        list_part_modified[-1] = 'N-Niveles'
                        values_parts = ", ".join('NULL' if value == '' else (str(value) if isinstance(value, (int, float)) else f"'{str(value)}'") for value in list_part_modified)
                        set_clause = ", ".join([f"{column} = {value}" for column, value in zip(columns_parts.split(", ")[1:], values_parts.split(", ")[1:])])
                        update_parts = f"UPDATE fabrication.parts SET {set_clause} WHERE code_part = '{list_part[0][0]}'"
                        with Database_Connection(config_database()) as conn:
                            with conn.cursor() as cur:
                                cur.execute(update_parts)
                            conn.commit()

                with Database_Connection(config_database()) as conn:
                    with conn.cursor() as cur:
                        cur.execute(commands_tags)
                    conn.commit()

            except (Exception, psycopg2.DatabaseError) as error:
                MessageHelper.show_message("Ha ocurrido el siguiente error:\n"
                            + str(error), "critical")

# Turn all lists in dataframe and grouped in order to sum same items
    data_lists = [
    (body_list, "df_body"),
    (cover_list, "df_cover"),
    (glass_list, "df_glass"),
    (gasket_list, "df_gasket"),
    (mica_list, "df_mica"),
    (bolts_list, "df_bolts"),
    (nipplehex_list, "df_nipplehex"),
    (valve_list, "df_valve"),
    (flangevalve_list, "df_flangevalve"),
    (nippletube_list, "df_nippletube"),
    (dv_list, "df_list"),
    (antifrost_list, "df_antifrost"),
    (expand_illuminators_from_list(illuminator_list), "df_illuminator"),
    (plug_list, "df_plug")]

    data_frames_with_data = []

    for data_list, df_name in data_lists:
        if data_list:
            sublists = [sublist[2:] for sublist in data_list]
            df = pd.DataFrame(sublists, columns=['descripción', 'modelo', 'diseño', 'proceso', 'material', 'cantidad', 'suministro'])
            df = df.groupby(['descripción', 'modelo', 'diseño', 'proceso', 'material', 'suministro'])['cantidad'].sum().reset_index()
            data_frames_with_data.append(df)

    if data_frames_with_data:
        df_final = pd.concat(data_frames_with_data, ignore_index=True)

        values_supplies = df_final['suministro'].dropna().unique().tolist()

        query = """
        SELECT reference AS suministro, physical_stock AS st_fisico, available_stock AS st_disponible, pending_stock as st_pend, virtual_stock AS st_virtual
        FROM purch_fact.supplies
        WHERE reference IN %(values)s
        """

        with Database_Connection(config_database()) as conn:
            df_supplies = pd.read_sql(query, config_sql_engine(), params={"values": tuple(values_supplies)})

        df_final = (df_final.merge(df_supplies, on='suministro', how='left')
                    [['descripción', 'modelo', 'diseño', 'proceso', 'material', 'cantidad', 'suministro',
                        'st_fisico', 'st_disponible', 'st_pend', 'st_virtual']])

        df_final['almacen_si'] = ''
        df_final['almacen_no'] = ''
        df_final['proveedor'] = ''
        df_final['fecha_pedido'] = ''
        df_final['fecha_prevista'] = ''

        df_final = df_final[
            ['descripción', 'modelo', 'diseño', 'proceso', 'material',
            'cantidad', 'almacen_si', 'almacen_no', 'suministro',
            'proveedor', 'fecha_pedido', 'fecha_prevista',
            'st_fisico', 'st_disponible', 'st_pend', 'st_virtual']
        ]

    commands_client_order = ("""
                SELECT orders."num_order",orders."num_offer",offers."client"
                FROM offers
                INNER JOIN orders ON (offers."num_offer"=orders."num_offer")
                WHERE UPPER(orders."num_order") LIKE UPPER('%%'||%s||'%%')
                """)

    commands_client_offer = ("""
                SELECT offers."num_offer",offers."client"
                FROM offers
                WHERE UPPER(offers."num_offer") LIKE UPPER('%%'||%s||'%%')
                """)

    try:
        with Database_Connection(config_database()) as conn:
            with conn.cursor() as cur:
                if state == 'Order':
                    cur.execute(commands_client_order,(numorder,))
                    results=cur.fetchone()
                    client=results[2]
                else:
                    cur.execute(commands_client_offer,(numorder,))
                    results=cur.fetchone()
                    client=results[1]

    except (Exception, psycopg2.DatabaseError) as error:
        MessageHelper.show_message("Ha ocurrido el siguiente error:\n"
                    + str(error), "critical")

    excel_mat_order = material_order(df_final, numorder_pedmat, client, variable, num_ot)
    excel_mat_order.save_excel()


def others_matorder(proxy, model, numorder, numorder_pedmat, variable, state):
    """
    Processes material raw orders for others items by inserting new entries into the fabrication orders database.

    Args:
        proxy (QAbstractProxyModel): The proxy model containing the current data view.
        model (QAbstractItemModel): The model containing the main data.
        numorder (str): The order number to process.
        numorder_pedmat (str): The base order number for material orders.
        variable (str): A variable that determines the type of processing to be done. The specific usage
                        of this variable is not detailed in the function.

    Returns:
        None: This function does not return a value but modifies the database state.
    """
    id_list = []

    list_valves_210 = ['V-9305','V-9575','V-9576','2V-210']

    for row in range(proxy.rowCount()):
        if proxy.data(proxy.index(row, 2)) not in ["DELETED", "FOR INVOICING"]:
            first_column_value = proxy.data(proxy.index(row, 0))
            description = proxy.data(proxy.index(row, 8))
            id_list.append(first_column_value)

    commands_numot = ("""SELECT "ot_num"
                        FROM fabrication.fab_order
                        WHERE NOT "ot_num" LIKE '90%'
                        ORDER BY "ot_num" ASC
                        """)

    check_otpedmat = f"SELECT * FROM fabrication.fab_order WHERE id = '{numorder_pedmat + '-PEDMAT'}'"

    commands_otpedmat = ("""
                            INSERT INTO fabrication.fab_order (
                            "id","tag","element","qty_element",
                            "ot_num","qty_ot","start_date")
                            VALUES (%s,%s,%s,%s,%s,%s,%s)
                            """)

    if state == 'Offer':
        num_ot = '0'
    else:
        try:
            with Database_Connection(config_database()) as conn:
                with conn.cursor() as cur:
                    cur.execute(commands_numot)
                    results=cur.fetchall()
                    num_ot=results[-1][0]

            excel_file_path = r"\\ERP-EIPSA-DATOS\Comunes\EIPSA Sistemas de Gestion\MasterCTF\Bases\Contador.xlsm"
            workbook = openpyxl.load_workbook(excel_file_path, keep_vba=True)
            worksheet = workbook.active
            num_ot = worksheet['B2'].value

            with Database_Connection(config_database()) as conn:
                with conn.cursor() as cur:
                    cur.execute(check_otpedmat)
                    results=cur.fetchall()

            if len(results) == 0:
                data_numot=(numorder_pedmat + '-PEDMAT', numorder_pedmat, 'PEDIDO DE MATERIALES', 1, '{:06}'.format(int(num_ot) + 1), len(id_list), date.today().strftime("%d/%m/%Y"))
                with Database_Connection(config_database()) as conn:
                    with conn.cursor() as cur:
                        cur.execute(commands_otpedmat, data_numot)
                    conn.commit()

                worksheet['B2'].value = '{:06}'.format(int(num_ot) + 1)
                workbook.save(excel_file_path)

        except (Exception, psycopg2.DatabaseError) as error:
            MessageHelper.show_message("Ha ocurrido el siguiente error:\n"
                        + str(error), "critical")

    if len(id_list) != 0:
        model_list = []
        valve_model_list = []
        list_1 = []
        list_2 = []
        list_3 = []
        list_4 = []
        list_5 = []
        list_6 = []
        list_7 = []
        list_8 = []
        list_9 = []
        list_10 = []
        list_11 = []
        list_12 = []
        list_13 = []
        list_14 = []
        list_15 = []
        list_16 = []
        list_17 = []
        list_18 = []

        for element in id_list:
            for row in range(model.rowCount()):
                if model.data(model.index(row, 0)) == element:
                    target_row = row
                    break
            if target_row is not None:
                description = model.data(model.index(target_row, 8))

            # Order for 2V-210 valves
                if any(valve in description for valve in list_valves_210):
                    model_valve = re.match(r'(V-\d+-[A-Za-z0-9]+)', description).group(0)
                    material_valve = (re.match(r'(V-\d+-[A-Za-z0-9]+)(.*)', description).group(2).lstrip(' - ').strip()).split(' / ')[0]
                    sch_valve = re.search(r'V-\d+-(\w+)', description).group(1)

                    tradcodvalve = 'VÁLVULA 2V-210 SCH ' + sch_valve + (' BRIDADA ' + description.split(' / ')[1].strip()) if '# RF' in description else ''
                    schvalve = 'MOD.: ' + model_valve
                    designvalve = ''
                    processvalve = ''
                    materialvalve = material_valve
                    qtyvalve = 1
                    valve_model_list.append([tradcodvalve,schvalve,designvalve,processvalve,materialvalve,qtyvalve])

                    list_1.append(['VOLANTE','VÁLVULA 2V-210 - 1500#','','',materialvalve,1])
                    list_2.append(['ARANDELA VÁLVULA','VÁLVULA 2V-210 - 1500#','','','AC. INOX',1])
                    list_3.append(['VÁSTAGO','VÁLVULA 2V-210 - 1500#','','','AISI-316 + STELLITE',1])
                    list_4.append(['GUÍA VÁSTAGO/TUERCA (ø25 x LONG 37 mm) (EXAG 22 ec/ x 6 mm)','VÁLVULA 2V-210 - 1500#','','','AC. INOX' if materialvalve == '316' else 'AC. CARBONO',1])
                    list_5.append(['CAPELLI (HORQUILLA)','VÁLVULA 2V-210 - 1500#','','',materialvalve,1])
                    list_6.append(['FLANGETE','VÁLVULA 2V-210 - 1500#','','',materialvalve,1])
                    list_7.append(['PRENSA (ø25 x LONG 20 mm)','VÁLVULA 2V-210 - 1500#','','','AC. INOX' if materialvalve == '316' else 'AC. CARBONO',1])
                    list_8.append(['EMPAQUETADURA','VÁLVULA 2V-210 - 1500#','','','GRAFITO',1])
                    list_9.append(['TORNILLO CUADRADO (2 ud. POR VÁLVULA)','VÁLVULA 2V-210 - 1500#','','','AC. INOX',2])
                    list_10.append(['TORNILLO REDONDO (4 ud. POR VÁLVULA)','VÁLVULA 2V-210 - 1500#','','','AC. INOX',4])
                    list_11.append(['TUERCAS M10 2H','VÁLVULA 2V-210 - 1500#','','','A1942H',4])
                    list_12.append(['JUNTA ESPIROMETÁLICA 42x30x3,2mm','VÁLVULA 2V-210 - 1500#','','','AISI-316 + GRAFITO',1])
                    list_13.append(['CUERPO VÁLVULA 2V-210 - 1500#','','','',materialvalve,1])
                    list_14.append(['ASIENTO (ø20 x 16 mm)','VÁLVULA 2V-210 - 1500#','','','AISI-316 + STELLITE',1])
                    list_15.append(['BRIDA VÁLVULA '+ description.split(' / ')[1].strip(),'VÁLVULA 2V-210 - 1500#','','',materialvalve,1]) if '# RF' in description else ''
                    list_16.append(['TAPÓN PURGADOR 1/2" NPT-M','','','',materialvalve,1])
                    list_17.append(['TORNILLO TAPÓN PURGADOR','','','','AC. INOX',1])
                    if len(description.split(' / ')) > 2: 
                        list_18.append(['NIPLO ' + description.split(' / ')[2],'','','','AC. INOX' if materialvalve == '316' else 'AC. CARBONO',1]) 

                    data_lists = [
                    (valve_model_list, "df_valvemodel"),
                    (list_1, "df_list1"),
                    (list_2, "df_list2"),
                    (list_3, "df_list3"),
                    (list_4, "df_list4"),
                    (list_5, "df_list5"),
                    (list_6, "df_list6"),
                    (list_7, "df_list7"),
                    (list_8, "df_list8"),
                    (list_9, "df_list9"),
                    (list_10, "df_list10"),
                    (list_11, "df_list11"),
                    (list_12, "df_list12"),
                    (list_13, "df_list13"),
                    (list_14, "df_list14"),
                    (list_15, "df_list15"),
                    (list_16, "df_list16"),
                    (list_17, "df_list17"),
                    (list_18, "df_list18")]

            # Order for CN-32219
                if 'CN-32219' in description:
                    list_1.append(['BRIDA BLIND 4" 900# RTJ', 'ø293 x 53 mm', '', '', '321', 1])
                    list_2.append(['TUBO 1/4" SCH 40S (ø13,5 x ESP 2,3 mm)','(2323x' + 1 +')+(1753x' + 1 + ')+(1183x' + 1 + ')', '', '', '321', str(round(5260 / 1000, 2))])
                    list_3.append(['TUBO 3" SCH 80S','2664 x ' * 1, '', '', '321', str(round(2680 / 1000, 2))])
                    list_4.append(['BARRA ø25 x LONG. 30 mm (1/4" NPT-H)','', '', '', '321', 1])
                    list_5.append(['BARRA ø25 x LONG. 40 mm (1/4" NPT-H x 1/4" SW)','REDUCCIÓN 1/4"SW A 1/4" NPT-H', '', '', '321', 3])
                    list_6.append(['ACCESORIO FIJACIÓN BARRA ø20 x LONG 34 mm', '', '', '', '321', 3])
                    list_7.append(['CAP SOLDADO BARRA ø90 x LONG 67 mm', '', '', '', '321', 1])
                    list_8.append(['EMPTAPÓN DE PURGA 1/4" NPT-M (EXAG. 17 e/c x LONG. 31 mm)AQUETADURA', 'EXAG. 17 e/c (ø20 mm)', '', '', '321', 1])

                    data_lists = [
                    (list_1, "df_list1"),
                    (list_2, "df_list2"),
                    (list_3, "df_list3"),
                    (list_4, "df_list4"),
                    (list_5, "df_list5"),
                    (list_6, "df_list6"),
                    (list_7, "df_list7"),
                    (list_8, "df_list8")]

                else:
                    tradcod = str(description)
                    sch = ''
                    design = ''
                    process = ''
                    material = ''
                    qty = 1
                    model_list.append([tradcod,sch,design,process,material,qty])

                    data_lists = [
                    (model_list, "df_model"),]

        data_frames_with_data = []

        for data_list, df_name in data_lists:
            if data_list:
                sublists = [sublist[2:] for sublist in data_list]
                df = pd.DataFrame(sublists, columns=['descripción', 'modelo', 'diseño', 'proceso', 'material', 'cantidad', 'suministro'])
                df = df.groupby(['descripción', 'modelo', 'diseño', 'proceso', 'material', 'suministro'])['cantidad'].sum().reset_index()
                data_frames_with_data.append(df)

        if data_frames_with_data:
            df_final = pd.concat(data_frames_with_data, ignore_index=True)

            values_supplies = df_final['suministro'].dropna().unique().tolist()

            query = """
            SELECT reference AS suministro, physical_stock AS st_fisico, available_stock AS st_disponible, pending_stock as st_pend, virtual_stock AS st_virtual
            FROM purch_fact.supplies
            WHERE reference IN %(values)s
            """

            with Database_Connection(config_database()) as conn:
                df_supplies = pd.read_sql(query, config_sql_engine(), params={"values": tuple(values_supplies)})

            df_final = (df_final.merge(df_supplies, on='suministro', how='left')
                        [['descripción', 'modelo', 'diseño', 'proceso', 'material', 'cantidad', 'suministro',
                            'st_fisico', 'st_disponible', 'st_pend', 'st_virtual']])

            df_final['almacen_si'] = ''
            df_final['almacen_no'] = ''
            df_final['proveedor'] = ''
            df_final['fecha_pedido'] = ''
            df_final['fecha_prevista'] = ''

            df_final = df_final[
                ['descripción', 'modelo', 'diseño', 'proceso', 'material',
                'cantidad', 'almacen_si', 'almacen_no', 'suministro',
                'proveedor', 'fecha_pedido', 'fecha_prevista',
                'st_fisico', 'st_disponible', 'st_pend', 'st_virtual']
            ]

    commands_client_order = ("""
                SELECT orders."num_order",orders."num_offer",offers."client"
                FROM offers
                INNER JOIN orders ON (offers."num_offer"=orders."num_offer")
                WHERE UPPER(orders."num_order") LIKE UPPER('%%'||%s||'%%')
                """)

    commands_client_offer = ("""
                SELECT offers."num_offer",offers."client"
                FROM offers
                WHERE UPPER(offers."num_offer") LIKE UPPER('%%'||%s||'%%')
                """)

    try:
        with Database_Connection(config_database()) as conn:
            with conn.cursor() as cur:
                if state == 'Order':
                    cur.execute(commands_client_order,(numorder,))
                    results=cur.fetchone()
                    client=results[2]
                else:
                    cur.execute(commands_client_offer,(numorder,))
                    results=cur.fetchone()
                    client=results[1]

    except (Exception, psycopg2.DatabaseError) as error:
        MessageHelper.show_message("Ha ocurrido el siguiente error:\n"
                    + str(error), "critical")

    excel_mat_order = material_order(df_final, numorder_pedmat, client, variable, num_ot)
    excel_mat_order.save_excel()


def material_list(proxy, model, variable, numoffer):
    id_list = []
    flow_list = []
    temp_list = []
    level_list = []

    query_offer_data = ("SELECT client, project FROM offers WHERE num_offer = %s")

    proxy_data = proxy.data
    proxy_index = proxy.index

    data = model.data
    index = model.index

    with Database_Connection(config_database()) as conn:
        with conn.cursor() as cur:
            cur.execute(query_offer_data,(numoffer,))
            results = cur.fetchall()
            client_value = results[0][0]
            project_value = results[0][1]

    id_list = [
        proxy_data(proxy_index(row, 0))
        for row in range(proxy.rowCount())
        if proxy_data(proxy_index(row, 2)) == "QUOTED" and str(proxy_data(proxy_index(row, 5))) == ''
    ]

    row_map = {
        data(index(row, 0)): row
        for row in range(model.rowCount())
    }

    for element in id_list:
        row = row_map.get(element)
        if row is None:
            continue

        if variable == 'Caudal':
            # appending [type_value, flange_material_value, element_material_value, size_value, rating_value, facing_value, schedule_value, qty_value]
            flow_list.append([
                data(index(row, 8)),
                data(index(row, 13)),
                data(index(row, 19)),
                data(index(row, 9)),
                int(data(index(row, 10))) if data(index(row, 10)) != 'N/A' else data(index(row, 10)),
                data(index(row, 11)),
                data(index(row, 12)),
                1 #* int(data(index(row, 34)))
                ])

        if variable == 'Temperatura':
            # [type_value, tw_type_value, tw_material_value, size_value, rating_value, facing_value, insertion_value, qty_value]
            temp_list.append([
                data(index(row, 8)),
                data(index(row, 9)),
                data(index(row, 14)),
                data(index(row, 10)),
                int(data(index(row, 11))) if data(index(row, 11)) != 'N/A' else data(index(row, 11)),
                data(index(row, 12)),
                data(index(row, 16)),
                1
                ])

        if variable == 'Nivel':
            type_value = data(index(row, 8))
            if type_value == 'Magnetic':
                # [type_value, body_material, conn_size, conn_rating, conn_facing, c-c_length, float_material, bolting_material, qty_value]
                level_list.append([
                    data(index(row, 8)),
                    data(index(row, 10)),
                    data(index(row, 12)),
                    int(data(index(row, 13))) if data(index(row, 13)) != 'N/A' else data(index(row, 13)),
                    data(index(row, 14)),
                    data(index(row, 17)),
                    data(index(row, 26)),
                    data(index(row, 24)),
                    1
                    ])
            else:
                # [type_value, body_material, conn_size, conn_rating, conn_facing, c-c_length, cover_material, bolting_material, qty_value]
                level_list.append([
                    data(index(row, 8)),
                    data(index(row, 10)),
                    data(index(row, 12)),
                    int(data(index(row, 13))) if data(index(row, 13)) != 'N/A' else data(index(row, 13)),
                    data(index(row, 14)),
                    data(index(row, 27)),
                    data(index(row, 26)),
                    data(index(row, 24)),
                    1
                    ])

    dfs = []
    if flow_list:
        cols = ['TIPO', 'MAT. BRIDA', 'MAT. ELEMENTO', 'TAMAÑO', 'RATING', 'FACING', 'SCHEDULE', 'Nº EQUIPOS']
        df = pd.DataFrame(flow_list, columns=cols)
        df = df.groupby(cols[:-1], as_index=False)['Nº EQUIPOS'].sum()
        df["tam_num"] = df["TAMAÑO"].apply(size_to_float)
        df_sorted = df.sort_values(["TIPO", "MAT. BRIDA", "tam_num"]).drop(columns="tam_num")
        dfs.append(df_sorted)

    if temp_list:
        cols = ['TIPO', 'TIPO TW', 'MAT. TW', 'TAMAÑO', 'RATING', 'FACING', 'INSERCIÓN', 'Nº EQUIPOS']
        df = pd.DataFrame(temp_list, columns=cols)
        df = df.groupby(cols[:-1], as_index=False)['Nº EQUIPOS'].sum()
        df["tam_num"] = df["TAMAÑO"].apply(size_to_float)
        df_sorted = df.sort_values(["TIPO", "TIPO TW", "MAT. TW", "tam_num"]).drop(columns="tam_num")
        dfs.append(df_sorted)

    if level_list:
        cols = ['TIPO', 'MAT. CUERPO', 'TAMAÑO', 'RATING', 'FACING', 'LONG. C-C', 'MAT. CUBIERTA / FLOTADOR', 'MAT. TORN', 'Nº EQUIPOS']
        df = pd.DataFrame(level_list, columns=cols)
        df = df.groupby(cols[:-1], as_index=False)['Nº EQUIPOS'].sum()
        df["tam_num"] = df["TAMAÑO"].apply(size_to_float)
        df_sorted = df.sort_values(["TIPO", "MAT. CUERPO", "MAT. CUBIERTA / FLOTADOR", "LONG. C-C", "tam_num"]).drop(columns="tam_num")
        dfs.append(df_sorted)

    df_final = pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame()

    path, _ = QFileDialog.getSaveFileName(
    None,
    "Guardar lista de materiales oferta",
    "",
    "Excel Files (*.xlsx)"
    )
    if path:
        with pd.ExcelWriter(path, engine="xlsxwriter") as writer:
            sheet = "Materiales"
            bold = writer.book.add_format({'bold': True})

            worksheet = writer.book.add_worksheet(sheet)
            writer.sheets[sheet] = worksheet

            worksheet.write(0, 0, numoffer, bold)
            worksheet.write(0, 1, client_value, bold)
            worksheet.write(0, 2, project_value, bold)
            worksheet.write(0, 3, variable, bold)
            worksheet.write(0, 6, date.today().strftime("%d/%m/%Y"), bold)

            row = 3
            for type_equipment, df_type in df_final.groupby("TIPO"):
                df_type.to_excel(writer, sheet_name=sheet, startrow=row, index=False)

                total = df_type["Nº EQUIPOS"].sum() #calculate total quantity for type dataframe

                row += len(df_type) + 1  # +1 header

                # write TOTAL and total value
                worksheet.write(row, len(df_type.columns)-2, "TOTAL", bold)
                worksheet.write(row, len(df_type.columns)-1, total, bold)

                row += 2  # total + blank row
        MessageHelper.show_message("Guardado con éxito", "info")


def get_number_before_mm(text):
    """
    Extracts the integer value that appears before the 'mm' substring in the given text.

    Args:
        text (str): The input string to search for the number.

    Returns:
        int or None: The extracted integer value if found, or None if no match is found.
    """
    match = re.search(r'(\d+)\s*mm', text)
    if match:
        return int(match.group(1))
    return None


def size_to_float(size):
    size = size.replace('"', '')

    if 'N/A' in size:
        return 'N/A'
    
    if '-' in size:              # ejemplo 1-1/2
        whole, frac = size.split('-')
        return float(whole) + float(Fraction(frac))

    if '/' in size:              # ejemplo 1/2
        return float(Fraction(size))

    return float(size) 


def flow_material_list(proxy, model):
    """
    Processes material raw orders for flow items by inserting new entries into the fabrication orders database.

    Args:
        proxy (QAbstractProxyModel): The proxy model containing the current data view.
        model (QAbstractItemModel): The model containing the main data.

    Returns:
        df_final: This function returns a dataframe with the resume of materials.
    """

    parts = []

    proxy_data = proxy.data
    proxy_index = proxy.index

    data = model.data
    index = model.index

    id_list = [
        proxy_data(proxy_index(row, 0))
        for row in range(proxy.rowCount())
        if proxy_data(proxy_index(row, 2)) == "QUOTED"
        and str(proxy_data(proxy_index(row, 5))) == ''
    ]

    row_map = {
        data(index(row, 0)): row
        for row in range(model.rowCount())
    }

    flange_code_cache = {}
    sheet_material_cache = {}
    pipe_thk_cache = {}

    for element in id_list:
        row = row_map.get(element)
        if row is None:
            continue

        flange_material = data(index(row, 13))
        sch = data(index(row, 12))
        design = str(data(index(row, 61))).replace('.', ',') # pipe internal diameter
        size = f"{data(index(row,9))} {data(index(row,10))} {data(index(row,11))}"

        code_orifice_flange = data(index(row, 162))
        if code_orifice_flange:
            parts.append([
                data(index(row, 198)), # trad orifice flange
                sch,
                design,
                "", #data(index(row, 37)),
                flange_material,
                int(data(index(row, 186))) * int(data(index(row, 35))), # quantity orifice flange per equipment * number of equipments
                data(index(row, 210)) # code purch orifice flange
            ])

        code_line_flange = data(index(row,163))
        if code_line_flange:
            parts.append([
                data(index(row, 199)), # trad line flange
                sch,
                design,
                "", #data(index(row, 37)),
                flange_material,
                int(data(index(row, 187))) * int(data(index(row, 35))), # quantity line flange per equipment * number of equipments
                data(index(row, 211)) # code purch line flange
            ])

        code_gasket = data(index(row,164))
        if code_gasket:
            parts.append([
                data(index(row, 200)), # trad gasket
                size,
                '',
                '',
                '',
                int(data(index(row, 42))) * int(data(index(row, 35))), # quantity gasket per equipment * number of equipments
                data(index(row, 212)) # code purch gasket
            ]) 

        code_bolts = data(index(row,165))
        if code_bolts:
            qty = int(data(index(row,44))) if data(index(row,44)) != '' else 0
            parts.append([
                data(index(row, 201)), # trad bolts
                size,
                ('esp. placa ' + data(index(row, 21))),
                '',
                data(index(row, 24)) + " / " + data(index(row, 25)),
                qty * int(data(index(row, 35))), # quantity bolts per equipment * number of equipments
                data(index(row, 213)) # code purch bolts
            ])

        code_plugs = data(index(row, 166))
        if code_plugs != '':
            parts.append([
                data(index(row, 202)), # trad plug
                '',
                '',
                '',
                data(index(row, 45)), # material plug
                (int(data(index(row, 46))) if data(index(row, 46)) != '' else 0) * int(data(index(row, 35))), # quantity plugs per equipment * quantity of equipment
                data(index(row, 214))
                ])

        code_extractor = data(index(row,167))
        if code_extractor:
            parts.append([
                data(index(row, 203)), # trad extractor
                size,
                ('esp. placa ' + data(index(row, 21))),
                '',
                data(index(row, 47)),
                int(data(index(row, 49))) * int(data(index(row, 35))), # quantity extractor per equipment * number of equipments
                data(index(row, 215)) # code purch extractor
            ])

        code_plate = data(index(row,168))
        if code_plate:
            qty = int(float(data(index(row,28)))) if data(index(row,8)) == "MULTISTAGE RO" else 1
            process = 'ARAMCO' if data(index(row,22)) == 'ARA' else ''

            parts.append([
                data(index(row, 204)), # trad plate
                ('ESP ' + data(index(row, 21)) + 'mm'),
                data(index(row, 62)),
                process,
                data(index(row, 19)),
                qty * int(data(index(row, 35))), # quantity of plates per equipment * number of equipments
                data(index(row, 216)) # code purch plate
            ])

        code_handle = data(index(row,170))
        if code_handle and data(index(row,21)) not in ['3', '1/8" (3)']:
            if data(index(row,11)) == 'RTJ':
                modelhandle = ''
                designhandle = ''

                parts.append([
                    'BARRA MANGO', # trad handle
                    '',
                    '',
                    '',
                    '316SS',
                    ((int(float(data(index(row, 64)))) - 30) if 'datos' not in data(index(row, 64)) else 0) * int(data(index(row, 35))), # length of bar handle per equipment * quantity of equipments
                    '' # code purch handle
                    ])
            else:
                modelhandle = f"{data(index(row,64))}x{data(index(row,65))}x{data(index(row,66))} mm"
                designhandle = data(index(row,22))

            parts.append([
                data(index(row,206)), # trad handle
                modelhandle,
                designhandle,
                '',
                '316SS',
                1 * int(data(index(row, 35))), # quantity of handles per equipment * number of equipments
                data(index(row,218)) # code purch handle
            ])

        code_ch_ring = data(index(row,171))
        if code_ch_ring:
            schchring = 'ESP ' if data(index(row,11)) == "RTJ" else 'ESP 38,5mm ACABADO'

            parts.append([
                data(index(row, 207)), # trad chring
                schchring,
                'ø' + str(data(index(row, 62))),
                '', #data(index(row, 37)),
                data(index(row, 19)),
                1 * int(data(index(row, 35))), # quantity chring per equipment * quantity of equipments
                data(index(row, 219)) # code purch chring
            ])

        code_tube = data(index(row,172))
        if code_tube:
            parts.append([
                data(index(row, 208)), # trad tube
                sch,
                design,
                '',
                data(index(row,15)),
                float(data(index(row, 196))) * int(data(index(row, 35))), # quantity tube per equipment (length of tube) * quantity of equipments
                data(index(row, 220)) # code purch tube
            ])

        code_piece2 = data(index(row,173))
        if code_piece2:
            line_size = data(index(row,9))
            sch = data(index(row,12))
            flange_material = data(index(row,13))

            thk_key = (line_size, sch)

            if thk_key not in pipe_thk_cache:
                try:
                    with Database_Connection(config_database()) as conn:
                        with conn.cursor() as cur:
                            cur.execute("""
                                SELECT wall_thk
                                FROM validation_data.pipe_diam
                                WHERE line_size = %s
                                AND sch = %s
                            """,(line_size, sch))

                            pipe_thk_cache[thk_key] = cur.fetchone()[0]

                except (Exception, psycopg2.DatabaseError) as error:
                    MessageHelper.show_message(
                        "Ha ocurrido el siguiente error:\n" + str(error),
                        "critical"
                    )
                    pipe_thk_cache[thk_key] = ''

            thkmin = pipe_thk_cache[thk_key]

            modelpiece2 = 'Th mín ' + str(thkmin) + 'mm'

            if flange_material not in flange_code_cache:
                try:
                    with Database_Connection(config_database()) as conn:
                        with conn.cursor() as cur:
                            cur.execute("""
                                SELECT code
                                FROM validation_data.flow_flange_material
                                WHERE flange_material = %s
                            """,(flange_material,))

                            flange_code_cache[flange_material] = cur.fetchone()[0]

                except (Exception, psycopg2.DatabaseError) as error:
                    MessageHelper.show_message(
                        "Ha ocurrido el siguiente error:\n" + str(error),
                        "critical"
                    )
                    flange_code_cache[flange_material] = None

            flange_code = flange_code_cache[flange_material]

            if flange_code not in sheet_material_cache:
                try:
                    with Database_Connection(config_database()) as conn:
                        with conn.cursor() as cur:
                            cur.execute("""
                                SELECT sheet_material
                                FROM validation_data.flow_sheet_material
                                WHERE code = %s
                            """,(flange_code,))

                            sheet_material_cache[flange_code] = cur.fetchone()[0]

                except (Exception, psycopg2.DatabaseError) as error:
                    MessageHelper.show_message(
                        "Ha ocurrido el siguiente error:\n" + str(error),
                        "critical"
                    )
                    sheet_material_cache[flange_code] = ''

            materialpiece2 = sheet_material_cache[flange_code]

            parts.append([
                data(index(row, 209)), # trad wedge
                modelpiece2,
                '',
                '',
                materialpiece2,
                1 * int(data(index(row, 35))), # quantity of wedge parts per equipment * quantity of equipments
                data(index(row, 221)) # code purch wedge
            ])

    df = pd.DataFrame(
        parts,
        columns=[
            'descripción','modelo','diseño','proceso',
            'material','cantidad','suministro'
        ]
    )

    df = (
        df.groupby(
            ['descripción','modelo','diseño','proceso','material','suministro']
        )['cantidad']
        .sum()
        .reset_index()
    )

    return df


def temp_material_list(proxy, model):
    """
    Processes material raw orders for temp items by inserting new entries into the fabrication orders database.

    Args:
        proxy (QAbstractProxyModel): The proxy model containing the current data view.
        model (QAbstractItemModel): The model containing the main data.

    Returns:
        df_final: This function returns a dataframe with the resume of materials.
    """
    parts = []

    proxy_data = proxy.data
    proxy_index = proxy.index

    data = model.data
    index = model.index

    id_list = [
        proxy_data(proxy_index(row, 0))
        for row in range(proxy.rowCount())
        if proxy_data(proxy_index(row, 2)) == "QUOTED"
        and str(proxy_data(proxy_index(row, 5))) == ''
    ]

    row_map = {
        data(index(row, 0)): row
        for row in range(model.rowCount())
    }

    for element in id_list:
        row = row_map.get(element)
        if row is None:
            continue

        tw_type = data(index(row, 9))

        code_bar = data(index(row, 142))
        if code_bar :
            parts.append([
                data(index(row, 178)) if 'Helical' not in tw_type else 'VAINA HELICOIDAL' + (' BRIDADA ' + data(index(row, 10)) + ' ' + data(index(row, 11)) + ' ' + data(index(row, 12)) if tw_type == 'Flanged Helical' else ''),
                'U=' + data(index(row, 16)) + ' /L=' + data(index(row, 15)) if 'Stone' in tw_type or 'Helical' in tw_type else 'Barra ø=' + (data(index(row, 50))),
                'RAÍZ ø=' + data(index(row, 17)) if tw_type == 'Van-Stone TW' else '',
                '',
                data(index(row, 14)),
                data(index(row, 166)) if 'Helical' not in tw_type else 1,
                data(index(row, 190))
                ])

        code_tube = data(index(row, 143))
        if code_tube:
            parts.append([
                data(index(row, 179)),
                data(index(row, 38)),
                '',
                '',
                data(index(row, 14)),
                data(index(row, 167)),
                data(index(row, 191))
                ])

        code_flange = data(index(row, 144))
        if code_flange:
            list_tw = ['Buttweld TW','Forged Flanged TW','Threaded Helical','Van-Stone Helical','VORTICRACK']
            parts.append([
                data(index(row, 180)) if tw_type not in list_tw else '',
                '',
                '',
                '',
                data(index(row, 35)) if tw_type == 'Van-Stone TW'
                    else (data(index(row, 14)) if tw_type not in list_tw else ''),
                1 if tw_type not in list_tw else '',
                data(index(row, 192))
                ])

        code_sensor = data(index(row, 145))
        if code_sensor:
            parts.append([
                data(index(row, 181)),
                data(index(row, 33)) + '-' + data(index(row, 32)) if code_sensor[:4] == 'Bime' else '',
                data(index(row, 27)) + '-' + data(index(row, 28)) if code_sensor[:4] == 'Bime' else '',
                'PLATINO' if data(index(row, 181))[:5] == 'PT100' else
                    ('AC. INOX.' if data(index(row, 24)) == 'St.Steel' else
                    data(index(row, 24))),
                1 if data(index(row, 181))[:5] == 'PT100' or code_sensor[:4] == 'Bime' else
                    (float(data(index(row, 56)))/1000) if data(index(row, 56)) != '' else
                    '',
                data(index(row, 193))
                ])

        code_head = data(index(row, 146))
        if code_head:
            parts.append([
                data(index(row, 182)),
                data(index(row, 31)),
                '',
                data(index(row, 33)),
                ('ALUMINIO' if data(index(row, 31))[-2:] == 'AL' 
                            else ('AC.CARBONO' if data(index(row, 31))[-2:] == 'CS' 
                            else ('AC.INOXIDABLE' if data(index(row, 31))[-2:] == 'SS' 
                            else 'MATERIAL CABEZA NO DEFINIDO'))),
                1,
                data(index(row, 194))
                ])

        code_btb = data(index(row, 147))
        if code_btb:
            parts.append([
                data(index(row, 183)),
                ("RANGO " + data(index(row, 27)) + '-' + data(index(row, 28))) if code_btb[:2] == 'BI' else '',
                '',
                '',
                data(index(row, 24)) if code_btb[:2] == 'BI' else ('CERÁMICO' if code_btb[:2] == 'CE' else ''),
                data(index(row, 171)),
                data(index(row, 195))
                ])

        code_nipple = data(index(row, 148))
        if code_nipple:
            trad = data(index(row, 184))
            model = data(index(row, 30))
            parts.append([
                trad,
                '' if model == 'N/A' or model =='' else model,
                '',
                '',
                'A-105/A106' if trad[trad.find('('):trad.find('(')+9] == '(CS)' else 'AISI-316',
                1,
                data(index(row, 196))
                ])

        code_spring = data(index(row, 149))
        if code_spring:
            parts.append([
                data(index(row, 185)),
                '',
                '',
                '',
                'AC.INOX',
                1,
                data(index(row, 197))
                ])

        code_plug = data(index(row, 151))
        if code_plug:
            trad = data(index(row, 187))
            parts.append([
                trad,
                '',
                '',
                '',
                trad[trad.find('('):trad.find('(')+9],
                1,
                data(index(row, 199))
                ])

        code_puntal = data(index(row, 150))
        if code_puntal:
            parts.append([
                data(index(row, 186)),
                '',
                '',
                '',
                data(index(row, 14)),
                float(code_puntal[1:8])/1000 if code_puntal not in ['N/A', 'HO'] else 0,
                data(index(row, 198))
                ])

        code_tw = data(index(row, 152))
        if code_tw and ('Van-Stone TW' in tw_type or 'Forged' in tw_type):
            parts.append([
                data(index(row, 188)),
                'U=' + data(index(row, 16)) + ' / L=' + data(index(row, 15)),
                '',
                '',
                data(index(row, 14)),
                data(index(row, 176)),
                data(index(row, 200))
                ])

        code_extcable = data(index(row, 153))
        if code_extcable:
            parts.append([
                data(index(row, 189)),
                '',
                '',
                '',
                'AC. INOX.' if data(index(row, 24)) in ['AISI-304', 'AISI-310', 'AISI-316', 'AISI-321', 'St.Steel'] else data(index(row, 24)),
                float(data(index(row, 177))) if data(index(row, 177)) != '' else 0,
                data(index(row, 201))
                ])

    df = pd.DataFrame(
        parts,
        columns=[
            'descripción','modelo','diseño','proceso',
            'material','cantidad','suministro'
        ]
    )

    df = (
        df.groupby(
            ['descripción','modelo','diseño','proceso','material','suministro']
        )['cantidad']
        .sum()
        .reset_index()
    )

    return df


def level_material_list(proxy, model):
    """
    Processes material raw orders for level items by inserting new entries into the fabrication orders database.

    Args:
        proxy (QAbstractProxyModel): The proxy model containing the current data view.
        model (QAbstractItemModel): The model containing the main data.

    Returns:
        df_final: This function returns a dataframe with the resume of materials.
    """

    def expand_illuminators_from_list(illuminator_list):
        """
        Expand combinations like '1 x 1-I 218-T + 2 x 1-I 228-T'
        mantaining fields and original structure.
        """
        rows = []
        for item in illuminator_list:
            tradcod, model, design, process, material, qty = item

            parts = [p.strip() for p in str(tradcod).split("+")]
            for p in parts:
                match = re.match(r"(\d+)\s*x\s*(1-I\s*\d+-T)", p)
                if match:
                    quantity = int(match.group(1)) * qty
                    illuminator = "ILUMINADOR " + match.group(2).replace(" ", "")
                    rows.append([
                        illuminator, model, design, process, material, quantity
                    ])
        return rows

    parts = []

    proxy_data = proxy.data
    proxy_index = proxy.index

    data = model.data
    index = model.index

    id_list = [
        proxy_data(proxy_index(row, 0))
        for row in range(proxy.rowCount())
        if proxy_data(proxy_index(row, 2)) == "QUOTED"
        and str(proxy_data(proxy_index(row, 5))) == ''
    ]

    row_map = {
        data(index(row, 0)): row
        for row in range(model.rowCount())
    }

    for element in id_list:
        row = row_map.get(element)
        if row is None:
            continue

        model_num = data(index(row,9))
        model_value = model_num[:6] if model_num[2:4] !='HH' else model_num[:7]
        level_type = data(index(row, 8))
        conn_type = data(index(row, 15))
        nipplehexdim = data(index(row, 32))[:8]
        nippletubedim = data(index(row, 33))[:8]
        cc_length = int(data(index(row, 17)))

        code_body = data(index(row, 69))
        if code_body:
            if level_type in ['Transparent', 'Reflex']:
                parts.append([
                    data(index(row, 121)),
                    nipplehexdim,
                    '40x40' if model_value[2:3] != 'H' else ('100x50'if model_value[2:4] != 'HH' else '80x40'),
                    (nipplehexdim + '-M'),
                    'A-105' if data(index(row, 10)) == 'Carbon Steel' else data(index(row, 10)),
                    data(index(row, 71)),
                    data(index(row, 181))
                    ])
            else:
                parts.append([
                    data(index(row, 121)),
                    '',
                    '',
                    '',
                    'A-105' if data(index(row, 10)) == 'Carbon Steel' else data(index(row, 10)),
                    data(index(row, 71)),
                    data(index(row, 181))
                    ])

        code_cover = data(index(row, 72))
        if code_cover:
            commands_coverdim = ("""
                SELECT *
                FROM validation_data.level_cover_dim
                WHERE cover = %s
                """)

            try:
                cover_num = model_value[2:6] if model_value[2:4] !='HH' else model_value[3:7]
                cover_num = cover_num[:2] + '1' + cover_num[3:]
                with Database_Connection(config_database()) as conn:
                    with conn.cursor() as cur:
                        cur.execute(commands_coverdim,(cover_num,))
                        results=cur.fetchone()
                        length=results[1]
                        bores=results[2]

            except (Exception, psycopg2.DatabaseError) as error:
                MessageHelper.show_message("Ha ocurrido el siguiente error:\n"
                            + str(error), "critical")

            parts.append([
                data(index(row, 122)),
                ('L=' + str(length)),
                '80x30' if model_value[2:4] != 'HH' else '90x40',
                (str(bores) + ' taladros'),
                'A-105' if data(index(row, 27)) == 'Carbon Steel' else data(index(row, 27)),
                data(index(row, 74)),
                data(index(row, 182))
                ])

        code_bolts = data(index(row, 75))
        if code_bolts :
            parts.append([
                data(index(row, 123)),
                'TRANSPARENCIA' if level_type == 'Transparent' else 'REFLEXIÓN',
                '' if model_value[2:4] == 'HH' else ('M10x132 mm' if level_type == 'Transparent' else ''),
                '' if model_value[2:4] == 'HH' else ('cabeza exag 17 e/c' if level_type == 'Transparent' else ''),
                'B7/2H' if level_type in ['Transparent','Reflex'] else data(index(row, 24)),
                data(index(row, 77)),
                data(index(row, 183))
                ])

        code_nipplehex = data(index(row, 78))
        if code_nipplehex:
            parts.append([
                data(index(row, 124)),
                (str((cc_length-int(get_number_before_mm(data(index(row, 121))))-72)/2+22) + ' mm'),
                '',
                '',
                'A-105' if data(index(row, 10)) == 'Carbon Steel' else data(index(row, 10)),
                data(index(row, 80)),
                data(index(row, 184))
                ])

        code_valve = data(index(row, 81))
        if code_valve:
            parts.append([
                data(index(row, 125)),
                nipplehexdim[:4] + ' x ' + data(index(row, 20)),
                nipplehexdim[-3:] + '-H',
                '',
                'A-105' if data(index(row, 18))[-2:] == 'NB' else '316 SS',
                data(index(row, 83)),
                data(index(row, 185))
                ])

        code_flangevalve = data(index(row, 84))
        if code_flangevalve:
            parts.append([
                data(index(row, 126)),
                '',
                '',
                '',
                'A-105' if data(index(row, 10)) == 'Carbon Steel' else data(index(row, 10)),
                data(index(row, 86)),
                data(index(row, 186))
                ])

        code_dv = data(index(row, 87))
        if code_dv:
            parts.append([
                data(index(row, 127)),
                '',
                '',
                '',
                'A-105' if data(index(row, 10)) == 'Carbon Steel' else data(index(row, 10)),
                data(index(row, 89)),
                data(index(row, 187))
                ])

            if data(index(row, 127))[:3] == 'VÁL':
                parts.append([
                    'TAPÓN NORMAL ' + data(index(row, 20)) + data(index(row, 21)),
                    '',
                    '',
                    '',
                    'A-105' if data(index(row, 10)) == 'Carbon Steel' else data(index(row, 10)),
                    2,
                    'TA ' + data(index(row, 20)) + data(index(row, 21))
                    ])

        code_gasket = data(index(row, 96))
        if code_gasket:
            parts.append([
                data(index(row, 130)),
                'TRANSPARENCIA' if level_type == 'Transparent' else 'REFLEXIÓN',
                '',
                '',
                'GRAFOIL',
                data(index(row, 98)),
                data(index(row, 190))
                ])

        code_glass = data(index(row, 99))
        if code_glass:
            parts.append([
                data(index(row, 131)),
                'TRANSPARENCIA' if level_type == 'Transparent' else 'REFLEXIÓN',
                '',
                '',
                'BOROSILICATO',
                data(index(row, 101)),
                data(index(row, 191))
                ])

        code_mica = data(index(row, 105))
        if code_mica:
            parts.append([
                data(index(row, 133)),
                'TRANSPARENCIA',
                '',
                '',
                'MICA',
                data(index(row, 107)),
                data(index(row, 193))
                ])

        code_nippletube = data(index(row, 114))
        if code_nippletube:
            parts.append([
                data(index(row, 136)),
                '80 mm',
                '',
                '',
                'A-106' if data(index(row, 10)) in ['Carbon Steel','ASTM A350 LF2 CL2'] else data(index(row, 10)),
                data(index(row, 116)),
                data(index(row, 196))
                ])

        code_antifrost = data(index(row, 117))
        if code_antifrost:
            parts.append([
                data(index(row, 137)),
                '',
                '',
                '',
                'METACRILATO',
                data(index(row, 119)),
                data(index(row, 197))
                ])

        code_illuminator = data(index(row, 93))
        if code_illuminator:
            item = [
                data(index(row, 129)),
                '',
                '',
                '',
                'HIERRO',
                data(index(row, 95)),
                data(index(row, 189))
                ]

            for r in expand_illuminators_from_list(item):
                parts.append(r)

        code_scale = data(index(row, 90))
        code_float = data(index(row, 102))

    df = pd.DataFrame(
        parts,
        columns=[
            'descripción','modelo','diseño','proceso',
            'material','cantidad','suministro'
        ]
    )

    df = (
        df.groupby(
            ['descripción','modelo','diseño','proceso','material','suministro']
        )['cantidad']
        .sum()
        .reset_index()
    )

    return df


def others_material_list(proxy, model):
    """
    Processes material raw orders for others items by inserting new entries into the fabrication orders database.

    Args:
        proxy (QAbstractProxyModel): The proxy model containing the current data view.
        model (QAbstractItemModel): The model containing the main data.
        numorder (str): The order number to process.
        numorder_pedmat (str): The base order number for material orders.
        variable (str): A variable that determines the type of processing to be done. The specific usage
                        of this variable is not detailed in the function.

    Returns:
        None: This function does not return a value but modifies the database state.
    """
    parts = []

    list_valves_210 = ['V-9305','V-9575','V-9576','2V-210']

    proxy_data = proxy.data
    proxy_index = proxy.index

    data = model.data
    index = model.index

    id_list = [
        proxy_data(proxy_index(row, 0))
        for row in range(proxy.rowCount())
        if proxy_data(proxy_index(row, 2)) == "QUOTED"
        and str(proxy_data(proxy_index(row, 5))) == ''
    ]

    row_map = {
        data(index(row, 0)): row
        for row in range(model.rowCount())
    }

    for element in id_list:
        row = row_map.get(element)
        if row is None:
            continue

        parts = []
        description = data(index(row, 8))

    # Order for 2V-210 valves
        if any(valve in description for valve in list_valves_210):
            model_valve = re.match(r'(V-\d+-[A-Za-z0-9]+)', description).group(0)
            material_valve = (re.match(r'(V-\d+-[A-Za-z0-9]+)(.*)', description).group(2).lstrip(' - ').strip()).split(' / ')[0]
            sch_valve = re.search(r'V-\d+-(\w+)', description).group(1)

            parts.append([
                'VÁLVULA 2V-210 SCH ' + sch_valve + (' BRIDADA ' + description.split(' / ')[1].strip()) if '# RF' in description else '',
                'MOD.: ' + model_valve,
                '',
                '',
                material_valve,
                1,
                ''])

            parts.append(['VOLANTE','VÁLVULA 2V-210 - 1500#','','',material_valve,1, ''])
            parts.append(['ARANDELA VÁLVULA','VÁLVULA 2V-210 - 1500#','','','AC. INOX',1, ''])
            parts.append(['VÁSTAGO','VÁLVULA 2V-210 - 1500#','','','AISI-316 + STELLITE',1, ''])
            parts.append(['GUÍA VÁSTAGO/TUERCA (ø25 x LONG 37 mm) (EXAG 22 ec/ x 6 mm)','VÁLVULA 2V-210 - 1500#','','','AC. INOX' if material_valve == '316' else 'AC. CARBONO',1, ''])
            parts.append(['CAPELLI (HORQUILLA)','VÁLVULA 2V-210 - 1500#','','',material_valve,1, ''])
            parts.append(['FLANGETE','VÁLVULA 2V-210 - 1500#','','',material_valve,1, ''])
            parts.append(['PRENSA (ø25 x LONG 20 mm)','VÁLVULA 2V-210 - 1500#','','','AC. INOX' if material_valve == '316' else 'AC. CARBONO',1, ''])
            parts.append(['EMPAQUETADURA','VÁLVULA 2V-210 - 1500#','','','GRAFITO',1, ''])
            parts.append(['TORNILLO CUADRADO (2 ud. POR VÁLVULA)','VÁLVULA 2V-210 - 1500#','','','AC. INOX',2, ''])
            parts.append(['TORNILLO REDONDO (4 ud. POR VÁLVULA)','VÁLVULA 2V-210 - 1500#','','','AC. INOX',4, ''])
            parts.append(['TUERCAS M10 2H','VÁLVULA 2V-210 - 1500#','','','A1942H',4, ''])
            parts.append(['JUNTA ESPIROMETÁLICA 42x30x3,2mm','VÁLVULA 2V-210 - 1500#','','','AISI-316 + GRAFITO',1, ''])
            parts.append(['CUERPO VÁLVULA 2V-210 - 1500#','','','',material_valve,1, ''])
            parts.append(['ASIENTO (ø20 x 16 mm)','VÁLVULA 2V-210 - 1500#','','','AISI-316 + STELLITE',1, ''])
            parts.append(['BRIDA VÁLVULA '+ description.split(' / ')[1].strip(),'VÁLVULA 2V-210 - 1500#','','',material_valve,1, '']) if '# RF' in description else ''
            parts.append(['TAPÓN PURGADOR 1/2" NPT-M','','','',material_valve,1, ''])
            parts.append(['TORNILLO TAPÓN PURGADOR','','','','AC. INOX',1, ''])
            if len(description.split(' / ')) > 2: 
                parts.append(['NIPLO ' + description.split(' / ')[2],'','','','AC. INOX' if material_valve == '316' else 'AC. CARBONO',1, '']) 

    # Order for CN-32219
        elif 'CN-32219' in description:
            parts.append(['BRIDA BLIND 4" 900# RTJ', 'ø293 x 53 mm', '', '', '321', 1, ''])
            parts.append(['TUBO 1/4" SCH 40S (ø13,5 x ESP 2,3 mm)','(2323x' + 1 +')+(1753x' + 1 + ')+(1183x' + 1 + ')', '', '', '321', str(round(5260 / 1000, 2)), ''])
            parts.append(['TUBO 3" SCH 80S','2664 x ' * 1, '', '', '321', str(round(2680 / 1000, 2)), ''])
            parts.append(['BARRA ø25 x LONG. 30 mm (1/4" NPT-H)','', '', '', '321', 1, ''])
            parts.append(['BARRA ø25 x LONG. 40 mm (1/4" NPT-H x 1/4" SW)','REDUCCIÓN 1/4"SW A 1/4" NPT-H', '', '', '321', 3, ''])
            parts.append(['ACCESORIO FIJACIÓN BARRA ø20 x LONG 34 mm', '', '', '', '321', 3, ''])
            parts.append(['CAP SOLDADO BARRA ø90 x LONG 67 mm', '', '', '', '321', 1, ''])
            parts.append(['EMPTAPÓN DE PURGA 1/4" NPT-M (EXAG. 17 e/c x LONG. 31 mm)AQUETADURA', 'EXAG. 17 e/c (ø20 mm)', '', '', '321', 1, ''])

        else:
            parts.append([
                str(description),
                '',
                '',
                '',
                '',
                1,
                ''
                ])

    df = pd.DataFrame(
        parts,
        columns=[
            'descripción','modelo','diseño','proceso',
            'material','cantidad','suministro'
        ]
    )

    df = (
        df.groupby(
            ['descripción','modelo','diseño','proceso','material','suministro']
        )['cantidad']
        .sum()
        .reset_index()
    )

    return df


def general_material_list(proxy, model, variable, numoffer):
    id_list = []
    flow_list = []
    temp_list = []
    level_list = []

    query_offer_data = ("SELECT client, project FROM offers WHERE num_offer = %s")

    proxy_data = proxy.data
    proxy_index = proxy.index

    data = model.data
    index = model.index

    with Database_Connection(config_database()) as conn:
        with conn.cursor() as cur:
            cur.execute(query_offer_data,(numoffer,))
            results = cur.fetchall()
            client_value = results[0][0]
            project_value = results[0][1]

    id_list = [
        proxy_data(proxy_index(row, 0))
        for row in range(proxy.rowCount())
        if proxy_data(proxy_index(row, 2)) == "QUOTED" and str(proxy_data(proxy_index(row, 5))) == ''
    ]

    row_map = {
        data(index(row, 0)): row
        for row in range(model.rowCount())
    }

    for element in id_list:
        row = row_map.get(element)
        if row is None:
            continue

        if variable == 'Caudal':
            # appending [type_value, flange_material_value, element_material_value, size_value, rating_value, facing_value, schedule_value, qty_value]
            flow_list.append([
                data(index(row, 8)),
                data(index(row, 13)),
                data(index(row, 19)),
                data(index(row, 9)),
                int(data(index(row, 10))) if data(index(row, 10)) != 'N/A' else data(index(row, 10)),
                data(index(row, 11)),
                data(index(row, 12)),
                1 #* int(data(index(row, 34)))
                ])

        if variable == 'Temperatura':
            # [type_value, tw_type_value, tw_material_value, size_value, rating_value, facing_value, insertion_value, qty_value]
            temp_list.append([
                data(index(row, 8)),
                data(index(row, 9)),
                data(index(row, 14)),
                data(index(row, 10)),
                data(index(row, 11)) if data(index(row, 11)) != 'N/A' else data(index(row, 11)),
                data(index(row, 12)),
                data(index(row, 16)),
                1
                ])

        if variable == 'Nivel':
            type_value = data(index(row, 8))
            if type_value == 'Magnetic':
                # [type_value, body_material, conn_size, conn_rating, conn_facing, c-c_length, float_material, bolting_material, qty_value]
                level_list.append([
                    data(index(row, 8)),
                    data(index(row, 10)),
                    data(index(row, 12)),
                    int(data(index(row, 13))) if data(index(row, 13)) != 'N/A' else data(index(row, 13)),
                    data(index(row, 14)),
                    data(index(row, 17)),
                    data(index(row, 26)),
                    data(index(row, 24)),
                    1
                    ])
            else:
                # [type_value, body_material, conn_size, conn_rating, conn_facing, c-c_length, cover_material, bolting_material, qty_value]
                level_list.append([
                    data(index(row, 8)),
                    data(index(row, 10)),
                    data(index(row, 12)),
                    int(data(index(row, 13))) if data(index(row, 13)) != 'N/A' else data(index(row, 13)),
                    data(index(row, 14)),
                    data(index(row, 27)),
                    data(index(row, 26)),
                    data(index(row, 24)),
                    1
                    ])

        if variable == 'Others':
            df = None

    if flow_list:
        cols = ['TIPO', 'MAT. BRIDA', 'MAT. ELEMENTO', 'TAMAÑO', 'RATING', 'FACING', 'SCHEDULE', 'Nº EQUIPOS']
        df = pd.DataFrame(flow_list, columns=cols)

    if temp_list:
        cols = ['TIPO', 'TIPO TW', 'MAT. TW', 'TAMAÑO', 'RATING', 'FACING', 'INSERCIÓN', 'Nº EQUIPOS']
        df = pd.DataFrame(temp_list, columns=cols)

    if level_list:
        cols = ['TIPO', 'MAT. CUERPO', 'TAMAÑO', 'RATING', 'FACING', 'LONG. C-C', 'MAT. CUBIERTA / FLOTADOR', 'MAT. TORN', 'Nº EQUIPOS']
        df = pd.DataFrame(level_list, columns=cols)

    return df
