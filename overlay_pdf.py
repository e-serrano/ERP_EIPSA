import os
from fpdf import FPDF
import io
import pandas as pd
from config import config, get_path
import psycopg2
from PySide6 import QtCore, QtGui, QtWidgets
import openpyxl
from datetime import date
from utils.Database_Manager import Database_Connection
from utils.Show_Message import MessageHelper
from pypdf import PdfReader, PdfWriter

class PDF(FPDF):
    def rotate(self, angle, x=None, y=None):
        if x is None:
            x = self.x
        if y is None:
            y = self.y
        angle = -angle
        self._out(f"q {angle:.2f} 0 0 {angle:.2f} {x * self.k:.2f} {-y * self.k:.2f} cm")

    def text_rotated(self, x, y, txt, angle):
        self._out("q")
        self.rotate(angle, x, y)
        self.set_xy(x, y)
        self.cell(20, 20, txt)
        self._out("Q") 

    def fixed_height_multicell(self, w, total_h, txt, align_mc, fill=False):
        """
        Creates a multi-line cell with a fixed total height, dividing text into lines.

        Parameters:
            w (float): The width of the cell.
            total_h (float): The total height of the cell.
            txt (str): The text to be placed in the cell.
            align_mc (str): The alignment of the text.
            border (str, optional): Border settings for the cell. Defaults to ''.
            fill (bool, optional): Whether to fill the cell with color. Defaults to False.
        """
        words = txt.split() # Divide text in words
        lines = []
        line = ''
        for word in words:
            if self.get_string_width(line + word + ' ') > w - 0.5:
                lines.append(line) # Add line to line list and starts a new one
                line = word + ' '
            else:
                line += word + ' ' # Add word to actual line
        lines.append(line) # Add last line to line list
        
        line_height = total_h / len(lines) # Calculate height of each line to get a total height = total_h

        x, y = self.get_x(), self.get_y() # Save actual position

        for line in lines:
            # Print each line with the calculated height
            self.multi_cell(w, line_height, line, align=align_mc, fill=fill)
            self.set_x(x)

        self.set_xy(x, y + total_h)


# Function to create PDF with specific note in a position
def new_content_notes(technical_note):
    """
    Generates a PDF containing a new content based on the specified value and equipment type.

    Args:
        technical_note (str): The content to be added to the PDF.

    Returns:
        io.BytesIO: A byte stream containing the generated PDF.
    """
    pdf = FPDF(unit='mm')
    pdf.add_font('COURIERTXT', '', str(get_path("Resources", "Iconos", "COURIERTXT.ttf")))
    pdf.set_font("courier", "", 10)
    pdf.set_text_color(0, 0, 0)

    pdf.add_page()
    pdf.set_xy(20,230)
    pdf.set_font("courier", "B", 10)
    pdf.cell(150, 5, "NOTES:")
    pdf.set_xy(20,235)
    pdf.set_font("courier", "", 10)
    pdf.multi_cell(150, 5, str(technical_note)) #x_position, y_position, technical_note)

    return io.BytesIO(pdf.output())


# Function to create PDF with specific text in a position
def new_content_tags(value, type_eq):
    """
    Generates a PDF containing a new content based on the specified value and equipment type.

    Args:
        value (str): The content to be added to the PDF.
        type_eq (str): The type of equipment, determining the positioning in the PDF.

    Returns:
        io.BytesIO: A byte stream containing the generated PDF.
    """
    pdf = FPDF(unit='mm')
    pdf.set_font("helvetica", "", 10)
    pdf.set_text_color(0, 0, 0)

    excel_file = r"\\erp-eipsa-datos\Comunes\EIPSA-ERP\Plantillas Importación\Importar Tags Cálculos.xlsx"
    df_data = pd.read_excel(excel_file, sheet_name='Posiciones')
    df_data = df_data.set_index('type')

    x_position = df_data['x(mm)'][type_eq]
    y_position = df_data['y(mm)'][type_eq]

    if type_eq == 'MUL':
        pdf.add_page()
        pdf.set_xy(x_position, y_position)
        with pdf.rotation(90):
            pdf.cell(10, 10, value)

    else:
        pdf.add_page()
        pdf.text(x_position, y_position, value)

    return io.BytesIO(pdf.output())


def general_dwg(num_ot, material=None):
    """
    Generates a PDF containing a new content based on the specified value and equipment type.

    Args:
        material (str): The material code.
    """
    pdf = FPDF(unit='mm')
    pdf.set_font("helvetica", "B", 12)
    pdf.set_text_color(49, 49, 229)

    pdf.add_page()

    pdf.image(str(get_path("Resources", "Iconos", "QualityStamp.png")), x = 3, y = 10, w = 19, h = 11)

    pdf.set_draw_color(255, 0, 0)

    if material is None:
        material = 'A105'

    query = ('''
        SELECT colors.bg_color_1, colors.bg_color_2, colors.border_color
        FROM validation_data.material_color_code AS colors
        WHERE UPPER (colors.material) LIKE UPPER('%%'||%s||'%%')
        ''')

    try:
        with Database_Connection(config()) as conn:
            with conn.cursor() as cur:
                cur.execute(query,(material,))
                results_colors=cur.fetchall()

    except (Exception, psycopg2.DatabaseError) as error:
        MessageHelper.show_message("Ha ocurrido el siguiente error:\n"
                    + str(error), "critical")

    first_color = results_colors[0][0]
    second_color = results_colors[0][1]
    border_color = results_colors[0][2]

    pdf.set_line_width(1)
    pdf.set_draw_color(*map(int, first_color.split(',')))
    pdf.rect(20, 5, 185, 286, style='D')

    pdf.set_draw_color(*map(int, second_color.split(',')))
    pdf.rect(19, 4, 187, 288, style='D')

    if border_color is not None:
        pdf.set_draw_color(*map(int, border_color.split(',')))
        pdf.rect(18, 3, 189, 290, style='D')

    pdf.add_font('IDAutomationHC39M', '', str(get_path("Resources", "Iconos", "IDAutomationHC39M_Free.ttf")))
    pdf.set_font("helvetica", "B", 12)
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("IDAutomationHC39M", size=16)
    pdf.set_x(-9)
    pdf.set_y(160)
    with pdf.rotation(90):
        pdf.cell(60, 10, "*" + num_ot + "*", align='C')

    return io.BytesIO(pdf.output())


def general_dwg_landscape(num_ot, material=None):
    """
    Generates a PDF containing a new content based on the specified value and equipment type."
    """
    pdf = FPDF(unit='mm')
    
    pdf.set_font("helvetica", "B", 12)
    pdf.set_text_color(49, 49, 229)

    x_center = pdf.w / 2
    y_center = pdf.h / 2

    query = ('''
        SELECT colors.bg_color_1, colors.bg_color_2, colors.border_color
        FROM validation_data.material_color_code AS colors
        WHERE UPPER (colors.material) LIKE UPPER('%%'||%s||'%%')
        ''')

    try:
        with Database_Connection(config()) as conn:
            with conn.cursor() as cur:
                cur.execute(query,(material,))
                results_colors=cur.fetchall()

    except (Exception, psycopg2.DatabaseError) as error:
        MessageHelper.show_message("Ha ocurrido el siguiente error:\n"
                    + str(error), "critical")

    first_color = results_colors[0][0]
    second_color = results_colors[0][1]
    border_color = results_colors[0][2]

    pdf.add_page()
    with pdf.rotation(angle=-90, x=52, y=1):
        pdf.image(str(get_path("Resources", "Iconos", "QualityStamp.png")), x = 52, y = 1, w = 13, h = 10)

    pdf.set_line_width(1)
    pdf.set_draw_color(*map(int, first_color.split(',')))
    pdf.rect(3, 3, 203, 291, style='D')

    pdf.set_draw_color(*map(int, second_color.split(',')))
    pdf.rect(2, 2, 205, 293, style='D')

    if border_color is not None:
        pdf.set_draw_color(*map(int, border_color.split(',')))
        pdf.rect(1, 1, 207, 295, style='D')

    pdf.set_xy(80, 10)

    pdf.set_text_color(0, 0, 0)
    pdf.add_font('IDAutomationHC39M', '', str(get_path("Resources", "Iconos", "IDAutomationHC39M_Free.ttf")))
    pdf.set_font("IDAutomationHC39M", size=16)
    # with pdf.rotation(180, x_center, y_center):
    pdf.cell(10, 10, "*" + num_ot + "*", align='C')

    pdf.set_font("helvetica", "B", 12)

    return io.BytesIO(pdf.output())


def drawing_number(num_order, info_drawing, counter):
    """
    Generates a PDF containing a new content based on the specified value and equipment type."
    """
    pdf = FPDF(unit='mm')
    pdf.add_font('IDAutomationHC39M', '', str(get_path("Resources", "Iconos", "IDAutomationHC39M_Free.ttf")))
    pdf.set_font("helvetica", "B", 12)
    pdf.set_text_color(49, 49, 229)

    pdf.add_page()

    pdf.set_xy(163, 262)

    pdf.cell(37, 7, f"{str(info_drawing[0].split('.')[0])}/{counter:02d}", align='C')

    order_id = f"{num_order} - {info_drawing[0].split('.')[0]}/{counter:02d} - {info_drawing[1]}"

    check_ot = f"SELECT * FROM fabrication.fab_order WHERE id = '{order_id}'"

    insert_ot = ("""INSERT INTO fabrication.fab_order (
                    "id","tag","element","qty_element",
                    "ot_num","qty_ot","start_date")
                    VALUES (%s,%s,%s,%s,%s,%s,%s)
                    """)

    try:
        excel_file_path = r"\\erp-eipsa-datos\Comunes\EIPSA Sistemas de Gestion\MasterCTF\Bases\Contador.xlsm"
        workbook = openpyxl.load_workbook(excel_file_path, keep_vba=True)
        worksheet = workbook.active
        num_ot = worksheet['B2'].value

        with Database_Connection(config()) as conn:
            with conn.cursor() as cur:
                cur.execute(check_ot)
                results=cur.fetchall()

        if len(results) == 0:
            with Database_Connection(config()) as conn:
                with conn.cursor() as cur:
                    data=(order_id, num_order, info_drawing[1], 1, '{:06}'.format(int(num_ot) + 1), int(info_drawing[2]), date.today().strftime("%d/%m/%Y"))
                    cur.execute(insert_ot, data)
                conn.commit()

            worksheet['B2'].value = '{:06}'.format(int(num_ot) + 1)
            workbook.save(excel_file_path)
            workbook.close()

            num_ot_text = '{:06}'.format(int(num_ot) + 1)

        else:
            num_ot = '{:06}'.format(int(results[0][4]))
            num_ot_text = '{:06}'.format(int(num_ot))

    except (Exception, psycopg2.DatabaseError) as error:
        MessageHelper.show_message("Ha ocurrido el siguiente error:\n"
                    + str(error), "critical")

    pdf.set_text_color(0, 0, 0)
    pdf.set_font("IDAutomationHC39M", size=16)
    pdf.set_x(-9)
    pdf.set_y(160)
    with pdf.rotation(90):
        pdf.cell(60, 10, "*" + num_ot_text + "*", align='C')
    
    pdf.set_font("helvetica", "B", 12)

    return io.BytesIO(pdf.output())

def drawing_number_x(num_order, info_drawing, counter):
    """
    Generates a PDF containing a new content based on the specified value and equipment type."
    """
    pdf = FPDF(unit='mm')
    pdf.add_font('IDAutomationHC39M', '', str(get_path("Resources", "Iconos", "IDAutomationHC39M_Free.ttf")))
    pdf.set_font("helvetica", "B", 12)
    pdf.set_text_color(49, 49, 229)

    pdf.add_page()

    pdf.set_xy(163, 262)

    pdf.cell(37, 7, f"{str(info_drawing[0].split('.')[0])}/{counter:02d}", align='C')

    # order_id = f"{num_order} - {info_drawing[0].split('.')[0]}/{counter:02d} - {info_drawing[1]}"

    # check_ot = f"SELECT * FROM fabrication.fab_order WHERE id = '{order_id}'"

    # insert_ot = ("""INSERT INTO fabrication.fab_order (
    #                 "id","tag","element","qty_element",
    #                 "ot_num","qty_ot","start_date")
    #                 VALUES (%s,%s,%s,%s,%s,%s,%s)
    #                 """)

    # try:
    #     excel_file_path = r"\\erp-eipsa-datos\Comunes\EIPSA Sistemas de Gestion\MasterCTF\Bases\Contador.xlsm"
    #     workbook = openpyxl.load_workbook(excel_file_path, keep_vba=True)
    #     worksheet = workbook.active
    #     num_ot = worksheet['B2'].value

    #     with Database_Connection(config()) as conn:
    #         with conn.cursor() as cur:
    #             cur.execute(check_ot)
    #             results=cur.fetchall()

    #     if len(results) == 0:
    #         with Database_Connection(config()) as conn:
    #             with conn.cursor() as cur:
    #                 data=(order_id, num_order, info_drawing[1], 1, '{:06}'.format(int(num_ot) + 1), int(info_drawing[2]), date.today().strftime("%d/%m/%Y"))
    #                 cur.execute(insert_ot, data)
    #             conn.commit()

    #         worksheet['B2'].value = '{:06}'.format(int(num_ot) + 1)
    #         workbook.save(excel_file_path)
    #         workbook.close()

    #         num_ot_text = '{:06}'.format(int(num_ot) + 1)

    #     else:
    #         num_ot = '{:06}'.format(int(results[0][4]))
    #         num_ot_text = '{:06}'.format(int(num_ot))

    # except (Exception, psycopg2.DatabaseError) as error:
    #     MessageHelper.show_message("Ha ocurrido el siguiente error:\n"
    #                 + str(error), "critical")

    pdf.set_text_color(0, 0, 0)
    pdf.set_font("IDAutomationHC39M", size=16)
    pdf.set_x(-9)
    pdf.set_y(160)
    # with pdf.rotation(90):
    #     pdf.cell(60, 10, "*" + num_ot_text + "*", align='C')
    
    pdf.set_font("helvetica", "B", 12)

    return io.BytesIO(pdf.output())


def drawing_number_landscape(num_order, info_drawing, counter):
    """
    Generates a PDF containing a new content based on the specified value and equipment type."
    """
    pdf = FPDF(unit='mm')
    pdf.add_font('IDAutomationHC39M', '', str(get_path("Resources", "Iconos", "IDAutomationHC39M_Free.ttf")))
    pdf.set_font("helvetica", "B", 12)
    pdf.set_text_color(49, 49, 229)

    pdf.add_page()

    pdf.set_xy(18, 263)
    with pdf.rotation(270):
        pdf.cell(23, 4, f"{str(info_drawing[0].split('.')[0])}/{counter:02d}", align='C')

    order_id = f"{num_order} - {info_drawing[0].split('.')[0]}/{counter:02d} - {info_drawing[1]}"

    check_ot = f"SELECT * FROM fabrication.fab_order WHERE id = '{order_id}'"

    insert_ot = ("""INSERT INTO fabrication.fab_order (
                    "id","tag","element","qty_element",
                    "ot_num","qty_ot","start_date")
                    VALUES (%s,%s,%s,%s,%s,%s,%s)
                    """)

    try:
        excel_file_path = r"\\erp-eipsa-datos\Comunes\EIPSA Sistemas de Gestion\MasterCTF\Bases\Contador.xlsm"
        workbook = openpyxl.load_workbook(excel_file_path, keep_vba=True)
        worksheet = workbook.active
        num_ot = worksheet['B2'].value

        with Database_Connection(config()) as conn:
            with conn.cursor() as cur:
                cur.execute(check_ot)
                results=cur.fetchall()

        if len(results) == 0:
            with Database_Connection(config()) as conn:
                with conn.cursor() as cur:
                    data=(order_id, num_order, info_drawing[1], 1, '{:06}'.format(int(num_ot) + 1), int(info_drawing[2]), date.today().strftime("%d/%m/%Y"))
                    cur.execute(insert_ot, data)
                conn.commit()

            worksheet['B2'].value = '{:06}'.format(int(num_ot) + 1)
            workbook.save(excel_file_path)

            num_ot_text = '{:06}'.format(int(num_ot) + 1)

        else:
            num_ot = '{:06}'.format(int(results[0][4]))
            num_ot_text = '{:06}'.format(int(num_ot))

    except (Exception, psycopg2.DatabaseError) as error:
        MessageHelper.show_message("Ha ocurrido el siguiente error:\n"
                    + str(error), "critical")

    pdf.set_text_color(0, 0, 0)
    pdf.set_font("IDAutomationHC39M", size=16)
    pdf.set_x(160)
    pdf.set_y(8)
    # with pdf.rotation(90):
    pdf.cell(60, 10, "*" + num_ot_text + "*", align='C')

    pdf.set_font("helvetica", "B", 12)

    return io.BytesIO(pdf.output())

def drawing_number_landscape_x(num_order, info_drawing, counter):
    """
    Generates a PDF containing a new content based on the specified value and equipment type."
    """
    pdf = FPDF(unit='mm')
    pdf.add_font('IDAutomationHC39M', '', str(get_path("Resources", "Iconos", "IDAutomationHC39M_Free.ttf")))
    pdf.set_font("helvetica", "B", 12)
    pdf.set_text_color(49, 49, 229)

    pdf.add_page()

    pdf.set_xy(18, 263)
    with pdf.rotation(270):
        pdf.cell(23, 4, f"{str(info_drawing[0].split('.')[0])}/{counter:02d}", align='C')

    # order_id = f"{num_order} - {info_drawing[0].split('.')[0]}/{counter:02d} - {info_drawing[1]}"

    # check_ot = f"SELECT * FROM fabrication.fab_order WHERE id = '{order_id}'"

    # insert_ot = ("""INSERT INTO fabrication.fab_order (
    #                 "id","tag","element","qty_element",
    #                 "ot_num","qty_ot","start_date")
    #                 VALUES (%s,%s,%s,%s,%s,%s,%s)
    #                 """)

    # try:
    #     excel_file_path = r"\\erp-eipsa-datos\Comunes\EIPSA Sistemas de Gestion\MasterCTF\Bases\Contador.xlsm"
    #     workbook = openpyxl.load_workbook(excel_file_path, keep_vba=True)
    #     worksheet = workbook.active
    #     num_ot = worksheet['B2'].value

    #     with Database_Connection(config()) as conn:
    #         with conn.cursor() as cur:
    #             cur.execute(check_ot)
    #             results=cur.fetchall()

    #     if len(results) == 0:
    #         with Database_Connection(config()) as conn:
    #             with conn.cursor() as cur:
    #                 data=(order_id, num_order, info_drawing[1], 1, '{:06}'.format(int(num_ot) + 1), int(info_drawing[2]), date.today().strftime("%d/%m/%Y"))
    #                 cur.execute(insert_ot, data)
    #             conn.commit()

    #         worksheet['B2'].value = '{:06}'.format(int(num_ot) + 1)
    #         workbook.save(excel_file_path)

    #         num_ot_text = '{:06}'.format(int(num_ot) + 1)

    #     else:
    #         num_ot = '{:06}'.format(int(results[0][4]))
    #         num_ot_text = '{:06}'.format(int(num_ot))

    # except (Exception, psycopg2.DatabaseError) as error:
    #     MessageHelper.show_message("Ha ocurrido el siguiente error:\n"
    #                 + str(error), "critical")

    pdf.set_text_color(0, 0, 0)
    pdf.set_font("IDAutomationHC39M", size=16)
    pdf.set_x(160)
    pdf.set_y(8)
    # with pdf.rotation(90):
    # pdf.cell(60, 10, "*" + num_ot_text + "*", align='C')

    pdf.set_font("helvetica", "B", 12)

    return io.BytesIO(pdf.output())


def general_dwg_m(num_order, item_data, material=None):
    """
    Generates a PDF containing a new content based on the specified value and equipment type.

    Args:
        num_order (str): The order number.
        material (str): The material code.
        item_data (list): The list of items to be included in the PDF.
    """
    pdf = FPDF(unit='mm')
    pdf.set_font("helvetica", "B", 12)
    pdf.set_text_color(49, 49, 229)

    pdf.add_page()

    item_data = list(item_data)
    cnt = item_data[0][0]

    pdf.set_draw_color(255, 0, 0)

    pdf.set_xy(27, 248)
    pdf.cell(20, 10, str(cnt), align='C')

    if material is not None:
        pdf.set_xy(47, 248)
        pdf.set_font("helvetica", "B", 8)
        pdf.cell(36, 9, str(material), align='C')

    if material is None:
        material = 'A105'

    query = ('''
        SELECT colors.bg_color_1, colors.bg_color_2, colors.border_color
        FROM validation_data.material_color_code AS colors
        WHERE UPPER (colors.material) LIKE UPPER('%%'||%s||'%%')
        ''')

    try:
        with Database_Connection(config()) as conn:
            with conn.cursor() as cur:
                cur.execute(query,(material,))
                results_colors=cur.fetchall()

    except (Exception, psycopg2.DatabaseError) as error:
        MessageHelper.show_message("Ha ocurrido el siguiente error:\n"
                    + str(error), "critical")

    first_color = results_colors[0][0]
    second_color = results_colors[0][1]
    border_color = results_colors[0][2]

    pdf.set_line_width(1)
    pdf.set_draw_color(*map(int, first_color.split(',')))
    pdf.rect(20, 8, 183, 280, style='D')

    pdf.set_draw_color(*map(int, second_color.split(',')))
    pdf.rect(19, 7, 185, 282, style='D')

    if border_color is not None:
        pdf.set_draw_color(*map(int, border_color.split(',')))
        pdf.rect(18, 6, 197, 284, style='D')

    pdf.set_xy(151, 248)
    pdf.set_font("helvetica", "B", 12)
    pdf.cell(49, 9, str(num_order), align='C')

    return io.BytesIO(pdf.output())


def flange_dwg_flangedTW(num_order, material, count):
    """
    Generates a PDF containing a new content based on the specified value and equipment type.

    Args:
        num_order (str): The order number.
        material (str): The material code.
        count (int): The number of items to be included in the PDF.
    """
    query = ('''
        SELECT colors.bg_color_1, colors.bg_color_2, colors.border_color
        FROM validation_data.material_color_code AS colors
        JOIN validation_data.temp_tw_material AS tw_materials ON tw_materials.code_material = colors.material
        WHERE UPPER (tw_materials.tw_material) LIKE UPPER('%%'||%s||'%%')
        ''')

    try:
        with Database_Connection(config()) as conn:
            with conn.cursor() as cur:
                cur.execute(query,(material,))
                results_colors=cur.fetchall()

    except (Exception, psycopg2.DatabaseError) as error:
        MessageHelper.show_message("Ha ocurrido el siguiente error:\n"
                    + str(error), "critical")

    first_color = results_colors[0][0]
    second_color = results_colors[0][1]
    border_color = results_colors[0][2]

    pdf = FPDF(unit='mm')
    pdf.set_font("helvetica", "B", 12)
    pdf.set_text_color(49, 49, 229)

    pdf.add_page()

    pdf.set_line_width(1)
    pdf.set_draw_color(*map(int, first_color.split(',')))
    pdf.rect(25, 8, 178, 280, style='D')

    pdf.set_draw_color(*map(int, second_color.split(',')))
    pdf.rect(23, 6, 182, 284, style='D')

    if border_color is not None:
        pdf.set_draw_color(*map(int, border_color.split(',')))
        pdf.rect(21, 4, 186, 288, style='D')

    pdf.set_xy(26, 248)
    pdf.cell(19, 9, str(count), align='C')

    pdf.set_xy(48, 248)
    pdf.cell(34, 9, str(material), align='C')

    pdf.set_xy(151, 248)
    pdf.cell(49, 9, str(num_order), align='C')

    return io.BytesIO(pdf.output())


def bar_dwg_flangedTW(num_order, material, base_diam, item_data):
    """
    Generates a PDF containing a new content based on the specified value and equipment type.""
    
    Args:
        num_order (str): The order number.
        material (str): The material code.
        base_diam (int): The base diameter of equipment
        item_data (list): The list of items to be included in the PDF.
    """

    query = ('''
        SELECT colors.bg_color_1, colors.bg_color_2, colors.border_color
        FROM validation_data.material_color_code AS colors
        JOIN validation_data.temp_tw_material AS tw_materials ON tw_materials.code_material = colors.material
        WHERE UPPER (tw_materials.tw_material) LIKE UPPER('%%'||%s||'%%')
        ''')

    try:
        with Database_Connection(config()) as conn:
            with conn.cursor() as cur:
                cur.execute(query,(material,))
                results_colors=cur.fetchall()

    except (Exception, psycopg2.DatabaseError) as error:
        MessageHelper.show_message("Ha ocurrido el siguiente error:\n"
                    + str(error), "critical")

    first_color = results_colors[0][0]
    second_color = results_colors[0][1]
    border_color = results_colors[0][2]

    pdf = FPDF(unit='mm')
    pdf.set_font("helvetica", "B", 12)
    pdf.set_text_color(49, 49, 229)

    pdf.add_page()

    pdf.set_line_width(1)
    pdf.set_draw_color(*map(int, first_color.split(',')))
    pdf.rect(25, 8, 178, 280, style='D')

    pdf.set_draw_color(*map(int, second_color.split(',')))
    pdf.rect(23, 6, 182, 284, style='D')

    if border_color is not None:
        pdf.set_draw_color(*map(int, border_color.split(',')))
        pdf.rect(21, 4, 186, 288, style='D')

    total_count = 0

    pdf.set_xy(27, 19)

    for bore, std_len, p_length, cnt in item_data:
        pdf.cell(15, 6.8, str(cnt), align='C')
        pdf.cell(15, 6.8, str(bore), align='C')
        if base_diam <= 35:
            pdf.cell(15, 6.8, str(int(p_length)), align='C')
        elif base_diam <= 40:
            pdf.cell(15, 6.8, str(int(p_length) + 3), align='C')
        else:
            pdf.cell(15, 6.8, str(int(p_length) + 5), align='C')

        if base_diam <= 35:
            pdf.cell(15, 6.8, str(int(std_len) + 10), align='C')
        elif base_diam <= 40:
            pdf.cell(15, 6.8, str(int(std_len) + 10 + 3), align='C')
        else:
            pdf.cell(15, 6.8, str(int(std_len) + 10 + 5), align='C')

        pdf.ln()
        y_pos = pdf.get_y()
        pdf.set_xy(27, y_pos)
        total_count += int(cnt)

    pdf.set_xy(172, 208)
    with pdf.rotation(90):
        pdf.cell(11, 7, str(int(std_len) + 10 - int(p_length)), align='C')
    
    pdf.set_xy(26, 248)
    pdf.cell(19, 9, str(total_count), align='C')

    pdf.set_xy(48, 248)
    pdf.cell(34, 9, str(material), align='C')

    pdf.set_xy(151, 248)
    pdf.cell(49, 9, str(num_order), align='C')

    return io.BytesIO(pdf.output())


def bar_dwg_notflangedTW(num_order, material, base_diam, item_data):
    """
    Generates a PDF containing a new content based on the specified value and equipment type.""
    
    Args:
        num_order (str): The order number.
        material (str): The material code.
        base_diam (str): The base diameter of equipment
        item_data (list): The list of items to be included in the PDF.
    """

    query = ('''
        SELECT colors.bg_color_1, colors.bg_color_2, colors.border_color
        FROM validation_data.material_color_code AS colors
        JOIN validation_data.temp_tw_material AS tw_materials ON tw_materials.code_material = colors.material
        WHERE UPPER (tw_materials.tw_material) LIKE UPPER('%%'||%s||'%%')
        ''')

    try:
        with Database_Connection(config()) as conn:
            with conn.cursor() as cur:
                cur.execute(query,(material,))
                results_colors=cur.fetchall()

    except (Exception, psycopg2.DatabaseError) as error:
        MessageHelper.show_message("Ha ocurrido el siguiente error:\n"
                    + str(error), "critical")

    first_color = results_colors[0][0]
    second_color = results_colors[0][1]
    border_color = results_colors[0][2]

    pdf = FPDF(unit='mm')
    pdf.set_font("helvetica", "B", 12)
    pdf.set_text_color(49, 49, 229)

    pdf.add_page()

    pdf.set_line_width(1)
    pdf.set_draw_color(*map(int, first_color.split(',')))
    pdf.rect(25, 8, 178, 280, style='D')

    pdf.set_draw_color(*map(int, second_color.split(',')))
    pdf.rect(23, 6, 182, 284, style='D')

    if border_color is not None:
        pdf.set_draw_color(*map(int, border_color.split(',')))
        pdf.rect(21, 4, 186, 288, style='D')

    total_count = 0

    pdf.set_xy(27, 19)

    for bore, std_len, p_length, cnt in item_data:
        pdf.cell(15, 6.8, str(cnt), align='C')
        pdf.cell(15, 6.8, str(bore), align='C')

        if base_diam < 38:
            pdf.cell(15, 6.8, str(int(p_length)), align='C')
        elif base_diam < 45:
            pdf.cell(15, 6.8, str(int(p_length) + 3), align='C')
        else:
            pdf.cell(15, 6.8, str(int(p_length) + 5), align='C')

        if base_diam < 38:
            pdf.cell(15, 6.8, str(int(std_len) + 10), align='C')
        elif base_diam < 45:
            pdf.cell(15, 6.8, str(int(std_len) + 10 + 3), align='C')
        else:
            pdf.cell(15, 6.8, str(int(std_len) + 10 + 5), align='C')

        pdf.ln()
        y_pos = pdf.get_y()
        pdf.set_xy(27, y_pos)
        total_count += cnt

    pdf.set_xy(172, 208)
    with pdf.rotation(90):
        pdf.cell(11, 7, str(int(std_len) + 10 - int(p_length)), align='C')

    pdf.set_xy(26, 248)
    pdf.cell(19, 9, str(total_count), align='C')

    pdf.set_xy(48, 248)
    pdf.cell(34, 9, str(material), align='C')

    pdf.set_xy(151, 248)
    pdf.cell(49, 9, str(num_order), align='C')

    return io.BytesIO(pdf.output())


def flange_dwg_orifice(num_order, type, material, schedule, tapping_size, tapping_num, tapping_orientation, gasket, type_flange, item_data):
    """
    Generates a PDF containing a new content based on the specified value and equipment type.

    Args:
        num_order (str): The order number.
        type (str): The type of equipment.
        material (str): The material code.
        schedule (str): The schedule of item
        tapping (str): The tapping configuration
        gasket (str): The type of gasket.
        type_flange (str): The type of flange.
        item_data (list): The list of items to be included in the PDF.
    """
    query = ('''
        SELECT colors.bg_color_1, colors.bg_color_2, colors.border_color
        FROM validation_data.material_color_code AS colors
        JOIN validation_data.flow_flange_material AS flange_materials ON flange_materials.code_material = colors.material
        WHERE UPPER (flange_materials.flange_material) LIKE UPPER('%%'||%s||'%%')
        ''')

    try:
        with Database_Connection(config()) as conn:
            with conn.cursor() as cur:
                cur.execute(query,(material,))
                results_colors=cur.fetchall()

    except (Exception, psycopg2.DatabaseError) as error:
        MessageHelper.show_message("Ha ocurrido el siguiente error:\n"
                    + str(error), "critical")

    first_color = results_colors[0][0]
    second_color = results_colors[0][1]
    border_color = results_colors[0][2]

    pdf = FPDF(unit='mm')
    pdf.set_font("helvetica", "B", 12)
    pdf.set_text_color(49, 49, 229)

    pdf.add_page()

    pdf.set_line_width(1)
    pdf.set_draw_color(*map(int, first_color.split(',')))
    pdf.rect(20, 8, 183, 280, style='D')

    pdf.set_draw_color(*map(int, second_color.split(',')))
    pdf.rect(19, 7, 185, 282, style='D')

    if border_color is not None:
        pdf.set_draw_color(*map(int, border_color.split(',')))
        pdf.rect(18, 6, 197, 284, style='D')

    item_data = list(item_data)

    if type == 'M.RUN':
        pipe_int_diam = item_data[0][0]
        flange_height = item_data[0][1]
        cnt = item_data[0][2]
    else:
        pipe_int_diam = item_data[0][0]
        cnt = item_data[0][1]

    pdf.set_draw_color(255, 0, 0)

    pdf.set_xy(29, 200)
    pdf.cell(12.5, 6.5, "ORIENTACIÓN TOMAS: " + str(tapping_orientation), align='L')
    
    pdf.set_xy(29, 210)
    pdf.cell(12.5, 6.5, "EXTRACTORES A UN TALADRO A LA IZQUIERDA DE LA TOMA", align='L')
    
    pdf.line(28, 234, 32, 238)
    pdf.line(28, 238, 32, 234)
    pdf.line(27, 231, 58, 231)

    if type == 'M.RUN':
        if type_flange == 'WN':
            pdf.set_line_width(0.5)
            pdf.set_draw_color(49, 49, 229)
            pdf.line(135, 31, 160, 14)

            pdf.circle(130, 31, 5, style='D')

            pdf.set_font('Helvetica', 'B', 12)
            pdf.set_xy(160, 14)
            pdf.cell(0.1, 0.1, "ACABADO TIPO A")

            pdf.set_draw_color(255, 0, 0)
            pdf.set_line_width(1)
            pdf.line(192, 68, 195, 75)

            pdf.set_xy(193, 87)
            with pdf.rotation(90):
                pdf.cell(0.1, 0.1, f"{flange_height:.1f}")
        else:
            pdf.set_font('Helvetica', 'B', 12)
            pdf.set_xy(160, 14)
            pdf.cell(0.1, 0.1, "ACABADO TIPO B")

    pdf.set_xy(111, 223.5)
    pdf.cell(30, 6, str(tapping_size), align='C')

    if 'SPW' in gasket:
        x_pos = 158.5
        y_pos = 229.5
        pdf.line(100, 240, 201, 240)
    elif 'RTJ' in gasket:
        x_pos = 162.5
        y_pos = 232.5
    else:
        x_pos = 158.5
        y_pos = 236
        pdf.line(100, 233, 201, 233)
    pdf.set_xy(x_pos, y_pos)
    pdf.cell(12.5, 6.5, str(2*cnt), align='C')
    pdf.cell(13, 6.5, str(schedule), align='C')
    pdf.cell(17, 6.5, str(pipe_int_diam), align='C')

    pdf.set_xy(26, 248)
    pdf.cell(19, 9, str(2*cnt), align='C')

    pdf.set_xy(48, 248)
    pdf.set_font("helvetica", "B", 8)
    pdf.cell(34, 9, str(material), align='C')

    pdf.set_xy(83, 248)
    pdf.set_font("helvetica", "B", 12)
    pdf.cell(47, 9, str(tapping_num) + " POR BRIDA", align='C')

    pdf.set_xy(151, 248)
    pdf.cell(49, 9, str(num_order), align='C')

    return io.BytesIO(pdf.output())


def flange_dwg_line(num_order, material, schedule, type_flange, reduction, connection, item_data):
    """
    Generates a PDF containing a new content based on the specified value and equipment type.

    Args:
        num_order (str): The order number.
        material (str): The material code.
        schedule (str): The schedule of item.
        type_flange (str): The type of flange.
        reduction (str): If reduction exists.
        connection (str): The type of connection.
        item_data (list): The list of items to be included in the PDF.
    """
    query = ('''
        SELECT colors.bg_color_1, colors.bg_color_2, colors.border_color
        FROM validation_data.material_color_code AS colors
        JOIN validation_data.flow_flange_material AS flange_materials ON flange_materials.code_material = colors.material
        WHERE UPPER (flange_materials.flange_material) LIKE UPPER('%%'||%s||'%%')
        ''')

    try:
        with Database_Connection(config()) as conn:
            with conn.cursor() as cur:
                cur.execute(query,(material,))
                results_colors=cur.fetchall()

    except (Exception, psycopg2.DatabaseError) as error:
        app = QtWidgets.QApplication.instance()
        if app is None:
            app = QtWidgets.QApplication([])

        MessageHelper.show_message("Ha ocurrido el siguiente error:\n"
                    + str(error), "critical")

    first_color = results_colors[0][0]
    second_color = results_colors[0][1]
    border_color = results_colors[0][2]

    pdf = FPDF(unit='mm')
    pdf.set_auto_page_break(auto=False)
    pdf.set_font("helvetica", "B", 12)
    pdf.set_text_color(49, 49, 229)
    pdf.set_draw_color(255, 0, 0)

    pdf.add_page()

    pdf.set_line_width(1)
    pdf.set_draw_color(*map(int, first_color.split(',')))
    pdf.rect(20, 8, 183, 280, style='D')

    pdf.set_draw_color(*map(int, second_color.split(',')))
    pdf.rect(19, 7, 185, 282, style='D')

    if border_color is not None:
        pdf.set_draw_color(*map(int, border_color.split(',')))
        pdf.rect(18, 6, 197, 284, style='D')

    item_data = list(item_data)
    cnt = item_data[0][2]


    if reduction != 'REDUCTION':
        pipe_int_diam = item_data[0][0]
        flange_height = item_data[0][1]

        if type_flange == 'WN':
            pdf.line(192, 68, 195, 75)

            pdf.set_font('Helvetica', 'B', 12)
            pdf.set_xy(56, 34)
            pdf.cell(20, 5, "A", align='C')

            pdf.set_xy(193, 87)
            with pdf.rotation(90):
                pdf.cell(0.1, 0.1, f"{flange_height:.1f}")
        elif type_flange == 'SW':
            pdf.set_font('Helvetica', 'B', 12)
            pdf.set_xy(160, 14)
            pdf.cell(0.1, 0.1, "ACABADO TIPO C")
        else:
            pdf.set_font('Helvetica', 'B', 12)
            pdf.set_xy(160, 14)
            pdf.cell(0.1, 0.1, "ACABADO TIPO B")

        pdf.set_xy(93, 250)
        pdf.set_font("helvetica", "B", 12)
        pdf.cell(10, 5, str(schedule), align='C')

        pdf.set_xy(108, 250)
        pdf.set_font("helvetica", "B", 12)
        pdf.cell(14, 5, str(pipe_int_diam), align='C')

    else:
        outter_size = connection.split(' ')[0].rsplit('"/',1)[0].strip() + '"'
        inner_size = connection.split(' ')[0].rsplit('"/',1)[1].strip()
        outter_sch = schedule.split(' / ')[0].strip() if '/' in schedule else schedule
        inner_sch = schedule.split(' / ')[1].strip() if '/' in schedule else schedule
        rating = connection.split('#')[0].split(' ')[1].strip()
        facing = connection.split(' ')[2].strip()

        query_outter_in_diam = ('''
        SELECT in_diam FROM validation_data.pipe_diam
        WHERE line_size = %s and sch = %s
        ''')

        query_inner_in_diam = ('''
        SELECT in_diam FROM validation_data.pipe_diam
        WHERE line_size = %s and sch = %s
        ''')

        query_outter_flange_size_rf = ('''
        SELECT dim_o, dim_w, dim_rf, dim_x, num_tal, dim_lrf, dim_tf, dim_h, dim_y FROM verification.flanges_verification
        WHERE code_flange = %s
        ''')

        query_inner_flange_size_rf = ('''
        SELECT dim_ah FROM verification.flanges_verification
        WHERE code_flange = %s
        ''')

        try:
            with Database_Connection(config()) as conn:
                with conn.cursor() as cur:
                    cur.execute(query_outter_in_diam, (outter_size,outter_sch,))
                    results_outter_in_diam=cur.fetchall()

                    cur.execute(query_inner_in_diam, (inner_size,inner_sch,))
                    results_inner_in_diam=cur.fetchall()

                    if facing in ['RF', 'FF']:
                        cur.execute(query_outter_flange_size_rf, ('B16.5-' + outter_size + '-' + rating,))
                        results_outter_flange=cur.fetchall()

                        cur.execute(query_inner_flange_size_rf, ('B16.5-' + inner_size + '-' + rating,))
                        results_inner_flange=cur.fetchall()

        except (Exception, psycopg2.DatabaseError) as error:
            app = QtWidgets.QApplication.instance()
            if app is None:
                app = QtWidgets.QApplication([])

            MessageHelper.show_message("Ha ocurrido el siguiente error:\n"
                    + str(error), "critical")

        j_diam_outter = results_outter_in_diam[0][0]
        j_diam_innner = results_inner_in_diam[0][0]
        dim_o = results_outter_flange[0][0]
        dim_w = results_outter_flange[0][1]
        dim_rf = results_outter_flange[0][2]
        dim_x = results_outter_flange[0][3]
        num_tal = results_outter_flange[0][4]
        dim_lrf = results_outter_flange[0][5]
        dim_tf = results_outter_flange[0][6]
        dim_h = results_outter_flange[0][7]
        dim_y = results_outter_flange[0][8]
        dim_ah = results_inner_flange[0][0]

        pdf.set_font('Helvetica', 'B', 8)

        pdf.set_xy(105, 223)
        pdf.cell(10, 5, str(dim_o), align='C')

        pdf.set_xy(104, 216)
        pdf.cell(10, 5, str(dim_rf), align='C')

        pdf.set_xy(105, 209)
        pdf.cell(10, 5, str(j_diam_outter), align='C')

        pdf.set_xy(104, 123)
        pdf.cell(10, 5, str(dim_x), align='C')

        pdf.set_xy(167, 69)
        pdf.cell(10, 5, f"{float(dim_w.replace(',','.')) / 2:.1f}", align='C')

        pdf.set_xy(161, 166)
        pdf.cell(4, 5, str(num_tal), align='C')

        pdf.set_xy(172, 166)
        pdf.cell(7, 5, str(dim_lrf), align='C')

        pdf.set_xy(184, 196)
        with pdf.rotation(90):
            pdf.cell(5, 5, f"{(float(dim_h.replace(',','.')) + float(dim_tf.replace(',','.'))):.1f}")

        pdf.set_xy(184, 174)
        with pdf.rotation(90):
            pdf.cell(5, 5, f"{round(float(dim_y.replace(',','.')) - float(dim_h.replace(',','.')) - float(dim_tf.replace(',','.')), 1):.1f}")

        pdf.set_xy(175, 148)
        with pdf.rotation(90):
            pdf.cell(5, 5, str('??'))

        pdf.set_xy(30, 184)
        with pdf.rotation(90):
            pdf.cell(5, 5, str(dim_y))

        pdf.set_xy(106, 137)
        pdf.cell(9, 5, str(j_diam_innner), align='C')

        pdf.set_xy(104, 130)
        pdf.cell(10, 5, str(dim_ah), align='C')

        pdf.set_xy(124, 266)
        pdf.cell(10, 5, str(rating), align='C')

        pdf.set_xy(113, 272)
        pdf.cell(10, 5, str(outter_size.split('"')[0]), align='C')

        pdf.set_xy(150, 272)
        pdf.cell(10, 5, str(inner_size.split('"')[0]), align='C')

    pdf.set_xy(26, 248)
    pdf.cell(19, 9, str(2*cnt), align='C')

    pdf.set_xy(48, 248)
    pdf.set_font("helvetica", "B", 8)
    pdf.cell(34, 9, str(material), align='C')

    pdf.set_xy(151, 248)
    pdf.cell(49, 9, str(num_order), align='C')

    return io.BytesIO(pdf.output())


def tube_dwg_meterrun(num_order, size, schedule, tube_material, calibrated, item_data):
    """
    Generates a PDF containing a new content based on the specified value and equipment type.

    Args:
        num_order (str): The order number.
        size (str): The size of the tube.
        schedule (str): The schedule of item
        tube_material (str): The material code of the tube.
        calibrated (bool): Indicates if the tube is calibrated.
        item_data (list): The list of items to be included in the PDF.
    """
    query = ('''
        SELECT colors.bg_color_1, colors.bg_color_2, colors.border_color
        FROM validation_data.material_color_code AS colors
        JOIN validation_data.flow_tube_material AS tube_materials ON tube_materials.code_material = colors.material
        WHERE UPPER (tube_materials.tube_material) LIKE UPPER('%%'||%s||'%%')
        ''')

    try:
        with Database_Connection(config()) as conn:
            with conn.cursor() as cur:
                cur.execute(query,(tube_material,))
                results_colors=cur.fetchall()

    except (Exception, psycopg2.DatabaseError) as error:
        MessageHelper.show_message("Ha ocurrido el siguiente error:\n"
                    + str(error), "critical")

    first_color = results_colors[0][0]
    second_color = results_colors[0][1]
    border_color = results_colors[0][2]

    pdf = FPDF(unit='mm')
    pdf.set_font("helvetica", "B", 12)
    pdf.set_text_color(49, 49, 229)

    pdf.add_page()

    pdf.set_line_width(1)
    pdf.set_draw_color(*map(int, first_color.split(',')))
    pdf.rect(20, 8, 183, 280, style='D')

    pdf.set_draw_color(*map(int, second_color.split(',')))
    pdf.rect(19, 7, 185, 282, style='D')

    if border_color is not None:
        pdf.set_draw_color(*map(int, border_color.split(',')))
        pdf.rect(18, 6, 197, 284, style='D')

    item_data = list(item_data)

    final_pipe_int_diam = item_data[0][0]
    pipe_ext_diam = item_data[0][1]
    length_1 = item_data[0][2]
    length_2 = item_data[0][3]
    welding_type_orifice = item_data[0][4]
    welding_type_line = item_data[0][5]
    cnt = item_data[0][6]

    pdf.set_draw_color(255, 0, 0)

    pdf.set_xy(65, 183)
    pdf.set_font("helvetica", "B", 12)
    if calibrated == 'YES':
        pdf.cell(40, 8, "TUBO CALIBRADO " + str(size), align='C')
    else:
        pdf.cell(40, 8, "TUBO " + str(size) + " SCH " + str(schedule), align='C')

    pdf.set_xy(39, 210)
    pdf.cell(14, 5, str(cnt), align='C')
    pdf.cell(45, 5, str(welding_type_orifice), align='C')
    pdf.cell(44, 5, str(welding_type_line), align='C')
    pdf.cell(14, 5, str(length_1), align='C')
    pdf.cell(15, 5, str(final_pipe_int_diam), align='C')
    pdf.cell(15, 5, str(pipe_ext_diam), align='C')

    pdf.set_xy(39, 216)
    pdf.cell(14, 5, str(cnt), align='C')
    pdf.cell(45, 5, str(welding_type_orifice), align='C')
    pdf.cell(44, 5, str(welding_type_line), align='C')
    pdf.cell(14, 5, str(length_2), align='C')
    pdf.cell(15, 5, str(final_pipe_int_diam), align='C')
    pdf.cell(15, 5, str(pipe_ext_diam), align='C')

    pdf.set_xy(27, 248)
    pdf.cell(20, 10, str(cnt), align='C')
    
    pdf.set_xy(47, 248)
    pdf.set_font("helvetica", "B", 8)
    pdf.cell(36, 9, str(tube_material), align='C')

    pdf.set_xy(151, 248)
    pdf.set_font("helvetica", "B", 12)
    pdf.cell(49, 9, str(num_order), align='C')

    return io.BytesIO(pdf.output())


def welding_dwg_meterrun(num_order, material, flange_type, item_data):
    """
    Generates a PDF containing a new content based on the specified value and equipment type.

    Args:
        num_order (str): The order number.
        material (str): The material code.
        flange_type (str): The type of flange.
        item_data (list): The list of items to be included in the PDF.
    """
    query = ('''
        SELECT colors.bg_color_1, colors.bg_color_2, colors.border_color
        FROM validation_data.material_color_code AS colors
        JOIN validation_data.flow_flange_material AS flange_materials ON flange_materials.code_material = colors.material
        WHERE UPPER (flange_materials.flange_material) LIKE UPPER('%%'||%s||'%%')
        ''')

    try:
        with Database_Connection(config()) as conn:
            with conn.cursor() as cur:
                cur.execute(query,(material,))
                results_colors=cur.fetchall()

    except (Exception, psycopg2.DatabaseError) as error:
        MessageHelper.show_message("Ha ocurrido el siguiente error:\n"
                    + str(error), "critical")

    first_color = results_colors[0][0]
    second_color = results_colors[0][1]
    border_color = results_colors[0][2]

    pdf = FPDF(unit='mm')
    pdf.set_font("helvetica", "B", 12)
    pdf.set_text_color(49, 49, 229)

    pdf.add_page()

    pdf.set_line_width(1)
    pdf.set_draw_color(*map(int, first_color.split(',')))
    pdf.rect(20, 8, 183, 280, style='D')

    pdf.set_draw_color(*map(int, second_color.split(',')))
    pdf.rect(19, 7, 185, 282, style='D')

    if border_color is not None:
        pdf.set_draw_color(*map(int, border_color.split(',')))
        pdf.rect(18, 6, 197, 284, style='D')

    item_data = list(item_data)

    cnt = item_data[0][0]

    pdf.set_draw_color(255, 0, 0)

    pdf.set_xy(27, 248)
    pdf.cell(20, 10, str(4*cnt), align='C')
    
    pdf.set_xy(47, 248)
    pdf.set_font("helvetica", "B", 8)
    pdf.cell(36, 9, str(material), align='C')

    pdf.set_xy(151, 248)
    pdf.set_font("helvetica", "B", 12)
    pdf.cell(49, 9, str(num_order), align='C')

    return io.BytesIO(pdf.output())


def loose_valves_dwg_dim(num_order, material, connection_1, connection_2, exterior_size, item_data):
    """
    Generates a PDF containing a new content based on the specified value and equipment type.

    Args:
        num_order (str): The order number.
        material (str): The material code.
        connection_1 (str): The first side connection type.
        connection_2 (str): The second side connection type.
        exterior_size (str): The process size connection
        item_data (list): The list of items to be included in the PDF.
    """
    query = ('''
        SELECT colors.bg_color_1, colors.bg_color_2, colors.border_color
        FROM validation_data.material_color_code AS colors
        WHERE UPPER (colors.material) LIKE UPPER('%%'||%s||'%%')
        ''')

    try:
        with Database_Connection(config()) as conn:
            with conn.cursor() as cur:
                cur.execute(query,(material,))
                results_colors=cur.fetchall()

    except (Exception, psycopg2.DatabaseError) as error:
        MessageHelper.show_message("Ha ocurrido el siguiente error:\n"
                    + str(error), "critical")

    first_color = results_colors[0][0]
    second_color = results_colors[0][1]
    border_color = results_colors[0][2]

    pdf = FPDF(unit='mm')
    pdf.set_font("helvetica", "B", 12)
    pdf.set_text_color(49, 49, 229)

    pdf.add_page()

    pdf.set_line_width(1)
    pdf.set_draw_color(*map(int, first_color.split(',')))
    pdf.rect(20, 8, 183, 280, style='D')

    pdf.set_draw_color(*map(int, second_color.split(',')))
    pdf.rect(19, 7, 185, 282, style='D')

    if border_color is not None:
        pdf.set_draw_color(*map(int, border_color.split(',')))
        pdf.rect(18, 6, 187, 284, style='D')

    item_data = list(item_data)

    cnt = item_data[0][0]

    pdf.set_draw_color(255, 0, 0)

    pdf.set_xy(44, 146)
    pdf.cell(27, 9, str(cnt), align='C')

    pdf.set_xy(72, 146)
    pdf.cell(27, 9, str(material), align='C')

    pdf.set_xy(100, 146)
    pdf.cell(27, 9, str(exterior_size), align='C')

    pdf.set_xy(128, 146)
    pdf.cell(27, 9, str(connection_1), align='C')

    pdf.set_xy(156, 146)
    pdf.cell(27, 9, str(connection_2), align='C')
    
    pdf.set_xy(27, 248)
    pdf.cell(20, 10, str(cnt), align='C')
    
    pdf.set_xy(47, 248)
    pdf.set_font("helvetica", "B", 8)
    pdf.cell(36, 9, str(material), align='C')

    pdf.set_xy(151, 248)
    pdf.set_font("helvetica", "B", 12)
    pdf.cell(49, 9, str(num_order), align='C')

    return io.BytesIO(pdf.output())


def dwg_m_landscape(num_order, item_data, material=None):
    """
    Generates a PDF containing a new content based on the specified value and equipment type.

    Args:
        num_order (str): The order number.
        material (str): The material code.
        item_data (list): The list of items to be included in the PDF.
    """
    if material is None:
        material = 'A105'

    query = ('''
        SELECT colors.bg_color_1, colors.bg_color_2, colors.border_color
        FROM validation_data.material_color_code AS colors
        WHERE UPPER (colors.material) LIKE UPPER('%%'||%s||'%%')
        ''')

    try:
        with Database_Connection(config()) as conn:
            with conn.cursor() as cur:
                cur.execute(query,(material,))
                results_colors=cur.fetchall()

    except (Exception, psycopg2.DatabaseError) as error:
        MessageHelper.show_message("Ha ocurrido el siguiente error:\n"
                    + str(error), "critical")

    first_color = results_colors[0][0]
    second_color = results_colors[0][1]
    border_color = results_colors[0][2]

    pdf = FPDF(unit='mm')
    pdf.set_font("helvetica", "B", 12)
    pdf.set_text_color(49, 49, 229)

    pdf.add_page()

    pdf.set_line_width(1)
    pdf.set_draw_color(*map(int, first_color.split(',')))
    pdf.rect(6, 19, 198, 270, style='D')

    pdf.set_draw_color(*map(int, second_color.split(',')))
    pdf.rect(5, 18, 200, 272, style='D')

    if border_color is not None:
        pdf.set_draw_color(*map(int, border_color.split(',')))
        pdf.rect(4, 17, 202, 274, style='D')

    item_data = list(item_data)
    cnt = item_data[0][0]

    pdf.set_draw_color(255, 0, 0)

    pdf.set_xy(35, 166)
    with pdf.rotation(270):
        pdf.cell(10, 10, str(cnt), align='C')

    pdf.set_xy(35, 254)
    with pdf.rotation(270):
        pdf.set_font("helvetica", "B", 12)
        pdf.cell(30, 6, str(num_order), align='C')

    return io.BytesIO(pdf.output())


def dwg_dim_32218_32219(num_order, material, item_data):
    """
    Generates a PDF containing a new content based on the specified value and equipment type.

    Args:
        num_order (str): The order number.
        material (str): The material code.
        item_data (list): The list of items to be included in the PDF.
    """
    query = ('''
        SELECT colors.bg_color_1, colors.bg_color_2, colors.border_color
        FROM validation_data.material_color_code AS colors
        WHERE UPPER (colors.material) LIKE UPPER('%%'||%s||'%%')
        ''')

    try:
        with Database_Connection(config()) as conn:
            with conn.cursor() as cur:
                cur.execute(query,(material,))
                results_colors=cur.fetchall()

    except (Exception, psycopg2.DatabaseError) as error:
        MessageHelper.show_message("Ha ocurrido el siguiente error:\n"
                    + str(error), "critical")

    first_color = results_colors[0][0]
    second_color = results_colors[0][1]
    border_color = results_colors[0][2]

    pdf = FPDF(unit='mm')
    pdf.set_font("helvetica", "B", 12)
    pdf.set_text_color(49, 49, 229)

    pdf.add_page()

    pdf.set_line_width(1)
    pdf.set_draw_color(*map(int, first_color.split(',')))
    pdf.rect(6, 19, 198, 270, style='D')

    pdf.set_draw_color(*map(int, second_color.split(',')))
    pdf.rect(5, 18, 200, 272, style='D')

    if border_color is not None:
        pdf.set_draw_color(*map(int, border_color.split(',')))
        pdf.rect(4, 17, 202, 274, style='D')

    item_data = list(item_data)

    cnt = item_data[0][0]

    pdf.set_draw_color(255, 0, 0)

    pdf.set_xy(33, 179)
    with pdf.rotation(270):
        pdf.cell(10, 10, str(cnt), align='C')

    pdf.set_xy(32, 257)
    with pdf.rotation(270):
        pdf.set_font("helvetica", "B", 12)
        pdf.cell(30, 6, str(num_order), align='C')

    return io.BytesIO(pdf.output())


def dwg_m_welding_32218_32219(num_order, material, item_data):
    """
    Generates a PDF containing a new content based on the specified value and equipment type.

    Args:
        num_order (str): The order number.
        material (str): The material code.
        item_data (list): The list of items to be included in the PDF.
    """
    query = ('''
        SELECT colors.bg_color_1, colors.bg_color_2, colors.border_color
        FROM validation_data.material_color_code AS colors
        WHERE UPPER (colors.material) LIKE UPPER('%%'||%s||'%%')
        ''')

    try:
        with Database_Connection(config()) as conn:
            with conn.cursor() as cur:
                cur.execute(query,(material,))
                results_colors=cur.fetchall()

    except (Exception, psycopg2.DatabaseError) as error:
        MessageHelper.show_message("Ha ocurrido el siguiente error:\n"
                    + str(error), "critical")

    first_color = results_colors[0][0]
    second_color = results_colors[0][1]
    border_color = results_colors[0][2]

    pdf = FPDF(unit='mm')
    pdf.set_font("helvetica", "B", 12)
    pdf.set_text_color(49, 49, 229)

    pdf.add_page()

    pdf.set_line_width(1)
    pdf.set_draw_color(*map(int, first_color.split(',')))
    pdf.rect(6, 19, 198, 270, style='D')

    pdf.set_draw_color(*map(int, second_color.split(',')))
    pdf.rect(5, 18, 200, 272, style='D')

    if border_color is not None:
        pdf.set_draw_color(*map(int, border_color.split(',')))
        pdf.rect(4, 17, 202, 274, style='D')

    item_data = list(item_data)

    cnt = item_data[0][0]

    pdf.set_draw_color(255, 0, 0)

    pdf.set_xy(33, 179)
    with pdf.rotation(270):
        pdf.cell(10, 10, str(cnt), align='C')

    pdf.set_xy(32, 257)
    with pdf.rotation(270):
        pdf.set_font("helvetica", "B", 12)
        pdf.cell(30, 6, str(num_order), align='C')

    return io.BytesIO(pdf.output())


def dwg_m_32218_32219(num_order, material, item_data):
    """
    Generates a PDF containing a new content based on the specified value and equipment type.

    Args:
        num_order (str): The order number.
        material (str): The material code.
        item_data (list): The list of items to be included in the PDF.
    """
    query = ('''
        SELECT colors.bg_color_1, colors.bg_color_2, colors.border_color
        FROM validation_data.material_color_code AS colors
        WHERE UPPER (colors.material) LIKE UPPER('%%'||%s||'%%')
        ''')

    try:
        with Database_Connection(config()) as conn:
            with conn.cursor() as cur:
                cur.execute(query,(material,))
                results_colors=cur.fetchall()

    except (Exception, psycopg2.DatabaseError) as error:
        MessageHelper.show_message("Ha ocurrido el siguiente error:\n"
                    + str(error), "critical")

    first_color = results_colors[0][0]
    second_color = results_colors[0][1]
    border_color = results_colors[0][2]

    pdf = FPDF(unit='mm')
    pdf.set_font("helvetica", "B", 12)
    pdf.set_text_color(49, 49, 229)

    pdf.add_page()

    pdf.set_line_width(1)
    pdf.set_draw_color(*map(int, first_color.split(',')))
    pdf.rect(20, 8, 183, 280, style='D')

    pdf.set_draw_color(*map(int, second_color.split(',')))
    pdf.rect(19, 7, 185, 282, style='D')

    if border_color is not None:
        pdf.set_draw_color(*map(int, border_color.split(',')))
        pdf.rect(18, 6, 197, 284, style='D')

    item_data = list(item_data)

    cnt = item_data[0][0]

    pdf.set_draw_color(255, 0, 0)

    pdf.set_xy(27, 248)
    pdf.cell(20, 10, str(cnt), align='C')

    pdf.set_xy(151, 248)
    pdf.set_font("helvetica", "B", 12)
    pdf.cell(49, 9, str(num_order), align='C')

    return io.BytesIO(pdf.output())


def dwg_dim_flange_plate(num_order: str, tag: str,
                        size: str, schedule: str, rating: str, facing: str,
                        pipe_int_diam: str, bore_diam: str, thickness: str, dv_diam: str, notes: str, tapping_size:str,
                        o_diam: str, a_diam: str, c_flange: str, y_flange: str, x_diam: str, r_diam: str, d_diam: str, t_flange: str, n_bores: str, diam_bores: str,
                        handle_height: str, handle_width: str, plate_c_dim: str, plate_ext_diam: str,
                        rtj_thickness: str, rtj_r_type: str, rtj_p_diam: str, rtj_e_diam: str, rtj_f_diam: str,
                        engineering: str, customer: str, project: str, po_number: str,
                        flange_material: str, stud_material: str, nut_material: str, plate_material: str, gasket_material: str, plug_qty: str, plug_material: str, screw_material: str,
                        notes_flange: str, notes_stud: str, notes_nut: str, notes_plate: str, notes_gasket: str, notes_plugs: str, notes_jack_screw: str,
                        drawing_number: str, drawing_date: str, counter: int) -> io.BytesIO:
    """
    Generates a PDF containing a new content based on the specified value and equipment type.
    """

    pdf = PDF()
    pdf.set_font("Helvetica", "", 6)
    pdf.set_text_color(255, 0, 0)

    pdf.add_page()

    # with pdf.rotation(angle=-90, x=52, y=45):
    #     pdf.image(str(get_path("Resources", "Iconos", "Logo Nobg.ico")), x = 52, y = 45, w = 40, h = 30)

    # Tapping Size
    pdf.set_xy(183, 40)
    with pdf.rotation(270):
        pdf.cell(21, 5, str(tapping_size), align='C')

    # Equipment tag and principal dimensions
    pdf.set_xy(195, 166)
    with pdf.rotation(270):
        pdf.cell(21, 5, str(tag), align='C')
        pdf.cell(23, 5, "", align='C')
        pdf.cell(11, 5, str(pipe_int_diam), align='C')
        pdf.cell(11, 5, str(bore_diam), align='C')
        pdf.cell(12, 5, str(thickness), align='C')
        pdf.cell(11, 5, str(dv_diam) if str(dv_diam) != 'None' else 'N/A', align='C')
        pdf.cell(11, 5, str(notes) if str(notes) not in  ['None', 'NO', 'N/A'] else '', align='C')

    pdf.set_xy(195, 187)
    with pdf.rotation(270):
        pdf.cell(23, 3, str(size) + " " + str(rating) + " WN " + str(facing), align='C')
    
    pdf.set_xy(192, 187)
    with pdf.rotation(270):
        pdf.cell(23, 3, "SCH " + str(schedule), align='C')

    # Flange dimensions
    pdf.set_xy(88, 25)
    with pdf.rotation(270):
        pdf.cell(8, 4, str(size), align='C')
        pdf.cell(12, 4, str(o_diam), align='C')
        pdf.cell(12, 4, str(a_diam), align='C')
        pdf.cell(12, 4, str(c_flange), align='C')
        pdf.cell(12, 4, str(y_flange), align='C')
        pdf.cell(12, 4, str(x_diam), align='C')
        pdf.cell(12, 4, str(r_diam), align='C')
        pdf.cell(12, 4, str(d_diam), align='C')
        pdf.cell(8, 4, str(t_flange), align='C')
        pdf.cell(9, 4, str(n_bores), align='C')
        pdf.cell(14, 4, str(diam_bores), align='C')

    # Plate dimensions
    pdf.set_xy(19, 111)
    with pdf.rotation(270):
        pdf.cell(12, 4, f"{size}", align='C')
        pdf.cell(10, 4, f"{handle_height}", align='C')
        pdf.cell(9, 4, f"{handle_width}", align='C')
        pdf.cell(8, 4, f"{plate_c_dim}", align='C')
        pdf.cell(11, 4, f"{plate_ext_diam}", align='C')

    pdf.set_xy(52, 66)
    with pdf.rotation(270):
        pdf.cell(28, 6, "SURFACE: 125-250 Ra" if facing == 'RF' else ("SURFACE: 63-125 Ra" if facing == 'RF63-125' else ''), align='C')

    # Project data
    pdf.set_xy(102, 222)
    with pdf.rotation(270):
        pdf.cell(38, 5, f"{engineering}", align='C')
        pdf.cell(31, 5, f"{customer}", align='C')

    pdf.set_xy(92, 222)
    with pdf.rotation(270):
        pdf.cell(38, 9, f"{project}", align='C')
        pdf.cell(31, 9, f"{po_number}", align='C')

    # Materials and quantities
    pdf.set_font("Helvetica", "", 6)
    pdf.set_xy(69, 215)
    with pdf.rotation(270):
        pdf.cell(23, 4, f"{screw_material}", align='C')
        pdf.set_font("Helvetica", "", 5)
        pdf.fixed_height_multicell(55, 4, notes_jack_screw if notes_jack_screw else '', 'L')

    pdf.set_font("Helvetica", "", 6)
    pdf.set_xy(65, 206.5)
    with pdf.rotation(270):
        pdf.cell(8.5, 4, f"{plug_qty}", align='C')
        pdf.cell(23, 4, f"{plug_material}", align='C')
        pdf.set_font("Helvetica", "", 5)
        pdf.fixed_height_multicell(55, 4, notes_plugs if notes_plugs else '', 'L')

    main_gasket, details = gasket_material.split(" (", 1)
    details_gasket = details.rstrip(")")

    pdf.set_font("Helvetica", "", 6)
    pdf.set_xy(61, 215)
    with pdf.rotation(270):
        pdf.cell(23, 4, f"{main_gasket}", align='C')
        pdf.cell(55, 4, f"{details_gasket}", align='L')
        # pdf.fixed_height_multicell(55, 4, notes_gasket if notes_gasket else '', 'L')

    pdf.set_font("Helvetica", "", 6)
    pdf.set_xy(57, 215)
    with pdf.rotation(270):
        pdf.cell(23, 4, f"{plate_material}", align='C')
        pdf.set_font("Helvetica", "", 5)
        pdf.fixed_height_multicell(55, 4, notes_plate if notes_plate else '', 'L')

    pdf.set_font("Helvetica", "", 6)
    pdf.set_xy(53, 215)
    with pdf.rotation(270):
        pdf.cell(23, 4, f"{nut_material}", align='C')
        pdf.set_font("Helvetica", "", 5)
        pdf.fixed_height_multicell(55, 4, notes_nut if notes_nut else '', 'L')

    pdf.set_font("Helvetica", "", 6)
    pdf.set_xy(49, 215)
    with pdf.rotation(270):
        pdf.cell(23, 5, f"{stud_material}", align='C')
        pdf.set_font("Helvetica", "", 5)
        pdf.fixed_height_multicell(55, 4, notes_stud if notes_stud else '', 'L')

    pdf.set_font("Helvetica", "", 6)
    pdf.set_xy(44, 215)
    with pdf.rotation(270):
        pdf.cell(23, 4, f"{flange_material}", align='C')
        pdf.set_font("Helvetica", "", 5)
        pdf.fixed_height_multicell(55, 4, notes_flange if notes_flange else '', 'L')

    # Drawing information
    pdf.set_xy(21, 177.75)
    with pdf.rotation(270):
        pdf.cell(10, 3, f"{drawing_date}", align='C')

    pdf.set_xy(31, 256)
    with pdf.rotation(270):
        pdf.cell(37, 6, f"{num_order}", align='C')

    pdf.set_xy(22, 265)
    with pdf.rotation(270):
        pdf.cell(28, 5, f"{num_order[2:] + "-" + str(drawing_number)}", align='C')

    pdf.set_xy(16, 265)
    with pdf.rotation(270):
        pdf.cell(28, 6, f"{drawing_number}/{counter:02d}", align='C')

    pdf.set_font("Helvetica", "U", 10)
    pdf.set_xy(105, 50)
    with pdf.rotation(270):
        pdf.cell(28, 6, "RANGE OF SURFACE: 125-250 Ra" if facing == 'RF' else ("RANGE OF SURFACE: 63-125 Ra" if facing == 'RF63-125' else ''), align='C')

    return io.BytesIO(pdf.output())


def dwg_of_orifice_plate(num_order, connection, element_material, of_drawing, of_drawing_date, handle_height, handle_width, plate_c_dim, plate_ext_diam, item_data):
    """
    Generates a PDF containing a new content based on the specified value and equipment type.
    """

    pdf = FPDF(unit='mm')
    pdf.set_font("Helvetica", "", 6)
    pdf.set_text_color(255, 0, 0)

    pdf.add_page()

    total_count = 0

    # Tapping Size
    pdf.set_xy(196, 156)
    with pdf.rotation(270):
        for tag, line_size, rating, facing, schedule, element_material, pipe_int_diam, orif_diam, plate_thk, dv_diam, w_diam, nace, cnt in item_data:
            y_pos = pdf.get_y()
            x_pos = pdf.get_x()

            size=line_size
            facing=facing
            pdf.cell(23, 6, str(tag), align='C')
            pdf.cell(18, 3, str(line_size) + " " + str(rating) + "# " + str(facing), align='C')
            pdf.set_xy(x_pos + 23, y_pos + 3)
            pdf.cell(18, 3, f"{"SCH " + str(schedule)}", align='C')
            pdf.set_xy(x_pos + 23 + 16.5, y_pos)
            pdf.cell(17.5, 6, str(element_material), align='C')
            pdf.cell(9.5, 6, str(pipe_int_diam), align='C')
            pdf.cell(9, 6, str(orif_diam), align='C')
            pdf.cell(9, 6, str(plate_thk), align='C')
            pdf.cell(10.5, 6, str(dv_diam), align='C')
            pdf.cell(10, 6, str(w_diam), align='C')
            pdf.cell(31.5, 6, str(nace) if str(nace) not in  ['None', 'NO', 'N/A'] else '', align='C')

            pdf.ln()
            pdf.set_xy(x_pos, y_pos + 6)

            total_count += int(cnt)

    # Plate dimensions
    pdf.set_xy(138, 75)
    with pdf.rotation(270):
        pdf.cell(12, 4, "SURFACE: 125-250 Ra" if facing == 'RF' else ("SURFACE: 63-125 Ra" if facing == 'RF63-125' else ''), align='C')

    pdf.set_xy(52, 28.5)
    with pdf.rotation(270):
        pdf.cell(12, 4, f"{size}", align='C')
        pdf.cell(10, 4, f"{handle_height}", align='C')
        pdf.cell(9, 4, f"{handle_width}", align='C')
        pdf.cell(8, 4, f"{plate_c_dim}", align='C')
        pdf.cell(11, 4, f"{plate_ext_diam}", align='C')

    pdf.set_xy(36, 166)
    with pdf.rotation(270):
        pdf.cell(14, 10, f"{total_count}", align='C')
        pdf.cell(41, 10, f"{element_material}", align='C')
        pdf.cell(35, 10, f"{connection}", align='C')

    # Drawing information
    pdf.set_xy(21, 177.75)
    with pdf.rotation(270):
        pdf.cell(10, 3, f"{of_drawing_date}", align='C')

    pdf.set_xy(36, 256)
    with pdf.rotation(270):
        pdf.cell(37, 10, f"{num_order}", align='C')

    pdf.set_xy(22, 265)
    with pdf.rotation(270):
        pdf.cell(28, 5, f"{of_drawing}", align='C')

    pdf.set_xy(16, 265)
    with pdf.rotation(270):
        pdf.cell(28, 6, f"{of_drawing}", align='C')

    return io.BytesIO(pdf.output())



# writer = PdfWriter()
# reader = PdfReader(r'M:\Comunes\EIPSA-ERP\4 PLANOS AUTOMATICOS\conjuntos bridas + placa\bridas RF 2 tomas\STD01 bridas RF 2 tomas 8 taladros (2-3-4 pulgadas).pdf')
# page_overlay = PdfReader(dwg_dim_flange_plate("P-XX/YYY-SZZ", "abcdefghijklmno",
#                                                         '24"', "40", "1500#", "RTJ",
#                                                         "49.22", "7.25", "3", "D=2.38", "NACE", '1/2" NPT-F',
#                                                         "11.11", "22.22", "33.33", "44.44", "55.55", "66.66", "77.77", "88.88", "16", "99.99",
#                                                         "1.11", "8", "1.11", "123.45",
#                                                         "engineering", "customer", "project", "10112210910",
#                                                         "ASTM A105", "ASTM A193 B7", "ASTM A194 2H", "A240 316SS", "Spiral Wound Gasket", "2", "ASTM A105", "ASTM A193 B7",
#                                                         "01", "45")).pages[0]

# reader.pages[0].merge_page(page2=page_overlay)
# writer.add_page(reader.pages[0])

# writer.write(r'M:\Comunes\EIPSA-ERP\4 PLANOS AUTOMATICOS\conjuntos bridas + placa\bridas RF 2 tomas\STD01 bridas RF 2 tomas 8 taladros (2-3-4 pulgadas) RELLENO.pdf')