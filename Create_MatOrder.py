from config import config
import psycopg2
from Excel_Export_Templates import material_order
import pandas as pd
from datetime import *
import PyQt6.QtCore
from PyQt6 import QtCore, QtGui, QtWidgets
import os
import openpyxl
import re

basedir = r"\\nas01\DATOS\Comunes\EIPSA-ERP"

def flow_matorder(proxy, model, numorder, numorder_pedmat, variable):
    """
    Processes material raw orders for flow items by inserting new entries into the fabrication orders database.

    Args:
        proxy (QAbstractProxyModel): The proxy model containing the current data view.
        model (QAbstractItemModel): The model containing the main data.
        numorder (str): The order number to process.
        numorder_pedmat (str): The base order number for material orders.
        variable (str): A variable that determines the type of processing to be done. The specific usage
                        of this variable is not detailed in the function.

    Returns:
        None: This function does not return a value but modifies the database state.
    """
    id_list=[]
    orifice_flange_list = []
    line_flange_list = []
    gasket_list = []
    bolts_list = []
    plugs_list = []
    extractor_list = []
    plate_list = []
    nipple_list = []
    handle_list = []
    chring_list = []
    tube_list = []
    piece2_list = []

    for row in range(proxy.rowCount()):
        first_column_value = proxy.data(proxy.index(row, 0))
        id_list.append(first_column_value)

    commands_numot = ("""SELECT "ot_num"
                        FROM fabrication.fab_order
                        WHERE NOT "ot_num" LIKE '90%'
                        ORDER BY "ot_num" ASC
                        """)
    check_otpedmat = f"SELECT * FROM fabrication.fab_order WHERE id = '{numorder_pedmat + '-PEDMAT'}'"
    commands_otpedmat = ("""
                            INSERT INTO fabrication.fab_order (
                            "id","tag","element","qty_element",
                            "ot_num","qty_ot","start_date")
                            VALUES (%s,%s,%s,%s,%s,%s,%s)
                            """)
    conn = None
    try:
    # read the connection parameters
        params = config()
    # connect to the PostgreSQL server
        conn = psycopg2.connect(**params)
        cur = conn.cursor()
    # execution of commands
        cur.execute(commands_numot)
        results=cur.fetchall()
        num_ot=results[-1][0]

        excel_file_path = r"\\nas01\DATOS\Comunes\EIPSA Sistemas de Gestion\MasterCTF\Bases\Contador.xlsm"
        workbook = openpyxl.load_workbook(excel_file_path, keep_vba=True)
        worksheet = workbook.active
        num_ot = worksheet['B2'].value
        cur.execute(check_otpedmat)
        results=cur.fetchall()
        if len(results) == 0:
            data=(numorder_pedmat + '-PEDMAT', numorder_pedmat, 'PEDIDO DE MATERIALES', 1, '{:06}'.format(int(num_ot) + 1), len(id_list), date.today().strftime("%d/%m/%Y"))
            cur.execute(commands_otpedmat, data)
            worksheet['B2'].value = '{:06}'.format(int(num_ot) + 1)
            workbook.save(excel_file_path)
    # close communication with the PostgreSQL database server
        cur.close()
    # commit the changes
        conn.commit()
    except (Exception, psycopg2.DatabaseError) as error:
        dlg = QtWidgets.QMessageBox()
        new_icon = QtGui.QIcon()
        new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
        dlg.setWindowIcon(new_icon)
        dlg.setWindowTitle("ERP EIPSA")
        dlg.setText("Ha ocurrido el siguiente error:\n"
                    + str(error))
        print(error)
        dlg.setIcon(QtWidgets.QMessageBox.Icon.Critical)
        dlg.exec()
        del dlg, new_icon
    finally:
        if conn is not None:
            conn.close()

    for element in id_list:
        for row in range(model.rowCount()):
            if model.data(model.index(row, 0)) == element:
                target_row = row
                break
        if target_row is not None:
            code_orifice_flange = model.data(model.index(target_row, 75))[:19] + model.data(model.index(target_row, 75))[23:]
            codefab_orifice_flange = model.data(model.index(target_row, 76))
            code_line_flange = model.data(model.index(target_row, 78))
            codefab_line_flange = model.data(model.index(target_row, 79))
            code_gasket = model.data(model.index(target_row, 81))
            codefab_gasket = model.data(model.index(target_row, 82))
            code_bolts = model.data(model.index(target_row, 84))
            codefab_bolts = model.data(model.index(target_row, 85))
            code_plugs = model.data(model.index(target_row, 87))
            codefab_plugs = model.data(model.index(target_row, 88))
            code_extractor = model.data(model.index(target_row, 90))
            codefab_extractor = model.data(model.index(target_row, 91))
            code_plate = model.data(model.index(target_row, 93))
            codefab_plate = model.data(model.index(target_row, 94))
            code_nipple = model.data(model.index(target_row, 96))
            codefab_nipple = model.data(model.index(target_row, 97))
            code_handle = model.data(model.index(target_row, 99))
            codefab_handle = model.data(model.index(target_row, 100))
            code_chring = model.data(model.index(target_row, 102))
            codefab_chring = model.data(model.index(target_row, 103))
            code_tube = model.data(model.index(target_row, 105))
            codefab_tube = model.data(model.index(target_row, 106))
            code_piece2 = model.data(model.index(target_row, 108))
            codefab_piece2 = model.data(model.index(target_row, 109))
            all_list_parts =[]

            if code_orifice_flange != '':
                tradcodbror = model.data(model.index(target_row, 113))
                schbror = model.data(model.index(target_row, 12))
                designbror = model.data(model.index(target_row, 111)).replace('.',',')
                processbror = "" #model.data(model.index(target_row, 37))
                materialbror = model.data(model.index(target_row, 13))
                qtybror = model.data(model.index(target_row, 77))
                orifice_flange_list.append([code_orifice_flange,codefab_orifice_flange,tradcodbror,schbror,designbror,processbror,materialbror,qtybror])
                all_list_parts.append(orifice_flange_list)

            if code_line_flange != '':
                tradcodbrline = model.data(model.index(target_row, 114))
                schbrline = model.data(model.index(target_row, 12))
                designbrline = model.data(model.index(target_row, 111)).replace('.',',')
                processbrline = "" #model.data(model.index(target_row, 37))
                materialbrline = model.data(model.index(target_row, 13))
                qtybrline = model.data(model.index(target_row, 80))
                line_flange_list.append([code_line_flange,codefab_line_flange,tradcodbrline,schbrline,designbrline,processbrline,materialbrline,qtybrline])
                all_list_parts.append(line_flange_list)

            if code_gasket != '':
                tradcodgasket = model.data(model.index(target_row, 115))
                schgasket = (model.data(model.index(target_row, 9)) + " " + 
                                model.data(model.index(target_row, 10)) + " " + 
                                model.data(model.index(target_row, 11)))
                designgasket = ''
                processgasket = ''
                materialgasket = ''
                qtygasket = model.data(model.index(target_row, 83))
                gasket_list.append([code_gasket,codefab_gasket,tradcodgasket,schgasket,designgasket,processgasket,materialgasket,qtygasket])
                all_list_parts.append(gasket_list)

            if code_bolts != '':
                tradcodbolts = model.data(model.index(target_row, 116))
                modelbolts = (model.data(model.index(target_row, 9)) + " " + 
                                model.data(model.index(target_row, 10)) + " " + 
                                model.data(model.index(target_row, 11)))
                designbolts = ('esp. placa ' + model.data(model.index(target_row, 19)))
                processbolts = ''
                materialbolts = model.data(model.index(target_row, 22))
                qtybolts = model.data(model.index(target_row, 64))
                bolts_list.append([code_bolts,codefab_bolts,tradcodbolts,modelbolts,designbolts,processbolts,materialbolts,qtybolts])
                all_list_parts.append(bolts_list)

            if code_extractor != '':
                tradcodextractor = model.data(model.index(target_row, 118))
                sizebrida = (model.data(model.index(target_row, 9)) + " " + 
                                model.data(model.index(target_row, 10)) + " " + 
                                model.data(model.index(target_row, 11)))
                designextractor = ('esp. placa ' + model.data(model.index(target_row, 19)))
                processextractor = ''
                material_extractor = re.search(r"^(.*?) / (\S+)(?:\s(\S+))?(?:\s)?(?:\s(.+))?$", model.data(model.index(target_row, 22))) if model.data(model.index(target_row, 22)) != 'N/A' else ""
                term_1 = material_extractor.group(1) if material_extractor != "" else ""
                term_4 = material_extractor.group(4)  if material_extractor and material_extractor.group(4) else ""
                materialextractor = f"{term_1} {term_4}" if material_extractor != "" else ""
                qtyextractor = model.data(model.index(target_row, 66))
                extractor_list.append([code_extractor,codefab_extractor,tradcodextractor,sizebrida,designextractor,processextractor,materialextractor,qtyextractor])
                all_list_parts.append(extractor_list)

            if code_plate != '':
                tradcodplate = model.data(model.index(target_row, 119))
                modelplate = ('ESP ' + model.data(model.index(target_row, 19)) + 'mm')
                diamextplate = model.data(model.index(target_row, 61))
                processplate = 'ARAMCO' if model.data(model.index(target_row, 20)) =='ARA' else ''
                materialplate = model.data(model.index(target_row, 17))
                qtyplate = model.data(model.index(target_row, 25)) if model.data(model.index(target_row, 8)) == "MULTISTAGE RO" else 1
                plate_list.append([code_plate,codefab_plate,tradcodplate,modelplate,diamextplate,processplate,materialplate,qtyplate])
                all_list_parts.append(plate_list)

            if code_nipple != '':
                tradcodnipple = model.data(model.index(target_row, 120))
                modelnipple = ''
                designnipple = ''
                processnipple = ''
                materialnipple = model.data(model.index(target_row, 13))
                qtynipple = model.data(model.index(target_row, 57))
                nipple_list.append([code_nipple,codefab_nipple,tradcodnipple,modelnipple,designnipple,processnipple,materialnipple,qtynipple])
                all_list_parts.append(nipple_list)

            if code_handle != '' and model.data(model.index(target_row, 19)) not in ['3', '1/8" (3)']:
                tradcodhandle = model.data(model.index(target_row, 121))
                modelhandle = (model.data(model.index(target_row, 62)) + 'mm')
                designhandle = model.data(model.index(target_row, 20))
                processhandle = ''
                materialhandle = '316SS'
                qtyhandle = 1
                handle_list.append([code_handle,codefab_handle,tradcodhandle,modelhandle,designhandle,processhandle,materialhandle,qtyhandle])
                all_list_parts.append(handle_list)

            if code_chring != '':
                tradcodchring = model.data(model.index(target_row, 122))
                schchring = 'ESP ' if model.data(model.index(target_row, 11)) == "RTJ" else 'ESP 38,5mm ACABADO'
                designchring = "ø" + str(model.data(model.index(target_row, 61)))
                processchring = "" #model.data(model.index(target_row, 37))
                materialchring = model.data(model.index(target_row, 17))
                qtychring = 1
                chring_list.append([code_chring,codefab_chring,tradcodchring,schchring,designchring,processchring,materialchring,qtychring])
                all_list_parts.append(chring_list)

            if code_plugs != '':
                tradcodplug = model.data(model.index(target_row, 117))
                modelplug = ''
                designplug = ''
                processplug = ''
                materialplug = 'ASTM A105' if model.data(model.index(target_row, 85))[-2:] == 'C1' else model.data(model.index(target_row, 13))
                qtyplug = int(model.data(model.index(target_row, 55))) if model.data(model.index(target_row, 55)) != '' else 0
                plugs_list.append([code_plugs,codefab_plugs,tradcodplug,modelplug,designplug,processplug,materialplug,qtyplug])
                all_list_parts.append(plugs_list)

            if code_tube != '':
                tradcodtube = model.data(model.index(target_row, 123))
                schtube = model.data(model.index(target_row, 12))
                designtube = model.data(model.index(target_row, 111)).replace('.',',')
                processtube = ''
                commands_flangecode = ("""
                    SELECT code
                    FROM validation_data.flow_flange_material
                    WHERE flange_material = %s
                    """)
                commands_tubematerial = ("""
                    SELECT tube_material
                    FROM validation_data.flow_tube_material
                    WHERE code = %s
                    """)
                conn = None
                try:
                # read the connection parameters
                    params = config()
                # connect to the PostgreSQL server
                    conn = psycopg2.connect(**params)
                    cur = conn.cursor()
                # execution of commands one by one
                    cur.execute(commands_flangecode,(model.data(model.index(target_row, 13)),))
                    results=cur.fetchone()
                    code=results[0]
                    cur.execute(commands_tubematerial,(code,))
                    results=cur.fetchall()
                    materialtube = results[0][0]
                # close communication with the PostgreSQL database server
                    cur.close()
                # commit the changes
                    conn.commit()
                except (Exception, psycopg2.DatabaseError) as error:
                    dlg = QtWidgets.QMessageBox()
                    new_icon = QtGui.QIcon()
                    new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
                    dlg.setWindowIcon(new_icon)
                    dlg.setWindowTitle("ERP EIPSA")
                    dlg.setText("Ha ocurrido el siguiente error:\n"
                                + str(error))
                    print(error)
                    dlg.setIcon(QtWidgets.QMessageBox.Icon.Critical)
                    dlg.exec()
                    del dlg, new_icon
                finally:
                    if conn is not None:
                        conn.close()
                qtytube = model.data(model.index(target_row, 107))
                tube_list.append([code_tube,codefab_tube,tradcodtube,schtube,designtube,processtube,materialtube,qtytube])
                all_list_parts.append(tube_list)

            if code_piece2 != '':
                tradcodpiece2 = model.data(model.index(target_row, 122))
                commands_thk = ("""
                    SELECT wall_thk
                    FROM validation_data.pipe_diam
                    WHERE (line_size = %s
                    AND
                    sch = %s)
                    """)
                conn = None
                try:
                # read the connection parameters
                    params = config()
                # connect to the PostgreSQL server
                    conn = psycopg2.connect(**params)
                    cur = conn.cursor()
                # execution of commands one by one
                    cur.execute(commands_thk,(model.data(model.index(target_row, 9)),model.data(model.index(target_row, 12)),))
                    results=cur.fetchone()
                    thkmin=results[0]
                # close communication with the PostgreSQL database server
                    cur.close()
                # commit the changes
                    conn.commit()
                except (Exception, psycopg2.DatabaseError) as error:
                    dlg = QtWidgets.QMessageBox()
                    new_icon = QtGui.QIcon()
                    new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
                    dlg.setWindowIcon(new_icon)
                    dlg.setWindowTitle("ERP EIPSA")
                    dlg.setText("Ha ocurrido el siguiente error:\n"
                                + str(error))
                    print(error)
                    dlg.setIcon(QtWidgets.QMessageBox.Icon.Critical)
                    dlg.exec()
                    del dlg, new_icon
                finally:
                    if conn is not None:
                        conn.close()
                modelpiece2 = ('Th mín ' + thkmin + 'mm')
                designpiece2 = ''
                processpiece2 = ''
                commands_flangecode = ("""
                    SELECT code
                    FROM validation_data.flow_flange_material
                    WHERE flange_material = %s
                    """)
                commands_sheetmaterial = ("""
                    SELECT sheet_material
                    FROM validation_data.flow_sheet_material
                    WHERE code = %s
                    """)
                conn = None
                try:
                # read the connection parameters
                    params = config()
                # connect to the PostgreSQL server
                    conn = psycopg2.connect(**params)
                    cur = conn.cursor()
                # execution of commands one by one
                    cur.execute(commands_flangecode,(model.data(model.index(target_row, 13)),))
                    results=cur.fetchone()
                    code=results[0]
                    cur.execute(commands_sheetmaterial,(code,))
                    results=cur.fetchall()
                    materialpiece2 = results[0]
                # close communication with the PostgreSQL database server
                    cur.close()
                # commit the changes
                    conn.commit()
                except (Exception, psycopg2.DatabaseError) as error:
                    dlg = QtWidgets.QMessageBox()
                    new_icon = QtGui.QIcon()
                    new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
                    dlg.setWindowIcon(new_icon)
                    dlg.setWindowTitle("ERP EIPSA")
                    dlg.setText("Ha ocurrido el siguiente error:\n"
                                + str(error))
                    print(error)
                    dlg.setIcon(QtWidgets.QMessageBox.Icon.Critical)
                    dlg.exec()
                    del dlg, new_icon
                finally:
                    if conn is not None:
                        conn.close()
                qtypiece2 = 1
                piece2_list.append([code_piece2,codefab_piece2,tradcodpiece2,modelpiece2,designpiece2,processpiece2,materialpiece2,qtypiece2])
                all_list_parts.append(piece2_list)

            columns_equipments = ["code_equipment", "code_fab_equipment", "translate_equipment", "section_type",
                                    "f_orifice_flange", "qty_f_orifice_flange", "f_line_flange", "qty_f_line_flange",
                                    "f_gasket", "qty_f_gasket", "f_bolts", "qty_f_bolts",
                                    "f_plug", "qty_f_plug", "f_extractor", "qty_f_extractor",
                                    "f_plate", "qty_f_plate", "f_nipple", "qty_f_nipple",
                                    "f_handle", "qty_f_handle", "f_chring", "qty_f_chring",
                                    "f_tube", "qty_f_tube", "f_piece2", "qty_f_piece2"]
            columns_parts = ["code_part", "code_fab_part", "code_element", "model", "design", "process", "material", "section_type"]
            columns_tags = ["code", "equipment", "num_order","order_material","contractual_date","inspection"]
            values_equipments = [model.data(model.index(target_row, 72)), model.data(model.index(target_row, 73)), model.data(model.index(target_row, 74)), "Q-CAUD",
                                model.data(model.index(target_row, 75)), model.data(model.index(target_row, 77)), model.data(model.index(target_row, 78)), model.data(model.index(target_row, 80)),
                                model.data(model.index(target_row, 81)), model.data(model.index(target_row, 83)), model.data(model.index(target_row, 84)), model.data(model.index(target_row, 86)),
                                model.data(model.index(target_row, 87)), model.data(model.index(target_row, 89)), model.data(model.index(target_row, 90)), model.data(model.index(target_row, 92)),
                                model.data(model.index(target_row, 96)), model.data(model.index(target_row, 95)), model.data(model.index(target_row, 96)), model.data(model.index(target_row, 98)),
                                model.data(model.index(target_row, 99)), model.data(model.index(target_row, 101)), model.data(model.index(target_row, 102)), model.data(model.index(target_row, 104)),
                                model.data(model.index(target_row, 105)), model.data(model.index(target_row, 107)), model.data(model.index(target_row, 108)), model.data(model.index(target_row, 110))]
            values_tags = [model.data(model.index(target_row, 4)) + "-" + model.data(model.index(target_row, 8)) + "-" + model.data(model.index(target_row, 1)), 
                            model.data(model.index(target_row, 72)), model.data(model.index(target_row, 4)), model.data(model.index(target_row, 44)),
                            model.data(model.index(target_row, 33)), model.data(model.index(target_row, 68))]

            columns_equipments  = ", ".join([f'"{column}"' for column in columns_equipments])
            values_equipments =  ", ".join(['NULL' if value == '' or value == 0 else (str(value) if isinstance(value, (int, float)) else f"'{str(value)}'") for value in values_equipments])

            columns_tags  = ", ".join([f'"{column}"' for column in columns_tags])
            values_tags =  ", ".join(['NULL' if value == '' or value == PyQt6.QtCore.QDate() else (str(value) if isinstance(value, (int, float)) else (f"'{value.toString('yyyy-MM-dd')}'" if isinstance(value, PyQt6.QtCore.QDate) else f"'{str(value)}'")) for value in values_tags])

            columns_parts = ", ".join([f'"{column}"' for column in columns_parts])

            commands_equipments = f"INSERT INTO fabrication.equipments ({columns_equipments}) VALUES ({values_equipments})"
            commands_tags = f"INSERT INTO fabrication.tags ({columns_tags}) VALUES ({values_tags})"

            check_equipments = f"SELECT * FROM fabrication.equipments WHERE code_equipment = '{model.data(model.index(target_row, 72))}'"

            conn = None
            try:
            # read the connection parameters
                params = config()
            # connect to the PostgreSQL server
                conn = psycopg2.connect(**params)
                cur = conn.cursor()
            # execution of commands
                cur.execute(check_equipments)
                results=cur.fetchall()
                if len(results) == 0:
                    cur.execute(commands_equipments)
                else:
                    set_clause = ", ".join([f"{column} = {value}" for column, value in zip(columns_equipments.split(", ")[1:], values_equipments.split(", ")[1:])])
                    update_equipments = f"UPDATE fabrication.equipments SET {set_clause} WHERE code_equipment = '{model.data(model.index(target_row, 72))}'"
                    cur.execute(update_equipments)

                for list_part in all_list_parts:
                    check_parts = f"SELECT * FROM fabrication.parts WHERE code_part = '{list_part[0][0]}'"
                    cur.execute(check_parts)
                    results=cur.fetchall()
                    if len(results) == 0:
                        list_part_modified = list_part[0].copy()
                        list_part_modified[-1] = 'Q-CAUD'
                        values_parts = ", ".join('NULL' if value == '' else (str(value) if isinstance(value, (int, float)) else f"'{str(value)}'") for value in list_part_modified)
                        commands_parts = f"INSERT INTO fabrication.parts ({columns_parts}) VALUES ({values_parts})"
                        cur.execute(commands_parts)
                    else:
                        list_part_modified = list_part[0].copy()
                        list_part_modified[-1] = 'Q-CAUD'
                        values_parts = ", ".join('NULL' if value == '' else (str(value) if isinstance(value, (int, float)) else f"'{str(value)}'") for value in list_part_modified)
                        set_clause = ", ".join([f"{column} = {value}" for column, value in zip(columns_parts.split(", ")[1:], values_parts.split(", ")[1:])])
                        update_parts = f"UPDATE fabrication.parts SET {set_clause} WHERE code_part = '{list_part[0][0]}'"
                        cur.execute(update_parts)

                cur.execute(commands_tags)
            # close communication with the PostgreSQL database server
                cur.close()
            # commit the changes
                conn.commit()
            except (Exception, psycopg2.DatabaseError) as error:
                dlg = QtWidgets.QMessageBox()
                new_icon = QtGui.QIcon()
                new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
                dlg.setWindowIcon(new_icon)
                dlg.setWindowTitle("ERP EIPSA")
                dlg.setText("Ha ocurrido el siguiente error:\n"
                            + str(error))
                print(error)
                dlg.setIcon(QtWidgets.QMessageBox.Icon.Critical)
                dlg.exec()
                del dlg, new_icon
            finally:
                if conn is not None:
                    conn.close()

# Turn all lists in dataframe and grouped in order to sum same items
    data_lists = [
    (orifice_flange_list, "df_orifice_flange"),
    (line_flange_list, "df_line_flange"),
    (gasket_list, "df_gasket"),
    (bolts_list, "df_bolts"),
    (plugs_list, "df_plugs"),
    (extractor_list, "df_extractor"),
    (plate_list, "df_plate"),
    (nipple_list, "df_nipple"),
    (handle_list, "df_handle"),
    (chring_list, "df_chring"),
    (tube_list, "df_tube"),
    (piece2_list, "df_piece2")]

    data_frames_with_data = []

    for data_list, df_name in data_lists:
        if data_list:
            sublists = [sublist[2:] for sublist in data_list]
            df = pd.DataFrame(sublists)
            df = df.groupby([0, 1, 2, 3, 4])[5].sum().reset_index()
            data_frames_with_data.append(df)

    if data_frames_with_data:
        df_combined = pd.concat(data_frames_with_data, ignore_index=True)

    commands_client = ("""
                SELECT orders."num_order",orders."num_offer",offers."client"
                FROM offers
                INNER JOIN orders ON (offers."num_offer"=orders."num_offer")
                WHERE UPPER(orders."num_order") LIKE UPPER('%%'||%s||'%%')
                """)
    conn = None
    try:
    # read the connection parameters
        params = config()
    # connect to the PostgreSQL server
        conn = psycopg2.connect(**params)
        cur = conn.cursor()
    # execution of commands one by one
        cur.execute(commands_client,(numorder,))
        results=cur.fetchone()
        client=results[2]
    # close communication with the PostgreSQL database server
        cur.close()
    # commit the changes
        conn.commit()
    except (Exception, psycopg2.DatabaseError) as error:
        dlg = QtWidgets.QMessageBox()
        new_icon = QtGui.QIcon()
        new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
        dlg.setWindowIcon(new_icon)
        dlg.setWindowTitle("ERP EIPSA")
        dlg.setText("Ha ocurrido el siguiente error:\n"
                    + str(error))
        print(error)
        dlg.setIcon(QtWidgets.QMessageBox.Icon.Critical)
        dlg.exec()
        del dlg, new_icon
    finally:
        if conn is not None:
            conn.close()
    excel_mat_order = material_order(df_combined,numorder_pedmat,client,variable,num_ot)
    excel_mat_order.save_excel()


def temp_matorder(proxy, model, numorder, numorder_pedmat, variable):
    """
    Processes material raw orders for temp items by inserting new entries into the fabrication orders database.

    Args:
        proxy (QAbstractProxyModel): The proxy model containing the current data view.
        model (QAbstractItemModel): The model containing the main data.
        numorder (str): The order number to process.
        numorder_pedmat (str): The base order number for material orders.
        variable (str): A variable that determines the type of processing to be done. The specific usage
                        of this variable is not detailed in the function.

    Returns:
        None: This function does not return a value but modifies the database state.
    """
    id_list=[]
    bar_list = []
    tube_list = []
    flange_list = []
    sensor_list = []
    head_list = []
    btb_list = []
    nipple_list = []
    spring_list = []
    plug_list = []
    puntal_list = []
    tw_list = []
    extcable_list = []

    for row in range(proxy.rowCount()):
        first_column_value = proxy.data(proxy.index(row, 0))
        id_list.append(first_column_value)

    commands_numot = ("""SELECT "ot_num"
                        FROM fabrication.fab_order
                        WHERE NOT "ot_num" LIKE '90%'
                        ORDER BY "ot_num" ASC
                        """)
    check_otpedmat = f"SELECT * FROM fabrication.fab_order WHERE id = '{numorder_pedmat + '-PEDMAT'}'"
    commands_otpedmat = ("""
                            INSERT INTO fabrication.fab_order (
                            "id","tag","element","qty_element",
                            "ot_num","qty_ot","start_date")
                            VALUES (%s,%s,%s,%s,%s,%s,%s)
                            """)
    conn = None
    try:
    # read the connection parameters
        params = config()
    # connect to the PostgreSQL server
        conn = psycopg2.connect(**params)
        cur = conn.cursor()
    # execution of commands
        cur.execute(commands_numot)
        results=cur.fetchall()
        num_ot=results[-1][0]

        excel_file_path = r"\\nas01\DATOS\Comunes\EIPSA Sistemas de Gestion\MasterCTF\Bases\Contador.xlsm"
        workbook = openpyxl.load_workbook(excel_file_path)
        worksheet = workbook.active
        num_ot = worksheet['B2'].value
        cur.execute(check_otpedmat)
        results=cur.fetchall()
        if len(results) == 0:
            data=(numorder_pedmat + '-PEDMAT', numorder_pedmat, 'PEDIDO DE MATERIALES', 1, '{:06}'.format(int(num_ot) + 1), len(id_list), date.today().strftime("%d/%m/%Y"))
            cur.execute(commands_otpedmat, data)
            worksheet['B2'].value = '{:06}'.format(int(num_ot) + 1)
            workbook.save(excel_file_path)
    # close communication with the PostgreSQL database server
        cur.close()
    # commit the changes
        conn.commit()
    except (Exception, psycopg2.DatabaseError) as error:
        dlg = QtWidgets.QMessageBox()
        new_icon = QtGui.QIcon()
        new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
        dlg.setWindowIcon(new_icon)
        dlg.setWindowTitle("ERP EIPSA")
        dlg.setText("Ha ocurrido el siguiente error:\n"
                    + str(error))
        dlg.setIcon(QtWidgets.QMessageBox.Icon.Critical)
        dlg.exec()
        del dlg, new_icon
    finally:
        if conn is not None:
            conn.close()

    for element in id_list:
        for row in range(model.rowCount()):
            if model.data(model.index(row, 0)) == element:
                target_row = row
                break
        if target_row is not None:
            code_bar = model.data(model.index(target_row, 83))
            codefab_bar = model.data(model.index(target_row, 84))
            code_tube = model.data(model.index(target_row, 86))
            codefab_tube = model.data(model.index(target_row, 87))
            code_flange = model.data(model.index(target_row, 89))
            codefab_flange = model.data(model.index(target_row, 90))
            code_sensor = model.data(model.index(target_row, 92))
            codefab_sensor = model.data(model.index(target_row, 93))
            code_head = model.data(model.index(target_row, 95))
            codefab_head = model.data(model.index(target_row, 96))
            code_btb = model.data(model.index(target_row, 98))
            codefab_btb = model.data(model.index(target_row, 99))
            code_nipple = model.data(model.index(target_row, 101))
            codefab_nipple = model.data(model.index(target_row, 102))
            code_spring = model.data(model.index(target_row, 104))
            codefab_spring = model.data(model.index(target_row, 105))
            code_puntal = model.data(model.index(target_row, 107))
            codefab_puntal = model.data(model.index(target_row, 108))
            code_plug = model.data(model.index(target_row, 110))
            codefab_plug = model.data(model.index(target_row, 111))
            code_tw = model.data(model.index(target_row, 113))
            codefab_tw = model.data(model.index(target_row, 114))
            code_extcable = model.data(model.index(target_row, 116))
            codefab_extcable = model.data(model.index(target_row, 117))
            all_list_parts =[]

            if code_bar != '':
                tradcodbar = model.data(model.index(target_row, 120)) if 'Helical' not in model.data(model.index(target_row, 9)) else 'VAINA HELICOIDAL' + (' BRIDADA ' + model.data(model.index(target_row, 10)) + ' ' + model.data(model.index(target_row, 11)) + ' ' + model.data(model.index(target_row, 12)) if model.data(model.index(target_row, 9)) == 'Flanged Helical' else '')
                modelbar = ('U=' + model.data(model.index(target_row, 16)) + ' /L=' + model.data(model.index(target_row, 15)) if 'Stone' in model.data(model.index(target_row, 9)) or 'Helical' in model.data(model.index(target_row, 9))
                            else 'Barra ø=' + '35' if float(model.data(model.index(target_row, 17)).replace(',','.'))<=33.5 else model.data(model.index(target_row, 17)))
                notesbar = ('RAÍZ ø=' + model.data(model.index(target_row, 17)) if model.data(model.index(target_row, 9)) == 'Van-Stone TW'
                            else '')
                processbar = ''
                materialbar = model.data(model.index(target_row, 14))
                qtybar = model.data(model.index(target_row, 85)) if 'Helical' not in model.data(model.index(target_row, 9)) else 1
                bar_list.append([code_bar,codefab_bar,tradcodbar,modelbar,notesbar,processbar,materialbar,qtybar])
                all_list_parts.append(bar_list)

            if code_tube != '':
                tradcodtube = model.data(model.index(target_row, 121))
                schtube = model.data(model.index(target_row, 33))
                notestube = ''
                processtube = ''
                materialtube = model.data(model.index(target_row, 14))
                qtytube = model.data(model.index(target_row, 88))
                tube_list.append([code_tube,codefab_tube,tradcodtube,schtube,notestube,processtube,materialtube,qtytube])
                all_list_parts.append(tube_list)

            if code_flange != '':
                tradcodflange = model.data(model.index(target_row, 122)) if model.data(model.index(target_row, 9)) not in ['Buttweld TW','Flanged Helical','Threaded Helical','Van-Stone Helical','VORTICRACK'] else ''
                modelflange = ''
                notesflange = ''
                processflange = ''
                materialflange = (model.data(model.index(target_row, 30)) if model.data(model.index(target_row, 9)) == 'Van-Stone TW'
                                    else model.data(model.index(target_row, 14))) if model.data(model.index(target_row, 9)) not in ['Buttweld TW','Flanged Helical','Threaded Helical','Van-Stone Helical','VORTICRACK'] else ''
                qtyflange = 1 if model.data(model.index(target_row, 9)) not in ['Buttweld TW','Flanged Helical','Threaded Helical','Van-Stone Helical','VORTICRACK'] else ''
                flange_list.append([code_flange,codefab_flange,tradcodflange,modelflange,notesflange,processflange,materialflange,qtyflange])
                all_list_parts.append(flange_list)

            print(code_sensor)
            if code_sensor != '':
                tradcodsensor = model.data(model.index(target_row, 123))
                modelsensor = (model.data(model.index(target_row, 28)) + '-' + model.data(model.index(target_row, 27)) if code_sensor[:4] == 'Bime'
                                else '')
                notesensor = (model.data(model.index(target_row, 23)) + '-' + model.data(model.index(target_row, 24)) if code_sensor[:4] == 'Bime'
                                else '')
                processsensor = ''
                materialsensor = ('PLATINO' if tradcodsensor[:5] == 'PT100'
                                else ('AC. INOX.' if model.data(model.index(target_row, 20)) == 'St.Steel' else model.data(model.index(target_row, 20))))
                qtysensor = (1 if tradcodsensor[:5] == 'PT100' or code_sensor[:4] == 'Bime'
                                else (float(model.data(model.index(target_row, 73)))/1000) if model.data(model.index(target_row, 73)) != '' else '')
                sensor_list.append([code_sensor,codefab_sensor,tradcodsensor,modelsensor,notesensor,processsensor,materialsensor,qtysensor])
                all_list_parts.append(sensor_list)

            if code_head != '':
                tradcodhead = model.data(model.index(target_row, 124))
                modelhead = model.data(model.index(target_row, 27))
                noteshead = ''
                processhead = model.data(model.index(target_row, 28))
                materialhead = ('ALUMINIO' if modelhead[-2:] == 'AL' 
                                else ('AC.CARBONO' if modelhead[-2:] == 'CS' 
                                else ('AC.INOXIDABLE' if modelhead[-2:] == 'SS' 
                                else 'MATERIAL CABEZA NO DEFINIDO')))
                qtyhead = 1
                head_list.append([code_head,codefab_head,tradcodhead,modelhead,noteshead,processhead,materialhead,qtyhead])
                all_list_parts.append(head_list)

            if code_btb != '':
                tradcodbtb = model.data(model.index(target_row, 125))
                modelbtb = "RANGO " + (model.data(model.index(target_row, 23)) + '-' + model.data(model.index(target_row, 24)) if code_btb[:2] == 'BI' 
                            else '')
                notesbtb = ''
                processbtb = ''
                materialbtb = (model.data(model.index(target_row, 20)) if code_btb[:2] == 'BI' 
                            else ('CERÁMICO' if code_btb[:2] == 'CE' else ''))
                qtybtb = model.data(model.index(target_row, 100))
                btb_list.append([code_btb,codefab_btb,tradcodbtb,modelbtb,notesbtb,processbtb,materialbtb,qtybtb])
                all_list_parts.append(btb_list)

            if code_nipple != '':
                tradcodnipple = model.data(model.index(target_row, 126))
                modelnipple = ('' if model.data(model.index(target_row, 26)) == 'N/A' or model.data(model.index(target_row, 26))=='' else model.data(model.index(target_row, 26)))
                notesnipple = ''
                processnipple = ''
                materialnipple = 'A-105/A106' if tradcodnipple[tradcodnipple.find('('):tradcodnipple.find('(')+9] == '(CS)' else 'AISI-316'
                qtynipple = 1
                nipple_list.append([code_nipple,codefab_nipple,tradcodnipple,modelnipple,notesnipple,processnipple,materialnipple,qtynipple])
                all_list_parts.append(nipple_list)

            if code_spring != '':
                tradcodspring = model.data(model.index(target_row, 127))
                modelspring = ''
                notesspring = ''
                processspring = ''
                materialspring = 'AC.INOX'
                qtyspring = 1
                spring_list.append([code_spring,codefab_spring,tradcodspring,modelspring,notesspring,processspring,materialspring,qtyspring])
                all_list_parts.append(spring_list)

            if code_plug != '':
                tradcodplug = model.data(model.index(target_row, 129))
                modelplug = ''
                notesplug = ''
                processplug = ''
                materialplug = tradcodplug[tradcodplug.find('('):tradcodplug.find('(')+9]
                qtyplug = 1
                plug_list.append([code_plug,codefab_plug,tradcodplug,modelplug,notesplug,processplug,materialplug,qtyplug])
                all_list_parts.append(plug_list)

            if code_puntal != '':
                tradcodpuntal = model.data(model.index(target_row, 128))
                modelpuntal = ''
                notespuntal = ''
                processpuntal = ''
                materialpuntal = model.data(model.index(target_row, 14))
                qtypuntal = float(code_puntal[1:8])/1000
                puntal_list.append([code_puntal,codefab_puntal,tradcodpuntal,modelpuntal,notespuntal,processpuntal,materialpuntal,qtypuntal])
                all_list_parts.append(puntal_list)

            if code_tw != '':
                tradcodtw = model.data(model.index(target_row, 130))
                modeltw = ''
                notestw = ''
                processtw = ''
                materialtw = ''
                qtytw = model.data(model.index(target_row, 115))
                tw_list.append([code_tw,codefab_tw,tradcodtw,modeltw,notestw,processtw,materialtw,qtytw])
                all_list_parts.append(tw_list)

            if code_extcable != '':
                tradcodextcable = model.data(model.index(target_row, 131))
                modelextcable = ''
                notesextcable = ''
                processextcable = ''
                materialextcable = 'AC. INOX.' if model.data(model.index(target_row, 20)) in ['AISI-304', 'AISI-310', 'AISI-316', 'AISI-321', 'St.Steel'] else model.data(model.index(target_row, 20))
                qtyextcable = float(model.data(model.index(target_row, 73)))/1000 if model.data(model.index(target_row, 73)) != '' else ''
                extcable_list.append([code_extcable,codefab_extcable,tradcodextcable,modelextcable,notesextcable,processextcable,materialextcable,qtyextcable])
                all_list_parts.append(extcable_list)

            columns_equipments = ["code_equipment", "code_fab_equipment", "translate_equipment", "section_type",
                                            "t_bar", "qty_t_bar", "t_tube", "qty_t_tube",
                                            "t_flange", "qty_t_flange", "t_sensor", "qty_t_sensor",
                                            "t_head", "qty_t_head", "t_btb", "qty_t_btb",
                                            "t_nippleextcomp", "qty_t_nippleextcomp", "t_spring", "qty_t_spring",
                                            "t_puntal", "qty_t_puntal", "t_plug", "qty_t_plug",
                                            "t_tw", "qty_t_tw", "t_extcable", "qty_t_extcable"]
            columns_parts = ["code_part", "code_fab_part", "code_element", "model", "design", "process", "material", "section_type"]
            columns_tags = ["code", "equipment", "num_order", "order_material", "contractual_date", "inspection"]
            values_equipments = [model.data(model.index(target_row, 80)), model.data(model.index(target_row, 81)), model.data(model.index(target_row, 82)), "T-TEMP",
                                model.data(model.index(target_row, 83)), model.data(model.index(target_row, 85)), model.data(model.index(target_row, 86)), model.data(model.index(target_row, 88)),
                                model.data(model.index(target_row, 89)), model.data(model.index(target_row, 91)), model.data(model.index(target_row, 92)), model.data(model.index(target_row, 94)),
                                model.data(model.index(target_row, 95)), model.data(model.index(target_row, 97)), model.data(model.index(target_row, 98)), model.data(model.index(target_row, 100)),
                                model.data(model.index(target_row, 101)), model.data(model.index(target_row, 103)), model.data(model.index(target_row, 104)), model.data(model.index(target_row, 106)),
                                model.data(model.index(target_row, 107)), model.data(model.index(target_row, 109)), model.data(model.index(target_row, 110)), model.data(model.index(target_row, 112)),
                                model.data(model.index(target_row, 113)), model.data(model.index(target_row, 115)), model.data(model.index(target_row, 116)), model.data(model.index(target_row, 118))]
            values_tags = [model.data(model.index(target_row, 4)) + "-" + model.data(model.index(target_row, 8)) + "-" + model.data(model.index(target_row, 1)), 
                            model.data(model.index(target_row, 80)), model.data(model.index(target_row, 4)), model.data(model.index(target_row, 54)),
                            model.data(model.index(target_row, 38)), model.data(model.index(target_row, 76))]

            columns_equipments  = ", ".join([f'"{column}"' for column in columns_equipments])
            values_equipments =  ", ".join(['NULL' if value == '' or value == 0 else (str(value) if isinstance(value, (int, float)) else f"'{str(value)}'") for value in values_equipments])

            columns_tags  = ", ".join([f'"{column}"' for column in columns_tags])
            values_tags =  ", ".join(['NULL' if value == '' or value == PyQt6.QtCore.QDate() else (str(value) if isinstance(value, (int, float)) else (f"'{value.toString('yyyy-MM-dd')}'" if isinstance(value, PyQt6.QtCore.QDate) else f"'{str(value)}'")) for value in values_tags])

            columns_parts = ", ".join([f'"{column}"' for column in columns_parts])

            commands_equipments = f"INSERT INTO fabrication.equipments ({columns_equipments}) VALUES ({values_equipments})"
            commands_tags = f"INSERT INTO fabrication.tags ({columns_tags}) VALUES ({values_tags})"

            check_equipments = f"SELECT * FROM fabrication.equipments WHERE code_equipment = '{model.data(model.index(target_row, 80))}'"

            conn = None
            try:
            # read the connection parameters
                params = config()
            # connect to the PostgreSQL server
                conn = psycopg2.connect(**params)
                cur = conn.cursor()
            # execution of commands
                cur.execute(check_equipments)
                results=cur.fetchall()
                if len(results) == 0:
                    cur.execute(commands_equipments)
                else:
                    set_clause = ", ".join([f"{column} = {value}" for column, value in zip(columns_equipments.split(", ")[1:], values_equipments.split(", ")[1:])])
                    update_equipments = f"UPDATE fabrication.equipments SET {set_clause} WHERE code_equipment = '{model.data(model.index(target_row, 80))}'"
                    cur.execute(update_equipments)

                for list_part in all_list_parts:
                    check_parts = f"SELECT * FROM fabrication.parts WHERE code_part = '{list_part[0][0]}'"
                    cur.execute(check_parts)
                    results=cur.fetchall()
                    if len(results) == 0:
                        list_part_modified = list_part[0].copy()
                        list_part_modified[-1] = 'T-TEMP'
                        values_parts = ", ".join('NULL' if value == '' else (str(value) if isinstance(value, (int, float)) else f"'{str(value)}'") for value in list_part_modified)
                        commands_parts = f"INSERT INTO fabrication.parts ({columns_parts}) VALUES ({values_parts})"
                        cur.execute(commands_parts)
                    else:
                        list_part_modified = list_part[0].copy()
                        list_part_modified[-1] = 'T-TEMP'
                        values_parts = ", ".join('NULL' if value == '' else (str(value) if isinstance(value, (int, float)) else f"'{str(value)}'") for value in list_part_modified)
                        set_clause = ", ".join([f"{column} = {value}" for column, value in zip(columns_parts.split(", ")[1:], values_parts.split(", ")[1:])])
                        update_parts = f"UPDATE fabrication.parts SET {set_clause} WHERE code_part = '{list_part[0][0]}'"
                        cur.execute(update_parts)

                cur.execute(commands_tags)
            # close communication with the PostgreSQL database server
                cur.close()
            # commit the changes
                conn.commit()
            except (Exception, psycopg2.DatabaseError) as error:
                dlg = QtWidgets.QMessageBox()
                new_icon = QtGui.QIcon()
                new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
                dlg.setWindowIcon(new_icon)
                dlg.setWindowTitle("ERP EIPSA")
                dlg.setText("Ha ocurrido el siguiente error:\n"
                            + str(error))
                dlg.setIcon(QtWidgets.QMessageBox.Icon.Critical)
                dlg.exec()
                del dlg, new_icon
            finally:
                if conn is not None:
                    conn.close()

# Turn all lists in dataframe and grouped in order to sum same items
    data_lists = [
    (bar_list, "df_bar"),
    (tube_list, "df_tube"),
    (flange_list, "df_flange"),
    (sensor_list, "df_sensor"),
    (head_list, "df_head"),
    (btb_list, "df_btb"),
    (nipple_list, "df_nipple"),
    (spring_list, "df_spring"),
    (plug_list, "df_plug"),
    (puntal_list, "df_puntal"),
    (extcable_list, "df_extcable")]

    data_frames_with_data = []

    for data_list, df_name in data_lists:
        if data_list:
            sublists = [sublist[2:] for sublist in data_list]
            df = pd.DataFrame(sublists)
            df = df.groupby([0, 1, 2, 3, 4])[5].sum().reset_index()
            data_frames_with_data.append(df)

    if data_frames_with_data:
        df_combined = pd.concat(data_frames_with_data, ignore_index=True)

    commands_client = ("""
                        SELECT orders."num_order",orders."num_offer",offers."client"
                        FROM offers
                        INNER JOIN orders ON (offers."num_offer"=orders."num_offer")
                        WHERE UPPER(orders."num_order") LIKE UPPER('%%'||%s||'%%')
                        """)
    conn = None
    try:
    # read the connection parameters
        params = config()
    # connect to the PostgreSQL server
        conn = psycopg2.connect(**params)
        cur = conn.cursor()
    # execution of commands one by one
        cur.execute(commands_client,(numorder,))
        results=cur.fetchone()
        client=results[2]
    # close communication with the PostgreSQL database server
        cur.close()
    # commit the changes
        conn.commit()
    except (Exception, psycopg2.DatabaseError) as error:
        dlg = QtWidgets.QMessageBox()
        new_icon = QtGui.QIcon()
        new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
        dlg.setWindowIcon(new_icon)
        dlg.setWindowTitle("ERP EIPSA")
        dlg.setText("Ha ocurrido el siguiente error:\n"
                    + str(error))
        dlg.setIcon(QtWidgets.QMessageBox.Icon.Critical)
        dlg.exec()
        del dlg, new_icon
    finally:
        if conn is not None:
            conn.close()
    excel_mat_order = material_order(df_combined,numorder_pedmat,client,variable,num_ot)
    excel_mat_order.save_excel()


def level_matorder(proxy, model, numorder, numorder_pedmat, variable):
    """
    Processes material raw orders for level items by inserting new entries into the fabrication orders database.

    Args:
        proxy (QAbstractProxyModel): The proxy model containing the current data view.
        model (QAbstractItemModel): The model containing the main data.
        numorder (str): The order number to process.
        numorder_pedmat (str): The base order number for material orders.
        variable (str): A variable that determines the type of processing to be done. The specific usage
                        of this variable is not detailed in the function.

    Returns:
        None: This function does not return a value but modifies the database state.
    """
    id_list = []
    body_list = []
    cover_list = []
    glass_list = []
    gasket_list = []
    mica_list = []
    bolts_list = []
    nipplehex_list = []
    valve_list = []
    flangevalve_list = []
    nippletube_list = []
    dv_list = []
    plug_list = []
    antifrost_list = []
    illuminator_list = []

    for row in range(proxy.rowCount()):
        first_column_value = proxy.data(proxy.index(row, 0))
        id_list.append(first_column_value)

    commands_numot = ("""SELECT "ot_num"
                        FROM fabrication.fab_order
                        WHERE NOT "ot_num" LIKE '90%'
                        ORDER BY "ot_num" ASC
                        """)
    check_otpedmat = f"SELECT * FROM fabrication.fab_order WHERE id = '{numorder_pedmat + '-PEDMAT'}'"
    commands_otpedmat = ("""
                            INSERT INTO fabrication.fab_order (
                            "id","tag","element","qty_element",
                            "ot_num","qty_ot","start_date")
                            VALUES (%s,%s,%s,%s,%s,%s,%s)
                            """)
    conn = None
    try:
    # read the connection parameters
        params = config()
    # connect to the PostgreSQL server
        conn = psycopg2.connect(**params)
        cur = conn.cursor()
    # execution of commands
        cur.execute(commands_numot)
        results=cur.fetchall()
        num_ot=results[-1][0]

        excel_file_path = r"\\nas01\DATOS\Comunes\EIPSA Sistemas de Gestion\MasterCTF\Bases\Contador.xlsm"
        workbook = openpyxl.load_workbook(excel_file_path)
        worksheet = workbook.active
        num_ot = worksheet['B2'].value
        cur.execute(check_otpedmat)
        results=cur.fetchall()
        if len(results) == 0:
            data=(numorder_pedmat + '-PEDMAT', numorder_pedmat, 'PEDIDO DE MATERIALES', 1, '{:06}'.format(int(num_ot) + 1), len(id_list), date.today().strftime("%d/%m/%Y"))
            cur.execute(commands_otpedmat, data)
            worksheet['B2'].value = '{:06}'.format(int(num_ot) + 1)
            workbook.save(excel_file_path)
    # close communication with the PostgreSQL database server
        cur.close()
    # commit the changes
        conn.commit()
    except (Exception, psycopg2.DatabaseError) as error:
        dlg = QtWidgets.QMessageBox()
        new_icon = QtGui.QIcon()
        new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
        dlg.setWindowIcon(new_icon)
        dlg.setWindowTitle("ERP EIPSA")
        dlg.setText("Ha ocurrido el siguiente error:\n"
                    + str(error))
        dlg.setIcon(QtWidgets.QMessageBox.Icon.Critical)
        dlg.exec()
        del dlg, new_icon
    finally:
        if conn is not None:
            conn.close()

    for element in id_list:
        for row in range(model.rowCount()):
            if model.data(model.index(row, 0)) == element:
                target_row = row
                break
        if target_row is not None:
            code_body = model.data(model.index(target_row, 69))
            codefab_body = model.data(model.index(target_row, 70))
            code_cover = model.data(model.index(target_row, 72))
            codefab_cover = model.data(model.index(target_row, 73))
            code_bolts = model.data(model.index(target_row, 75))
            codefab_bolts = model.data(model.index(target_row, 76))
            code_nipplehex = model.data(model.index(target_row, 78))
            codefab_nipplehex = model.data(model.index(target_row, 79))
            code_valve = model.data(model.index(target_row, 81))
            codefab_valve = model.data(model.index(target_row, 82))
            code_flangevalve = model.data(model.index(target_row, 84))
            codefab_flangevalve = model.data(model.index(target_row, 85))
            code_dv = model.data(model.index(target_row, 87))
            codefab_dv = model.data(model.index(target_row, 88))
            code_scale = model.data(model.index(target_row, 90))
            code_fab_scale = model.data(model.index(target_row, 91))
            code_illuminator = model.data(model.index(target_row, 93))
            codefab_illuminator = model.data(model.index(target_row, 94))
            code_gasket = model.data(model.index(target_row, 96))
            codefab_gasket = model.data(model.index(target_row, 97))
            code_glass = model.data(model.index(target_row, 99))
            codefab_glass = model.data(model.index(target_row, 100))
            code_float = model.data(model.index(target_row, 102))
            codefab_float = model.data(model.index(target_row, 103))
            code_mica = model.data(model.index(target_row, 105))
            codefab_mica = model.data(model.index(target_row, 106))
            code_nippletube = model.data(model.index(target_row, 114))
            codefab_nippletube = model.data(model.index(target_row, 115))
            code_antifrost = model.data(model.index(target_row, 117))
            codefab_antifrost = model.data(model.index(target_row, 118))
            all_list_parts = []

            model_num = model.data(model.index(target_row, 9))[:6] if model.data(model.index(target_row, 9))[2:4] !='HH' else model.data(model.index(target_row, 9))[:7]
            level_type = model.data(model.index(target_row, 8))
            conn_type = model.data(model.index(target_row, 15))
            nipplehexdim = model.data(model.index(target_row, 32))[:8]
            nippletubedim = model.data(model.index(target_row, 33))[:8]
            cc_length = int(model.data(model.index(target_row, 17)))

            if code_body != '':
                tradcodbody = model.data(model.index(target_row, 121))
                modelbody = nipplehexdim
                designbody = '40x40' if model_num[2:3] != 'H' else ('100x50'if model_num[2:4] != 'HH' else '80x40')
                processbody = (nipplehexdim + '-M')
                materialbody = 'A-105' if model.data(model.index(target_row, 10)) == 'Carbon Steel' else model.data(model.index(target_row, 10))
                qtybody = model.data(model.index(target_row, 71))
                body_list.append([code_body,codefab_body,tradcodbody,modelbody,designbody,processbody,materialbody,qtybody])
                all_list_parts.append(body_list)

            if code_cover != '':
                commands_coverdim = ("""
                    SELECT *
                    FROM validation_data.level_cover_dim
                    WHERE cover = %s
                    """)
                conn = None
                try:
                # read the connection parameters
                    params = config()
                # connect to the PostgreSQL server
                    conn = psycopg2.connect(**params)
                    cur = conn.cursor()
                # execution of commands one by one
                    cover_num = model_num[2:6] if model_num[2:4] !='HH' else model_num[3:7]
                    cover_num = cover_num[:2] + '1' + cover_num[3:]
                    cur.execute(commands_coverdim,(cover_num,))
                    results=cur.fetchone()
                    length=results[1]
                    bores=results[2]
                # close communication with the PostgreSQL database server
                    cur.close()
                # commit the changes
                    conn.commit()
                except (Exception, psycopg2.DatabaseError) as error:
                    dlg = QtWidgets.QMessageBox()
                    new_icon = QtGui.QIcon()
                    new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
                    dlg.setWindowIcon(new_icon)
                    dlg.setWindowTitle("ERP EIPSA")
                    dlg.setText("Ha ocurrido el siguiente error:\n"
                                + str(error))
                    dlg.setIcon(QtWidgets.QMessageBox.Icon.Critical)
                    dlg.exec()
                    del dlg, new_icon
                finally:
                    if conn is not None:
                        conn.close()

                tradcodcover = model.data(model.index(target_row, 122))
                modelcover = ('L=' + str(length))
                designcover = '80x30' if model_num[2:4] != 'HH' else '90x40'
                processcover = (str(bores) + ' taladros')
                materialcover = 'A-105' if model.data(model.index(target_row, 10)) == 'Carbon Steel' else model.data(model.index(target_row, 10))
                qtycover = model.data(model.index(target_row, 74))
                cover_list.append([code_cover,codefab_cover,tradcodcover,modelcover,designcover,processcover,materialcover,qtycover])
                all_list_parts.append(cover_list)

            if code_glass != '':
                tradcodglass = model.data(model.index(target_row, 131))
                modelglass = 'TRANSPARENCIA' if level_type == 'Transparent' else 'REFLEXIÓN'
                designglass = ''
                processglass = ''
                materialglass ='BOROSILICATO'
                qtyglass = model.data(model.index(target_row, 101))
                glass_list.append([code_glass,codefab_glass,tradcodglass,modelglass,designglass,processglass,materialglass,qtyglass])
                all_list_parts.append(glass_list)

            if code_gasket != '':
                tradcodgasket = model.data(model.index(target_row, 130))
                modelgasket = 'TRANSPARENCIA' if level_type == 'Transparent' else 'REFLEXIÓN'
                designgasket = ''
                processgasket = ''
                materialgasket ='GRAFOIL'
                qtygasket = model.data(model.index(target_row, 98))
                gasket_list.append([code_gasket,codefab_gasket,tradcodgasket,modelgasket,designgasket,processgasket,materialgasket,qtygasket])
                all_list_parts.append(gasket_list)

            if code_mica != '':
                tradcodmica = model.data(model.index(target_row, 133))
                modelmica = 'TRANSPARENCIA'
                designmica = ''
                processmica = ''
                materialmica ='MICA'
                qtymica = model.data(model.index(target_row, 107))
                mica_list.append([code_mica,codefab_mica,tradcodmica,modelmica,designmica,processmica,materialmica,qtymica])
                all_list_parts.append(mica_list)

            if code_bolts != '':
                tradcodbolts = model.data(model.index(target_row, 123))
                modelbolts = 'TRANSPARENCIA' if level_type == 'Transparent' else 'REFLEXIÓN'
                designbolts = '' if model_num[2:4] == 'HH' else ('M10x132 mm' if level_type == 'Transparent' else '')
                processbolts = '' if model_num[2:4] == 'HH' else ('cabeza exag 17 e/c' if level_type == 'Transparent' else '')
                materialbolts = 'B7/2H' if level_type in ['Transparent','Reflex'] else model.data(model.index(target_row, 24))
                qtybolts = model.data(model.index(target_row, 77))
                bolts_list.append([code_bolts,codefab_bolts,tradcodbolts,modelbolts,designbolts,processbolts,materialbolts,qtybolts])
                all_list_parts.append(bolts_list)

            if code_nipplehex != '':
                tradcodnipplehex = model.data(model.index(target_row, 124))
                modelnipplehex = (str((cc_length-int(get_number_before_mm(tradcodbody))-72)/2+22) + ' mm')
                designnipplehex = ''
                processnipplehex = ''
                materialnipplehex = 'A-105' if model.data(model.index(target_row, 10)) == 'Carbon Steel' else model.data(model.index(target_row, 10))
                qtynipplehex = model.data(model.index(target_row, 80))
                nipplehex_list.append([code_nipplehex,codefab_nipplehex,tradcodnipplehex,modelnipplehex,designnipplehex,processnipplehex,materialnipplehex,qtynipplehex])
                all_list_parts.append(nipplehex_list)

            if code_valve != '':
                tradcodvalve = model.data(model.index(target_row, 125))
                modelvalve = nipplehexdim[:4] + ' x ' + model.data(model.index(target_row, 20))
                designvalve = nipplehexdim[-3:] + '-H'
                processvalve = ''
                materialvalve = 'A-105' if model.data(model.index(target_row, 18))[-2:] == 'NB' else '316 SS'
                qtyvalve = model.data(model.index(target_row, 83))
                valve_list.append([code_valve,codefab_valve,tradcodvalve,modelvalve,designvalve,processvalve,materialvalve,qtyvalve])
                all_list_parts.append(valve_list)

            if code_flangevalve != '':
                tradcodflangevalve = model.data(model.index(target_row, 126))
                modelflangevalve = ''
                designflangevalve = ''
                processflangevalve = ''
                materialflangevalve = 'A-105' if model.data(model.index(target_row, 10)) == 'Carbon Steel' else model.data(model.index(target_row, 10))
                qtyflangevalve = model.data(model.index(target_row, 86))
                flangevalve_list.append([code_flangevalve,codefab_flangevalve,tradcodflangevalve,modelflangevalve,designflangevalve,processflangevalve,materialflangevalve,qtyflangevalve])
                all_list_parts.append(flangevalve_list)

            if code_dv != '':
                tradcoddv = model.data(model.index(target_row, 127))
                modeldv = ''
                designdv = ''
                processdv = ''
                materialdv = 'A-105' if model.data(model.index(target_row, 10)) == 'Carbon Steel' else model.data(model.index(target_row, 10))
                qtydv = model.data(model.index(target_row, 89))
                dv_list.append([code_dv,codefab_dv,tradcoddv,modeldv,designdv,processdv,materialdv,qtydv])
                all_list_parts.append(dv_list)

            if tradcoddv[:3] == 'VÁL':
                tradcodplug = 'TAPÓN NORMAL ' + model.data(model.index(target_row, 20)) + model.data(model.index(target_row, 21))
                modelplug = ''
                designplug = ''
                processplug = ''
                materialplug = 'A-105' if model.data(model.index(target_row, 10)) == 'Carbon Steel' else model.data(model.index(target_row, 10))
                qtyplug = 2
                plug_list.append([tradcodplug,modelplug,designplug,processplug,materialplug,qtyplug])

            if code_nippletube != '':
                tradcodnippletube = model.data(model.index(target_row, 136))
                modelnippletube = '80 mm'
                designnippletube = ''
                processnippletube = ''
                materialnippletube = 'A-106' if model.data(model.index(target_row, 10)) in ['Carbon Steel','ASTM A350 LF2 CL2'] else model.data(model.index(target_row, 10))
                qtynippletube = model.data(model.index(target_row, 116))
                nippletube_list.append([code_nippletube,codefab_nippletube,tradcodnippletube,modelnippletube,designnippletube,processnippletube,materialnippletube,qtynippletube])
                all_list_parts.append(nippletube_list)

            if code_illuminator != '':
                tradcodilluminator = model.data(model.index(target_row, 129))
                modelilluminator = model_num[:6].replace('S','I') if model_num[2:4] != 'HH' else model_num[:7].replace('HH','I')
                designilluminator = ''
                processilluminator = ''
                materialilluminator = 'HIERRO'
                qtyilluminator = model.data(model.index(target_row, 95))
                illuminator_list.append([code_illuminator,codefab_illuminator,tradcodilluminator,modelilluminator,designilluminator,processilluminator,materialilluminator,qtyilluminator])
                all_list_parts.append(illuminator_list)

            if code_antifrost != '':
                tradcodantifrost = model.data(model.index(target_row, 137))
                modelantifrost = ''
                designantifrost = ''
                processantifrost = ''
                materialantifrost = 'METACRILATO'
                qtyantifrost = model.data(model.index(target_row, 119))
                antifrost_list.append([code_antifrost,codefab_antifrost,tradcodantifrost,modelantifrost,designantifrost,processantifrost,materialantifrost,qtyantifrost])
                all_list_parts.append(antifrost_list)

            columns_equipments = ["code_equipment", "code_fab_equipment", "translate_equipment", "section_type",
                                            "l_body", "qty_l_body", "l_cover", "qty_l_cover",
                                            "l_studs", "qty_l_studs", "l_nipplehex", "qty_l_nipplehex",
                                            "l_valve", "qty_l_valve", "l_flange", "qty_l_flange",
                                            "l_dv", "qty_l_dv", "l_scale", "qty_l_scale",
                                            "l_illuminator", "qty_l_illuminator", "l_gasketglass", "qty_l_gasketglass",
                                            "l_glass", "qty_l_glass", "l_float", "qty_l_float",
                                            "l_mica", "qty_l_mica", "l_flags", "qty_l_flags",
                                            "l_gasketflange", "qty_l_gasketflange", "l_nippletub", "qty_l_nippletub",
                                            "l_antifrost", "qty_l_antifrost"]
            columns_parts = ["code_part", "code_fab_part", "code_element", "model", "design", "process", "material", "section_type"]
            columns_tags = ["code", "equipment", "num_order","order_material","contractual_date","inspection"]
            values_equipments = [model.data(model.index(target_row, 66)), model.data(model.index(target_row, 67)), model.data(model.index(target_row, 68)), "N-Niveles",
                                model.data(model.index(target_row, 69)), model.data(model.index(target_row, 71)), model.data(model.index(target_row, 72)), model.data(model.index(target_row, 74)),
                                model.data(model.index(target_row, 75)), model.data(model.index(target_row, 77)), model.data(model.index(target_row, 78)), model.data(model.index(target_row, 80)),
                                model.data(model.index(target_row, 81)), model.data(model.index(target_row, 83)), model.data(model.index(target_row, 84)), model.data(model.index(target_row, 86)),
                                model.data(model.index(target_row, 87)), model.data(model.index(target_row, 89)), model.data(model.index(target_row, 90)), model.data(model.index(target_row, 92)),
                                model.data(model.index(target_row, 93)), model.data(model.index(target_row, 95)), model.data(model.index(target_row, 96)), model.data(model.index(target_row, 98)),
                                model.data(model.index(target_row, 99)), model.data(model.index(target_row, 101)), model.data(model.index(target_row, 102)), model.data(model.index(target_row, 104)),
                                model.data(model.index(target_row, 105)), model.data(model.index(target_row, 107)), model.data(model.index(target_row, 108)), model.data(model.index(target_row, 110)),
                                model.data(model.index(target_row, 111)), model.data(model.index(target_row, 113)), model.data(model.index(target_row, 114)), model.data(model.index(target_row, 116)),
                                model.data(model.index(target_row, 117)), model.data(model.index(target_row, 119))]
            values_tags = [model.data(model.index(target_row, 4)) + "-" + model.data(model.index(target_row, 8)) + "-" + model.data(model.index(target_row, 1)), 
                            model.data(model.index(target_row, 66)), model.data(model.index(target_row, 4)), model.data(model.index(target_row, 48)),
                            model.data(model.index(target_row, 39)), model.data(model.index(target_row, 62))]

            columns_equipments  = ", ".join([f'"{column}"' for column in columns_equipments])
            values_equipments =  ", ".join(['NULL' if value == '' or value == 0 else (str(value) if isinstance(value, (int, float)) else f"'{str(value)}'") for value in values_equipments])

            columns_tags  = ", ".join([f'"{column}"' for column in columns_tags])
            values_tags =  ", ".join(['NULL' if value == '' or value == PyQt6.QtCore.QDate() else (str(value) if isinstance(value, (int, float)) else (f"'{value.toString('yyyy-MM-dd')}'" if isinstance(value, PyQt6.QtCore.QDate) else f"'{str(value)}'")) for value in values_tags])

            columns_parts = ", ".join([f'"{column}"' for column in columns_parts])

            commands_equipments = f"INSERT INTO fabrication.equipments ({columns_equipments}) VALUES ({values_equipments})"
            commands_tags = f"INSERT INTO fabrication.tags ({columns_tags}) VALUES ({values_tags})"

            check_equipments = f"SELECT * FROM fabrication.equipments WHERE code_equipment = '{model.data(model.index(target_row, 66))}'"

            conn = None
            try:
            # read the connection parameters
                params = config()
            # connect to the PostgreSQL server
                conn = psycopg2.connect(**params)
                cur = conn.cursor()
            # execution of commands
                cur.execute(check_equipments)
                results=cur.fetchall()
                if len(results) == 0:
                    cur.execute(commands_equipments)
                else:
                    set_clause = ", ".join([f"{column} = {value}" for column, value in zip(columns_equipments.split(", ")[1:], values_equipments.split(", ")[1:])])
                    update_equipments = f"UPDATE fabrication.equipments SET {set_clause} WHERE code_equipment = '{model.data(model.index(target_row, 66))}'"
                    cur.execute(update_equipments)

                for list_part in all_list_parts:
                    check_parts = f"SELECT * FROM fabrication.parts WHERE code_part = '{list_part[0][0]}'"
                    cur.execute(check_parts)
                    results=cur.fetchall()
                    if len(results) == 0:
                        list_part_modified = list_part[0].copy()
                        list_part_modified[-1] = 'N-Niveles'
                        values_parts = ", ".join('NULL' if value == '' else (str(value) if isinstance(value, (int, float)) else f"'{str(value)}'") for value in list_part_modified)
                        commands_parts = f"INSERT INTO fabrication.parts ({columns_parts}) VALUES ({values_parts})"
                        cur.execute(commands_parts)
                    else:
                        list_part_modified = list_part[0].copy()
                        list_part_modified[-1] = 'N-Niveles'
                        values_parts = ", ".join('NULL' if value == '' else (str(value) if isinstance(value, (int, float)) else f"'{str(value)}'") for value in list_part_modified)
                        set_clause = ", ".join([f"{column} = {value}" for column, value in zip(columns_parts.split(", ")[1:], values_parts.split(", ")[1:])])
                        update_parts = f"UPDATE fabrication.parts SET {set_clause} WHERE code_part = '{list_part[0][0]}'"
                        cur.execute(update_parts)

                cur.execute(commands_tags)
            # close communication with the PostgreSQL database server
                cur.close()
            # commit the changes
                conn.commit()
            except (Exception, psycopg2.DatabaseError) as error:
                dlg = QtWidgets.QMessageBox()
                new_icon = QtGui.QIcon()
                new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
                dlg.setWindowIcon(new_icon)
                dlg.setWindowTitle("ERP EIPSA")
                dlg.setText("Ha ocurrido el siguiente error:\n"
                            + str(error))
                dlg.setIcon(QtWidgets.QMessageBox.Icon.Critical)
                dlg.exec()
                del dlg, new_icon
            finally:
                if conn is not None:
                    conn.close()

# Turn all lists in dataframe and grouped in order to sum same items
    data_lists = [
    (body_list, "df_body"),
    (cover_list, "df_cover"),
    (glass_list, "df_glass"),
    (gasket_list, "df_gasket"),
    (mica_list, "df_mica"),
    (bolts_list, "df_bolts"),
    (nipplehex_list, "df_nipplehex"),
    (valve_list, "df_valve"),
    (flangevalve_list, "df_flangevalve"),
    (nippletube_list, "df_nippletube"),
    (dv_list, "df_list"),
    (antifrost_list, "df_antifrost"),
    (illuminator_list, "df_illuminator")]

    data_frames_with_data = []

    for data_list, df_name in data_lists:
        if data_list:
            sublists = [sublist[2:] for sublist in data_list]
            df = pd.DataFrame(sublists)
            df = df.groupby([0, 1, 2, 3, 4])[5].sum().reset_index()
            data_frames_with_data.append(df)

    if data_frames_with_data:
        df_combined = pd.concat(data_frames_with_data, ignore_index=True)

    commands_client = ("""
                        SELECT orders."num_order",orders."num_offer",offers."client"
                        FROM offers
                        INNER JOIN orders ON (offers."num_offer"=orders."num_offer")
                        WHERE UPPER(orders."num_order") LIKE UPPER('%%'||%s||'%%')
                        """)
    conn = None
    try:
    # read the connection parameters
        params = config()
    # connect to the PostgreSQL server
        conn = psycopg2.connect(**params)
        cur = conn.cursor()
    # execution of commands one by one
        cur.execute(commands_client,(numorder,))
        results=cur.fetchone()
        client=results[2]
    # close communication with the PostgreSQL database server
        cur.close()
    # commit the changes
        conn.commit()
    except (Exception, psycopg2.DatabaseError) as error:
        dlg = QtWidgets.QMessageBox()
        new_icon = QtGui.QIcon()
        new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
        dlg.setWindowIcon(new_icon)
        dlg.setWindowTitle("ERP EIPSA")
        dlg.setText("Ha ocurrido el siguiente error:\n"
                    + str(error))
        dlg.setIcon(QtWidgets.QMessageBox.Icon.Critical)
        dlg.exec()
        del dlg, new_icon
    finally:
        if conn is not None:
            conn.close()
    excel_mat_order = material_order(df_combined,numorder_pedmat,client,variable,num_ot)
    excel_mat_order.save_excel()


def others_matorder(proxy, model, numorder, numorder_pedmat, variable):
    """
    Processes material raw orders for others items by inserting new entries into the fabrication orders database.

    Args:
        proxy (QAbstractProxyModel): The proxy model containing the current data view.
        model (QAbstractItemModel): The model containing the main data.
        numorder (str): The order number to process.
        numorder_pedmat (str): The base order number for material orders.
        variable (str): A variable that determines the type of processing to be done. The specific usage
                        of this variable is not detailed in the function.

    Returns:
        None: This function does not return a value but modifies the database state.
    """
    id_list = []

    list_valves_210 = ['V-9305','V-9575','V-9576','2V-210']

    for row in range(proxy.rowCount()):
        first_column_value = proxy.data(proxy.index(row, 0))
        description = proxy.data(proxy.index(row, 8))

        id_list.append(first_column_value)

    commands_numot = ("""SELECT "ot_num"
                        FROM fabrication.fab_order
                        WHERE NOT "ot_num" LIKE '90%'
                        ORDER BY "ot_num" ASC
                        """)
    check_otpedmat = f"SELECT * FROM fabrication.fab_order WHERE id = '{numorder_pedmat + '-PEDMAT'}'"
    commands_otpedmat = ("""
                            INSERT INTO fabrication.fab_order (
                            "id","tag","element","qty_element",
                            "ot_num","qty_ot","start_date")
                            VALUES (%s,%s,%s,%s,%s,%s,%s)
                            """)
    conn = None
    try:
    # read the connection parameters
        params = config()
    # connect to the PostgreSQL server
        conn = psycopg2.connect(**params)
        cur = conn.cursor()
    # execution of commands
        cur.execute(commands_numot)
        results=cur.fetchall()
        num_ot=results[-1][0]

        excel_file_path = r"\\nas01\DATOS\Comunes\EIPSA Sistemas de Gestion\MasterCTF\Bases\Contador.xlsm"
        workbook = openpyxl.load_workbook(excel_file_path)
        worksheet = workbook.active
        num_ot = worksheet['B2'].value
        cur.execute(check_otpedmat)
        results=cur.fetchall()
        if len(results) == 0:
            data=(numorder_pedmat + '-PEDMAT', numorder_pedmat, 'PEDIDO DE MATERIALES', 1, '{:06}'.format(int(num_ot) + 1), len(id_list), date.today().strftime("%d/%m/%Y"))
            cur.execute(commands_otpedmat, data)
            worksheet['B2'].value = '{:06}'.format(int(num_ot) + 1)
            workbook.save(excel_file_path)
    # close communication with the PostgreSQL database server
        cur.close()
    # commit the changes
        conn.commit()
    except (Exception, psycopg2.DatabaseError) as error:
        dlg = QtWidgets.QMessageBox()
        new_icon = QtGui.QIcon()
        new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
        dlg.setWindowIcon(new_icon)
        dlg.setWindowTitle("ERP EIPSA")
        dlg.setText("Ha ocurrido el siguiente error:\n"
                    + str(error))
        dlg.setIcon(QtWidgets.QMessageBox.Icon.Critical)
        dlg.exec()
        del dlg, new_icon
    finally:
        if conn is not None:
            conn.close()

    if len(id_list) != 0:
        model_list = []
        valve_model_list = []
        list_1 = []
        list_2 = []
        list_3 = []
        list_4 = []
        list_5 = []
        list_6 = []
        list_7 = []
        list_8 = []
        list_9 = []
        list_10 = []
        list_11 = []
        list_12 = []
        list_13 = []
        list_14 = []
        list_15 = []
        list_16 = []
        list_17 = []
        list_18 = []
        for element in id_list:
            for row in range(model.rowCount()):
                if model.data(model.index(row, 0)) == element:
                    target_row = row
                    break
            if target_row is not None:
                description = model.data(model.index(target_row, 8))
                print(description)

                if any(valve in description for valve in list_valves_210):
                    model_valve = re.match(r'(V-\d+-[A-Za-z0-9]+)', description).group(0)
                    material_valve = (re.match(r'(V-\d+-[A-Za-z0-9]+)(.*)', description).group(2).lstrip(' - ').strip()).split(' / ')[0]
                    sch_valve = re.search(r'V-\d+-(\w+)', description).group(1)

                    tradcodvalve = 'VÁLVULA 2V-210 SCH ' + sch_valve + (' BRIDADA ' + description.split(' / ')[1].strip()) if '# RF' in description else ''
                    schvalve = 'MOD.: ' + model_valve
                    designvalve = ''
                    processvalve = ''
                    materialvalve = material_valve
                    qtyvalve = 1
                    valve_model_list.append([tradcodvalve,schvalve,designvalve,processvalve,materialvalve,qtyvalve])

                    list_1.append(['VOLANTE','VÁLVULA 2V-210 - 1500#','','',materialvalve,1])
                    list_2.append(['ARANDELA VÁLVULA','VÁLVULA 2V-210 - 1500#','','','AC. INOX',1])
                    list_3.append(['VÁSTAGO','VÁLVULA 2V-210 - 1500#','','','AISI-316 + STELLITE',1])
                    list_4.append(['GUÍA VÁSTAGO/TUERCA (ø25 x LONG 37 mm) (EXAG 22 ec/ x 6 mm)','VÁLVULA 2V-210 - 1500#','','','AC. INOX' if materialvalve == '316' else 'AC. CARBONO',1])
                    list_5.append(['CAPELLI (HORQUILLA)','VÁLVULA 2V-210 - 1500#','','',materialvalve,1])
                    list_6.append(['FLANGETE','VÁLVULA 2V-210 - 1500#','','',materialvalve,1])
                    list_7.append(['PRENSA (ø25 x LONG 20 mm)','VÁLVULA 2V-210 - 1500#','','','AC. INOX' if materialvalve == '316' else 'AC. CARBONO',1])
                    list_8.append(['EMPAQUETADURA','VÁLVULA 2V-210 - 1500#','','','GRAFITO',1])
                    list_9.append(['TORNILLO CUADRADO (2 ud. POR VÁLVULA)','VÁLVULA 2V-210 - 1500#','','','AC. INOX',2])
                    list_10.append(['TORNILLO REDONDO (4 ud. POR VÁLVULA)','VÁLVULA 2V-210 - 1500#','','','AC. INOX',4])
                    list_11.append(['TUERCAS M10 2H','VÁLVULA 2V-210 - 1500#','','','A1942H',4])
                    list_12.append(['JUNTA ESPIROMETÁLICA 42x30x3,2mm','VÁLVULA 2V-210 - 1500#','','','AISI-316 + GRAFITO',1])
                    list_13.append(['CUERPO VÁLVULA 2V-210 - 1500#','','','',materialvalve,1])
                    list_14.append(['ASIENTO (ø20 x 16 mm)','VÁLVULA 2V-210 - 1500#','','','AISI-316 + STELLITE',1])
                    list_15.append(['BRIDA VÁLVULA '+ description.split(' / ')[1].strip(),'VÁLVULA 2V-210 - 1500#','','',materialvalve,1]) if '# RF' in description else ''
                    list_16.append(['TAPÓN PURGADOR 1/2" NPT-M','','','',materialvalve,1])
                    list_17.append(['TORNILLO TAPÓN PURGADOR','','','','AC. INOX',1])
                    if len(description.split(' / ')) > 2: 
                        list_18.append(['NIPLO ' + description.split(' / ')[2],'','','','AC. INOX' if materialvalve == '316' else 'AC. CARBONO',1]) 

                    data_lists = [
                    (valve_model_list, "df_valvemodel"),
                    (list_1, "df_list1"),
                    (list_2, "df_list2"),
                    (list_3, "df_list3"),
                    (list_4, "df_list4"),
                    (list_5, "df_list5"),
                    (list_6, "df_list6"),
                    (list_7, "df_list7"),
                    (list_8, "df_list8"),
                    (list_9, "df_list9"),
                    (list_10, "df_list10"),
                    (list_11, "df_list11"),
                    (list_12, "df_list12"),
                    (list_13, "df_list13"),
                    (list_14, "df_list14"),
                    (list_15, "df_list15"),
                    (list_16, "df_list16"),
                    (list_17, "df_list17"),
                    (list_18, "df_list18")]

                else:
                    tradcod = str(description)
                    sch = ''
                    design = ''
                    process = ''
                    material = ''
                    qty = 1
                    model_list.append([tradcod,sch,design,process,material,qty])

                    data_lists = [
                    (model_list, "df_model"),]

                    print(model_list)

        data_frames_with_data = []

        for data_list, df_name in data_lists:
            if data_list:
                sublists = [sublist for sublist in data_list]
                df = pd.DataFrame(sublists)
                df = df.groupby([0, 1, 2, 3, 4])[5].sum().reset_index()
                data_frames_with_data.append(df)

        if data_frames_with_data:
            df_combined = pd.concat(data_frames_with_data, ignore_index=True)

        print(df_combined)

        commands_client = ("""
                            SELECT orders."num_order",orders."num_offer",offers."client"
                            FROM offers
                            INNER JOIN orders ON (offers."num_offer"=orders."num_offer")
                            WHERE UPPER(orders."num_order") LIKE UPPER('%%'||%s||'%%')
                            """)
        conn = None
        try:
        # read the connection parameters
            params = config()
        # connect to the PostgreSQL server
            conn = psycopg2.connect(**params)
            cur = conn.cursor()
        # execution of commands one by one
            cur.execute(commands_client,(numorder,))
            results=cur.fetchone()
            client=results[2]
        # close communication with the PostgreSQL database server
            cur.close()
        # commit the changes
            conn.commit()
        except (Exception, psycopg2.DatabaseError) as error:
            dlg = QtWidgets.QMessageBox()
            new_icon = QtGui.QIcon()
            new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
            dlg.setWindowIcon(new_icon)
            dlg.setWindowTitle("ERP EIPSA")
            dlg.setText("Ha ocurrido el siguiente error:\n"
                        + str(error))
            dlg.setIcon(QtWidgets.QMessageBox.Icon.Critical)
            dlg.exec()
            del dlg, new_icon
        finally:
            if conn is not None:
                conn.close()
        excel_mat_order = material_order(df_combined,numorder_pedmat,client,variable,num_ot)
        excel_mat_order.save_excel()


def get_number_before_mm(text):
    """
    Extracts the integer value that appears before the 'mm' substring in the given text.

    Args:
        text (str): The input string to search for the number.

    Returns:
        int or None: The extracted integer value if found, or None if no match is found.
    """
    match = re.search(r'(\d+)\s*mm', text)
    if match:
        return int(match.group(1))
    return None