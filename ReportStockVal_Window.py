# Form implementation generated from reading ui file 'ReportArtMov_Window.ui'
#
# Created by: PyQt6 UI code generator 6.4.2
#
# WARNING: Any manual changes made to this file will be lost when pyuic6 is
# run again.  Do not edit this file unless you know what you are doing.


from PyQt6 import QtCore, QtGui, QtWidgets
from config import config
import psycopg2
import locale
import os
import pandas as pd
from PDF_Viewer import PDF_Viewer
from PyQt6.QtCore import QUrl
from tkinter.filedialog import asksaveasfilename
from openpyxl import Workbook
from openpyxl.styles import NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows

basedir = r"\\nas01\DATOS\Comunes\EIPSA-ERP"


class AlignDelegate(QtWidgets.QStyledItemDelegate):
    def initStyleOption(self, option, index):
        super(AlignDelegate, self).initStyleOption(option, index)
        option.displayAlignment = QtCore.Qt.AlignmentFlag.AlignCenter


class Ui_StockVal_Window(object):
    def __init__(self):
        self.pdf_viewer = PDF_Viewer()


    def setupUi(self, ReportStockVal):
        ReportStockVal.setObjectName("ReportStockVal")
        ReportStockVal.resize(1165, 945)
        ReportStockVal.setMinimumSize(QtCore.QSize(1165, 945))
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap("//nas01/DATOS/Comunes/EIPSA-ERP/Resources/Iconos/icon.ico"), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
        ReportStockVal.setWindowIcon(icon)
        ReportStockVal.setStyleSheet("QWidget {\n"
"background-color: rgb(255, 255, 255);\n"
"}\n"
"\n"
".QFrame {\n"
"    border: 2px solid black;\n"
"}\n"
"\n"
"QPushButton {\n"
"background-color: #33bdef;\n"
"  border: 1px solid transparent;\n"
"  border-radius: 3px;\n"
"  color: #fff;\n"
"  font-family: -apple-system,system-ui,\"Segoe UI\",\"Liberation Sans\",sans-serif;\n"
"  font-size: 12px;\n"
"  font-weight: 800;\n"
"  line-height: 1.15385;\n"
"  margin: 0;\n"
"  outline: none;\n"
"  padding: 2px .2em;\n"
"  text-align: center;\n"
"  text-decoration: none;\n"
"  vertical-align: baseline;\n"
"  white-space: nowrap;\n"
"}\n"
"\n"
"QPushButton:hover {\n"
"    background-color: #019ad2;\n"
"    border-color: rgb(0, 0, 0);\n"
"}\n"
"\n"
"QPushButton:focus {\n"
"    background-color: #019ad2;\n"
"    border-color: rgb(0, 0, 0);\n"
"}\n"
"\n"
"QPushButton:pressed {\n"
"    background-color: rgb(1, 140, 190);\n"
"    border-color: rgb(255, 255, 255)\n"
"}\n"
"\n"
"QPushButton:focus:pressed {\n"
"    background-color: rgb(1, 140, 190);\n"
"    border-color: rgb(255, 255, 255);\n"
"}")
        self.centralwidget = QtWidgets.QWidget(parent=ReportStockVal)
        self.centralwidget.setObjectName("centralwidget")
        self.gridLayout = QtWidgets.QGridLayout(self.centralwidget)
        self.gridLayout.setObjectName("gridLayout")
        self.frame = QtWidgets.QFrame(parent=self.centralwidget)
        self.frame.setFrameShape(QtWidgets.QFrame.Shape.StyledPanel)
        self.frame.setFrameShadow(QtWidgets.QFrame.Shadow.Raised)
        self.frame.setObjectName("frame")
        self.gridLayout_2 = QtWidgets.QGridLayout(self.frame)
        self.gridLayout_2.setObjectName("gridLayout_2")
        spacerItem = QtWidgets.QSpacerItem(20, 10, QtWidgets.QSizePolicy.Policy.Minimum, QtWidgets.QSizePolicy.Policy.Fixed)
        self.gridLayout_2.addItem(spacerItem, 0, 0, 1, 1)
        self.gridLayout1 = QtWidgets.QGridLayout()
        self.gridLayout1.setSpacing(0)
        self.gridLayout1.setObjectName("gridLayout1")
        self.Button_PDF = QtWidgets.QPushButton(parent=self.frame)
        self.Button_PDF.setMinimumSize(QtCore.QSize(int(175//1.5), int(35//1.5)))
        self.Button_PDF.setMaximumSize(QtCore.QSize(int(175//1.5), int(35//1.5)))
        self.Button_PDF.setFocusPolicy(QtCore.Qt.FocusPolicy.NoFocus)
        self.Button_PDF.setObjectName("Button_PDF")
        self.gridLayout1.addWidget(self.Button_PDF, 0, 2, 1, 1)
        self.Button_Excel = QtWidgets.QPushButton(parent=self.frame)
        self.Button_Excel.setMinimumSize(QtCore.QSize(int(175//1.5), int(35//1.5)))
        self.Button_Excel.setMaximumSize(QtCore.QSize(int(175//1.5), int(35//1.5)))
        self.Button_Excel.setFocusPolicy(QtCore.Qt.FocusPolicy.NoFocus)
        self.Button_Excel.setObjectName("Button_Excel")
        self.gridLayout1.addWidget(self.Button_Excel, 0, 3, 1, 1)
        self.gridLayout_2.addLayout(self.gridLayout1, 1, 0, 1, 1)
        spacerItem1 = QtWidgets.QSpacerItem(20, 10, QtWidgets.QSizePolicy.Policy.Minimum, QtWidgets.QSizePolicy.Policy.Fixed)
        self.gridLayout_2.addItem(spacerItem1, 2, 0, 1, 1)
        self.tableWidget = QtWidgets.QTableWidget(parent=self.frame)
        self.tableWidget.setObjectName("tableWidget")
        self.tableWidget.setColumnCount(6)
        self.tableWidget.setRowCount(0)
        for i in range(6):
            item = QtWidgets.QTableWidgetItem()
            font = QtGui.QFont()
            font.setPointSize(int(14//1.5))
            font.setBold(True)
            item.setFont(font)
            self.tableWidget.setHorizontalHeaderItem(i, item)
        self.gridLayout_2.addWidget(self.tableWidget, 3, 0, 1, 1)
        self.horizontalLayout = QtWidgets.QHBoxLayout()
        self.horizontalLayout.setObjectName("horizontalLayout")
        spacerItem2 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Policy.Expanding, QtWidgets.QSizePolicy.Policy.Minimum)
        self.horizontalLayout.addItem(spacerItem2)
        self.label_Total_Euro = QtWidgets.QLabel(parent=self.frame)
        self.label_Total_Euro.setMinimumSize(QtCore.QSize(0, int(50//1.5)))
        self.label_Total_Euro.setMaximumSize(QtCore.QSize(16777215, int(50//1.5)))
        font = QtGui.QFont()
        font.setPointSize(int(16//1.5))
        self.label_Total_Euro.setFont(font)
        self.label_Total_Euro.setObjectName("label_Total_Euro")
        self.horizontalLayout.addWidget(self.label_Total_Euro)
        self.label_TotalValue_Euro = QtWidgets.QLabel(parent=self.frame)
        self.label_TotalValue_Euro.setMinimumSize(QtCore.QSize(int(150//1.5), int(50//1.5)))
        self.label_TotalValue_Euro.setMaximumSize(QtCore.QSize(int(150//1.5), int(50//1.5)))
        font = QtGui.QFont()
        font.setPointSize(int(16//1.5))
        self.label_TotalValue_Euro.setFont(font)
        self.label_TotalValue_Euro.setAlignment(QtCore.Qt.AlignmentFlag.AlignRight|QtCore.Qt.AlignmentFlag.AlignTrailing|QtCore.Qt.AlignmentFlag.AlignVCenter)
        self.label_TotalValue_Euro.setObjectName("label_TotalValue_Euro")
        self.horizontalLayout.addWidget(self.label_TotalValue_Euro)
        self.label_Total = QtWidgets.QLabel(parent=self.frame)
        self.label_Total.setMinimumSize(QtCore.QSize(0, int(50//1.5)))
        self.label_Total.setMaximumSize(QtCore.QSize(16777215, int(50//1.5)))
        font = QtGui.QFont()
        font.setPointSize(int(16//1.5))
        self.label_Total.setFont(font)
        self.label_Total.setObjectName("label_Total")
        self.horizontalLayout.addWidget(self.label_Total)
        self.label_TotalValue = QtWidgets.QLabel(parent=self.frame)
        self.label_TotalValue.setMinimumSize(QtCore.QSize(int(100//1.5), int(50//1.5)))
        self.label_TotalValue.setMaximumSize(QtCore.QSize(int(100//1.5), int(50//1.5)))
        font = QtGui.QFont()
        font.setPointSize(int(16//1.5))
        self.label_TotalValue.setFont(font)
        self.label_TotalValue.setAlignment(QtCore.Qt.AlignmentFlag.AlignRight|QtCore.Qt.AlignmentFlag.AlignTrailing|QtCore.Qt.AlignmentFlag.AlignVCenter)
        self.label_TotalValue.setObjectName("label_TotalValue")
        self.horizontalLayout.addWidget(self.label_TotalValue)
        self.gridLayout_2.addLayout(self.horizontalLayout, 4, 0, 1, 1)
        self.gridLayout.addWidget(self.frame, 0, 0, 1, 1)
        ReportStockVal.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(parent=ReportStockVal)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 1165, 22))
        self.menubar.setObjectName("menubar")
        ReportStockVal.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(parent=ReportStockVal)
        self.statusbar.setObjectName("statusbar")
        ReportStockVal.setStatusBar(self.statusbar)
        self.tableWidget.verticalHeader().hide()
        self.tableWidget.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.ResizeMode.Stretch)
        self.tableWidget.horizontalHeader().setStyleSheet("QHeaderView::section {background-color: #33bdef; border: 1px solid black;}")

        self.retranslateUi(ReportStockVal)
        QtCore.QMetaObject.connectSlotsByName(ReportStockVal)

        self.Button_PDF.clicked.connect(self.generate_pdf)
        self.Button_Excel.clicked.connect(self.generate_excel)

        self.loaddata()

    def retranslateUi(self, ReportStockVal):
        _translate = QtCore.QCoreApplication.translate
        ReportStockVal.setWindowTitle(_translate("ReportStockVal", "Resumen Cliente"))
        item = self.tableWidget.horizontalHeaderItem(0)
        item.setText(_translate("ReportStockVal", "Referencia"))
        item = self.tableWidget.horizontalHeaderItem(1)
        item.setText(_translate("ReportStockVal", "Descripción"))
        item = self.tableWidget.horizontalHeaderItem(2)
        item.setText(_translate("ReportStockVal", "Val. Unit."))
        item = self.tableWidget.horizontalHeaderItem(3)
        item.setText(_translate("ReportStockVal", "Stock"))
        item = self.tableWidget.horizontalHeaderItem(4)
        item.setText(_translate("ReportStockVal", "Un."))
        item = self.tableWidget.horizontalHeaderItem(5)
        item.setText(_translate("ReportStockVal", "Subtotal"))
        self.label_Total.setText(_translate("ReportStockVal", "Total Cantidad:"))
        self.label_Total_Euro.setText(_translate("ReportStockVal", "Total Importe:"))
        self.Button_PDF.setText(_translate("ReportStockVal", "PDF"))
        self.Button_Excel.setText(_translate("ReportStockVal", "Excel"))

# Function to load data in table
    def loaddata(self):
        commands_supplies = f"""
                            SELECT supplies."reference", supplies."description", supplies."unit_value", ROUND(CAST(supplies."physical_stock" AS NUMERIC), 2) AS rounded_stock, measure_unit."measure_unit", supplies."unit_value" * supplies."physical_stock" AS subtotal
                            FROM purch_fact.supplies AS supplies
                            INNER JOIN purch_fact.measure_units AS measure_unit ON supplies."m_unit_id" = measure_unit."id"
                            WHERE supplies."physical_stock">0
                            ORDER BY supplies."reference"
                            """
        conn = None

        try:
        # read the connection parameters
            params = config()
        # connect to the PostgreSQL server
            conn = psycopg2.connect(**params)
            cur = conn.cursor()
        # execution of commands one by one
            cur.execute(commands_supplies)
            results=cur.fetchall()
        # close communication with the PostgreSQL database server
            cur.close()
        # commit the changes
            conn.commit()

            self.tableWidget.setRowCount(len(results))
            tablerow=0

            font = QtGui.QFont()
            font.setPointSize(int(14//1.5))

        # fill the Qt Table with the query results
            for row in results:
                for column in range(6):
                    value = row[column]
                    if value is None:
                        value = ''
                    it = QtWidgets.QTableWidgetItem(str(value))
                    it.setFlags(it.flags() & ~QtCore.Qt.ItemFlag.ItemIsEditable)
                    it.setFont(font)
                    self.tableWidget.setItem(tablerow, column, it)

                self.tableWidget.setItemDelegateForRow(tablerow, AlignDelegate(self.tableWidget))
                tablerow+=1

            self.tableWidget.verticalHeader().hide()
            self.tableWidget.setSortingEnabled(False)
            self.tableWidget.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.ResizeMode.Stretch)

            self.df = pd.DataFrame(results, columns=["Referencia", "Descripción", "Val. Un.", "Stock", "Un.", "Subtotal"])

        except (Exception, psycopg2.DatabaseError) as error:
            dlg = QtWidgets.QMessageBox()
            new_icon = QtGui.QIcon()
            new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
            dlg.setWindowIcon(new_icon)
            dlg.setWindowTitle("ERP EIPSA")
            dlg.setText("Ha ocurrido el siguiente error:\n"
                        + str(error))
            dlg.setIcon(QtWidgets.QMessageBox.Icon.Critical)
            dlg.exec()
            del dlg, new_icon
        finally:
            if conn is not None:
                conn.close()

        self.calculate_totalqty()

# Function to calculate total
    def calculate_totalqty(self):
            locale.setlocale(locale.LC_ALL, '')
            total = 0
            for row in range(self.tableWidget.rowCount()):
                item = self.tableWidget.item(row, 3)
                if item is not None:
                    value = item.text()
                    total += float(value)
            total = locale.format_string("%.2f", total, grouping=True)
            self.label_TotalValue.setText(total)


            total_euro = 0
            for row in range(self.tableWidget.rowCount()):
                item = self.tableWidget.item(row, 5)
                if item is not None:
                    value = item.text()
                    value=value.replace(".","")
                    value=value.replace(",",".")
                    value=value[:value.find(" €")]
                    total_euro += float(value)
            total_euro = locale.format_string("%.2f", total_euro, grouping=True)
            total_euro = total_euro + " €"
            self.label_TotalValue_Euro.setText(total_euro)

# Function to generate PDF
    def generate_pdf(self):
        from PDF_Styles import stock_valoration
        from tkinter.filedialog import asksaveasfilename

        pdf = stock_valoration()
        pdf.add_font('DejaVuSansCondensed', '', os.path.abspath(os.path.join(basedir, "Resources/Iconos/DejaVuSansCondensed.ttf")))
        pdf.add_font('DejaVuSansCondensed-Bold', '', os.path.abspath(os.path.join(basedir, "Resources/Iconos/DejaVuSansCondensed-Bold.ttf")))
        pdf.set_auto_page_break(auto=True, margin=0.5)
        pdf.add_page()
        pdf.alias_nb_pages()
        pdf.set_font('DejaVuSansCondensed', '', 8)

        for row in range(self.tableWidget.rowCount()):
            reference_text=self.tableWidget.item(row, 0).text()
            description_text=self.tableWidget.item(row, 1).text()
            val_text=self.tableWidget.item(row, 2).text()
            stock_text=self.tableWidget.item(row, 3).text()
            unit_text=self.tableWidget.item(row, 4).text()
            subtotal_text=self.tableWidget.item(row, 5).text()

            y_position = pdf.get_y()
            pdf.set_line_width(0.01)
            pdf.line(1, y_position, 20, y_position)

            if len(reference_text) > 30:
                x_position = pdf.get_x()
                pdf.multi_cell(5, 0.25, reference_text, align='L')
                pdf.set_y(y_position)
                pdf.set_x(x_position + 5)
            else:
                pdf.cell(5, 0.5, reference_text, align='L')

            if len(description_text) > 50:
                x_position = pdf.get_x()
                pdf.multi_cell(7.5, 0.25, description_text, align='L')
                pdf.set_y(y_position)
                pdf.set_x(x_position + 7.5)
            else:
                pdf.cell(7.5, 0.5, description_text, align='L')

            pdf.cell(1.5, 0.5, val_text, align='R')
            pdf.cell(1.5, 0.5, stock_text, align='C')
            pdf.cell(0.5, 0.5, unit_text, align='C')
            pdf.cell(3, 0.5, subtotal_text, align='R')
            pdf.ln(0.5)

        pdf.ln(0.5)
        pdf.set_font('DejaVuSansCondensed', '', 14)
        pdf.cell(5.5, 0.5, '')
        pdf.set_fill_color(247, 181, 128)
        pdf.cell(7, 0.5, 'Total Valorado: ' + self.label_TotalValue_Euro.text(), fill=True, align='C')
        
        pdf_buffer = pdf.output()

        temp_file_path = os.path.abspath(os.path.join(os.path.abspath(os.path.join(basedir, "Resources/pdfviewer/temp", "temp.pdf"))))

        with open(temp_file_path, "wb") as temp_file:
            temp_file.write(pdf_buffer)

        pdf.close()

        self.pdf_viewer.open(QUrl.fromLocalFile(temp_file_path))  # Open PDF on viewer
        self.pdf_viewer.showMaximized()

# Function to generate Excel
    def generate_excel(self):
        output_path = asksaveasfilename(defaultextension=".xlsx", filetypes=[("Archivos Excel", "*.xlsx")], title="Guardar Excel")

        self.df["Val. Un."] = self.df["Val. Un."].apply(self.euros_to_float)
        self.df["Subtotal"] = self.df["Subtotal"].apply(self.euros_to_float)

        if output_path:
            wb = Workbook()
            ws = wb.active

            # Add data to Excel
            for r_idx, row in enumerate(dataframe_to_rows(self.df, index=False, header=True), 1):
                ws.append(row)

            # Currency Style
            currency_style = NamedStyle(name='currency', number_format='#,##0.00 €')

            # Apply Currency Style
            for cell in ws['C']:
                cell.style = currency_style

            for cell in ws['F']:
                cell.style = currency_style

            # Save Excel
            wb.save(output_path)

# Function to transform euros to float values
    def euros_to_float(self, value):
        value = value.replace(".", "")
        value = value.replace(",", ".")
        value = value[: value.find(" €")]
        return float(value)



if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    ReportStockVal = QtWidgets.QMainWindow()
    ui = Ui_StockVal_Window()
    ui.setupUi(ReportStockVal)
    ReportStockVal.show()
    sys.exit(app.exec())
