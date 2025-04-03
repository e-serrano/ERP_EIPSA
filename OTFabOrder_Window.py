# Form implementation generated from reading ui file 'OTFabOrder_Window.ui'
#
# Created by: PyQt6 UI code generator 6.4.2
#
# WARNING: Any manual changes made to this file will be lost when pyuic6 is
# run again.  Do not edit this file unless you know what you are doing.


from PyQt6 import QtCore, QtGui, QtWidgets
from config import config
import psycopg2
from datetime import *
import os
import pandas as pd
from PDF_Styles import fab_order
from openpyxl import Workbook, load_workbook
from openpyxl.styles import NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows
from tkinter.filedialog import asksaveasfilename

basedir = r"\\nas01\DATOS\Comunes\EIPSA-ERP"


class AlignDelegate(QtWidgets.QStyledItemDelegate):
    """
    A custom item delegate for aligning cell content in a QTableView or QTableWidget to the center.

    Inherits from:
        QtWidgets.QStyledItemDelegate: Provides custom rendering and editing for table items.

    """
    def initStyleOption(self, option, index):
        """
        Initializes the style option for the item, setting its display alignment to center.

        Args:
            option (QtWidgets.QStyleOptionViewItem): The style option to initialize.
            index (QtCore.QModelIndex): The model index of the item.
        """
        super(AlignDelegate, self).initStyleOption(option, index)
        option.displayAlignment = QtCore.Qt.AlignmentFlag.AlignCenter


class Ui_OTFabOrder_Window(object):
    """
    UI class for the OT order for fabrication window.
    """
    def __init__(self, df_data, id_list, model, variable):
        """
        Initializes the fabrication order window UI with the given data, ID list, model, and variable.

        Parameters:
        - df_data (DataFrame): Data related to the orders.
        - id_list (list): List of order IDs.
        - model (Any): The UI model used for displaying the orders.
        - variable (Any): A variable to manage the logic for this specific window.
        """
        self.df_data = df_data
        self.id_list = id_list
        self.model = model
        self.variable = variable

    def setupUi(self, OTFabOrder_Window):
        """
        Sets up the user interface for the OTFabOrder_Window.

        Args:
            OTFabOrder_Window (QtWidgets.QMainWindow): The main window for the UI setup.
        """
        OTFabOrder_Window.setObjectName("OTFabOrder_Window")
        OTFabOrder_Window.resize(400, 561)
        OTFabOrder_Window.setMinimumSize(QtCore.QSize(600, 575))
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
        OTFabOrder_Window.setWindowIcon(icon)
        OTFabOrder_Window.setStyleSheet("QWidget {\n"
"background-color: rgb(255, 255, 255);\n"
"}\n"
"\n"
".QFrame {\n"
"    border: 2px solid black;\n"
"}\n"
"\n"
"QPushButton {\n"
"background-color: #33bdef;\n"
"  border: 1px solid transparent;\n"
"  border-radius: 3px;\n"
"  color: #fff;\n"
"  font-family: -apple-system,system-ui,\"Segoe UI\",\"Liberation Sans\",sans-serif;\n"
"  font-size: 15px;\n"
"  font-weight: 800;\n"
"  line-height: 1.15385;\n"
"  margin: 0;\n"
"  outline: none;\n"
"  padding: 8px .8em;\n"
"  text-align: center;\n"
"  text-decoration: none;\n"
"  vertical-align: baseline;\n"
"  white-space: nowrap;\n"
"}\n"
"\n"
"QPushButton:hover {\n"
"    background-color: #019ad2;\n"
"    border-color: rgb(0, 0, 0);\n"
"}\n"
"\n"
"QPushButton:pressed {\n"
"    background-color: rgb(1, 140, 190);\n"
"    border-color: rgb(255, 255, 255);\n"
"}")
        self.centralwidget = QtWidgets.QWidget(parent=OTFabOrder_Window)
        self.centralwidget.setObjectName("centralwidget")
        self.gridLayout = QtWidgets.QGridLayout(self.centralwidget)
        self.gridLayout.setObjectName("gridLayout")
        self.frame = QtWidgets.QFrame(parent=self.centralwidget)
        self.frame.setFrameShape(QtWidgets.QFrame.Shape.StyledPanel)
        self.frame.setFrameShadow(QtWidgets.QFrame.Shadow.Raised)
        self.frame.setObjectName("frame")
        self.gridLayout_2 = QtWidgets.QGridLayout(self.frame)
        self.gridLayout_2.setObjectName("gridLayout_2")
        spacerItem2 = QtWidgets.QSpacerItem(20, 10, QtWidgets.QSizePolicy.Policy.Minimum, QtWidgets.QSizePolicy.Policy.Fixed)
        self.gridLayout_2.addItem(spacerItem2, 0, 0, 1, 2)
        self.Button_Cancel = QtWidgets.QPushButton(parent=self.frame)
        self.Button_Cancel.setMinimumSize(QtCore.QSize(100, 35))
        self.Button_Cancel.setMaximumSize(QtCore.QSize(100, 35))
        self.Button_Cancel.setObjectName("Button_Cancel")
        self.gridLayout_2.addWidget(self.Button_Cancel, 1, 0, 1, 1)
        self.Button_Create = QtWidgets.QPushButton(parent=self.frame)
        self.Button_Create.setMinimumSize(QtCore.QSize(120, 35))
        self.Button_Create.setMaximumSize(QtCore.QSize(120, 35))
        self.Button_Create.setObjectName("Button_Create")
        self.gridLayout_2.addWidget(self.Button_Create, 1, 1, 1, 1)
        self.Button_Launch = QtWidgets.QPushButton(parent=self.frame)
        self.Button_Launch.setMinimumSize(QtCore.QSize(120, 35))
        self.Button_Launch.setMaximumSize(QtCore.QSize(120, 35))
        self.Button_Launch.setObjectName("Button_Launch")
        self.gridLayout_2.addWidget(self.Button_Launch, 1, 2, 1, 1)
        spacerItem3 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Policy.Expanding, QtWidgets.QSizePolicy.Policy.Minimum)
        self.gridLayout_2.addItem(spacerItem3, 1, 3, 1, 1)
        spacerItem4 = QtWidgets.QSpacerItem(20, 10, QtWidgets.QSizePolicy.Policy.Minimum, QtWidgets.QSizePolicy.Policy.Fixed)
        self.gridLayout_2.addItem(spacerItem4, 2, 0, 1, 2)
        headers_labels = ["Id OT", "Pedido Tipo Tag", "Elemento", "Cantidad", "OT",
                            "Fecha", "CantxOT", "Trad", "Plano Fabricación"]
        self.hLayout2 = QtWidgets.QHBoxLayout()
        self.hLayout2.setObjectName("hLayout2")
        self.hLayout2.setSpacing(0)
        self.gridLayout_2.addLayout(self.hLayout2, 3, 0, 1, 4)
        self.tableOT = QtWidgets.QTableWidget(parent=self.frame)
        self.tableOT.setObjectName("tableWidget")
        self.tableOT.setColumnCount(9)
        self.tableOT.setRowCount(0)
        for i in range(9):
            item = QtWidgets.QTableWidgetItem()
            font = QtGui.QFont()
            font.setPointSize(10)
            font.setBold(True)
            item.setFont(font)
            self.tableOT.setHorizontalHeaderItem(i, item)
        self.gridLayout_2.addWidget(self.tableOT, 4, 0, 1, 4)
        self.gridLayout.addWidget(self.frame, 0, 0, 1, 1)
        OTFabOrder_Window.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(parent=OTFabOrder_Window)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 400, 22))
        self.menubar.setObjectName("menubar")
        OTFabOrder_Window.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(parent=OTFabOrder_Window)
        self.statusbar.setObjectName("statusbar")
        OTFabOrder_Window.setStatusBar(self.statusbar)
        self.tableOT.setSortingEnabled(True)
        self.tableOT.horizontalHeader().setStyleSheet("QHeaderView::section {background-color: #33bdef; border: 1px solid black;}")
        # OTFabOrder_Window.setWindowFlag(QtCore.Qt.WindowType.WindowCloseButtonHint, False)

        self.retranslateUi(OTFabOrder_Window)
        QtCore.QMetaObject.connectSlotsByName(OTFabOrder_Window)

        self.tableOT.setRowCount(self.df_data.shape[0])
        tablerow=0

        for row_idx, row_data in self.df_data.iterrows():
            for col_idx, cell_data in enumerate(row_data):
                item = QtWidgets.QTableWidgetItem(str(cell_data))
                item.setFlags(item.flags() & ~QtCore.Qt.ItemFlag.ItemIsEditable)
                self.tableOT.setItem(row_idx, col_idx, item)

            self.tableOT.setItemDelegateForRow(tablerow, AlignDelegate(self.tableOT))
            tablerow+=1

        self.tableOT.verticalHeader().hide()
        self.tableOT.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.ResizeMode.ResizeToContents)
        self.tableOT.horizontalHeader().setSectionResizeMode(8, QtWidgets.QHeaderView.ResizeMode.Stretch)
        self.tableOT.setHorizontalHeaderLabels(headers_labels)

        self.Button_Cancel.clicked.connect(OTFabOrder_Window.close)
        self.Button_Create.clicked.connect(self.create_ot)
        self.Button_Launch.clicked.connect(self.launch_ot)


# Function to translate and updates the text of various UI elements
    def retranslateUi(self, OTFabOrder_Window):
        """
        Translates and updates the text of various UI elements.
        """
        _translate = QtCore.QCoreApplication.translate
        OTFabOrder_Window.setWindowTitle(_translate("OTFabOrder_Window", "Orden de Fabricación"))
        self.Button_Cancel.setText(_translate("OTFabOrder_Window", "Salir"))
        self.Button_Create.setText(_translate("OTFabOrder_Window", "Crear OT"))
        self.Button_Launch.setText(_translate("OTFabOrder_Window", "Lanzar OT"))

# Function to create OT and asign number
    def create_ot(self):
        """
        Creates a new fabrication order (OT) by generating the next order number 
        and populating the order table with relevant data.
        """
        commands_numot = ("""SELECT "ot_num"
                        FROM fabrication.fab_order
                        WHERE NOT "ot_num" LIKE '90%'
                        ORDER BY "ot_num" ASC
                        """)
        conn = None
        try:
        # read the connection parameters
            params = config()
        # connect to the PostgreSQL server
            conn = psycopg2.connect(**params)
            cur = conn.cursor()
        # execution of commands
            cur.execute(commands_numot)
            results=cur.fetchall()
            self.num_ot=results[-1][0]

            excel_file_path = r"\\nas01\DATOS\Comunes\EIPSA Sistemas de Gestion\MasterCTF\Bases\Contador.xlsm"
            workbook = load_workbook(excel_file_path, keep_vba=True)
            worksheet = workbook.active
            self.num_ot = worksheet['B2'].value
            self.num_ot = '{:06}'.format(int(self.num_ot) + 1)
        # close communication with the PostgreSQL database server
            cur.close()
        # commit the changes
            conn.commit()
        except (Exception, psycopg2.DatabaseError) as error:
            dlg = QtWidgets.QMessageBox()
            new_icon = QtGui.QIcon()
            new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
            dlg.setWindowIcon(new_icon)
            dlg.setWindowTitle("ERP EIPSA")
            dlg.setText("Ha ocurrido el siguiente error:\n"
                        + str(error))
            dlg.setIcon(QtWidgets.QMessageBox.Icon.Critical)
            dlg.exec()
            del dlg, new_icon
        finally:
            if conn is not None:
                conn.close()

        it = QtWidgets.QTableWidgetItem(str(self.num_ot))
        it.setFlags(it.flags() & ~QtCore.Qt.ItemFlag.ItemIsEditable)
        self.tableOT.setItem(0, 4, it)
        it = QtWidgets.QTableWidgetItem(str(date.today().strftime("%d/%m/%Y")))
        it.setFlags(it.flags() & ~QtCore.Qt.ItemFlag.ItemIsEditable)
        self.tableOT.setItem(0, 5, it)

        for i in range(1,self.tableOT.rowCount()):
            if self.tableOT.item(i, 2).text() == self.tableOT.item(i - 1, 2).text():
                it = QtWidgets.QTableWidgetItem(str(self.num_ot))
                it.setFlags(it.flags() & ~QtCore.Qt.ItemFlag.ItemIsEditable)
                self.tableOT.setItem(i, 4, it)
            else:
                self.num_ot = '{:06}'.format(int(self.num_ot) + 1)
                it = QtWidgets.QTableWidgetItem(str(self.num_ot))
                it.setFlags(it.flags() & ~QtCore.Qt.ItemFlag.ItemIsEditable)
                self.tableOT.setItem(i, 4, it)
            it = QtWidgets.QTableWidgetItem(str(date.today().strftime("%d/%m/%Y")))
            it.setFlags(it.flags() & ~QtCore.Qt.ItemFlag.ItemIsEditable)
            self.tableOT.setItem(i, 5, it)

        data_elements = []
        for row in range(self.tableOT.rowCount()):
            data_elements.append([self.tableOT.item(row, 2).text(), float(self.tableOT.item(row, 3).text())])

        df = pd.DataFrame(data_elements)
        df = df.groupby([0])[1].sum().reset_index()

        for row in range(self.tableOT.rowCount()):
            cant_ot = df[df.iloc[:, 0] == self.tableOT.item(row, 2).text()].iloc[:, 1].values[0]
            it = QtWidgets.QTableWidgetItem(str(round(cant_ot,3)))
            it.setFlags(it.flags() & ~QtCore.Qt.ItemFlag.ItemIsEditable)
            self.tableOT.setItem(row, 6, it)

        self.tableOT.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.ResizeMode.ResizeToContents)
        self.tableOT.horizontalHeader().setSectionResizeMode(8, QtWidgets.QHeaderView.ResizeMode.Stretch)

# Function to launch OT to database and generate document
    def launch_ot(self):
        """
        Launches a fabrication order (OT) by processing data from the model and 
        populating the order table with relevant attributes. Exports the final order table data to an Excel file, applying specific styles and formatting
        """
        data_of = []
        data_trad = []

        for element in self.id_list:
            for row in range(self.model.rowCount()):
                if self.model.data(self.model.index(row, 0)) == element:
                    target_row = row
                    break
            if target_row is not None:
                ped_type_tag = self.model.data(self.model.index(target_row, 4)) + '-' + self.model.data(self.model.index(target_row, 8)) + '-' + self.model.data(self.model.index(target_row, 1))
                if self.variable == 'Caudal':
                    num_of_plate = self.model.data(self.model.index(target_row, 48)) 
                    codefab_eq = self.model.data(self.model.index(target_row, 73))
                    trad_eq = self.model.data(self.model.index(target_row, 74))
                    codefab_orifice_flange = self.model.data(self.model.index(target_row, 76))
                    trad_orifice_flange = self.model.data(self.model.index(target_row, 113))
                    codefab_line_flange = self.model.data(self.model.index(target_row, 79))
                    trad_line_flange = self.model.data(self.model.index(target_row, 114))
                    codefab_gasket = self.model.data(self.model.index(target_row, 82))
                    trad_gasket = self.model.data(self.model.index(target_row, 115))
                    codefab_bolts = self.model.data(self.model.index(target_row, 85))
                    trad_bolts = self.model.data(self.model.index(target_row, 116))
                    codefab_plugs = self.model.data(self.model.index(target_row, 88))
                    trad_plugs = self.model.data(self.model.index(target_row, 117))
                    codefab_extractor = self.model.data(self.model.index(target_row, 91))
                    trad_extractor = self.model.data(self.model.index(target_row, 118))
                    codefab_plate = self.model.data(self.model.index(target_row, 94))
                    trad_plate = self.model.data(self.model.index(target_row, 119))
                    codefab_nipple = self.model.data(self.model.index(target_row, 97))
                    trad_nipple = self.model.data(self.model.index(target_row, 120))
                    codefab_handle = self.model.data(self.model.index(target_row, 100))
                    trad_handle = self.model.data(self.model.index(target_row, 121))
                    codefab_chring = self.model.data(self.model.index(target_row, 104))
                    trad_chring = self.model.data(self.model.index(target_row, 122))
                    codefab_tube = self.model.data(self.model.index(target_row, 106))
                    trad_tube = self.model.data(self.model.index(target_row, 123))
                    codefab_piece2 = self.model.data(self.model.index(target_row, 109))
                    trad_piece2 = self.model.data(self.model.index(target_row, 124))
                    list_of = [num_of_plate]
                    list_trad = [codefab_eq, trad_eq, codefab_orifice_flange, trad_orifice_flange, codefab_line_flange,
                                        trad_line_flange, codefab_gasket, trad_gasket, codefab_bolts, trad_bolts,
                                        codefab_plugs, trad_plugs, codefab_extractor, trad_extractor, codefab_plate,
                                        trad_plate, codefab_nipple, trad_nipple, codefab_handle, trad_handle,
                                        codefab_chring, trad_chring, codefab_tube, trad_tube, codefab_piece2, trad_piece2]
                elif self.variable == 'Temperatura':
                    num_of_sensor = self.model.data(self.model.index(target_row, 58))
                    num_of = self.model.data(self.model.index(target_row, 62))
                    codefab_eq = self.model.data(self.model.index(target_row, 81))
                    trad_eq = self.model.data(self.model.index(target_row,82))
                    codefab_bar = self.model.data(self.model.index(target_row, 84))
                    trad_bar = self.model.data(self.model.index(target_row, 120))
                    codefab_tube = self.model.data(self.model.index(target_row, 87))
                    trad_tube = self.model.data(self.model.index(target_row, 121))
                    codefab_flange = self.model.data(self.model.index(target_row, 90))
                    trad_flange = self.model.data(self.model.index(target_row, 122))
                    codefab_sensor = self.model.data(self.model.index(target_row, 93))
                    trad_sensor = self.model.data(self.model.index(target_row, 123))
                    codefab_head = self.model.data(self.model.index(target_row, 96))
                    trad_head = self.model.data(self.model.index(target_row, 124))
                    codefab_btb = self.model.data(self.model.index(target_row, 99))
                    trad_btb = self.model.data(self.model.index(target_row, 125))
                    codefab_nipple = self.model.data(self.model.index(target_row, 102))
                    trad_nipple = self.model.data(self.model.index(target_row, 126))
                    codefab_spring = self.model.data(self.model.index(target_row, 105))
                    trad_spring = self.model.data(self.model.index(target_row, 127))
                    codefab_puntal = self.model.data(self.model.index(target_row, 108))
                    trad_puntal = self.model.data(self.model.index(target_row, 128))
                    codefab_plug = self.model.data(self.model.index(target_row, 111))
                    trad_plug = self.model.data(self.model.index(target_row, 129))
                    codefab_tw = self.model.data(self.model.index(target_row, 114))
                    trad_tw = self.model.data(self.model.index(target_row, 130))
                    codefab_cable = self.model.data(self.model.index(target_row, 117))
                    trad_cable = self.model.data(self.model.index(target_row, 131))
                    list_of = [num_of, num_of_sensor]
                    list_trad = [codefab_eq, trad_eq, codefab_bar, trad_bar, codefab_tube,
                                        trad_tube, codefab_flange, trad_flange, codefab_sensor, trad_sensor,
                                        codefab_head, trad_head, codefab_btb, trad_btb, codefab_nipple,
                                        trad_nipple, codefab_spring, trad_spring, codefab_puntal, trad_puntal,
                                        codefab_plug, trad_plug, codefab_tw, trad_tw, codefab_cable, trad_cable]

                elif self.variable == 'Nivel':
                    num_of = self.model.data(self.model.index(target_row, 52))
                    codefab_eq = self.model.data(self.model.index(target_row, 67))
                    trad_eq = self.model.data(self.model.index(target_row, 68))
                    codefab_body = self.model.data(self.model.index(target_row, 70))
                    trad_body = self.model.data(self.model.index(target_row, 121))
                    codefab_cover = self.model.data(self.model.index(target_row, 73))
                    trad_cover = self.model.data(self.model.index(target_row, 122))
                    codefab_stud = self.model.data(self.model.index(target_row, 76))
                    trad_stud = self.model.data(self.model.index(target_row, 123))
                    codefab_niphex = self.model.data(self.model.index(target_row, 79))
                    trad_niphex= self.model.data(self.model.index(target_row, 124))
                    codefab_valve = self.model.data(self.model.index(target_row, 82))
                    trad_valve = self.model.data(self.model.index(target_row, 125))
                    codefab_flange = self.model.data(self.model.index(target_row, 85))
                    trad_flange = self.model.data(self.model.index(target_row, 126))
                    codefab_dv = self.model.data(self.model.index(target_row, 88))
                    trad_dv = self.model.data(self.model.index(target_row, 127))
                    codefab_scale = self.model.data(self.model.index(target_row, 91))
                    trad_scale = self.model.data(self.model.index(target_row, 128))
                    codefab_illum = self.model.data(self.model.index(target_row, 94))
                    trad_illum = self.model.data(self.model.index(target_row, 129))
                    codefab_gasketglass = self.model.data(self.model.index(target_row, 97))
                    trad_gasketglass = self.model.data(self.model.index(target_row, 130))
                    codefab_glass = self.model.data(self.model.index(target_row, 100))
                    trad_glass = self.model.data(self.model.index(target_row, 131))
                    codefab_float = self.model.data(self.model.index(target_row, 103))
                    trad_float = self.model.data(self.model.index(target_row, 132))
                    codefab_mica = self.model.data(self.model.index(target_row, 106))
                    trad_mica = self.model.data(self.model.index(target_row, 133))
                    codefab_flags = self.model.data(self.model.index(target_row, 109))
                    trad_flags = self.model.data(self.model.index(target_row, 134))
                    codefab_gasketflange = self.model.data(self.model.index(target_row, 112))
                    trad_gasketflange = self.model.data(self.model.index(target_row, 135))
                    codefab_niptub = self.model.data(self.model.index(target_row, 115))
                    trad_niptub = self.model.data(self.model.index(target_row, 136))
                    codefab_antifrost = self.model.data(self.model.index(target_row, 118))
                    trad_antifrost = self.model.data(self.model.index(target_row, 137))
                    list_of = [num_of]
                    list_trad = [codefab_eq, trad_eq, codefab_body, trad_body, codefab_cover,
                                        trad_cover, codefab_stud, trad_stud, codefab_niphex, trad_niphex,
                                        codefab_valve, trad_valve, codefab_flange, trad_flange, codefab_dv,
                                        trad_dv, codefab_scale, trad_scale, codefab_illum, trad_illum,
                                        codefab_gasketglass, trad_gasketglass, codefab_glass, trad_glass, codefab_float,
                                        trad_float, codefab_mica, trad_mica, codefab_flags, trad_flags,
                                        codefab_gasketflange, trad_gasketflange, codefab_niptub, trad_niptub, codefab_antifrost, trad_antifrost]

                list_of.insert(0, ped_type_tag)
                data_of.append(list_of)
                list_trad.insert(0, ped_type_tag)
                data_trad.append(list_trad)

        df_trad = pd.DataFrame(data_trad)
        for row in range (self.tableOT.rowCount()):
            rows_pedtypetag = df_trad[df_trad.iloc[:, 0] == self.tableOT.item(row, 1).text()]
            for idx, row_df in rows_pedtypetag.iterrows():
                for column_codefab, value in row_df.items():
                    if value == self.tableOT.item(row, 2).text():
                        trad_column = df_trad.columns[df_trad.columns.get_loc(column_codefab) + 1]
                        code_trad = row_df[trad_column]
                it = QtWidgets.QTableWidgetItem(str(code_trad))
                it.setFlags(it.flags() & ~QtCore.Qt.ItemFlag.ItemIsEditable)
                self.tableOT.setItem(row, 7, it)

        df_of = pd.DataFrame(data_of)
        for row in range (self.tableOT.rowCount()):
            if any(value in self.tableOT.item(row, 7).text() for value in ['TE', 'PT100']):
                of_drawing = df_of[df_of.iloc[:, 0] == self.tableOT.item(row, 1).text()].iloc[:, 2].values[0]
            else:
                of_drawing = df_of[df_of.iloc[:, 0] == self.tableOT.item(row, 1).text()].iloc[:, 1].values[0]
            it = QtWidgets.QTableWidgetItem(str(of_drawing))
            it.setFlags(it.flags() & ~QtCore.Qt.ItemFlag.ItemIsEditable)
            self.tableOT.setItem(row, 8, it)

        for row in range (self.tableOT.rowCount()):
            check_ot = f"SELECT * FROM fabrication.fab_order WHERE id = '{self.tableOT.item(row, 0).text()}'"
            commands_newot = ("""
                            INSERT INTO fabrication.fab_order (
                            "id","tag","element","qty_element",
                            "ot_num","qty_ot","start_date")
                            VALUES (%s,%s,%s,%s,%s,%s,%s)
                            """)
            commands_updateot = ("""
                                UPDATE fabrication.fab_order SET
                                "tag" = %s, "element" = %s, "qty_element" = %s,
                                "ot_num" = %s, "qty_ot" = %s,"start_date" = %s
                                WHERE "id" = %s
                                """)
            conn = None
            try:
            # read the connection parameters
                params = config()
            # connect to the PostgreSQL server
                conn = psycopg2.connect(**params)
                cur = conn.cursor()
            # execution of commands
                cur.execute(check_ot)
                results=cur.fetchall()
                if len(results) == 0:
                    data=(self.tableOT.item(row, 0).text(), self.tableOT.item(row, 1).text() + " // " + self.tableOT.item(row, 8).text(), self.tableOT.item(row, 2).text(),
                        self.tableOT.item(row, 3).text(), self.tableOT.item(row, 4).text(), self.tableOT.item(row, 6).text(), self.tableOT.item(row, 5).text())
                    cur.execute(commands_newot, data)
                else:
                    data=(self.tableOT.item(row, 1).text() + " // " + self.tableOT.item(row, 8).text(), self.tableOT.item(row, 2).text(), self.tableOT.item(row, 3).text(),
                        self.tableOT.item(row, 4).text(), self.tableOT.item(row, 6).text(), self.tableOT.item(row, 5).text(), self.tableOT.item(row, 0).text())
                    cur.execute(commands_updateot, data)

            # close communication with the PostgreSQL database server
                cur.close()
            # commit the changes
                conn.commit()
            except (Exception, psycopg2.DatabaseError) as error:
                dlg = QtWidgets.QMessageBox()
                new_icon = QtGui.QIcon()
                new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
                dlg.setWindowIcon(new_icon)
                dlg.setWindowTitle("ERP EIPSA")
                dlg.setText("Ha ocurrido el siguiente error:\n"
                            + str(error))
                dlg.setIcon(QtWidgets.QMessageBox.Icon.Critical)
                dlg.exec()
                del dlg, new_icon
            finally:
                if conn is not None:
                    conn.close()

        table_data = []

        for row in range(self.tableOT.rowCount()):
            row_data = []
            for col in range(self.tableOT.columnCount()):
                item = self.tableOT.item(row, col)
                row_data.append(item.text() if item is not None else "")
            table_data.append(row_data)

        df_toexport = pd.DataFrame(table_data, columns=['ID', 'TAG','ELEMENTO','CANT','OT','FECHA','CANTxOT','TRAD COD','OF'])

        output_path = asksaveasfilename(defaultextension=".xlsx", filetypes=[("Archivos Excel", "*.xlsx")], title="Guardar Excel")

        if output_path:
            wb = Workbook()
            ws = wb.active

            # Add data to Excel
            for index, row in df_toexport.iterrows():
                fecha_str = row['FECHA']
                if fecha_str is not None:
                    fecha_obj = datetime.strptime(fecha_str, '%d/%m/%Y').date()
                    df_toexport.at[index, 'FECHA'] = fecha_obj

                float_str = row['CANTxOT']
                if float_str is not None:
                    float_obj = float(float_str)
                    df_toexport.at[index, 'CANTxOT'] = float_obj

                float_str = row['CANT']
                if float_str is not None:
                    float_obj = float(float_str)
                    df_toexport.at[index, 'CANT'] = float_obj

            for r_idx, row in enumerate(dataframe_to_rows(df_toexport, index=False, header=True), 1):
                ws.append(row)

            # Currency Style
            currency_style = NamedStyle(name='currency', number_format='#,##0.00 €')
            date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
            otnum_style = NamedStyle(name='otnum_style', number_format='000000')

            # Apply Styles
            for cell in ws['F']:
                cell.style = date_style

            for cell in ws['E']:
                cell.style = otnum_style

            wb.save(output_path)

            excel_file_path = r"\\nas01\DATOS\Comunes\EIPSA Sistemas de Gestion\MasterCTF\Bases\Contador.xlsm"
            workbook = load_workbook(excel_file_path)
            worksheet = workbook.active
            worksheet['B2'].value = self.num_ot
            workbook.save(excel_file_path)



if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    OTFabOrder_Window = QtWidgets.QMainWindow()
    ui = Ui_OTFabOrder_Window()
    ui.setupUi(OTFabOrder_Window)
    OTFabOrder_Window.show()
    sys.exit(app.exec())