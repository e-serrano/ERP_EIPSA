import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, NamedStyle
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.utils import get_column_letter
from copy import deepcopy
from tkinter.filedialog import asksaveasfilename
from tkinter import Tk
from datetime import *
from config import config
import psycopg2
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from PyQt6 import QtGui, QtWidgets
import os
from math import exp
import re
import numpy as np


basedir = r"\\nas01\DATOS\Comunes\EIPSA-ERP"


# Templates for orders

class offer_flow:
    """
    A class to manage export offer details for flow equipments.
    
    Attributes:
        numoffer (str): Offer number.
        username (str): Name of the user creating the offer.
        rev (str): Revision number of the offer.
        project (str): Name of the project.
        delivery_term (str): Delivery terms for the offer.
        delivery_time (str): Expected delivery time.
        validity (str): Validity period of the offer.
        pay_term (str): Payment terms.
        testinspection (str): Information about testing and inspection.
        revchanges (str): Details of changes made in the revision.
        notes (str): Additional notes, split by line.
    """
    def __init__(self, numoffer, username, rev, project, delivery_term, delivery_time, validity, pay_term, testinspection, revchanges, notes):
        """
        Initializes the offer.

        Args:
            numoffer (str): Offer number.
            username (str): Name of the user creating the offer.
            rev (str): Revision number of the offer.
            project (str): Name of the project.
            delivery_term (str): Delivery terms for the offer.
            delivery_time (str): Expected delivery time.
            validity (str): Validity period of the offer.
            pay_term (str): Payment terms.
            testinspection (str): Information about testing and inspection.
            revchanges (str): Details of changes made in the revision.
            notes (str): Additional notes, split by line.
        """
        notes = notes.split('\n')
        
        date_offer = date.today().strftime("%d/%m/%Y")
        offername_commercial = numoffer + "-" + "Commercial.Rev" + rev
        offername_technical = numoffer + "-" + "Technical.Rev" + rev

        query_commercial = """
                    SELECT name, surname, email
                    FROM users_data.registration
                    WHERE username = %s
                    """
        query_dataoffer = """
                        SELECT client, num_ref_offer
                        FROM offers
                        WHERE UPPER (num_offer) LIKE UPPER('%%'||%s||'%%')
                        """
        query_tagsdata = """
                        SELECT *
                        FROM tags_data.tags_flow
                        WHERE (
                        UPPER ("num_offer") LIKE UPPER('%%'||%s||'%%')
                        AND
                        "tag_state" NOT IN ('PURCHASED','DELETED')
                        )
                        """
        conn = None
        try:
            # read the connection parameters
            params = config()
            # connect to the PostgreSQL server
            conn = psycopg2.connect(**params)
            cur = conn.cursor()

            cur.execute(query_dataoffer, (numoffer,))
            results_offer = cur.fetchall()
            client = results_offer[0][0]
            num_ref = results_offer[0][1]

            cur.execute(query_commercial, (username,))
            results_commercial = cur.fetchall()
            responsible = results_commercial[0][0] + " " + results_commercial[0][1]
            email = results_commercial[0][2]

            cur.execute(query_tagsdata, (numoffer,))
            data_tags = cur.fetchall()

            if len(data_tags) == 0:
                dlg = QtWidgets.QMessageBox()
                new_icon = QtGui.QIcon()
                new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
                dlg.setWindowIcon(new_icon)
                dlg.setWindowTitle("ERP EIPSA")
                dlg.setText("No hay TAGS importados en la oferta")
                dlg.setIcon(QtWidgets.QMessageBox.Icon.Warning)
                dlg.exec()
                del dlg, new_icon
            else:
                columns = []
                for elt in cur.description:
                    columns.append(elt[0])

                value_type_dict = {
                    "A. Chamber": 1,
                    "C. RING": 1,
                    "F": 1,
                    "F+C.RING": 1,
                    "F+P": 1,
                    "IFO": 1,
                    "M.RUN": 1,
                    "P": 1,
                    "NOZZLE BF": 2,
                    "NOZZLE BW": 2,
                    "NOZZLE F": 2,
                    "PTC-6": 2,
                    "VFM": 3,
                    "VFW": 3,
                    "VWM": 3,
                    "VWW": 3,
                    "WEDGE": 4,
                    "PITOT": 5,
                    "RO": 6,
                    "MULTISTAGE RO": 7,
                }

                # Setting the dataframe with the equipment data
                df = pd.DataFrame(data=data_tags, columns=columns)
                df = df.iloc[:, 1:32]
                df["value_type"] = df["item_type"].map(value_type_dict)
                df = df.sort_values(by=["value_type", "tag"])
                df["amount"] = df["amount"].apply(self.euros_to_float)
                total_amount_material = df["amount"].sum()
                df = df.drop([
                        "tag_state",
                        "num_offer",
                        "num_order",
                        "num_po",
                        "position",
                        "subposition",
                        "flange_type",
                        "plate_std",
                        "pipe_spec",
                        "aprox_weight"
                    ],
                    axis=1,)

                number_items = df.shape[0]
                documentation = number_items * 70

                # Loading Excel Template
                self.wb_commercial = load_workbook(
                    r"\\nas01\DATOS\Comunes\EIPSA-ERP\Plantillas Exportación\PLANTILLA OFERTA CAUDAL.xlsx"
                )

                # Editing sheet COVER
                sheet_name = "COVER"
                ws = self.wb_commercial[sheet_name]
                ws["E4"] = client
                ws["E6"] = offername_commercial
                ws["E8"] = num_ref
                ws["E10"] = project
                ws["E12"] = date_offer
                ws["E14"] = delivery_term
                ws["E16"] = validity + " days"
                ws["C43"] = responsible
                ws["C45"] = email

                # Editing sheet EQUIPMENT DATA
                sheets_confirmed = ["COVER", "1.2", "1.3", "NOTES"]

                if int(rev) == 0:
                    for item_type in df["item_type"].unique():
                        sheets_confirmed.append(item_type)

                dict_sheets_data = {}

                for value_type in df["value_type"].unique():
                    df_toexport = df[df["value_type"] == value_type]
                    df_toexport.insert(0, "N°", range(1, len(df_toexport) + 1))
                    df_toexport.set_index("N°", inplace=True, drop=False)
                    df_toexport.index.name = None
                    df_toexport = df_toexport.drop(["value_type"], axis=1)

                    eq_type = (
                        "FLOW ELEMENTS DATA"
                        if value_type == 1
                        else (
                            "NOZZLE ELEMENTS DATA"
                            if value_type == 2
                            else (
                                "VENTURI ELEMENTS DATA"
                                if value_type == 3
                                else (
                                    "WEDGE ELEMENTS DATA"
                                    if value_type == 4
                                    else (
                                        "PITOT ELEMENTS DATA"
                                        if value_type == 5
                                        else (
                                            "RO ELEMENTS DATA"
                                            if value_type == 6
                                            else "MULTISTAGE RO ELEMENTS DATA"
                                        )
                                    )
                                )
                            )
                        )
                    )

                    if eq_type not in sheets_confirmed:
                        sheets_confirmed.append(eq_type)

                    if eq_type == "FLOW ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "tube_material",
                                "plate_type",
                                "valve_conn",
                                "valve_material_body",
                                "stages_number",
                                "aprox_length"],
                            axis=1,)
                    elif eq_type == "NOZZLE ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "plate_type",
                                "plate_thk",
                                "gasket_material",
                                "valve_conn",
                                "valve_material_body",
                                "stages_number"],
                            axis=1,)
                    elif (eq_type == "VENTURI ELEMENTS DATA" or eq_type == "WEDGE ELEMENTS DATA"):
                        df_toexport = df_toexport.drop([
                                "plate_type",
                                "plate_thk",
                                "gasket_material",
                                "bolts_nuts_material",
                                "valve_conn",
                                "valve_material_body",
                                "stages_number"],
                            axis=1,)
                    elif eq_type == "PITOT ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "flange_material",
                                "tube_material",
                                "tapping_num_size",
                                "plate_type",
                                "plate_thk",
                                "gasket_material",
                                "stages_number",
                                "aprox_length"],
                            axis=1,)
                    elif eq_type == "RO ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "flange_material",
                                "tube_material",
                                "tapping_num_size",
                                "gasket_material",
                                "bolts_nuts_material",
                                "valve_conn",
                                "valve_material_body",
                                "stages_number",
                                "aprox_length"],
                            axis=1,)
                    elif eq_type == "MULTISTAGE RO ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "tapping_num_size",
                                "gasket_material",
                                "bolts_nuts_material",
                                "valve_conn",
                                "valve_material_body"],
                            axis=1,)

                    ws = self.wb_commercial[eq_type]
                    ws["G2"] = date_offer
                    ws["G3"] = num_ref
                    ws["G4"] = offername_commercial
                    if revchanges != "":
                        ws["G5"] = rev + " " + revchanges
                        ws["G5"].font = Font(name="Calibri", size=14, bold=True)
                        ws["G5"].fill = PatternFill("solid", fgColor="FFFF00")

                    if int(rev) > 0:
                        for row in ws.iter_rows(min_row=2, max_row=4, min_col=6, max_col=7):
                            for cell in row:
                                cell.value = None
                                cell._style = ws["F1"]._style

                    last_row = ws.max_row

                    num_column_amount = df_toexport.columns.get_loc("amount") + 1

                    for index, row in df_toexport.iterrows():  # Data in desired row
                        for col_num, value in enumerate(row, start=1):
                            cell = ws.cell(row=last_row + 1, column=col_num)
                            cell.value = value
                            if col_num == num_column_amount:
                                cell._style = ws["X1"]._style
                            else:
                                cell._style = ws["S1"]._style

                        last_row = ws.max_row

                    if eq_type == "VENTURI ELEMENTS DATA":
                        ws[f"A{last_row+3}"] = "PRICES INCLUDE MACHINED INTEGRAL CENTRE SECTION AND ALL STRUCTURAL WELDS 100% RADIOGRAPHED"
                        ws[f"A{last_row+3}"]._style = ws["Z2"]._style
                    ws[f"A{last_row+4}"] = "OFFER VALIDITY: " + validity + " DAYS"
                    ws[f"A{last_row+4}"]._style = ws["Z1"]._style
                    ws[f"A{last_row+5}"] = (
                        "DELIVERY TIME: "
                        + delivery_time
                        + " WEEKS SINCE DRAWING / CALCULATION APPROVAL (AUGUST AND LAST TWO DECEMBER WEEKS EXCLUDED)"
                    )
                    ws[f"A{last_row+5}"]._style = ws["Z1"]._style

                    if notes != "":
                        if isinstance(notes, list):
                            line = last_row + 6
                            for note in notes:
                                ws[f"A{line}"] = note
                                ws[f"A{line}"]._style = ws["Z1"]._style
                                line += 1
                        else:
                            line = last_row + 6
                            ws[f"A{line}"] = notes
                            ws[f"A{line}"]._style = ws["Z1"]._style

                    dict_sheets_data[eq_type] = [last_row, num_column_amount, df_toexport["amount"].sum(), df_toexport.shape[0]]

                ws.cell(row=last_row + 3, column=num_column_amount - 1).value = "QTY. TOTAL"
                ws.cell(row=last_row + 3, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=last_row + 3, column=num_column_amount).value = number_items

                row_amount = last_row + 4
                for key, value in dict_sheets_data.items():
                    parts_key = key.split(" ")
                    ws.cell(row=row_amount + 2, column=num_column_amount - 1).value = "TOTAL AMOUNT OF " + parts_key[0] + " " + parts_key[1] + " (QTY: " + str(value[3]) + ")"
                    ws.cell(row=row_amount + 2, column=num_column_amount).value = value[2]
                    ws.cell(row=row_amount + 2, column=num_column_amount - 1)._style = ws["R1"]._style
                    ws.cell(row=row_amount + 2, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                    ws.cell(row=row_amount + 2, column=num_column_amount)._style = ws["T1"]._style

                    row_amount += 2

                ws.cell(row=row_amount + 2, column=num_column_amount - 1).value = "TOTAL AMOUNT OF MATERIAL"
                ws.cell(row=row_amount + 2, column=num_column_amount).value = total_amount_material
                ws.cell(row=row_amount + 4, column=num_column_amount - 1).value = "PACKING AND TRANSPORT"
                ws.cell(row=row_amount + 4, column=num_column_amount).value = (f"=MROUND({get_column_letter(num_column_amount)}{row_amount + 2}*0.03,10)" if total_amount_material > 6700 else 200 )
                ws.cell(row=row_amount + 5, column=num_column_amount - 1).value = "TESTS & INSPECTION"
                ws.cell(row=row_amount + 5, column=num_column_amount).value = float(testinspection)
                ws.cell(row=row_amount + 6, column=num_column_amount - 1).value = "DOCUMENTATION"
                ws.cell(row=row_amount + 6, column=num_column_amount).value = documentation
                ws.cell(row=row_amount + 8, column=num_column_amount - 1).value = "TOTAL AMOUNT OF BID"
                ws.cell(row=row_amount + 8, column=num_column_amount).value = f"=SUM({get_column_letter(num_column_amount)}{row_amount + 2}:{get_column_letter(num_column_amount)}{row_amount + 6})"

                ws.cell(row=last_row + 3, column=num_column_amount - 1)._style = ws["R1"]._style
                ws.cell(row=last_row + 3, column=num_column_amount).font = Font(name="Calibri", size=14)
                ws.cell(row=row_amount + 2, column=num_column_amount - 1)._style = ws["R1"]._style
                ws.cell(row=row_amount + 2, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=row_amount + 2, column=num_column_amount)._style = ws["T1"]._style
                ws.cell(row=row_amount + 4, column=num_column_amount - 1).font = Font(name="Calibri", size=14)
                ws.cell(row=row_amount + 4, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=row_amount + 4, column=num_column_amount)._style = ws["T1"]._style
                ws.cell(row=row_amount + 5, column=num_column_amount - 1).font = Font(name="Calibri", size=14)
                ws.cell(row=row_amount + 5, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=row_amount + 5, column=num_column_amount)._style = ws["T1"]._style
                ws.cell(row=row_amount + 6, column=num_column_amount - 2)._style = ws["U1"]._style
                ws.cell(row=row_amount + 6, column=num_column_amount - 1)._style = ws["U1"]._style
                ws.cell(row=row_amount + 6, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=row_amount + 6, column=num_column_amount)._style = ws["V1"]._style
                ws.cell(row=row_amount + 8, column=num_column_amount - 1)._style = ws["R1"]._style
                ws.cell(row=row_amount + 8, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=row_amount + 8, column=num_column_amount)._style = ws["W1"]._style

            # Editing sheet NOTES
                sheet_name = "NOTES"  # Selecting  sheet
                ws = self.wb_commercial[sheet_name]

                rich_string = CellRichText(
                'We are only offering measuring flow elements. Please be informed that our product range includes temperature elements, and glass and magnetic level indicators; all with european certification. (https://www.eipsa.es/en/products)\n',
                TextBlock(InlineFont(b=True), 'The prices quoted could be reduced in case of purchasing our full range of products.'))
                ws["B6"] = rich_string

                rich_string = CellRichText(
                TextBlock(InlineFont(i=True), 'Estamos ofertando solamente elementos de medida de caudal, les informamos que en nuestra gama de fabricación con certificación europea, incluye también elementos de temperatura e indicadores de nivel de vidrio y magnéticos. (https://www.eipsa.es/productos)\n'),
                TextBlock(InlineFont(b=True, i=True), 'Los precios ofertados podrían reducirse en caso de compra de toda nuestra gama.'))
                ws["B7"] = rich_string

                rich_string = CellRichText(
                'Delivery time ' + delivery_time + ' weeks since drawing / calculation approval of ',
                TextBlock(InlineFont(b=True), 'all equipment'),
                ' in the contract, as well as critical documentation (August and last two December weeks excluded).')
                ws["B12"] = rich_string

                rich_string = CellRichText(
                'Plazo de entrega ' + delivery_time + ' semanas desde aprobación de planos y cálculos de la ',
                TextBlock(InlineFont(b=True, i=True), 'totalidad de los equipos'),
                TextBlock(InlineFont(i=True),' amparados por el contrato, asi como la documentación crítica (Agosto y las dos últimas semanas de diciembre excluidos).'))
                ws["B13"] = rich_string

                rich_string = CellRichText(
                'Quotation prepared according to the information provided in the datasheet corresponding to each TAG. ',
                TextBlock(InlineFont(u='single'), 'EIPSA does not hold the final responsibility regarding selection of equipment material neither analyze process data.\n'),
                'The datasheet will be considered the only technical/contractual document, any other documentation will be considered as complementary documentation with informative purpose.')
                ws["B21"] = rich_string

                rich_string = CellRichText(
                'Esta oferta ha sido elaborada en base a la información que figura en la hoja de datos correspondiente a cada TAG, ',
                TextBlock(InlineFont(i=True, u='single'), 'no siendo responsabilidad final de EIPSA la elección del material a utilizar en los equipos ni analizar datos de proceso.\n'),
                TextBlock(InlineFont(i=True),'La hoja de datos será el único documento técnico/contractual, cualquier otra documentación recibida será considerada como documentación complementaria a efectos informativos.'))
                ws["B22"] = rich_string

                if pay_term == "100_delivery":
                    ws["B45"] = (
                        "100% of total amount of purchase order upon delivery of material according to Incoterms 2020, FCA (our facilities, Spain).\n"
                        "Payment method: bank transfer"
                    )
                    ws["B46"] = (
                        "Pago del 100% del valor total de la orden de compra a la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España).\n"
                        "Método de pago: Transferencia bancaria."
                    )
                elif pay_term == "100_order":
                    ws["B45"] = (
                        "100 % of the total amount of purchase order upon receipt of purchase order.\n"
                        "Payment method: bank transfer"
                    )
                    ws["B46"] = (
                        "Pago del 100% del valor total de la orden de compra a la recepción de la orden.\n"
                        "Método de pago: Transferencia bancaria"
                    )
                elif pay_term == "90_10":
                    ws["B45"] = (
                        "PAYMENT TERMS:\n"
                        "90 % of the total amount of PO upon delivery of material according to Incoterms 2020, FCA (our facilities, Spain) and 10% when final documentation is approved. \n"
                        "Bank Transfer: 60 days since invoice issue date."
                    )
                    ws["B46"] = (
                        "TERMINOS DE PAGO:\n"
                        "Pago del 90% del Valor total de la orden de compra a la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España) y el 10% restante cuando la documentación final sea aprobada.\n"
                        "Transferencia Bancaria: 60 días desde emisión de factura."
                    )
                elif pay_term == "50_50":
                    ws["B45"] = (
                        "50 % of the total amount of purchase order upon receipt of purchase order. Remaining 50% before material be delivered according to Incoterms 2020, FCA (our facilities, Spain).\n"
                        "Payment method: bank transfer."
                    )
                    ws["B46"] = (
                        "Pago del 50% del valor total de la orden de compra a la recepción de la orden. El 50% restante antes de la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España).\n"
                        "Método de pago: Transferencia bancaria."
                    )
                elif pay_term == "Others":
                    ws["B45"] = "PAYMENT TERMS TO BE DEFINED"
                    ws["B45"].font = Font(name="Calibri", size=11, bold=True, color="FF0000")
                    ws["B46"] = "TERMINOS DE PAGO POR DEFINIR"
                    ws["B46"].font = Font(name="Calibri", size=11, bold=True, italic=True, color="FF0000")

                rich_string = CellRichText(
                'For amounts greater than 30,000.00 € we can issue a warranty bond (if required) valid until the end of the indicated warranty period.\nBond warranty of 10% will be issued with the invoice of the last supplement.\n',
                TextBlock(InlineFont(b=True), 'For lower amounts no warranty bond is issued.'))
                ws["B48"] = rich_string

                rich_string = CellRichText(
                'Para importes superiores a 30.000,00 €, si es requerido, podremos emitir aval de garantía y estará vigente hasta el final del periodo de garantía indicado.\nEl aval del 10% será emitido con la factura del último suplemento.\n',
                TextBlock(InlineFont(i=True, b=True), 'Por debajo de dicha cantidad, no se emitirán avales.'))
                ws["B49"] = rich_string

                ws["A58"] = (
                    "If you require further information related with this offer, please do not hesitate to contact:\n"
                    + responsible
                    + "\n"
                    + email
                    + "\n"
                    "Telf.: (+34) 916.582.118"
                )

                for sheet in self.wb_commercial.sheetnames:
                    if sheet not in sheets_confirmed:
                        sheet_to_delete = self.wb_commercial[sheet]
                        self.wb_commercial.remove(sheet_to_delete)

                if int(rev) > 0:
                    sheets_confirmed = ["COVER", "1.2", "1.3", "NOTES"]
                    for sheet in sheets_confirmed:
                        sheet_to_delete = self.wb_commercial[sheet]
                        self.wb_commercial.remove(sheet_to_delete)

                left_text = "Fecha/Date: " + date_offer
                right_text = "Petición nº/Inquiry: " + num_ref

                for sheet in self.wb_commercial.worksheets:
                    sheet.oddFooter.left.text = left_text
                    sheet.oddFooter.right.text = right_text
                    sheet.oddFooter.center.text = "Page &P de &N"

                    sheet.oddFooter.left.size = 9
                    sheet.oddFooter.right.size = 9
                    sheet.oddFooter.center.size = 9
                
                path = self.save_excel_commercial()

                # Creating the technical offer using the commercial one as template
                self.wb_technical = load_workbook(path)

                if int(rev) == 0:
                    sheet_name = "COVER"
                    ws = self.wb_technical[sheet_name]
                    ws["E6"] = offername_technical

                for value_type in df["value_type"].unique():
                    eq_type = (
                        "FLOW ELEMENTS DATA"
                        if value_type == 1
                        else (
                            "NOZZLE ELEMENTS DATA"
                            if value_type == 2
                            else (
                                "VENTURI ELEMENTS DATA"
                                if value_type == 3
                                else (
                                    "WEDGE ELEMENTS DATA"
                                    if value_type == 4
                                    else (
                                        "PITOT ELEMENTS DATA"
                                        if value_type == 5
                                        else (
                                            "RO ELEMENTS DATA"
                                            if value_type == 6
                                            else "MULTISTAGE RO ELEMENTS DATA"
                                        )
                                    )
                                )
                            )
                        )
                    )

                    ws = self.wb_technical[eq_type]
                    if int(rev) == 0:
                        ws["G4"] = offername_technical

                    last_row = dict_sheets_data[eq_type][0]
                    num_column_amount = dict_sheets_data[eq_type][1]

                    # self.wb_technical[eq_type].delete_rows(last_row + 8, 20)
                    ws[f"M{last_row+5}"] = ""
                    ws[f"N{last_row+5}"] = ""

                    ws.cell(row=row_amount + 2, column=num_column_amount).value = ""
                    ws.cell(row=row_amount + 4, column=num_column_amount).value = ""
                    ws.cell(row=row_amount + 5, column=num_column_amount).value = ""
                    ws.cell(row=row_amount + 6, column=num_column_amount).value = ""
                    ws.cell(row=row_amount + 8, column=num_column_amount).value = ""

                    ws.cell(row=row_amount + 2, column=num_column_amount - 1).value = ""
                    ws.cell(row=row_amount + 4, column=num_column_amount - 1).value = ""
                    ws.cell(row=row_amount + 5, column=num_column_amount - 1).value = ""
                    ws.cell(row=row_amount + 6, column=num_column_amount - 1).value = ""
                    ws.cell(row=row_amount + 8, column=num_column_amount - 1).value = ""

                # Deleting "Amount" column
                    self.wb_technical[eq_type].delete_cols(num_column_amount, 1)

                # Adjusting the print area
                    new_last_column = num_column_amount
                    last_print_row = 40
                    nuevo_rango_impresion = f'A1:{get_column_letter(new_last_column)}{last_print_row}'
                    self.wb_technical[eq_type].print_area = nuevo_rango_impresion

                    stamp_1 = self.wb_technical[eq_type]._images[1]
                    anchor_actual = stamp_1.anchor

                    from_col, from_col_off = anchor_actual._from.col, anchor_actual._from.colOff
                    from_row, from_row_off = anchor_actual._from.row, anchor_actual._from.rowOff
                    to_col, to_col_off = anchor_actual.to.col, anchor_actual.to.colOff
                    to_row, to_row_off = anchor_actual.to.row, anchor_actual.to.rowOff

                    from_cell = AnchorMarker(col=from_col - 1, colOff=from_col_off, row=from_row, rowOff=from_row_off)
                    to_cell = AnchorMarker(col=to_col - 1, colOff=to_col_off, row=to_row, rowOff=to_row_off)

                    new_anchor = TwoCellAnchor(_from = from_cell, to = to_cell, editAs='absolute')
                    stamp_1.anchor = new_anchor

                    stamp_2 = self.wb_technical[eq_type]._images[2]
                    anchor_actual = stamp_2.anchor

                    from_col, from_col_off = anchor_actual._from.col, anchor_actual._from.colOff
                    from_row, from_row_off = anchor_actual._from.row, anchor_actual._from.rowOff
                    to_col, to_col_off = anchor_actual.to.col, anchor_actual.to.colOff
                    to_row, to_row_off = anchor_actual.to.row, anchor_actual.to.rowOff

                    from_cell = AnchorMarker(col=from_col - 1, colOff=from_col_off, row=from_row, rowOff=from_row_off)
                    to_cell = AnchorMarker(col=to_col - 1, colOff=to_col_off, row=to_row, rowOff=to_row_off)

                    new_anchor = TwoCellAnchor(_from = from_cell, to = to_cell, editAs='absolute')
                    stamp_2.anchor = new_anchor

                if int(rev) == 0:
                    ws = self.wb_technical[self.wb_technical.sheetnames[-2]]
                    ws.cell(row=last_row + 3, column=num_column_amount).value = number_items
                    ws.cell(row=last_row + 3, column=num_column_amount).font = Font(name="Calibri", size=14)

            # Editing sheet NOTES
                if int(rev) == 0:
                    sheet_name = "NOTES"  # Selecting  sheet
                    ws = self.wb_technical[sheet_name]

                    rich_string = CellRichText(
                    'We are only offering measuring flow elements. Please be informed that our product range includes temperature elements, and glass and magnetic level indicators; all with european certification. (https://www.eipsa.es/en/products)\n',
                    TextBlock(InlineFont(b=True), 'The prices quoted could be reduced in case of purchasing our full range of products.'))
                    ws["B6"] = rich_string

                    rich_string = CellRichText(
                    TextBlock(InlineFont(i=True), 'Estamos ofertando solamente elementos de medida de caudal, les informamos que en nuestra gama de fabricación con certificación europea, incluye también elementos de temperatura e indicadores de nivel de vidrio y magnéticos. (https://www.eipsa.es/productos)\n'),
                    TextBlock(InlineFont(b=True, i=True), 'Los precios ofertados podrían reducirse en caso de compra de toda nuestra gama.'))
                    ws["B7"] = rich_string

                    rich_string = CellRichText(
                    'Delivery time ' + delivery_time + ' weeks since drawing / calculation approval of ',
                    TextBlock(InlineFont(b=True), 'all equipment'),
                    ' in the contract, as well as critical documentation (August and last two December weeks excluded).')
                    ws["B12"] = rich_string

                    rich_string = CellRichText(
                    'Plazo de entrega ' + delivery_time + ' semanas desde aprobación de planos y cálculos de la ',
                    TextBlock(InlineFont(b=True, i=True), 'totalidad de los equipos'),
                    TextBlock(InlineFont(i=True),' amparados por el contrato, asi como la documentación crítica (Agosto y las dos últimas semanas de diciembre excluidos).'))
                    ws["B13"] = rich_string

                    rich_string = CellRichText(
                    'Quotation prepared according to the information provided in the datasheet corresponding to each TAG. ',
                    TextBlock(InlineFont(u='single'), 'EIPSA does not hold the final responsibility regarding selection of equipment material neither analyze process data.\n'),
                    'The datasheet will be considered the only technical/contractual document, any other documentation will be considered as complementary documentation with informative purpose.')
                    ws["B21"] = rich_string

                    rich_string = CellRichText(
                    'Esta oferta ha sido elaborada en base a la información que figura en la hoja de datos correspondiente a cada TAG, ',
                    TextBlock(InlineFont(i=True, u='single'), 'no siendo responsabilidad final de EIPSA la elección del material a utilizar en los equipos ni analizar datos de proceso.\n'),
                    TextBlock(InlineFont(i=True),'La hoja de datos será el único documento técnico/contractual, cualquier otra documentación recibida será considerada como documentación complementaria a efectos informativos.'))
                    ws["B22"] = rich_string

                    if pay_term == "100_delivery":
                        ws["B45"] = (
                            "100% of total amount of purchase order upon delivery of material according to Incoterms 2020, FCA (our facilities, Spain).\n"
                            "Payment method: bank transfer"
                        )
                        ws["B46"] = (
                            "Pago del 100% del valor total de la orden de compra a la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España).\n"
                            "Método de pago: Transferencia bancaria."
                        )
                    elif pay_term == "100_order":
                        ws["B45"] = (
                            "100 % of the total amount of purchase order upon receipt of purchase order.\n"
                            "Payment method: bank transfer"
                        )
                        ws["B46"] = (
                            "Pago del 100% del valor total de la orden de compra a la recepción de la orden.\n"
                            "Método de pago: Transferencia bancaria"
                        )
                    elif pay_term == "90_10":
                        ws["B45"] = (
                            "PAYMENT TERMS:\n"
                            "90 % of the total amount of PO upon delivery of material according to Incoterms 2020, FCA (our facilities, Spain) and 10% when final documentation is approved. \n"
                            "Bank Transfer: 60 days since invoice issue date."
                        )
                        ws["B46"] = (
                            "TERMINOS DE PAGO:\n"
                            "Pago del 90% del Valor total de la orden de compra a la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España) y el 10% restante cuando la documentación final sea aprobada.\n"
                            "Transferencia Bancaria: 60 días desde emisión de factura."
                        )
                    elif pay_term == "50_50":
                        ws["B45"] = (
                            "50 % of the total amount of purchase order upon receipt of purchase order. Remaining 50% before material be delivered according to Incoterms 2020, FCA (our facilities, Spain).\n"
                            "Payment method: bank transfer."
                        )
                        ws["B46"] = (
                            "Pago del 50% del valor total de la orden de compra a la recepción de la orden. El 50% restante antes de la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España).\n"
                            "Método de pago: Transferencia bancaria."
                        )
                    elif pay_term == "Others":
                        ws["B45"] = "PAYMENT TERMS TO BE DEFINED"
                        ws["B45"].font = Font(name="Calibri", size=11, bold=True, color="FF0000")
                        ws["B46"] = "TERMINOS DE PAGO POR DEFINIR"
                        ws["B46"].font = Font(name="Calibri", size=11, bold=True, italic=True, color="FF0000")

                    rich_string = CellRichText(
                    'For amounts greater than 30,000.00 € we can issue a warranty bond (if required) valid until the end of the indicated warranty period.\nBond warranty of 10% will be issued with the invoice of the last supplement.\n',
                    TextBlock(InlineFont(b=True), 'For lower amounts no warranty bond is issued.'))
                    ws["B48"] = rich_string

                    rich_string = CellRichText(
                    'Para importes superiores a 30.000,00 €, si es requerido, podremos emitir aval de garantía y estará vigente hasta el final del periodo de garantía indicado.\nEl aval del 10% será emitido con la factura del último suplemento.\n',
                    TextBlock(InlineFont(i=True, b=True), 'Por debajo de dicha cantidad, no se emitirán avales.'))
                    ws["B49"] = rich_string

                    ws["A58"] = (
                        "If you require further information related with this offer, please do not hesitate to contact:\n"
                        + responsible
                        + "\n"
                        + email
                        + "\n"
                        "Telf.: (+34) 916.582.118"
                    )
                    
                    std = self.wb_technical["1.3"]
                    self.wb_technical.remove(std)

                left_text = "Fecha/Date: " + date_offer
                right_text = "Petición nº/Inquiry: " + num_ref

                for sheet in self.wb_technical.worksheets:
                    sheet.oddFooter.left.text = left_text
                    sheet.oddFooter.right.text = right_text
                    sheet.oddFooter.center.text = "Page &P de &N"

                    sheet.oddFooter.left.size = 9
                    sheet.oddFooter.right.size = 9
                    sheet.oddFooter.center.size = 9

                self.save_excel_technical()

                root = Tk()
                root.withdraw()  # Hiding main window Tkinter

                # close communication with the PostgreSQL database server
                # commit the changes
                conn.commit()
        except (Exception, psycopg2.DatabaseError) as error:
            dlg = QtWidgets.QMessageBox()
            new_icon = QtGui.QIcon()
            new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
            dlg.setWindowIcon(new_icon)
            dlg.setWindowTitle("ERP EIPSA")
            dlg.setText("Ha ocurrido el siguiente error:\n"
                        + str(error))
            dlg.setIcon(QtWidgets.QMessageBox.Icon.Critical)
            dlg.exec()
            del dlg, new_icon
        finally:
            if conn is not None:
                conn.close()

    def euros_to_float(self, value):
        """
        Converts a string value representing an amount in euros to a float.
        
        Args:
            value (str): The string representation of an amount in euros, with commas for decimal separation and ' €' for currency indication.
        
        Returns:
            float: The numeric value of the amount in euros.
        """
        value = value.replace(".", "")
        value = value.replace(",", ".")
        value = value[: value.find(" €")]
        return float(value)

    def save_excel_commercial(self):
        """Saves the populated Excel workbook to a specified location.
        Opens a dialog window for the user to select the file path and name.
        """
        # Dialog window to select folder and file name; if path is selected, excel file is saved
        output_path_commercial = asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Archivos de Excel", "*.xlsx")],
            title="Guardar Oferta comercial",
        )
        if output_path_commercial:
            self.wb_commercial.save(output_path_commercial)
            return output_path_commercial

    def save_excel_technical(self):
        """Saves the populated Excel workbook to a specified location.
        Opens a dialog window for the user to select the file path and name.
        """
        output_path_technical = asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Archivos de Excel", "*.xlsx")],
            title="Guardar Oferta técnica",
        )
        if output_path_technical:
            self.wb_technical.save(output_path_technical)

    def adjust_images(self, sheet):
        """
        Adjusts the width of all images in the provided spreadsheet sheet by decreasing each by 22 units.
        
        Args:
            sheet: The spreadsheet sheet containing images to be adjusted.
        """
        for image in sheet._images:
            image.width -= 22

class offer_short_flow_spanish:
    """
    A class to manage export offer (short format in spanish) details for flow equipments.
    
    Attributes:
        numoffer (str): Offer number.
        username (str): Name of the user creating the offer.
        rev (str): Revision number of the offer.
        project (str): Name of the project.
        delivery_term (str): Delivery terms for the offer.
        delivery_time (str): Expected delivery time.
        validity (str): Validity period of the offer.
        pay_term (str): Payment terms.
        testinspection (str): Information about testing and inspection.
        revchanges (str): Details of changes made in the revision.
        notes (str): Additional notes, split by line.
    """
    def __init__(self, numoffer, username, rev, project, delivery_term, delivery_time, validity, pay_term, testinspection, revchanges, notes):
        """
        Initializes the offer.

        Args:
            numoffer (str): Offer number.
            username (str): Name of the user creating the offer.
            rev (str): Revision number of the offer.
            project (str): Name of the project.
            delivery_term (str): Delivery terms for the offer.
            delivery_time (str): Expected delivery time.
            validity (str): Validity period of the offer.
            pay_term (str): Payment terms.
            testinspection (str): Information about testing and inspection.
            revchanges (str): Details of changes made in the revision.
            notes (str): Additional notes, split by line.
        """
        notes = notes.split('\n')
        date_offer = date.today().strftime("%d/%m/%Y")
        offername_commercial = numoffer + "-" + "Commercial.Rev" + rev
        offername_technical = numoffer + "-" + "Technical.Rev" + rev

        query_commercial = """
                    SELECT name, surname, email
                    FROM users_data.registration
                    WHERE username = %s
                    """
        query_dataoffer = """
                        SELECT client, num_ref_offer
                        FROM offers
                        WHERE UPPER (num_offer) LIKE UPPER('%%'||%s||'%%')
                        """
        query_tagsdata = """
                        SELECT *
                        FROM tags_data.tags_flow
                        WHERE (
                        UPPER ("num_offer") LIKE UPPER('%%'||%s||'%%')
                        AND
                        "tag_state" NOT IN ('PURCHASED','DELETED')
                        )
                        """
        conn = None
        try:
            # read the connection parameters
            params = config()
            # connect to the PostgreSQL server
            conn = psycopg2.connect(**params)
            cur = conn.cursor()

            cur.execute(query_dataoffer, (numoffer,))
            results_offer = cur.fetchall()
            client = results_offer[0][0]
            num_ref = results_offer[0][1]

            cur.execute(query_commercial, (username,))
            results_commercial = cur.fetchall()
            responsible = results_commercial[0][0] + " " + results_commercial[0][1]
            email = results_commercial[0][2]

            cur.execute(query_tagsdata, (numoffer,))
            data_tags = cur.fetchall()

            if len(data_tags) == 0:
                dlg = QtWidgets.QMessageBox()
                new_icon = QtGui.QIcon()
                new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
                dlg.setWindowIcon(new_icon)
                dlg.setWindowTitle("ERP EIPSA")
                dlg.setText("No hay TAGS importados en la oferta")
                dlg.setIcon(QtWidgets.QMessageBox.Icon.Warning)
                dlg.exec()
                del dlg, new_icon
            else:
                columns = []
                for elt in cur.description:
                    columns.append(elt[0])

                value_type_dict = {
                    "A. Chamber": 1,
                    "C.RING": 1,
                    "F": 1,
                    "F+C.RING": 1,
                    "F+P": 1,
                    "IFO": 1,
                    "M.RUN": 1,
                    "P": 1,
                    "NOZZLE BF": 2,
                    "NOZZLE BW": 2,
                    "NOZZLE F": 2,
                    "PTC-6": 2,
                    "VFM": 3,
                    "VFW": 3,
                    "VWM": 3,
                    "VWW": 3,
                    "WEDGE": 4,
                    "PITOT": 5,
                    "RO": 6,
                    "MULTISTAGE RO": 7,
                }

                # Setting the dataframe with the equipment data
                df = pd.DataFrame(data=data_tags, columns=columns)
                df = df.iloc[:, 1:32]
                df["value_type"] = df["item_type"].map(value_type_dict)
                df = df.sort_values(by=["value_type", "tag"])
                df["amount"] = df["amount"].apply(self.euros_to_float)
                total_amount_material = df["amount"].sum()
                df = df.drop([
                        "tag_state",
                        "num_offer",
                        "num_order",
                        "num_po",
                        "position",
                        "subposition",
                        "flange_type",
                        "plate_std",
                        "pipe_spec",
                        "aprox_weight"
                    ],
                    axis=1,)

                number_items = df.shape[0]
                documentation = number_items * 70

                # Loading Excel Template
                self.wb_commercial = load_workbook(
                    r"\\nas01\DATOS\Comunes\EIPSA-ERP\Plantillas Exportación\PLANTILLA CORTA OFERTA CAUDAL.xlsx"
                )

                # Editing sheet COVER
                sheet_name = "COVER"
                ws = self.wb_commercial[sheet_name]
                ws["E4"] = client
                ws["E6"] = offername_commercial
                ws["E8"] = num_ref
                ws["E10"] = project
                ws["E12"] = date_offer
                ws["E14"] = delivery_term
                ws["E16"] = validity + " days"
                ws["C43"] = responsible
                ws["C45"] = email

                # Editing sheet EQUIPMENT DATA
                sheets_confirmed = ["COVER", "1.2", "1.3", "NOTES"]

                if int(rev) == 0:
                    for item_type in df["item_type"].unique():
                        sheets_confirmed.append(item_type)

                dict_sheets_data = {}

                for value_type in df["value_type"].unique():
                    df_toexport = df[df["value_type"] == value_type]
                    df_toexport.insert(0, "N°", range(1, len(df_toexport) + 1))
                    df_toexport.set_index("N°", inplace=True, drop=False)
                    df_toexport.index.name = None
                    df_toexport = df_toexport.drop(["value_type"], axis=1)

                    eq_type = (
                        "FLOW ELEMENTS DATA"
                        if value_type == 1
                        else (
                            "NOZZLE ELEMENTS DATA"
                            if value_type == 2
                            else (
                                "VENTURI ELEMENTS DATA"
                                if value_type == 3
                                else (
                                    "WEDGE ELEMENTS DATA"
                                    if value_type == 4
                                    else (
                                        "PITOT ELEMENTS DATA"
                                        if value_type == 5
                                        else (
                                            "RO ELEMENTS DATA"
                                            if value_type == 6
                                            else "MULTISTAGE RO ELEMENTS DATA"
                                        )
                                    )
                                )
                            )
                        )
                    )

                    if eq_type not in sheets_confirmed:
                        sheets_confirmed.append(eq_type)

                    if eq_type == "FLOW ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "tube_material",
                                "plate_type",
                                "valve_conn",
                                "valve_material_body",
                                "stages_number",
                                "aprox_length"],
                            axis=1,)
                    elif eq_type == "NOZZLE ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "plate_type",
                                "plate_thk",
                                "gasket_material",
                                "valve_conn",
                                "valve_material_body",
                                "stages_number"],
                            axis=1,)
                    elif (eq_type == "VENTURI ELEMENTS DATA" or eq_type == "WEDGE ELEMENTS DATA"):
                        df_toexport = df_toexport.drop([
                                "plate_type",
                                "plate_thk",
                                "gasket_material",
                                "bolts_nuts_material",
                                "valve_conn",
                                "valve_material_body",
                                "stages_number"],
                            axis=1,)
                    elif eq_type == "PITOT ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "flange_material",
                                "tube_material",
                                "tapping_num_size",
                                "plate_type",
                                "plate_thk",
                                "gasket_material",
                                "stages_number",
                                "aprox_length"],
                            axis=1,)
                    elif eq_type == "RO ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "flange_material",
                                "tube_material",
                                "tapping_num_size",
                                "gasket_material",
                                "bolts_nuts_material",
                                "valve_conn",
                                "valve_material_body",
                                "stages_number",
                                "aprox_length"],
                            axis=1,)
                    elif eq_type == "MULTISTAGE RO ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "tapping_num_size",
                                "gasket_material",
                                "bolts_nuts_material",
                                "valve_conn",
                                "valve_material_body"],
                            axis=1,)

                    ws = self.wb_commercial[eq_type]
                    ws["G2"] = date_offer
                    ws["G3"] = num_ref
                    ws["G4"] = offername_commercial
                    if revchanges != "":
                        ws["G5"] = rev + " " + revchanges
                        ws["G5"].font = Font(name="Calibri", size=14, bold=True)
                        ws["G5"].fill = PatternFill("solid", fgColor="FFFF00")

                    if int(rev) > 0:
                        for row in ws.iter_rows(min_row=2, max_row=4, min_col=6, max_col=7):
                            for cell in row:
                                cell.value = None
                                cell._style = ws["F1"]._style

                    last_row = ws.max_row

                    num_column_amount = df_toexport.columns.get_loc("amount") + 1

                    for index, row in df_toexport.iterrows():  # Data in desired row
                        for col_num, value in enumerate(row, start=1):
                            cell = ws.cell(row=last_row + 1, column=col_num)
                            cell.value = value
                            if col_num == num_column_amount:
                                cell._style = ws["X1"]._style
                            else:
                                cell._style = ws["S1"]._style

                        last_row = ws.max_row

                    if eq_type == "VENTURI ELEMENTS DATA":
                        ws[f"A{last_row+3}"] = "LOS PRECIOS INCLUYEN LA SECCIÓN CENTRAL INTEGRAL MECANIZADA Y TODAS LAS SOLDADURAS ESTRUCTURALES 100% RADIOGRAFIADAS"
                        ws[f"A{last_row+3}"]._style = ws["Z2"]._style
                    ws[f"A{last_row+4}"] = "VALIDEZ DE LA OFERTA: " + validity + " DÍAS"
                    ws[f"A{last_row+4}"]._style = ws["Z1"]._style
                    ws[f"A{last_row+5}"] = (
                        "PLAZO DE ENTREGA: "
                        + delivery_time
                        + " SEMANAS DESDE APROBACIÓN DE PLANOS / CÁLCULOS (AGOSTO Y ÚLTIMAS DOS SEMANAS DE DICIEMBRE EXCLUIDOS)"
                    )
                    ws[f"A{last_row+5}"]._style = ws["Z1"]._style

                    if notes != "":
                        if isinstance(notes, list):
                            line = last_row + 6
                            for note in notes:
                                ws[f"A{line}"] = note
                                ws[f"A{line}"]._style = ws["Z1"]._style
                                line += 1
                        else:
                            line = last_row + 6
                            ws[f"A{line}"] = notes
                            ws[f"A{line}"]._style = ws["Z1"]._style

                    dict_sheets_data[eq_type] = [last_row, num_column_amount, df_toexport["amount"].sum(), df_toexport.shape[0]]

                ws.cell(row=last_row + 3, column=num_column_amount - 1).value = "CANTIDAD TOTAL"
                ws.cell(row=last_row + 3, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=last_row + 3, column=num_column_amount).value = number_items
                
                row_amount = last_row + 4
                for key, value in dict_sheets_data.items():
                    parts_key = key.split(" ")
                    ws.cell(row=row_amount + 2, column=num_column_amount - 1).value = "IMPORTE TOTAL DE " + parts_key[0] + " " + parts_key[1] + " (CANTIDAD: " + str(value[3]) + ")"
                    ws.cell(row=row_amount + 2, column=num_column_amount).value = value[2]
                    ws.cell(row=row_amount + 2, column=num_column_amount - 1)._style = ws["R1"]._style
                    ws.cell(row=row_amount + 2, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                    ws.cell(row=row_amount + 2, column=num_column_amount)._style = ws["T1"]._style

                    row_amount += 2

                ws.cell(row=row_amount + 2, column=num_column_amount - 1).value = "IMPORTE TOTAL DEL MATERIAL"
                ws.cell(row=row_amount + 2, column=num_column_amount).value = total_amount_material
                ws.cell(row=row_amount + 4, column=num_column_amount - 1).value = "PACKING Y TRANSPORTE"
                ws.cell(row=row_amount + 4, column=num_column_amount).value = (f"=MROUND({get_column_letter(num_column_amount)}{row_amount + 2}*0.03,10)" if total_amount_material > 6700 else 200)
                ws.cell(row=row_amount + 5, column=num_column_amount - 1).value = "PRUEBAS E INSPECCIÓN"
                ws.cell(row=row_amount + 5, column=num_column_amount).value = float(testinspection)
                ws.cell(row=row_amount + 6, column=num_column_amount - 1).value = "DOCUMENTACIÓN"
                ws.cell(row=row_amount + 6, column=num_column_amount).value = documentation
                ws.cell(row=row_amount + 8, column=num_column_amount - 1).value = "IMPORTE TOTAL DE LA OFERTA"
                ws.cell(row=row_amount + 8, column=num_column_amount).value = f"=SUM({get_column_letter(num_column_amount)}{row_amount + 2}:{get_column_letter(num_column_amount)}{row_amount + 6})"

                ws.cell(row=last_row + 3, column=num_column_amount - 1)._style = ws["R1"]._style
                ws.cell(row=last_row + 3, column=num_column_amount).font = Font(name="Calibri", size=14)
                ws.cell(row=row_amount + 2, column=num_column_amount - 1)._style = ws["R1"]._style
                ws.cell(row=row_amount + 2, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=row_amount + 2, column=num_column_amount)._style = ws["T1"]._style
                ws.cell(row=row_amount + 4, column=num_column_amount - 1).font = Font(name="Calibri", size=14)
                ws.cell(row=row_amount + 4, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=row_amount + 4, column=num_column_amount)._style = ws["T1"]._style
                ws.cell(row=row_amount + 5, column=num_column_amount - 1).font = Font(name="Calibri", size=14)
                ws.cell(row=row_amount + 5, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=row_amount + 5, column=num_column_amount)._style = ws["T1"]._style
                ws.cell(row=row_amount + 6, column=num_column_amount - 2)._style = ws["U1"]._style
                ws.cell(row=row_amount + 6, column=num_column_amount - 1)._style = ws["U1"]._style
                ws.cell(row=row_amount + 6, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=row_amount + 6, column=num_column_amount)._style = ws["V1"]._style
                ws.cell(row=row_amount + 8, column=num_column_amount - 1)._style = ws["R1"]._style
                ws.cell(row=row_amount + 8, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=row_amount + 8, column=num_column_amount)._style = ws["W1"]._style

            # Editing sheet NOTES
                sheet_name = "NOTES"  # Selecting  sheet
                ws = self.wb_commercial[sheet_name]

                rich_string = CellRichText(
                TextBlock(InlineFont(i=True), 'Estamos ofertando solamente elementos de medida de caudal, les informamos que en nuestra gama de fabricación con certificación europea, incluye también elementos de temperatura e indicadores de nivel de vidrio y magnéticos. (https://www.eipsa.es/productos)\n'),
                TextBlock(InlineFont(b=True, i=True), 'Los precios ofertados podrían reducirse en caso de compra de toda nuestra gama.'))
                ws["B6"] = rich_string

                rich_string = CellRichText(
                'Plazo de entrega ' + delivery_time + ' semanas desde aprobación de planos y cálculos de la ',
                TextBlock(InlineFont(b=True, i=True), 'totalidad de los equipos'),
                TextBlock(InlineFont(i=True),' amparados por el contrato, asi como la documentación crítica (Agosto y las dos últimas semanas de diciembre excluidos).'))
                ws["B10"] = rich_string

                rich_string = CellRichText(
                'Esta oferta ha sido elaborada en base a la información que figura en la hoja de datos correspondiente a cada TAG, ',
                TextBlock(InlineFont(i=True, u='single'), 'no siendo responsabilidad final de EIPSA la elección del material a utilizar en los equipos ni analizar datos de proceso.\n'),
                TextBlock(InlineFont(i=True),'La hoja de datos será el único documento técnico/contractual, cualquier otra documentación recibida será considerada como documentación complementaria a efectos informativos.'))
                ws["B16"] = rich_string

                if pay_term == "100_delivery":
                    ws["B30"] = (
                        "Pago del 100% del valor total de la orden de compra a la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España).\n"
                        "Método de pago: Transferencia bancaria."
                    )
                elif pay_term == "100_order":
                    ws["B30"] = (
                        "Pago del 100% del valor total de la orden de compra a la recepción de la orden.\n"
                        "Método de pago: Transferencia bancaria"
                    )
                elif pay_term == "90_10":
                    ws["B30"] = (
                        "TERMINOS DE PAGO:\n"
                        "Pago del 90% del Valor total de la orden de compra a la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España) y el 10% restante cuando la documentación final sea aprobada.\n"
                        "Transferencia Bancaria: 60 días desde emisión de factura."
                    )
                elif pay_term == "50_50":
                    ws["B30"] = (
                        "Pago del 50% del valor total de la orden de compra a la recepción de la orden. El 50% restante antes de la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España).\n"
                        "Método de pago: Transferencia bancaria."
                    )
                elif pay_term == "Others":
                    ws["B30"] = "TERMINOS DE PAGO POR DEFINIR"
                    ws["B30"].font = Font(name="Calibri", size=11, bold=True, italic=True, color="FF0000")

                rich_string = CellRichText(
                'Para importes superiores a 30.000,00 €, si es requerido, podremos emitir aval de garantía y estará vigente hasta el final del periodo de garantía indicado.\nEl aval del 10% será emitido con la factura del último suplemento.\n',
                TextBlock(InlineFont(i=True, b=True), 'Por debajo de dicha cantidad, no se emitirán avales.'))
                ws["B28"] = rich_string

                ws["A36"] = (
                    "Si necesita más información relacionada con esta oferta, no dude en ponerse en contacto con:\n"
                    + responsible
                    + "\n"
                    + email
                    + "\n"
                    "Telf.: (+34) 916.582.118"
                )

                for sheet in self.wb_commercial.sheetnames:
                    if sheet not in sheets_confirmed:
                        sheet_to_delete = self.wb_commercial[sheet]
                        self.wb_commercial.remove(sheet_to_delete)

                if int(rev) > 0:
                    sheets_confirmed = ["COVER", "1.2", "1.3", "NOTES"]
                    for sheet in sheets_confirmed:
                        sheet_to_delete = self.wb_commercial[sheet]
                        self.wb_commercial.remove(sheet_to_delete)

                left_text = "Fecha/Date: " + date_offer
                right_text = "Petición nº/Inquiry: " + num_ref

                for sheet in self.wb_commercial.worksheets:
                    sheet.oddFooter.left.text = left_text
                    sheet.oddFooter.right.text = right_text
                    sheet.oddFooter.center.text = "Page &P de &N"

                    sheet.oddFooter.left.size = 9
                    sheet.oddFooter.right.size = 9
                    sheet.oddFooter.center.size = 9
                
                path = self.save_excel_commercial()

                # Creating the technical offer using the commercial one as template
                self.wb_technical = load_workbook(path)

                if int(rev) == 0:
                    sheet_name = "COVER"
                    ws = self.wb_technical[sheet_name]
                    ws["E6"] = offername_technical

                for value_type in df["value_type"].unique():
                    eq_type = (
                        "FLOW ELEMENTS DATA"
                        if value_type == 1
                        else (
                            "NOZZLE ELEMENTS DATA"
                            if value_type == 2
                            else (
                                "VENTURI ELEMENTS DATA"
                                if value_type == 3
                                else (
                                    "WEDGE ELEMENTS DATA"
                                    if value_type == 4
                                    else (
                                        "PITOT ELEMENTS DATA"
                                        if value_type == 5
                                        else (
                                            "RO ELEMENTS DATA"
                                            if value_type == 6
                                            else "MULTISTAGE RO ELEMENTS DATA"
                                        )
                                    )
                                )
                            )
                        )
                    )

                    ws = self.wb_technical[eq_type]
                    if int(rev) == 0:
                        ws["G4"] = offername_technical

                    last_row = dict_sheets_data[eq_type][0]
                    num_column_amount = dict_sheets_data[eq_type][1]

                    # self.wb_technical[eq_type].delete_rows(last_row + 5, 20)

                    ws.cell(row=row_amount + 2, column=num_column_amount).value = ""
                    ws.cell(row=row_amount + 4, column=num_column_amount).value = ""
                    ws.cell(row=row_amount + 5, column=num_column_amount).value = ""
                    ws.cell(row=row_amount + 6, column=num_column_amount).value = ""
                    ws.cell(row=row_amount + 8, column=num_column_amount).value = ""

                    ws.cell(row=row_amount + 2, column=num_column_amount - 1).value = ""
                    ws.cell(row=row_amount + 4, column=num_column_amount - 1).value = ""
                    ws.cell(row=row_amount + 5, column=num_column_amount - 1).value = ""
                    ws.cell(row=row_amount + 6, column=num_column_amount - 1).value = ""
                    ws.cell(row=row_amount + 8, column=num_column_amount - 1).value = ""

                # Deleting "Amount" column
                    self.wb_technical[eq_type].delete_cols(num_column_amount, 1)

                # Adjusting the print area
                    new_last_column = num_column_amount
                    last_print_row = 40
                    nuevo_rango_impresion = f'A1:{get_column_letter(new_last_column)}{last_print_row}'
                    self.wb_technical[eq_type].print_area = nuevo_rango_impresion

                    stamp_1 = self.wb_technical[eq_type]._images[1]
                    anchor_actual = stamp_1.anchor

                    from_col, from_col_off = anchor_actual._from.col, anchor_actual._from.colOff
                    from_row, from_row_off = anchor_actual._from.row, anchor_actual._from.rowOff
                    to_col, to_col_off = anchor_actual.to.col, anchor_actual.to.colOff
                    to_row, to_row_off = anchor_actual.to.row, anchor_actual.to.rowOff

                    from_cell = AnchorMarker(col=from_col - 1, colOff=from_col_off, row=from_row, rowOff=from_row_off)
                    to_cell = AnchorMarker(col=to_col - 1, colOff=to_col_off, row=to_row, rowOff=to_row_off)

                    new_anchor = TwoCellAnchor(_from = from_cell, to = to_cell, editAs='absolute')
                    stamp_1.anchor = new_anchor

                    stamp_2 = self.wb_technical[eq_type]._images[2]
                    anchor_actual = stamp_2.anchor

                    from_col, from_col_off = anchor_actual._from.col, anchor_actual._from.colOff
                    from_row, from_row_off = anchor_actual._from.row, anchor_actual._from.rowOff
                    to_col, to_col_off = anchor_actual.to.col, anchor_actual.to.colOff
                    to_row, to_row_off = anchor_actual.to.row, anchor_actual.to.rowOff

                    from_cell = AnchorMarker(col=from_col - 1, colOff=from_col_off, row=from_row, rowOff=from_row_off)
                    to_cell = AnchorMarker(col=to_col - 1, colOff=to_col_off, row=to_row, rowOff=to_row_off)

                    new_anchor = TwoCellAnchor(_from = from_cell, to = to_cell, editAs='absolute')
                    stamp_2.anchor = new_anchor

                if int(rev) == 0:
                    ws = self.wb_technical[self.wb_technical.sheetnames[-2]]
                    ws.cell(row=last_row + 3, column=num_column_amount).value = number_items
                    ws.cell(row=last_row + 3, column=num_column_amount).font = Font(name="Calibri", size=14)

            # Editing sheet NOTES
                if int(rev) == 0:
                    sheet_name = "NOTES"  # Selecting  sheet
                    ws = self.wb_technical[sheet_name]

                    rich_string = CellRichText(
                    TextBlock(InlineFont(i=True), 'Estamos ofertando solamente elementos de medida de caudal, les informamos que en nuestra gama de fabricación con certificación europea, incluye también elementos de temperatura e indicadores de nivel de vidrio y magnéticos. (https://www.eipsa.es/productos)\n'),
                    TextBlock(InlineFont(b=True, i=True), 'Los precios ofertados podrían reducirse en caso de compra de toda nuestra gama.'))
                    ws["B6"] = rich_string

                    rich_string = CellRichText(
                    'Plazo de entrega ' + delivery_time + ' semanas desde aprobación de planos y cálculos de la ',
                    TextBlock(InlineFont(b=True, i=True), 'totalidad de los equipos'),
                    TextBlock(InlineFont(i=True),' amparados por el contrato, asi como la documentación crítica (Agosto y las dos últimas semanas de diciembre excluidos).'))
                    ws["B10"] = rich_string

                    rich_string = CellRichText(
                    'Esta oferta ha sido elaborada en base a la información que figura en la hoja de datos correspondiente a cada TAG, ',
                    TextBlock(InlineFont(i=True, u='single'), 'no siendo responsabilidad final de EIPSA la elección del material a utilizar en los equipos ni analizar datos de proceso.\n'),
                    TextBlock(InlineFont(i=True),'La hoja de datos será el único documento técnico/contractual, cualquier otra documentación recibida será considerada como documentación complementaria a efectos informativos.'))
                    ws["B16"] = rich_string

                    if pay_term == "100_delivery":
                        ws["B30"] = (
                            "Pago del 100% del valor total de la orden de compra a la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España).\n"
                            "Método de pago: Transferencia bancaria."
                        )
                    elif pay_term == "100_order":
                        ws["B30"] = (
                            "Pago del 100% del valor total de la orden de compra a la recepción de la orden.\n"
                            "Método de pago: Transferencia bancaria"
                        )
                    elif pay_term == "90_10":
                        ws["B30"] = (
                            "TERMINOS DE PAGO:\n"
                            "Pago del 90% del Valor total de la orden de compra a la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España) y el 10% restante cuando la documentación final sea aprobada.\n"
                            "Transferencia Bancaria: 60 días desde emisión de factura."
                        )
                    elif pay_term == "50_50":
                        ws["B30"] = (
                            "Pago del 50% del valor total de la orden de compra a la recepción de la orden. El 50% restante antes de la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España).\n"
                            "Método de pago: Transferencia bancaria."
                        )
                    elif pay_term == "Others":
                        ws["B30"] = "TERMINOS DE PAGO POR DEFINIR"
                        ws["B30"].font = Font(name="Calibri", size=11, bold=True, italic=True, color="FF0000")

                    rich_string = CellRichText(
                    'Para importes superiores a 30.000,00 €, si es requerido, podremos emitir aval de garantía y estará vigente hasta el final del periodo de garantía indicado.\nEl aval del 10% será emitido con la factura del último suplemento.\n',
                    TextBlock(InlineFont(i=True, b=True), 'Por debajo de dicha cantidad, no se emitirán avales.'))
                    ws["B28"] = rich_string

                    ws["A36"] = (
                        "Si necesita más información relacionada con esta oferta, no dude en ponerse en contacto con:\n"
                        + responsible
                        + "\n"
                        + email
                        + "\n"
                        "Telf.: (+34) 916.582.118"
                    )

                    std = self.wb_technical["1.3"]
                    self.wb_technical.remove(std)

                left_text = "Fecha/Date: " + date_offer
                right_text = "Petición nº/Inquiry: " + num_ref

                for sheet in self.wb_technical.worksheets:
                    sheet.oddFooter.left.text = left_text
                    sheet.oddFooter.right.text = right_text
                    sheet.oddFooter.center.text = "Page &P de &N"

                    sheet.oddFooter.left.size = 9
                    sheet.oddFooter.right.size = 9
                    sheet.oddFooter.center.size = 9
                
                self.save_excel_technical()

                root = Tk()
                root.withdraw()  # Hiding main window Tkinter

                # close communication with the PostgreSQL database server
                # commit the changes
                conn.commit()
        except (Exception, psycopg2.DatabaseError) as error:
            dlg = QtWidgets.QMessageBox()
            new_icon = QtGui.QIcon()
            new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
            dlg.setWindowIcon(new_icon)
            dlg.setWindowTitle("ERP EIPSA")
            dlg.setText("Ha ocurrido el siguiente error:\n"
                        + str(error))
            dlg.setIcon(QtWidgets.QMessageBox.Icon.Critical)
            dlg.exec()
            del dlg, new_icon
        finally:
            if conn is not None:
                conn.close()

    def euros_to_float(self, value):
        """
        Converts a string value representing an amount in euros to a float.
        
        Args:
            value (str): The string representation of an amount in euros, with commas for decimal separation and ' €' for currency indication.
        
        Returns:
            float: The numeric value of the amount in euros.
        """
        value = value.replace(".", "")
        value = value.replace(",", ".")
        value = value[: value.find(" €")]
        return float(value)

    def save_excel_commercial(self):
        """Saves the populated Excel workbook to a specified location.
        Opens a dialog window for the user to select the file path and name.
        """
        # Dialog window to select folder and file name; if path is selected, excel file is saved
        output_path_commercial = asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Archivos de Excel", "*.xlsx")],
            title="Guardar Oferta comercial",
        )
        if output_path_commercial:
            self.wb_commercial.save(output_path_commercial)
            return output_path_commercial

    def save_excel_technical(self):
        """Saves the populated Excel workbook to a specified location.
        Opens a dialog window for the user to select the file path and name.
        """
        output_path_technical = asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Archivos de Excel", "*.xlsx")],
            title="Guardar Oferta técnica",
        )
        if output_path_technical:
            self.wb_technical.save(output_path_technical)

    def adjust_images(self, sheet):
        """
        Adjusts the width of all images in the provided spreadsheet sheet by decreasing each by 22 units.
        
        Args:
            sheet: The spreadsheet sheet containing images to be adjusted.
        """
        for image in sheet._images:
            image.width -= 22

class offer_short_flow_english:
    """
    A class to manage export offer (short format in english) details for flow equipments.
    
    Attributes:
        numoffer (str): Offer number.
        username (str): Name of the user creating the offer.
        rev (str): Revision number of the offer.
        project (str): Name of the project.
        delivery_term (str): Delivery terms for the offer.
        delivery_time (str): Expected delivery time.
        validity (str): Validity period of the offer.
        pay_term (str): Payment terms.
        testinspection (str): Information about testing and inspection.
        revchanges (str): Details of changes made in the revision.
        notes (str): Additional notes, split by line.
    """
    def __init__(self, numoffer, username, rev, project, delivery_term, delivery_time, validity, pay_term, testinspection, revchanges, notes):
        """
        Initializes the offer.

        Args:
            numoffer (str): Offer number.
            username (str): Name of the user creating the offer.
            rev (str): Revision number of the offer.
            project (str): Name of the project.
            delivery_term (str): Delivery terms for the offer.
            delivery_time (str): Expected delivery time.
            validity (str): Validity period of the offer.
            pay_term (str): Payment terms.
            testinspection (str): Information about testing and inspection.
            revchanges (str): Details of changes made in the revision.
            notes (str): Additional notes, split by line.
        """
        notes = notes.split('\n')
        date_offer = date.today().strftime("%d/%m/%Y")
        offername_commercial = numoffer + "-" + "Commercial.Rev" + rev
        offername_technical = numoffer + "-" + "Technical.Rev" + rev

        query_commercial = """
                    SELECT name, surname, email
                    FROM users_data.registration
                    WHERE username = %s
                    """
        query_dataoffer = """
                        SELECT client, num_ref_offer
                        FROM offers
                        WHERE UPPER (num_offer) LIKE UPPER('%%'||%s||'%%')
                        """
        query_tagsdata = """
                        SELECT *
                        FROM tags_data.tags_flow
                        WHERE (
                        UPPER ("num_offer") LIKE UPPER('%%'||%s||'%%')
                        AND
                        "tag_state" NOT IN ('PURCHASED','DELETED')
                        )
                        """
        conn = None
        try:
            # read the connection parameters
            params = config()
            # connect to the PostgreSQL server
            conn = psycopg2.connect(**params)
            cur = conn.cursor()

            cur.execute(query_dataoffer, (numoffer,))
            results_offer = cur.fetchall()
            client = results_offer[0][0]
            num_ref = results_offer[0][1]

            cur.execute(query_commercial, (username,))
            results_commercial = cur.fetchall()
            responsible = results_commercial[0][0] + " " + results_commercial[0][1]
            email = results_commercial[0][2]

            cur.execute(query_tagsdata, (numoffer,))
            data_tags = cur.fetchall()

            if len(data_tags) == 0:
                dlg = QtWidgets.QMessageBox()
                new_icon = QtGui.QIcon()
                new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
                dlg.setWindowIcon(new_icon)
                dlg.setWindowTitle("ERP EIPSA")
                dlg.setText("No hay TAGS importados en la oferta")
                dlg.setIcon(QtWidgets.QMessageBox.Icon.Warning)
                dlg.exec()
                del dlg, new_icon
            else:
                columns = []
                for elt in cur.description:
                    columns.append(elt[0])

                value_type_dict = {
                    "A. Chamber": 1,
                    "C.RING": 1,
                    "F": 1,
                    "F+C.RING": 1,
                    "F+P": 1,
                    "IFO": 1,
                    "M.RUN": 1,
                    "P": 1,
                    "NOZZLE BF": 2,
                    "NOZZLE BW": 2,
                    "NOZZLE F": 2,
                    "PTC-6": 2,
                    "VFM": 3,
                    "VFW": 3,
                    "VWM": 3,
                    "VWW": 3,
                    "WEDGE": 4,
                    "PITOT": 5,
                    "RO": 6,
                    "MULTISTAGE RO": 7,
                }

                # Setting the dataframe with the equipment data
                df = pd.DataFrame(data=data_tags, columns=columns)
                df = df.iloc[:, 1:32]
                df["value_type"] = df["item_type"].map(value_type_dict)
                df = df.sort_values(by=["value_type", "tag"])
                df["amount"] = df["amount"].apply(self.euros_to_float)
                total_amount_material = df["amount"].sum()
                df = df.drop([
                        "tag_state",
                        "num_offer",
                        "num_order",
                        "num_po",
                        "position",
                        "subposition",
                        "flange_type",
                        "plate_std",
                        "pipe_spec",
                        "aprox_weight"
                    ],
                    axis=1,)

                number_items = df.shape[0]
                documentation = number_items * 70

                # Loading Excel Template
                self.wb_commercial = load_workbook(
                    r"\\nas01\DATOS\Comunes\EIPSA-ERP\Plantillas Exportación\PLANTILLA CORTA OFERTA CAUDAL - ingles.xlsx"
                )

                # Editing sheet COVER
                sheet_name = "COVER"
                ws = self.wb_commercial[sheet_name]
                ws["E4"] = client
                ws["E6"] = offername_commercial
                ws["E8"] = num_ref
                ws["E10"] = project
                ws["E12"] = date_offer
                ws["E14"] = delivery_term
                ws["E16"] = validity + " days"
                ws["C43"] = responsible
                ws["C45"] = email

                # Editing sheet EQUIPMENT DATA
                sheets_confirmed = ["COVER", "1.2", "1.3", "NOTES"]

                if int(rev) == 0:
                    for item_type in df["item_type"].unique():
                        sheets_confirmed.append(item_type)

                dict_sheets_data = {}

                for value_type in df["value_type"].unique():
                    df_toexport = df[df["value_type"] == value_type]
                    df_toexport.insert(0, "N°", range(1, len(df_toexport) + 1))
                    df_toexport.set_index("N°", inplace=True, drop=False)
                    df_toexport.index.name = None
                    df_toexport = df_toexport.drop(["value_type"], axis=1)

                    eq_type = (
                        "FLOW ELEMENTS DATA"
                        if value_type == 1
                        else (
                            "NOZZLE ELEMENTS DATA"
                            if value_type == 2
                            else (
                                "VENTURI ELEMENTS DATA"
                                if value_type == 3
                                else (
                                    "WEDGE ELEMENTS DATA"
                                    if value_type == 4
                                    else (
                                        "PITOT ELEMENTS DATA"
                                        if value_type == 5
                                        else (
                                            "RO ELEMENTS DATA"
                                            if value_type == 6
                                            else "MULTISTAGE RO ELEMENTS DATA"
                                        )
                                    )
                                )
                            )
                        )
                    )

                    if eq_type not in sheets_confirmed:
                        sheets_confirmed.append(eq_type)

                    if eq_type == "FLOW ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "tube_material",
                                "plate_type",
                                "valve_conn",
                                "valve_material_body",
                                "stages_number",
                                "aprox_length"],
                            axis=1,)
                    elif eq_type == "NOZZLE ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "plate_type",
                                "plate_thk",
                                "gasket_material",
                                "valve_conn",
                                "valve_material_body",
                                "stages_number"],
                            axis=1,)
                    elif (eq_type == "VENTURI ELEMENTS DATA" or eq_type == "WEDGE ELEMENTS DATA"):
                        df_toexport = df_toexport.drop([
                                "plate_type",
                                "plate_thk",
                                "gasket_material",
                                "bolts_nuts_material",
                                "valve_conn",
                                "valve_material_body",
                                "stages_number"],
                            axis=1,)
                    elif eq_type == "PITOT ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "flange_material",
                                "tube_material",
                                "tapping_num_size",
                                "plate_type",
                                "plate_thk",
                                "gasket_material",
                                "stages_number",
                                "aprox_length"],
                            axis=1,)
                    elif eq_type == "RO ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "flange_material",
                                "tube_material",
                                "tapping_num_size",
                                "gasket_material",
                                "bolts_nuts_material",
                                "valve_conn",
                                "valve_material_body",
                                "stages_number",
                                "aprox_length"],
                            axis=1,)
                    elif eq_type == "MULTISTAGE RO ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "tapping_num_size",
                                "gasket_material",
                                "bolts_nuts_material",
                                "valve_conn",
                                "valve_material_body"],
                            axis=1,)

                    ws = self.wb_commercial[eq_type]
                    ws["G2"] = date_offer
                    ws["G3"] = num_ref
                    ws["G4"] = offername_commercial
                    if revchanges != "":
                        ws["G5"] = rev + " " + revchanges
                        ws["G5"].font = Font(name="Calibri", size=14, bold=True)
                        ws["G5"].fill = PatternFill("solid", fgColor="FFFF00")

                    if int(rev) > 0:
                        for row in ws.iter_rows(min_row=2, max_row=4, min_col=6, max_col=7):
                            for cell in row:
                                cell.value = None
                                cell._style = ws["F1"]._style

                    last_row = ws.max_row

                    num_column_amount = df_toexport.columns.get_loc("amount") + 1

                    for index, row in df_toexport.iterrows():  # Data in desired row
                        for col_num, value in enumerate(row, start=1):
                            cell = ws.cell(row=last_row + 1, column=col_num)
                            cell.value = value
                            if col_num == num_column_amount:
                                cell._style = ws["X1"]._style
                            else:
                                cell._style = ws["S1"]._style

                        last_row = ws.max_row

                    if eq_type == "VENTURI ELEMENTS DATA":
                        ws[f"A{last_row+3}"] = "PRICES INCLUDE MACHINED INTEGRAL CENTRE SECTION AND ALL STRUCTURAL WELDS 100% RADIOGRAPHED"
                        ws[f"A{last_row+3}"]._style = ws["Z2"]._style
                    ws[f"A{last_row+4}"] = "OFFER VALIDITY: " + validity + " DAYS"
                    ws[f"A{last_row+4}"]._style = ws["Z1"]._style
                    ws[f"A{last_row+5}"] = (
                        "DELIVERY TIME: "
                        + delivery_time
                        + " WEEKS SINCE DRAWING / CALCULATION APPROVAL (AUGUST AND LAST TWO DECEMBER WEEKS EXCLUDED)"
                    )
                    ws[f"A{last_row+5}"]._style = ws["Z1"]._style

                    if notes != "":
                        if isinstance(notes, list):
                            line = last_row + 6
                            for note in notes:
                                ws[f"A{line}"] = note
                                ws[f"A{line}"]._style = ws["Z1"]._style
                                line += 1
                        else:
                            line = last_row + 6
                            ws[f"A{line}"] = notes
                            ws[f"A{line}"]._style = ws["Z1"]._style

                    dict_sheets_data[eq_type] = [last_row, num_column_amount, df_toexport["amount"].sum(), df_toexport.shape[0]]

                ws.cell(row=last_row + 3, column=num_column_amount - 1).value = "QTY. TOTAL"
                ws.cell(row=last_row + 3, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=last_row + 3, column=num_column_amount).value = number_items
                
                row_amount = last_row + 4
                for key, value in dict_sheets_data.items():
                    parts_key = key.split(" ")
                    ws.cell(row=row_amount + 2, column=num_column_amount - 1).value = "TOTAL AMOUNT OF " + parts_key[0] + " " + parts_key[1] + " (QTY: " + str(value[3]) + ")"
                    ws.cell(row=row_amount + 2, column=num_column_amount).value = value[2]
                    ws.cell(row=row_amount + 2, column=num_column_amount - 1)._style = ws["R1"]._style
                    ws.cell(row=row_amount + 2, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                    ws.cell(row=row_amount + 2, column=num_column_amount)._style = ws["T1"]._style

                    row_amount += 2

                ws.cell(row=row_amount + 2, column=num_column_amount - 1).value = "TOTAL AMOUNT OF MATERIAL"
                ws.cell(row=row_amount + 2, column=num_column_amount).value = total_amount_material
                ws.cell(row=row_amount + 4, column=num_column_amount - 1).value = "PACKING AND TRANSPORT"
                ws.cell(row=row_amount + 4, column=num_column_amount).value = (f"=MROUND({get_column_letter(num_column_amount)}{row_amount + 2}*0.03,10)" if total_amount_material > 6700 else 200)
                ws.cell(row=row_amount + 5, column=num_column_amount - 1).value = "TESTS & INSPECTION"
                ws.cell(row=row_amount + 5, column=num_column_amount).value = float(testinspection)
                ws.cell(row=row_amount + 6, column=num_column_amount - 1).value = "DOCUMENTATION"
                ws.cell(row=row_amount + 6, column=num_column_amount).value = documentation
                ws.cell(row=row_amount + 8, column=num_column_amount - 1).value = "TOTAL AMOUNT OF BID"
                ws.cell(row=row_amount + 8, column=num_column_amount).value = f"=SUM({get_column_letter(num_column_amount)}{row_amount + 2}:{get_column_letter(num_column_amount)}{row_amount + 6})"

                ws.cell(row=last_row + 3, column=num_column_amount - 1)._style = ws["R1"]._style
                ws.cell(row=last_row + 3, column=num_column_amount).font = Font(name="Calibri", size=14)
                ws.cell(row=row_amount + 2, column=num_column_amount - 1)._style = ws["R1"]._style
                ws.cell(row=row_amount + 2, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=row_amount + 2, column=num_column_amount)._style = ws["T1"]._style
                ws.cell(row=row_amount + 4, column=num_column_amount - 1).font = Font(name="Calibri", size=14)
                ws.cell(row=row_amount + 4, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=row_amount + 4, column=num_column_amount)._style = ws["T1"]._style
                ws.cell(row=row_amount + 5, column=num_column_amount - 1).font = Font(name="Calibri", size=14)
                ws.cell(row=row_amount + 5, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=row_amount + 5, column=num_column_amount)._style = ws["T1"]._style
                ws.cell(row=row_amount + 6, column=num_column_amount - 2)._style = ws["U1"]._style
                ws.cell(row=row_amount + 6, column=num_column_amount - 1)._style = ws["U1"]._style
                ws.cell(row=row_amount + 6, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=row_amount + 6, column=num_column_amount)._style = ws["V1"]._style
                ws.cell(row=row_amount + 8, column=num_column_amount - 1)._style = ws["R1"]._style
                ws.cell(row=row_amount + 8, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=row_amount + 8, column=num_column_amount)._style = ws["W1"]._style

            # Editing sheet NOTES
                sheet_name = "NOTES"  # Selecting  sheet
                ws = self.wb_commercial[sheet_name]

                rich_string = CellRichText(
                'We are only offering measuring flow elements. Please be informed that our product range includes temperature elements, and glass and magnetic level indicators; all with european certification. (https://www.eipsa.es/en/products)\n',
                TextBlock(InlineFont(b=True), 'The prices quoted could be reduced in case of purchasing our full range of products.'))
                ws["B6"] = rich_string

                rich_string = CellRichText(
                TextBlock(InlineFont(i=True), 'Estamos ofertando solamente elementos de medida de caudal, les informamos que en nuestra gama de fabricación con certificación europea, incluye también elementos de temperatura e indicadores de nivel de vidrio y magnéticos. (https://www.eipsa.es/productos)\n'),
                TextBlock(InlineFont(b=True, i=True), 'Los precios ofertados podrían reducirse en caso de compra de toda nuestra gama.'))
                ws["B7"] = rich_string

                rich_string = CellRichText(
                'Delivery time ' + delivery_time + ' weeks since drawing / calculation approval of ',
                TextBlock(InlineFont(b=True), 'all equipment'),
                ' in the contract, as well as critical documentation (August and last two December weeks excluded).')
                ws["B12"] = rich_string

                rich_string = CellRichText(
                'Plazo de entrega ' + delivery_time + ' semanas desde aprobación de planos y cálculos de la ',
                TextBlock(InlineFont(b=True, i=True), 'totalidad de los equipos'),
                TextBlock(InlineFont(i=True),' amparados por el contrato, asi como la documentación crítica (Agosto y las dos últimas semanas de diciembre excluidos).'))
                ws["B13"] = rich_string

                rich_string = CellRichText(
                'Quotation prepared according to the information provided in the datasheet corresponding to each TAG. ',
                TextBlock(InlineFont(u='single'), 'EIPSA does not hold the final responsibility regarding selection of equipment material neither analyze process data.\n'),
                'The datasheet will be considered the only technical/contractual document, any other documentation will be considered as complementary documentation with informative purpose.')
                ws["B21"] = rich_string

                rich_string = CellRichText(
                'Esta oferta ha sido elaborada en base a la información que figura en la hoja de datos correspondiente a cada TAG, ',
                TextBlock(InlineFont(i=True, u='single'), 'no siendo responsabilidad final de EIPSA la elección del material a utilizar en los equipos ni analizar datos de proceso.\n'),
                TextBlock(InlineFont(i=True),'La hoja de datos será el único documento técnico/contractual, cualquier otra documentación recibida será considerada como documentación complementaria a efectos informativos.'))
                ws["B22"] = rich_string

                if pay_term == "100_delivery":
                    ws["B45"] = (
                        "100% of total amount of purchase order upon delivery of material according to Incoterms 2020, FCA (our facilities, Spain).\n"
                        "Payment method: bank transfer"
                    )
                    ws["B46"] = (
                        "Pago del 100% del valor total de la orden de compra a la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España).\n"
                        "Método de pago: Transferencia bancaria."
                    )
                elif pay_term == "100_order":
                    ws["B45"] = (
                        "100 % of the total amount of purchase order upon receipt of purchase order.\n"
                        "Payment method: bank transfer"
                    )
                    ws["B46"] = (
                        "Pago del 100% del valor total de la orden de compra a la recepción de la orden.\n"
                        "Método de pago: Transferencia bancaria"
                    )
                elif pay_term == "90_10":
                    ws["B45"] = (
                        "PAYMENT TERMS:\n"
                        "90 % of the total amount of PO upon delivery of material according to Incoterms 2020, FCA (our facilities, Spain) and 10% when final documentation is approved. \n"
                        "Bank Transfer: 60 days since invoice issue date."
                    )
                    ws["B46"] = (
                        "TERMINOS DE PAGO:\n"
                        "Pago del 90% del Valor total de la orden de compra a la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España) y el 10% restante cuando la documentación final sea aprobada.\n"
                        "Transferencia Bancaria: 60 días desde emisión de factura."
                    )
                elif pay_term == "50_50":
                    ws["B45"] = (
                        "50 % of the total amount of purchase order upon receipt of purchase order. Remaining 50% before material be delivered according to Incoterms 2020, FCA (our facilities, Spain).\n"
                        "Payment method: bank transfer."
                    )
                    ws["B46"] = (
                        "Pago del 50% del valor total de la orden de compra a la recepción de la orden. El 50% restante antes de la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España).\n"
                        "Método de pago: Transferencia bancaria."
                    )
                elif pay_term == "Others":
                    ws["B45"] = "PAYMENT TERMS TO BE DEFINED"
                    ws["B45"].font = Font(name="Calibri", size=11, bold=True, color="FF0000")
                    ws["B46"] = "TERMINOS DE PAGO POR DEFINIR"
                    ws["B46"].font = Font(name="Calibri", size=11, bold=True, italic=True, color="FF0000")

                rich_string = CellRichText(
                'For amounts greater than 30,000.00 € we can issue a warranty bond (if required) valid until the end of the indicated warranty period.\nBond warranty of 10% will be issued with the invoice of the last supplement.\n',
                TextBlock(InlineFont(b=True), 'For lower amounts no warranty bond is issued.'))
                ws["B42"] = rich_string

                rich_string = CellRichText(
                'Para importes superiores a 30.000,00 €, si es requerido, podremos emitir aval de garantía y estará vigente hasta el final del periodo de garantía indicado.\nEl aval del 10% será emitido con la factura del último suplemento.\n',
                TextBlock(InlineFont(i=True, b=True), 'Por debajo de dicha cantidad, no se emitirán avales.'))
                ws["B43"] = rich_string

                ws["A52"] = (
                    "Si necesita más información relacionada con esta oferta, no dude en ponerse en contacto con:\n"
                    + responsible
                    + "\n"
                    + email
                    + "\n"
                    "Telf.: (+34) 916.582.118"
                )

                for sheet in self.wb_commercial.sheetnames:
                    if sheet not in sheets_confirmed:
                        sheet_to_delete = self.wb_commercial[sheet]
                        self.wb_commercial.remove(sheet_to_delete)

                if int(rev) > 0:
                    sheets_confirmed = ["COVER", "1.2", "1.3", "NOTES"]
                    for sheet in sheets_confirmed:
                        sheet_to_delete = self.wb_commercial[sheet]
                        self.wb_commercial.remove(sheet_to_delete)

                left_text = "Fecha/Date: " + date_offer
                right_text = "Petición nº/Inquiry: " + num_ref

                for sheet in self.wb_commercial.worksheets:
                    sheet.oddFooter.left.text = left_text
                    sheet.oddFooter.right.text = right_text
                    sheet.oddFooter.center.text = "Page &P de &N"

                    sheet.oddFooter.left.size = 9
                    sheet.oddFooter.right.size = 9
                    sheet.oddFooter.center.size = 9
                
                path = self.save_excel_commercial()

                # Creating the technical offer using the commercial one as template
                self.wb_technical = load_workbook(path)

                if int(rev) == 0:
                    sheet_name = "COVER"
                    ws = self.wb_technical[sheet_name]
                    ws["E6"] = offername_technical

                for value_type in df["value_type"].unique():
                    eq_type = (
                        "FLOW ELEMENTS DATA"
                        if value_type == 1
                        else (
                            "NOZZLE ELEMENTS DATA"
                            if value_type == 2
                            else (
                                "VENTURI ELEMENTS DATA"
                                if value_type == 3
                                else (
                                    "WEDGE ELEMENTS DATA"
                                    if value_type == 4
                                    else (
                                        "PITOT ELEMENTS DATA"
                                        if value_type == 5
                                        else (
                                            "RO ELEMENTS DATA"
                                            if value_type == 6
                                            else "MULTISTAGE RO ELEMENTS DATA"
                                        )
                                    )
                                )
                            )
                        )
                    )

                    ws = self.wb_technical[eq_type]
                    if int(rev) == 0:
                        ws["G4"] = offername_technical

                    last_row = dict_sheets_data[eq_type][0]
                    num_column_amount = dict_sheets_data[eq_type][1]

                    # self.wb_technical[eq_type].delete_rows(last_row + 6, 20)
                    ws[f"M{last_row+5}"] = ""
                    ws[f"N{last_row+5}"] = ""

                    ws.cell(row=row_amount + 2, column=num_column_amount).value = ""
                    ws.cell(row=row_amount + 4, column=num_column_amount).value = ""
                    ws.cell(row=row_amount + 5, column=num_column_amount).value = ""
                    ws.cell(row=row_amount + 6, column=num_column_amount).value = ""
                    ws.cell(row=row_amount + 8, column=num_column_amount).value = ""

                    ws.cell(row=row_amount + 2, column=num_column_amount - 1).value = ""
                    ws.cell(row=row_amount + 4, column=num_column_amount - 1).value = ""
                    ws.cell(row=row_amount + 5, column=num_column_amount - 1).value = ""
                    ws.cell(row=row_amount + 6, column=num_column_amount - 1).value = ""
                    ws.cell(row=row_amount + 8, column=num_column_amount - 1).value = ""

                # Deleting "Amount" column
                    self.wb_technical[eq_type].delete_cols(num_column_amount, 1)

                # Adjusting the print area
                    new_last_column = num_column_amount
                    last_print_row = 40
                    nuevo_rango_impresion = f'A1:{get_column_letter(new_last_column)}{last_print_row}'
                    self.wb_technical[eq_type].print_area = nuevo_rango_impresion

                    stamp_1 = self.wb_technical[eq_type]._images[1]
                    anchor_actual = stamp_1.anchor

                    from_col, from_col_off = anchor_actual._from.col, anchor_actual._from.colOff
                    from_row, from_row_off = anchor_actual._from.row, anchor_actual._from.rowOff
                    to_col, to_col_off = anchor_actual.to.col, anchor_actual.to.colOff
                    to_row, to_row_off = anchor_actual.to.row, anchor_actual.to.rowOff

                    from_cell = AnchorMarker(col=from_col - 1, colOff=from_col_off, row=from_row, rowOff=from_row_off)
                    to_cell = AnchorMarker(col=to_col - 1, colOff=to_col_off, row=to_row, rowOff=to_row_off)

                    new_anchor = TwoCellAnchor(_from = from_cell, to = to_cell, editAs='absolute')
                    stamp_1.anchor = new_anchor

                    stamp_2 = self.wb_technical[eq_type]._images[2]
                    anchor_actual = stamp_2.anchor

                    from_col, from_col_off = anchor_actual._from.col, anchor_actual._from.colOff
                    from_row, from_row_off = anchor_actual._from.row, anchor_actual._from.rowOff
                    to_col, to_col_off = anchor_actual.to.col, anchor_actual.to.colOff
                    to_row, to_row_off = anchor_actual.to.row, anchor_actual.to.rowOff

                    from_cell = AnchorMarker(col=from_col - 1, colOff=from_col_off, row=from_row, rowOff=from_row_off)
                    to_cell = AnchorMarker(col=to_col - 1, colOff=to_col_off, row=to_row, rowOff=to_row_off)

                    new_anchor = TwoCellAnchor(_from = from_cell, to = to_cell, editAs='absolute')
                    stamp_2.anchor = new_anchor

                if int(rev) == 0:
                    ws = self.wb_technical[self.wb_technical.sheetnames[-2]]
                    ws.cell(row=last_row + 3, column=num_column_amount).value = number_items
                    ws.cell(row=last_row + 3, column=num_column_amount).font = Font(name="Calibri", size=14)

            # Editing sheet NOTES
                if int(rev) == 0:
                    sheet_name = "NOTES"  # Selecting  sheet
                    ws = self.wb_technical[sheet_name]

                    rich_string = CellRichText(
                    'We are only offering measuring flow elements. Please be informed that our product range includes temperature elements, and glass and magnetic level indicators; all with european certification. (https://www.eipsa.es/en/products)\n',
                    TextBlock(InlineFont(b=True), 'The prices quoted could be reduced in case of purchasing our full range of products.'))
                    ws["B6"] = rich_string

                    rich_string = CellRichText(
                    TextBlock(InlineFont(i=True), 'Estamos ofertando solamente elementos de medida de caudal, les informamos que en nuestra gama de fabricación con certificación europea, incluye también elementos de temperatura e indicadores de nivel de vidrio y magnéticos. (https://www.eipsa.es/productos)\n'),
                    TextBlock(InlineFont(b=True, i=True), 'Los precios ofertados podrían reducirse en caso de compra de toda nuestra gama.'))
                    ws["B7"] = rich_string

                    rich_string = CellRichText(
                    'Delivery time ' + delivery_time + ' weeks since drawing / calculation approval of ',
                    TextBlock(InlineFont(b=True), 'all equipment'),
                    ' in the contract, as well as critical documentation (August and last two December weeks excluded).')
                    ws["B12"] = rich_string

                    rich_string = CellRichText(
                    'Plazo de entrega ' + delivery_time + ' semanas desde aprobación de planos y cálculos de la ',
                    TextBlock(InlineFont(b=True, i=True), 'totalidad de los equipos'),
                    TextBlock(InlineFont(i=True),' amparados por el contrato, asi como la documentación crítica (Agosto y las dos últimas semanas de diciembre excluidos).'))
                    ws["B13"] = rich_string

                    rich_string = CellRichText(
                    'Quotation prepared according to the information provided in the datasheet corresponding to each TAG. ',
                    TextBlock(InlineFont(u='single'), 'EIPSA does not hold the final responsibility regarding selection of equipment material neither analyze process data.\n'),
                    'The datasheet will be considered the only technical/contractual document, any other documentation will be considered as complementary documentation with informative purpose.')
                    ws["B21"] = rich_string

                    rich_string = CellRichText(
                    'Esta oferta ha sido elaborada en base a la información que figura en la hoja de datos correspondiente a cada TAG, ',
                    TextBlock(InlineFont(i=True, u='single'), 'no siendo responsabilidad final de EIPSA la elección del material a utilizar en los equipos ni analizar datos de proceso.\n'),
                    TextBlock(InlineFont(i=True),'La hoja de datos será el único documento técnico/contractual, cualquier otra documentación recibida será considerada como documentación complementaria a efectos informativos.'))
                    ws["B22"] = rich_string

                    if pay_term == "100_delivery":
                        ws["B45"] = (
                            "100% of total amount of purchase order upon delivery of material according to Incoterms 2020, FCA (our facilities, Spain).\n"
                            "Payment method: bank transfer"
                        )
                        ws["B46"] = (
                            "Pago del 100% del valor total de la orden de compra a la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España).\n"
                            "Método de pago: Transferencia bancaria."
                        )
                    elif pay_term == "100_order":
                        ws["B45"] = (
                            "100 % of the total amount of purchase order upon receipt of purchase order.\n"
                            "Payment method: bank transfer"
                        )
                        ws["B46"] = (
                            "Pago del 100% del valor total de la orden de compra a la recepción de la orden.\n"
                            "Método de pago: Transferencia bancaria"
                        )
                    elif pay_term == "90_10":
                        ws["B45"] = (
                            "PAYMENT TERMS:\n"
                            "90 % of the total amount of PO upon delivery of material according to Incoterms 2020, FCA (our facilities, Spain) and 10% when final documentation is approved. \n"
                            "Bank Transfer: 60 days since invoice issue date."
                        )
                        ws["B46"] = (
                            "TERMINOS DE PAGO:\n"
                            "Pago del 90% del Valor total de la orden de compra a la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España) y el 10% restante cuando la documentación final sea aprobada.\n"
                            "Transferencia Bancaria: 60 días desde emisión de factura."
                        )
                    elif pay_term == "50_50":
                        ws["B45"] = (
                            "50 % of the total amount of purchase order upon receipt of purchase order. Remaining 50% before material be delivered according to Incoterms 2020, FCA (our facilities, Spain).\n"
                            "Payment method: bank transfer."
                        )
                        ws["B46"] = (
                            "Pago del 50% del valor total de la orden de compra a la recepción de la orden. El 50% restante antes de la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España).\n"
                            "Método de pago: Transferencia bancaria."
                        )
                    elif pay_term == "Others":
                        ws["B45"] = "PAYMENT TERMS TO BE DEFINED"
                        ws["B45"].font = Font(name="Calibri", size=11, bold=True, color="FF0000")
                        ws["B46"] = "TERMINOS DE PAGO POR DEFINIR"
                        ws["B46"].font = Font(name="Calibri", size=11, bold=True, italic=True, color="FF0000")

                    rich_string = CellRichText(
                    'For amounts greater than 30,000.00 € we can issue a warranty bond (if required) valid until the end of the indicated warranty period.\nBond warranty of 10% will be issued with the invoice of the last supplement.\n',
                    TextBlock(InlineFont(b=True), 'For lower amounts no warranty bond is issued.'))
                    ws["B42"] = rich_string

                    rich_string = CellRichText(
                    'Para importes superiores a 30.000,00 €, si es requerido, podremos emitir aval de garantía y estará vigente hasta el final del periodo de garantía indicado.\nEl aval del 10% será emitido con la factura del último suplemento.\n',
                    TextBlock(InlineFont(i=True, b=True), 'Por debajo de dicha cantidad, no se emitirán avales.'))
                    ws["B43"] = rich_string

                    ws["A52"] = (
                        "Si necesita más información relacionada con esta oferta, no dude en ponerse en contacto con:\n"
                        + responsible
                        + "\n"
                        + email
                        + "\n"
                        "Telf.: (+34) 916.582.118"
                    )

                    std = self.wb_technical["1.3"]
                    self.wb_technical.remove(std)

                left_text = "Fecha/Date: " + date_offer
                right_text = "Petición nº/Inquiry: " + num_ref

                for sheet in self.wb_technical.worksheets:
                    sheet.oddFooter.left.text = left_text
                    sheet.oddFooter.right.text = right_text
                    sheet.oddFooter.center.text = "Page &P de &N"

                    sheet.oddFooter.left.size = 9
                    sheet.oddFooter.right.size = 9
                    sheet.oddFooter.center.size = 9
                
                self.save_excel_technical()

                root = Tk()
                root.withdraw()  # Hiding main window Tkinter

                # close communication with the PostgreSQL database server
                # commit the changes
                conn.commit()
        except (Exception, psycopg2.DatabaseError) as error:
            dlg = QtWidgets.QMessageBox()
            new_icon = QtGui.QIcon()
            new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
            dlg.setWindowIcon(new_icon)
            dlg.setWindowTitle("ERP EIPSA")
            dlg.setText("Ha ocurrido el siguiente error:\n"
                        + str(error))
            dlg.setIcon(QtWidgets.QMessageBox.Icon.Critical)
            dlg.exec()
            del dlg, new_icon
        finally:
            if conn is not None:
                conn.close()

    def euros_to_float(self, value):
        """
        Converts a string value representing an amount in euros to a float.
        
        Args:
            value (str): The string representation of an amount in euros, with commas for decimal separation and ' €' for currency indication.
        
        Returns:
            float: The numeric value of the amount in euros.
        """
        value = value.replace(".", "")
        value = value.replace(",", ".")
        value = value[: value.find(" €")]
        return float(value)

    def save_excel_commercial(self):
        """Saves the populated Excel workbook to a specified location.
        Opens a dialog window for the user to select the file path and name.
        """
        # Dialog window to select folder and file name; if path is selected, excel file is saved
        output_path_commercial = asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Archivos de Excel", "*.xlsx")],
            title="Guardar Oferta comercial",
        )
        if output_path_commercial:
            self.wb_commercial.save(output_path_commercial)
            return output_path_commercial

    def save_excel_technical(self):
        """Saves the populated Excel workbook to a specified location.
        Opens a dialog window for the user to select the file path and name.
        """
        output_path_technical = asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Archivos de Excel", "*.xlsx")],
            title="Guardar Oferta técnica",
        )
        if output_path_technical:
            self.wb_technical.save(output_path_technical)

    def adjust_images(self, sheet):
        """
        Adjusts the width of all images in the provided spreadsheet sheet by decreasing each by 22 units.
        
        Args:
            sheet: The spreadsheet sheet containing images to be adjusted.
        """
        for image in sheet._images:
            image.width -= 22

class offer_temp:
    """
    A class to manage export offer details for temp equipments.
    
    Attributes:
        numoffer (str): Offer number.
        username (str): Name of the user creating the offer.
        rev (str): Revision number of the offer.
        project (str): Name of the project.
        delivery_term (str): Delivery terms for the offer.
        delivery_time (str): Expected delivery time.
        validity (str): Validity period of the offer.
        pay_term (str): Payment terms.
        testinspection (str): Information about testing and inspection.
        revchanges (str): Details of changes made in the revision.
        notes (str): Additional notes, split by line.
    """
    def __init__(self, numoffer, username, rev, project, delivery_term, delivery_time, validity, pay_term, testinspection, revchanges, notes):
        """
        Initializes the offer.

        Args:
            numoffer (str): Offer number.
            username (str): Name of the user creating the offer.
            rev (str): Revision number of the offer.
            project (str): Name of the project.
            delivery_term (str): Delivery terms for the offer.
            delivery_time (str): Expected delivery time.
            validity (str): Validity period of the offer.
            pay_term (str): Payment terms.
            testinspection (str): Information about testing and inspection.
            revchanges (str): Details of changes made in the revision.
            notes (str): Additional notes, split by line.
        """
        notes = notes.split('\n')
        date_offer = date.today().strftime("%d/%m/%Y")
        offername_commercial = numoffer + "-" + "Commercial.Rev" + rev
        offername_technical = numoffer + "-" + "Technical.Rev" + rev

        query_commercial = """
                    SELECT name, surname, email
                    FROM users_data.registration
                    WHERE username = %s
                    """
        query_dataoffer = """
                        SELECT client, num_ref_offer
                        FROM offers
                        WHERE UPPER (num_offer) LIKE UPPER('%%'||%s||'%%')
                        """
        query_tagsdata = """
                        SELECT *
                        FROM tags_data.tags_temp
                        WHERE (
                        UPPER ("num_offer") LIKE UPPER('%%'||%s||'%%')
                        AND
                        "tag_state" NOT IN ('PURCHASED','DELETED')
                        )
                        """
        conn = None
        try:
            # read the connection parameters
            params = config()
            # connect to the PostgreSQL server
            conn = psycopg2.connect(**params)
            cur = conn.cursor()

            cur.execute(query_dataoffer, (numoffer,))
            results_offer = cur.fetchall()
            client = results_offer[0][0]
            num_ref = results_offer[0][1]

            cur.execute(query_commercial, (username,))
            results_commercial = cur.fetchall()
            responsible = results_commercial[0][0] + " " + results_commercial[0][1]
            email = results_commercial[0][2]

            cur.execute(query_tagsdata, (numoffer,))
            data_tags = cur.fetchall()

            if len(data_tags) == 0:
                dlg = QtWidgets.QMessageBox()
                new_icon = QtGui.QIcon()
                new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
                dlg.setWindowIcon(new_icon)
                dlg.setWindowTitle("ERP EIPSA")
                dlg.setText("No hay TAGS importados en la oferta")
                dlg.setIcon(QtWidgets.QMessageBox.Icon.Warning)
                dlg.exec()
                del dlg, new_icon
            else:
                columns = []
                for elt in cur.description:
                    columns.append(elt[0])

                value_type_dict = {
                    "TW": 1,
                    "TW+TE": 2,
                    "TW+TE+TIT": 2,
                    "RETAINING FLANGE":2,
                    "TW+BIM": 3,
                    "TE": 4,
                    "BIM": 5,
                    "TIT": 6,
                    "SKIN+TT": 7,
                    "SKIN POINT": 7,
                    "Multi-T": 8
                }

                df = pd.DataFrame(data=data_tags, columns=columns)
                df["value_type"] = df["item_type"].map(value_type_dict)
                df = df.sort_values(by=["tag", "value_type"])
                df = df.iloc[:, 1:37]
                df["value_type"] = df["item_type"].map(value_type_dict)
                df = df.sort_values(by=["value_type", "tag"])
                df["amount"] = df["amount"].apply(self.euros_to_float)
                total_amount_material = df["amount"].sum()
                df = df.drop([
                        "tag_state",
                        "num_offer",
                        "num_order",
                        "num_po",
                        "position",
                        "subposition",
                        "std_tw",
                        "insulation"
                    ],
                    axis=1,)

                number_items = df.shape[0]
                documentation = number_items * 70

                # Loading Excel Template
                self.wb_commercial = load_workbook(
                    r"\\nas01\DATOS\Comunes\EIPSA-ERP\Plantillas Exportación\PLANTILLA OFERTA TEMPERATURA.xlsx"
                )

                # Editing sheet COVER
                sheet_name = "COVER"
                ws = self.wb_commercial[sheet_name]
                ws["E4"] = client
                ws["E6"] = offername_commercial
                ws["E8"] = num_ref
                ws["E10"] = project
                ws["E12"] = date_offer
                ws["E14"] = delivery_term
                ws["E16"] = validity + " days"
                ws["C43"] = responsible
                ws["C45"] = email

                # Editing sheet EQUIPMENT DATA
                sheets_confirmed = ["COVER", "1.2", "1.3", "NOTES"]

                if int(rev) == 0:
                    for item_type in df["item_type"].unique():
                        sheets_confirmed.append(item_type)

                dict_sheets_data = {}

                for value_type in df["value_type"].unique():
                    df_toexport = df[df["value_type"] == value_type]
                    df_toexport.insert(0, "N°", range(1, len(df_toexport) + 1))
                    df_toexport.set_index("N°", inplace=True, drop=False)
                    df_toexport.index.name = None
                    df_toexport = df_toexport.drop(["value_type"], axis=1)

                    eq_type = (
                        "TW ELEMENTS DATA"
                        if value_type == 1
                        else (
                            "TW+TE ELEMENTS DATA"
                            if value_type == 2
                            else (
                                "TW+BIM ELEMENTS DATA"
                                if value_type == 3
                                else (
                                    "TE ELEMENTS DATA"
                                    if value_type == 4
                                    else (
                                        "BIM ELEMENTS DATA"
                                        if value_type == 5
                                        else (
                                            "TIT ELEMENTS DATA"
                                            if value_type == 6
                                            else (
                                                "SKIN POINT ELEMENTS DATA"
                                                if value_type == 7
                                                else "MULTI-T ELEMENTS DATA"
                                            )
                                        )
                                    )
                                )
                            )
                        )
                    )

                    if eq_type not in sheets_confirmed:
                        sheets_confirmed.append(eq_type)

                    if eq_type == "TW ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "std_length",
                                "sensor_element",
                                "sheath_stem_material",
                                "sheath_stem_diam",
                                "temp_inf",
                                "temp_sup",
                                "nipple_ext_material",
                                "nipple_ext_length",
                                "head_case_material",
                                "elec_conn_case_diam",
                                "tt_cerblock"],
                            axis=1,)
                    elif eq_type == "TW+TE ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "std_length",
                                "temp_inf",
                                "temp_sup",
                                "nipple_ext_length"],
                            axis=1,)
                    elif eq_type == "TW+BIM ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "std_length",
                                "nipple_ext_length",
                                "tt_cerblock",
                                "puntal",
                                "tube_t"],
                            axis=1,)
                    elif eq_type == "TE ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "tw_type",
                                "size",
                                "rating",
                                "facing",
                                "material_tw",
                                "std_length",
                                "root_diam",
                                "tip_diam",
                                "temp_inf",
                                "temp_sup",
                                "nipple_ext_material",
                                "nipple_ext_length",
                                "head_case_material",
                                "elec_conn_case_diam",
                                "tt_cerblock",
                                "material_flange_lj",
                                "gasket_material",
                                "puntal",
                                "tube_t"],
                            axis=1,)
                    elif eq_type == "BIM ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "tw_type",
                                "size",
                                "rating",
                                "facing",
                                "material_tw",
                                "std_length",
                                "root_diam",
                                "tip_diam",
                                "nipple_ext_length",
                                "tt_cerblock",
                                "material_flange_lj",
                                "gasket_material",
                                "puntal",
                                "tube_t"],
                            axis=1,)
                    elif eq_type == "TIT ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "tw_type",
                                "size",
                                "rating",
                                "facing",
                                "material_tw",
                                "std_length",
                                "ins_length",
                                "root_diam",
                                "tip_diam",
                                "sensor_element",
                                "sheath_stem_material",
                                "sheath_stem_diam",
                                "nipple_ext_material",
                                "nipple_ext_length",
                                "head_case_material",
                                "material_flange_lj",
                                "gasket_material",
                                "puntal",
                                "tube_t"],
                            axis=1,)
                    elif (eq_type == "SKIN POINT ELEMENTS DATA" or eq_type == "SKIN+TT ELEMENTS DATA"):
                        df_toexport = df_toexport.drop([
                                "tw_type",
                                "size",
                                "rating",
                                "facing",
                                "material_tw",
                                "std_length",
                                "root_diam",
                                "tip_diam",
                                "temp_inf",
                                "temp_sup",
                                "material_flange_lj",
                                "gasket_material",
                                "puntal",
                                "tube_t"],
                            axis=1,)
                    elif eq_type == "MULTI-T ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "material_tw",
                                "root_diam",
                                "tip_diam",
                                "temp_inf",
                                "temp_sup",
                                "tt_cerblock",
                                "material_flange_lj",
                                "puntal",
                                "tube_t"],
                            axis=1,)

                    ws = self.wb_commercial[eq_type]
                    ws["G2"] = date_offer
                    ws["G3"] = num_ref
                    ws["G4"] = offername_commercial
                    if revchanges != "":
                        ws["G5"] = rev + " " + revchanges
                        ws["G5"].font = Font(name="Calibri", size=14, bold=True)
                        ws["G5"].fill = PatternFill("solid", fgColor="FFFF00")

                    if int(rev) > 0:
                        for row in ws.iter_rows(min_row=2, max_row=4, min_col=6, max_col=7):
                            for cell in row:
                                cell.value = None
                                cell._style = ws["F1"]._style

                    last_row = ws.max_row

                    num_column_amount = df_toexport.columns.get_loc("amount") + 1

                    for index, row in df_toexport.iterrows():  # Data in desired row
                        for col_num, value in enumerate(row, start=1):
                            cell = ws.cell(row=last_row + 1, column=col_num)
                            cell.value = value
                            if col_num == num_column_amount:
                                cell._style = ws["AG1"]._style
                            else:
                                cell._style = ws["AB1"]._style

                        last_row = ws.max_row

                    ws[f"A{last_row+3}"] = "OFFER VALIDITY: " + validity + " DAYS"
                    ws[f"A{last_row+3}"]._style = ws["AI1"]._style
                    ws[f"A{last_row+4}"] = (
                        "DELIVERY TIME: "
                        + delivery_time
                        + " WEEKS SINCE DRAWING / CALCULATION APPROVAL (AUGUST AND LAST TWO DECEMBER WEEKS EXCLUDED)"
                    )
                    ws[f"A{last_row+4}"]._style = ws["AI1"]._style

                    if notes != "":
                        if isinstance(notes, list):
                            line = last_row + 5
                            for note in notes:
                                ws[f"A{line}"] = note
                                ws[f"A{line}"]._style = ws["AI1"]._style
                                line += 1
                        else:
                            line = last_row + 5
                            ws[f"A{line}"] = notes
                            ws[f"A{line}"]._style = ws["AI1"]._style
                    dict_sheets_data[eq_type] = [last_row, num_column_amount, df_toexport["amount"].sum(), df_toexport.shape[0]]

                ws.cell(row=last_row + 3, column=num_column_amount - 1).value = "QTY. TOTAL"
                ws.cell(row=last_row + 3, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=last_row + 3, column=num_column_amount).value = number_items
                
                row_amount = last_row + 4
                for key, value in dict_sheets_data.items():
                    parts_key = key.split(" ")
                    ws.cell(row=row_amount + 2, column=num_column_amount - 1).value = "TOTAL AMOUNT OF " + parts_key[0] + " " + parts_key[1] + " (QTY: " + str(value[3]) + ")"
                    ws.cell(row=row_amount + 2, column=num_column_amount).value = value[2]
                    ws.cell(row=row_amount + 2, column=num_column_amount - 1)._style = ws["AA1"]._style
                    ws.cell(row=row_amount + 2, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                    ws.cell(row=row_amount + 2, column=num_column_amount)._style = ws["AC1"]._style

                    row_amount += 2

                ws.cell(row=row_amount + 2, column=num_column_amount - 1).value = "TOTAL AMOUNT OF MATERIAL"
                ws.cell(row=row_amount + 2, column=num_column_amount).value = total_amount_material
                ws.cell(row=row_amount + 4, column=num_column_amount - 1).value = "PACKING AND TRANSPORT"
                ws.cell(row=row_amount + 4, column=num_column_amount).value = (f"=MROUND({get_column_letter(num_column_amount)}{row_amount + 2}*0.03,10)" if total_amount_material > 6700 else 200)
                ws.cell(row=row_amount + 5, column=num_column_amount - 1).value = "TESTS & INSPECTION"
                ws.cell(row=row_amount + 5, column=num_column_amount).value = float(testinspection)
                ws.cell(row=row_amount + 6, column=num_column_amount - 1).value = "DOCUMENTATION"
                ws.cell(row=row_amount + 6, column=num_column_amount).value = documentation
                ws.cell(row=row_amount + 8, column=num_column_amount - 1).value = "TOTAL AMOUNT OF BID"
                ws.cell(row=row_amount + 8, column=num_column_amount).value = f"=SUM({get_column_letter(num_column_amount)}{row_amount + 2}:{get_column_letter(num_column_amount)}{row_amount + 6})"

                ws.cell(row=last_row + 3, column=num_column_amount - 1)._style = ws["AA1"]._style
                ws.cell(row=last_row + 3, column=num_column_amount).font = Font(name="Calibri", size=14)
                ws.cell(row=row_amount + 2, column=num_column_amount - 1)._style = ws["AA1"]._style
                ws.cell(row=row_amount + 2, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=row_amount + 2, column=num_column_amount)._style = ws["AC1"]._style
                ws.cell(row=row_amount + 4, column=num_column_amount - 1).font = Font(name="Calibri", size=14)
                ws.cell(row=row_amount + 4, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=row_amount + 4, column=num_column_amount)._style = ws["AC1"]._style
                ws.cell(row=row_amount + 5, column=num_column_amount - 1).font = Font(name="Calibri", size=14)
                ws.cell(row=row_amount + 5, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=row_amount + 5, column=num_column_amount)._style = ws["AC1"]._style
                ws.cell(row=row_amount + 6, column=num_column_amount - 2)._style = ws["AD1"]._style
                ws.cell(row=row_amount + 6, column=num_column_amount - 1)._style = ws["AD1"]._style
                ws.cell(row=row_amount + 6, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=row_amount + 6, column=num_column_amount)._style = ws["AE1"]._style
                ws.cell(row=row_amount + 8, column=num_column_amount - 1)._style = ws["AA1"]._style
                ws.cell(row=row_amount + 8, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=row_amount + 8, column=num_column_amount)._style = ws["AF1"]._style

            # Editing sheet NOTES
                sheet_name = "NOTES"  # Selecting  sheet
                ws = self.wb_commercial[sheet_name]

                rich_string = CellRichText(
                'We are only offering measuring temperature elements. Please be informed that our product range includes flow elements, and glass and magentic level indicators; all with european certification. (https://www.eipsa.es/en/products)\n',
                TextBlock(InlineFont(b=True), 'The prices quoted could be reduced in case of purchasing our full range of products.'))
                ws["B6"] = rich_string

                rich_string = CellRichText(
                TextBlock(InlineFont(i=True), 'Estamos ofertando solamente elementos de medida de temperatura, les informamos que en nuestra gama de fabricación con certificación europea, incluye también elementos de caudal e indicadores de nivel de vidrio y magnéticos. (https://www.eipsa.es/productos)\n'),
                TextBlock(InlineFont(b=True, i=True), 'Los precios ofertados podrían reducirse en caso de compra de toda nuestra gama.'))
                ws["B7"] = rich_string

                rich_string = CellRichText(
                'Delivery time ' + delivery_time + ' weeks since drawing / calculation approval of ',
                TextBlock(InlineFont(b=True), 'all equipment'),
                ' in the contract, as well as critical documentation (August and last two December weeks excluded).')
                ws["B12"] = rich_string

                rich_string = CellRichText(
                'Plazo de entrega ' + delivery_time + ' semanas desde aprobación de planos y cálculos de la ',
                TextBlock(InlineFont(b=True, i=True), 'totalidad de los equipos'),
                TextBlock(InlineFont(i=True),' amparados por el contrato, asi como la documentación crítica (Agosto y las dos últimas semanas de diciembre excluidos).'))
                ws["B13"] = rich_string

                rich_string = CellRichText(
                'Quotation prepared according to the information provided in the datasheet corresponding to each TAG. ',
                TextBlock(InlineFont(u='single'), 'EIPSA does not hold the final responsibility regarding selection of equipment material neither analyze process data.\n'),
                'The datasheet will be considered the only technical/contractual document, any other documentation will be considered as complementary documentation with informative purpose.')
                ws["B21"] = rich_string

                rich_string = CellRichText(
                'Esta oferta ha sido elaborada en base a la información que figura en la hoja de datos correspondiente a cada TAG, ',
                TextBlock(InlineFont(i=True, u='single'), 'no siendo responsabilidad final de EIPSA la elección del material a utilizar en los equipos ni analizar datos de proceso.\n'),
                TextBlock(InlineFont(i=True),'La hoja de datos será el único documento técnico/contractual, cualquier otra documentación recibida será considerada como documentación complementaria a efectos informativos.'))
                ws["B22"] = rich_string

                if pay_term == "100_delivery":
                    ws["B45"] = (
                        "100% of total amount of purchase order upon delivery of material according to Incoterms 2020, FCA (our facilities, Spain).\n"
                        "Payment method: bank transfer"
                    )
                    ws["B46"] = (
                        "Pago del 100% del valor total de la orden de compra a la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España).\n"
                        "Método de pago: Transferencia bancaria."
                    )
                elif pay_term == "100_order":
                    ws["B45"] = (
                        "100 % of the total amount of purchase order upon receipt of purchase order.\n"
                        "Payment method: bank transfer"
                    )
                    ws["B46"] = (
                        "Pago del 100% del valor total de la orden de compra a la recepción de la orden.\n"
                        "Método de pago: Transferencia bancaria"
                    )
                elif pay_term == "90_10":
                    ws["B45"] = (
                        "PAYMENT TERMS:\n"
                        "90 % of the total amount of PO upon delivery of material according to Incoterms 2020, FCA (our facilities, Spain) and 10% when final documentation is approved. \n"
                        "Bank Transfer: 60 days since invoice issue date."
                    )
                    ws["B46"] = (
                        "TERMINOS DE PAGO:\n"
                        "Pago del 90% del Valor total de la orden de compra a la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España) y el 10% restante cuando la documentación final sea aprobada.\n"
                        "Transferencia Bancaria: 60 días desde emisión de factura."
                    )
                elif pay_term == "50_50":
                    ws["B45"] = (
                        "50 % of the total amount of purchase order upon receipt of purchase order. Remaining 50% before material be delivered according to Incoterms 2020, FCA (our facilities, Spain).\n"
                        "Payment method: bank transfer."
                    )
                    ws["B46"] = (
                        "Pago del 50% del valor total de la orden de compra a la recepción de la orden. El 50% restante antes de la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España).\n"
                        "Método de pago: Transferencia bancaria."
                    )
                elif pay_term == "Others":
                    ws["B45"] = "PAYMENT TERMS TO BE DEFINED"
                    ws["B45"].font = Font(name="Calibri", size=11, bold=True, color="FF0000")
                    ws["B46"] = "TERMINOS DE PAGO POR DEFINIR"
                    ws["B46"].font = Font(name="Calibri", size=11, bold=True, italic=True, color="FF0000")

                rich_string = CellRichText(
                'For amounts greater than 30,000.00 € we can issue a warranty bond (if required) valid until the end of the indicated warranty period.\nBond warranty of 10% will be issued with the invoice of the last supplement.\n',
                TextBlock(InlineFont(b=True), 'For lower amounts no warranty bond is issued.'))
                ws["B48"] = rich_string

                rich_string = CellRichText(
                'Para importes superiores a 30.000,00 €, si es requerido, podremos emitir aval de garantía y estará vigente hasta el final del periodo de garantía indicado.\nEl aval del 10% será emitido con la factura del último suplemento.\n',
                TextBlock(InlineFont(i=True, b=True), 'Por debajo de dicha cantidad, no se emitirán avales.'))
                ws["B49"] = rich_string

                ws["A61"] = (
                    "If you require further information related with this offer, please do not hesitate to contact:\n"
                    + responsible
                    + "\n"
                    + email
                    + "\n"
                    "Telf.: (+34) 916.582.118"
                )

                for sheet in self.wb_commercial.sheetnames:
                    if sheet not in sheets_confirmed:
                        sheet_to_delete = self.wb_commercial[sheet]
                        self.wb_commercial.remove(sheet_to_delete)

                if int(rev) > 0:
                    sheets_confirmed = ["COVER", "1.2", "1.3", "NOTES"]
                    for sheet in sheets_confirmed:
                        sheet_to_delete = self.wb_commercial[sheet]
                        self.wb_commercial.remove(sheet_to_delete)

                left_text = "Fecha/Date: " + date_offer
                right_text = "Petición nº/Inquiry: " + num_ref

                for sheet in self.wb_commercial.worksheets:
                    sheet.oddFooter.left.text = left_text
                    sheet.oddFooter.right.text = right_text
                    sheet.oddFooter.center.text = "Page &P de &N"

                    sheet.oddFooter.left.size = 9
                    sheet.oddFooter.right.size = 9
                    sheet.oddFooter.center.size = 9
                
                path = self.save_excel_commercial()

                # Creating the technical offer using the commercial one as template
                self.wb_technical = load_workbook(path)

                if int(rev) == 0:
                    sheet_name = "COVER"
                    ws = self.wb_technical[sheet_name]
                    ws["E6"] = offername_technical

                for value_type in df["value_type"].unique():
                    eq_type = (
                        "TW ELEMENTS DATA"
                        if value_type == 1
                        else (
                            "TW+TE ELEMENTS DATA"
                            if value_type == 2
                            else (
                                "TW+BIM ELEMENTS DATA"
                                if value_type == 3
                                else (
                                    "TE ELEMENTS DATA"
                                    if value_type == 4
                                    else (
                                        "BIM ELEMENTS DATA"
                                        if value_type == 5
                                        else (
                                            "TIT ELEMENTS DATA"
                                            if value_type == 6
                                            else (
                                                "SKIN POINT ELEMENTS DATA"
                                                if value_type == 7
                                                else "MULTI-T ELEMENTS DATA"
                                            )
                                        )
                                    )
                                )
                            )
                        )
                    )

                    ws = self.wb_technical[eq_type]
                    if int(rev) == 0:
                        ws["G4"] = offername_technical

                    last_row = dict_sheets_data[eq_type][0]
                    num_column_amount = dict_sheets_data[eq_type][1]

                    # self.wb_technical[eq_type].delete_rows(last_row + 5, 20)

                    ws.cell(row=row_amount + 2, column=num_column_amount).value = ""
                    ws.cell(row=row_amount + 4, column=num_column_amount).value = ""
                    ws.cell(row=row_amount + 5, column=num_column_amount).value = ""
                    ws.cell(row=row_amount + 6, column=num_column_amount).value = ""
                    ws.cell(row=row_amount + 8, column=num_column_amount).value = ""

                    ws.cell(row=row_amount + 2, column=num_column_amount - 1).value = ""
                    ws.cell(row=row_amount + 4, column=num_column_amount - 1).value = ""
                    ws.cell(row=row_amount + 5, column=num_column_amount - 1).value = ""
                    ws.cell(row=row_amount + 6, column=num_column_amount - 1).value = ""
                    ws.cell(row=row_amount + 8, column=num_column_amount - 1).value = ""

                # Deleting "Amount" column
                    self.wb_technical[eq_type].delete_cols(num_column_amount, 1)

                # Adjusting the print area
                    new_last_column = num_column_amount
                    last_print_row = 40
                    nuevo_rango_impresion = f'A1:{get_column_letter(new_last_column)}{last_print_row}'
                    self.wb_technical[eq_type].print_area = nuevo_rango_impresion

                    stamp_1 = self.wb_technical[eq_type]._images[1]
                    anchor_actual = stamp_1.anchor

                    from_col, from_col_off = anchor_actual._from.col, anchor_actual._from.colOff
                    from_row, from_row_off = anchor_actual._from.row, anchor_actual._from.rowOff
                    to_col, to_col_off = anchor_actual.to.col, anchor_actual.to.colOff
                    to_row, to_row_off = anchor_actual.to.row, anchor_actual.to.rowOff

                    from_cell = AnchorMarker(col=from_col - 1, colOff=from_col_off, row=from_row, rowOff=from_row_off)
                    to_cell = AnchorMarker(col=to_col - 1, colOff=to_col_off, row=to_row, rowOff=to_row_off)

                    new_anchor = TwoCellAnchor(_from = from_cell, to = to_cell, editAs='absolute')
                    stamp_1.anchor = new_anchor

                    stamp_2 = self.wb_technical[eq_type]._images[2]
                    anchor_actual = stamp_2.anchor

                    from_col, from_col_off = anchor_actual._from.col, anchor_actual._from.colOff
                    from_row, from_row_off = anchor_actual._from.row, anchor_actual._from.rowOff
                    to_col, to_col_off = anchor_actual.to.col, anchor_actual.to.colOff
                    to_row, to_row_off = anchor_actual.to.row, anchor_actual.to.rowOff

                    from_cell = AnchorMarker(col=from_col - 1, colOff=from_col_off, row=from_row, rowOff=from_row_off)
                    to_cell = AnchorMarker(col=to_col - 1, colOff=to_col_off, row=to_row, rowOff=to_row_off)

                    new_anchor = TwoCellAnchor(_from = from_cell, to = to_cell, editAs='absolute')
                    stamp_2.anchor = new_anchor

                if int(rev) == 0:
                    ws = self.wb_technical[self.wb_technical.sheetnames[-2]]
                    ws.cell(row=last_row + 3, column=num_column_amount).value = number_items
                    ws.cell(row=last_row + 3, column=num_column_amount).font = Font(name="Calibri", size=14)

            # Editing sheet NOTES
                if int(rev) == 0:
                    sheet_name = "NOTES"
                    ws = self.wb_technical[sheet_name]

                    rich_string = CellRichText(
                    'We are only offering measuring temperature elements. Please be informed that our product range includes flow elements, and glass and magentic level indicators; all with european certification. (https://www.eipsa.es/en/products)\n',
                    TextBlock(InlineFont(b=True), 'The prices quoted could be reduced in case of purchasing our full range of products.'))
                    ws["B6"] = rich_string

                    rich_string = CellRichText(
                    TextBlock(InlineFont(i=True), 'Estamos ofertando solamente elementos de medida de temperatura, les informamos que en nuestra gama de fabricación con certificación europea, incluye también elementos de caudal e indicadores de nivel de vidrio y magnéticos. (https://www.eipsa.es/productos)\n'),
                    TextBlock(InlineFont(b=True, i=True), 'Los precios ofertados podrían reducirse en caso de compra de toda nuestra gama.'))
                    ws["B7"] = rich_string

                    rich_string = CellRichText(
                    'Delivery time ' + delivery_time + ' weeks since drawing / calculation approval of ',
                    TextBlock(InlineFont(b=True), 'all equipment'),
                    ' in the contract, as well as critical documentation (August and last two December weeks excluded).')
                    ws["B12"] = rich_string

                    rich_string = CellRichText(
                    'Plazo de entrega ' + delivery_time + ' semanas desde aprobación de planos y cálculos de la ',
                    TextBlock(InlineFont(b=True, i=True), 'totalidad de los equipos'),
                    TextBlock(InlineFont(i=True),' amparados por el contrato, asi como la documentación crítica (Agosto y las dos últimas semanas de diciembre excluidos).'))
                    ws["B13"] = rich_string

                    rich_string = CellRichText(
                    'Quotation prepared according to the information provided in the datasheet corresponding to each TAG. ',
                    TextBlock(InlineFont(u='single'), 'EIPSA does not hold the final responsibility regarding selection of equipment material neither analyze process data.\n'),
                    'The datasheet will be considered the only technical/contractual document, any other documentation will be considered as complementary documentation with informative purpose.')
                    ws["B21"] = rich_string

                    rich_string = CellRichText(
                    'Esta oferta ha sido elaborada en base a la información que figura en la hoja de datos correspondiente a cada TAG, ',
                    TextBlock(InlineFont(i=True, u='single'), 'no siendo responsabilidad final de EIPSA la elección del material a utilizar en los equipos ni analizar datos de proceso.\n'),
                    TextBlock(InlineFont(i=True),'La hoja de datos será el único documento técnico/contractual, cualquier otra documentación recibida será considerada como documentación complementaria a efectos informativos.'))
                    ws["B22"] = rich_string

                    if pay_term == "100_delivery":
                        ws["B45"] = (
                            "100% of total amount of purchase order upon delivery of material according to Incoterms 2020, FCA (our facilities, Spain).\n"
                            "Payment method: bank transfer"
                        )
                        ws["B46"] = (
                            "Pago del 100% del valor total de la orden de compra a la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España).\n"
                            "Método de pago: Transferencia bancaria."
                        )
                    elif pay_term == "100_order":
                        ws["B45"] = (
                            "100 % of the total amount of purchase order upon receipt of purchase order.\n"
                            "Payment method: bank transfer"
                        )
                        ws["B46"] = (
                            "Pago del 100% del valor total de la orden de compra a la recepción de la orden.\n"
                            "Método de pago: Transferencia bancaria"
                        )
                    elif pay_term == "90_10":
                        ws["B45"] = (
                            "PAYMENT TERMS:\n"
                            "90 % of the total amount of PO upon delivery of material according to Incoterms 2020, FCA (our facilities, Spain) and 10% when final documentation is approved. \n"
                            "Bank Transfer: 60 days since invoice issue date."
                        )
                        ws["B46"] = (
                            "TERMINOS DE PAGO:\n"
                            "Pago del 90% del Valor total de la orden de compra a la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España) y el 10% restante cuando la documentación final sea aprobada.\n"
                            "Transferencia Bancaria: 60 días desde emisión de factura."
                        )
                    elif pay_term == "50_50":
                        ws["B45"] = (
                            "50 % of the total amount of purchase order upon receipt of purchase order. Remaining 50% before material be delivered according to Incoterms 2020, FCA (our facilities, Spain).\n"
                            "Payment method: bank transfer."
                        )
                        ws["B46"] = (
                            "Pago del 50% del valor total de la orden de compra a la recepción de la orden. El 50% restante antes de la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España).\n"
                            "Método de pago: Transferencia bancaria."
                        )
                    elif pay_term == "Others":
                        ws["B45"] = "PAYMENT TERMS TO BE DEFINED"
                        ws["B45"].font = Font(name="Calibri", size=11, bold=True, color="FF0000")
                        ws["B46"] = "TERMINOS DE PAGO POR DEFINIR"
                        ws["B46"].font = Font(name="Calibri", size=11, bold=True, italic=True, color="FF0000")

                    rich_string = CellRichText(
                    'For amounts greater than 30,000.00 € we can issue a warranty bond (if required) valid until the end of the indicated warranty period.\nBond warranty of 10% will be issued with the invoice of the last supplement.\n',
                    TextBlock(InlineFont(b=True), 'For lower amounts no warranty bond is issued.'))
                    ws["B48"] = rich_string

                    rich_string = CellRichText(
                    'Para importes superiores a 30.000,00 €, si es requerido, podremos emitir aval de garantía y estará vigente hasta el final del periodo de garantía indicado.\nEl aval del 10% será emitido con la factura del último suplemento.\n',
                    TextBlock(InlineFont(i=True, b=True), 'Por debajo de dicha cantidad, no se emitirán avales.'))
                    ws["B49"] = rich_string

                    ws["A61"] = (
                        "If you require further information related with this offer, please do not hesitate to contact:\n"
                        + responsible
                        + "\n"
                        + email
                        + "\n"
                        "Telf.: (+34) 916.582.118"
                    )
                    
                    std = self.wb_technical["1.3"]
                    self.wb_technical.remove(std)

                left_text = "Fecha/Date: " + date_offer
                right_text = "Petición nº/Inquiry: " + num_ref

                for sheet in self.wb_technical.worksheets:
                    sheet.oddFooter.left.text = left_text
                    sheet.oddFooter.right.text = right_text
                    sheet.oddFooter.center.text = "Page &P de &N"

                    sheet.oddFooter.left.size = 9
                    sheet.oddFooter.right.size = 9
                    sheet.oddFooter.center.size = 9
                
                self.save_excel_technical()

                root = Tk()
                root.withdraw()  # Hiding main window Tkinter

                # close communication with the PostgreSQL database server
                # commit the changes
                conn.commit()
        except (Exception, psycopg2.DatabaseError) as error:
            dlg = QtWidgets.QMessageBox()
            new_icon = QtGui.QIcon()
            new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
            dlg.setWindowIcon(new_icon)
            dlg.setWindowTitle("ERP EIPSA")
            dlg.setText("Ha ocurrido el siguiente error:\n"
                        + str(error))
            dlg.setIcon(QtWidgets.QMessageBox.Icon.Critical)
            dlg.exec()
            del dlg, new_icon
        finally:
            if conn is not None:
                conn.close()

    def euros_to_float(self, value):
        """
        Converts a string value representing an amount in euros to a float.
        
        Args:
            value (str): The string representation of an amount in euros, with commas for decimal separation and ' €' for currency indication.
        
        Returns:
            float: The numeric value of the amount in euros.
        """
        value = value.replace(".", "")
        value = value.replace(",", ".")
        value = value[: value.find(" €")]
        return float(value)

    def save_excel_commercial(self):
        """Saves the populated Excel workbook to a specified location.
        Opens a dialog window for the user to select the file path and name.
        """
        # Dialog window to select folder and file name; if path is selected, excel file is saved
        output_path_commercial = asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Archivos de Excel", "*.xlsx")],
            title="Guardar Oferta comercial",
        )
        if output_path_commercial:
            self.wb_commercial.save(output_path_commercial)
            return output_path_commercial

    def save_excel_technical(self):
        """Saves the populated Excel workbook to a specified location.
        Opens a dialog window for the user to select the file path and name.
        """
        output_path_technical = asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Archivos de Excel", "*.xlsx")],
            title="Guardar Oferta técnica",
        )
        if output_path_technical:
            self.wb_technical.save(output_path_technical)

class offer_short_temp_spanish:
    """
    A class to manage export offer (short format in spanish) details for temp equipments.
    
    Attributes:
        numoffer (str): Offer number.
        username (str): Name of the user creating the offer.
        rev (str): Revision number of the offer.
        project (str): Name of the project.
        delivery_term (str): Delivery terms for the offer.
        delivery_time (str): Expected delivery time.
        validity (str): Validity period of the offer.
        pay_term (str): Payment terms.
        testinspection (str): Information about testing and inspection.
        revchanges (str): Details of changes made in the revision.
        notes (str): Additional notes, split by line.
    """
    def __init__(self, numoffer, username, rev, project, delivery_term, delivery_time, validity, pay_term, testinspection, revchanges, notes):
        """
        Initializes the offer.

        Args:
            numoffer (str): Offer number.
            username (str): Name of the user creating the offer.
            rev (str): Revision number of the offer.
            project (str): Name of the project.
            delivery_term (str): Delivery terms for the offer.
            delivery_time (str): Expected delivery time.
            validity (str): Validity period of the offer.
            pay_term (str): Payment terms.
            testinspection (str): Information about testing and inspection.
            revchanges (str): Details of changes made in the revision.
            notes (str): Additional notes, split by line.
        """
        date_offer = date.today().strftime("%d/%m/%Y")
        offername_commercial = numoffer + "-" + "Commercial.Rev" + rev
        offername_technical = numoffer + "-" + "Technical.Rev" + rev

        query_commercial = """
                    SELECT name, surname, email
                    FROM users_data.registration
                    WHERE username = %s
                    """
        query_dataoffer = """
                        SELECT client, num_ref_offer
                        FROM offers
                        WHERE UPPER (num_offer) LIKE UPPER('%%'||%s||'%%')
                        """
        query_tagsdata = """
                        SELECT *
                        FROM tags_data.tags_temp
                        WHERE (
                        UPPER ("num_offer") LIKE UPPER('%%'||%s||'%%')
                        AND
                        "tag_state" NOT IN ('PURCHASED','DELETED')
                        )
                        """
        conn = None
        try:
            # read the connection parameters
            params = config()
            # connect to the PostgreSQL server
            conn = psycopg2.connect(**params)
            cur = conn.cursor()

            cur.execute(query_dataoffer, (numoffer,))
            results_offer = cur.fetchall()
            client = results_offer[0][0]
            num_ref = results_offer[0][1]

            cur.execute(query_commercial, (username,))
            results_commercial = cur.fetchall()
            responsible = results_commercial[0][0] + " " + results_commercial[0][1]
            email = results_commercial[0][2]

            cur.execute(query_tagsdata, (numoffer,))
            data_tags = cur.fetchall()

            if len(data_tags) == 0:
                dlg = QtWidgets.QMessageBox()
                new_icon = QtGui.QIcon()
                new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
                dlg.setWindowIcon(new_icon)
                dlg.setWindowTitle("ERP EIPSA")
                dlg.setText("No hay TAGS importados en la oferta")
                dlg.setIcon(QtWidgets.QMessageBox.Icon.Warning)
                dlg.exec()
                del dlg, new_icon
            else:
                columns = []
                for elt in cur.description:
                    columns.append(elt[0])

                value_type_dict = {
                    "TW": 1,
                    "TW+TE": 2,
                    "TW+TE+TIT": 2,
                    "RETAINING FLANGE":2,
                    "TW+BIM": 3,
                    "TE": 4,
                    "BIM": 5,
                    "TIT": 6,
                    "SKIN+TT": 7,
                    "SKIN POINT": 7,
                    "Multi-T": 8
                }

                df = pd.DataFrame(data=data_tags, columns=columns)
                df["value_type"] = df["item_type"].map(value_type_dict)
                df = df.sort_values(by=["tag", "value_type"])
                df = df.iloc[:, 1:37]
                df["value_type"] = df["item_type"].map(value_type_dict)
                df = df.sort_values(by=["value_type", "tag"])
                df["amount"] = df["amount"].apply(self.euros_to_float)
                total_amount_material = df["amount"].sum()
                df = df.drop([
                        "tag_state",
                        "num_offer",
                        "num_order",
                        "num_po",
                        "position",
                        "subposition",
                        "std_tw",
                        "insulation"
                    ],
                    axis=1,)

                number_items = df.shape[0]
                documentation = number_items * 70

                # Loading Excel Template
                self.wb_commercial = load_workbook(
                    r"\\nas01\DATOS\Comunes\EIPSA-ERP\Plantillas Exportación\PLANTILLA CORTA OFERTA TEMPERATURA.xlsx"
                )

                # Editing sheet COVER
                sheet_name = "COVER"
                ws = self.wb_commercial[sheet_name]
                ws["E4"] = client
                ws["E6"] = offername_commercial
                ws["E8"] = num_ref
                ws["E10"] = project
                ws["E12"] = date_offer
                ws["E14"] = delivery_term
                ws["E16"] = validity + " days"
                ws["C43"] = responsible
                ws["C45"] = email

                # Editing sheet EQUIPMENT DATA
                sheets_confirmed = ["COVER", "1.2", "1.3", "NOTES"]

                if int(rev) == 0:
                    for item_type in df["item_type"].unique():
                        sheets_confirmed.append(item_type)

                dict_sheets_data = {}

                for value_type in df["value_type"].unique():
                    df_toexport = df[df["value_type"] == value_type]
                    df_toexport.insert(0, "N°", range(1, len(df_toexport) + 1))
                    df_toexport.set_index("N°", inplace=True, drop=False)
                    df_toexport.index.name = None
                    df_toexport = df_toexport.drop(["value_type"], axis=1)

                    eq_type = (
                        "TW ELEMENTS DATA"
                        if value_type == 1
                        else (
                            "TW+TE ELEMENTS DATA"
                            if value_type == 2
                            else (
                                "TW+BIM ELEMENTS DATA"
                                if value_type == 3
                                else (
                                    "TE ELEMENTS DATA"
                                    if value_type == 4
                                    else (
                                        "BIM ELEMENTS DATA"
                                        if value_type == 5
                                        else (
                                            "TIT ELEMENTS DATA"
                                            if value_type == 6
                                            else (
                                                "SKIN POINT ELEMENTS DATA"
                                                if value_type == 7
                                                else "MULTI-T RO ELEMENTS DATA"
                                            )
                                        )
                                    )
                                )
                            )
                        )
                    )

                    if eq_type not in sheets_confirmed:
                        sheets_confirmed.append(eq_type)

                    if eq_type == "TW ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "std_length",
                                "sensor_element",
                                "sheath_stem_material",
                                "sheath_stem_diam",
                                "temp_inf",
                                "temp_sup",
                                "nipple_ext_material",
                                "nipple_ext_length",
                                "head_case_material",
                                "elec_conn_case_diam",
                                "tt_cerblock"],
                            axis=1,)
                    elif eq_type == "TW+TE ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "std_length",
                                "temp_inf",
                                "temp_sup",
                                "nipple_ext_length"],
                            axis=1,)
                    elif eq_type == "TW+BIM ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "std_length",
                                "nipple_ext_length",
                                "tt_cerblock",
                                "puntal",
                                "tube_t"],
                            axis=1,)
                    elif eq_type == "TE ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "tw_type",
                                "size",
                                "rating",
                                "facing",
                                "material_tw",
                                "std_length",
                                "root_diam",
                                "tip_diam",
                                "temp_inf",
                                "temp_sup",
                                "nipple_ext_material",
                                "nipple_ext_length",
                                "head_case_material",
                                "elec_conn_case_diam",
                                "tt_cerblock",
                                "material_flange_lj",
                                "gasket_material",
                                "puntal",
                                "tube_t"],
                            axis=1,)
                    elif eq_type == "BIM ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "tw_type",
                                "size",
                                "rating",
                                "facing",
                                "material_tw",
                                "std_length",
                                "root_diam",
                                "tip_diam",
                                "nipple_ext_length",
                                "tt_cerblock",
                                "material_flange_lj",
                                "gasket_material",
                                "puntal",
                                "tube_t"],
                            axis=1,)
                    elif eq_type == "TIT ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "tw_type",
                                "size",
                                "rating",
                                "facing",
                                "material_tw",
                                "std_length",
                                "root_diam",
                                "tip_diam",
                                "sensor_element",
                                "sheath_stem_material",
                                "sheath_stem_diam",
                                "nipple_ext_material",
                                "nipple_ext_length",
                                "head_case_material",
                                "material_flange_lj",
                                "gasket_material",
                                "puntal",
                                "tube_t"],
                            axis=1,)
                    elif (eq_type == "SKIN POINT ELEMENTS DATA" or eq_type == "SKIN+TT ELEMENTS DATA"):
                        df_toexport = df_toexport.drop([
                                "tw_type",
                                "size",
                                "rating",
                                "facing",
                                "material_tw",
                                "std_length",
                                "root_diam",
                                "tip_diam",
                                "temp_inf",
                                "temp_sup",
                                "material_flange_lj",
                                "gasket_material",
                                "puntal",
                                "tube_t"],
                            axis=1,)
                    elif eq_type == "MULTI-T ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "material_tw",
                                "root_diam",
                                "tip_diam",
                                "temp_inf",
                                "temp_sup",
                                "tt_cerblock",
                                "material_flange_lj",
                                "puntal",
                                "tube_t"],
                            axis=1,)

                    ws = self.wb_commercial[eq_type]
                    ws["G2"] = date_offer
                    ws["G3"] = num_ref
                    ws["G4"] = offername_commercial
                    if revchanges != "":
                        ws["G5"] = rev + " " + revchanges
                        ws["G5"].font = Font(name="Calibri", size=14, bold=True)
                        ws["G5"].fill = PatternFill("solid", fgColor="FFFF00")

                    if int(rev) > 0:
                        for row in ws.iter_rows(min_row=2, max_row=4, min_col=6, max_col=7):
                            for cell in row:
                                cell.value = None
                                cell._style = ws["F1"]._style

                    last_row = ws.max_row

                    num_column_amount = df_toexport.columns.get_loc("amount") + 1

                    for index, row in df_toexport.iterrows():  # Data in desired row
                        for col_num, value in enumerate(row, start=1):
                            cell = ws.cell(row=last_row + 1, column=col_num)
                            cell.value = value
                            if col_num == num_column_amount:
                                cell._style = ws["AG1"]._style
                            else:
                                cell._style = ws["AB1"]._style

                        last_row = ws.max_row

                    ws[f"A{last_row+3}"] = "VALIDEZ DE LA OFERTA: " + validity + " DÍAS"
                    ws[f"A{last_row+3}"]._style = ws["AI1"]._style
                    ws[f"A{last_row+4}"] = (
                        "PLAZO DE ENTREGA: "
                        + delivery_time
                        + " SEMANAS DESDE APROBACIÓN DE PLANOS / CÁLCULOS (AGOSTO Y ÚLTIMAS DOS SEMANAS DE DICIEMBRE EXCLUIDOS)"
                    )
                    ws[f"A{last_row+4}"]._style = ws["AI1"]._style


                    if notes != "":
                        if isinstance(notes, list):
                            line = last_row + 5
                            for note in notes:
                                ws[f"A{line}"] = note
                                ws[f"A{line}"]._style = ws["AI1"]._style
                                line += 1
                        else:
                            line = last_row + 5
                            ws[f"A{line}"] = notes
                            ws[f"A{line}"]._style = ws["AI1"]._style

                    dict_sheets_data[eq_type] = [last_row, num_column_amount, df_toexport["amount"].sum(), df_toexport.shape[0]]

                ws.cell(row=last_row + 3, column=num_column_amount - 1).value = "CANTIDAD TOTAL"
                ws.cell(row=last_row + 3, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=last_row + 3, column=num_column_amount).value = number_items
                
                row_amount = last_row + 4
                for key, value in dict_sheets_data.items():
                    parts_key = key.split(" ")
                    ws.cell(row=row_amount + 2, column=num_column_amount - 1).value = "IMPORTE TOTAL DE " + parts_key[0] + " " + parts_key[1] + " (CANTIDAD: " + str(value[3]) + ")"
                    ws.cell(row=row_amount + 2, column=num_column_amount).value = value[2]
                    ws.cell(row=row_amount + 2, column=num_column_amount - 1)._style = ws["AA1"]._style
                    ws.cell(row=row_amount + 2, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                    ws.cell(row=row_amount + 2, column=num_column_amount)._style = ws["AC1"]._style

                    row_amount += 2

                ws.cell(row=row_amount + 2, column=num_column_amount - 1).value = "IMPORTE TOTAL DEL MATERIAL"
                ws.cell(row=row_amount + 2, column=num_column_amount).value = total_amount_material
                ws.cell(row=row_amount + 4, column=num_column_amount - 1).value = "PACKING Y TRANSPORTE"
                ws.cell(row=row_amount + 4, column=num_column_amount).value = (f"=MROUND({get_column_letter(num_column_amount)}{row_amount + 2}*0.03,10)" if total_amount_material > 6700 else 200)
                ws.cell(row=row_amount + 5, column=num_column_amount - 1).value = "PRUEBAS E INSPECCIÓN"
                ws.cell(row=row_amount + 5, column=num_column_amount).value = float(testinspection)
                ws.cell(row=row_amount + 6, column=num_column_amount - 1).value = "DOCUMENTACIÓN"
                ws.cell(row=row_amount + 6, column=num_column_amount).value = documentation
                ws.cell(row=row_amount + 8, column=num_column_amount - 1).value = "IMPORTE TOTAL DE LA OFERTA"
                ws.cell(row=row_amount + 8, column=num_column_amount).value = f"=SUM({get_column_letter(num_column_amount)}{row_amount + 2}:{get_column_letter(num_column_amount)}{row_amount + 6})"

                ws.cell(row=last_row + 3, column=num_column_amount - 1)._style = ws["AA1"]._style
                ws.cell(row=last_row + 3, column=num_column_amount).font = Font(name="Calibri", size=14)
                ws.cell(row=row_amount + 2, column=num_column_amount - 1)._style = ws["AA1"]._style
                ws.cell(row=row_amount + 2, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=row_amount + 2, column=num_column_amount)._style = ws["AC1"]._style
                ws.cell(row=row_amount + 4, column=num_column_amount - 1).font = Font(name="Calibri", size=14)
                ws.cell(row=row_amount + 4, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=row_amount + 4, column=num_column_amount)._style = ws["AC1"]._style
                ws.cell(row=row_amount + 5, column=num_column_amount - 1).font = Font(name="Calibri", size=14)
                ws.cell(row=row_amount + 5, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=row_amount + 5, column=num_column_amount)._style = ws["AC1"]._style
                ws.cell(row=row_amount + 6, column=num_column_amount - 2)._style = ws["AD1"]._style
                ws.cell(row=row_amount + 6, column=num_column_amount - 1)._style = ws["AD1"]._style
                ws.cell(row=row_amount + 6, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=row_amount + 6, column=num_column_amount)._style = ws["AE1"]._style
                ws.cell(row=row_amount + 8, column=num_column_amount - 1)._style = ws["AA1"]._style
                ws.cell(row=row_amount + 8, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=row_amount + 8, column=num_column_amount)._style = ws["AF1"]._style

            # Editing sheet NOTES
                sheet_name = "NOTES"  # Selecting  sheet
                ws = self.wb_commercial[sheet_name]

                rich_string = CellRichText(
                TextBlock(InlineFont(i=True), 'Estamos ofertando solamente elementos de medida de temperatura, les informamos que en nuestra gama de fabricación con certificación europea, incluye también elementos de caudal e indicadores de nivel de vidrio y magnéticos. (https://www.eipsa.es/productos)\n'),
                TextBlock(InlineFont(b=True, i=True), 'Los precios ofertados podrían reducirse en caso de compra de toda nuestra gama.'))
                ws["B6"] = rich_string

                rich_string = CellRichText(
                'Plazo de entrega ' + delivery_time + ' semanas desde aprobación de planos y cálculos de la ',
                TextBlock(InlineFont(b=True, i=True), 'totalidad de los equipos'),
                TextBlock(InlineFont(i=True),' amparados por el contrato, asi como la documentación crítica (Agosto y las dos últimas semanas de diciembre excluidos).'))
                ws["B10"] = rich_string

                rich_string = CellRichText(
                'Esta oferta ha sido elaborada en base a la información que figura en la hoja de datos correspondiente a cada TAG, ',
                TextBlock(InlineFont(i=True, u='single'), 'no siendo responsabilidad final de EIPSA la elección del material a utilizar en los equipos ni analizar datos de proceso.\n'),
                TextBlock(InlineFont(i=True),'La hoja de datos será el único documento técnico/contractual, cualquier otra documentación recibida será considerada como documentación complementaria a efectos informativos.'))
                ws["B16"] = rich_string

                if pay_term == "100_delivery":
                    ws["B35"] = (
                        "Pago del 100% del valor total de la orden de compra a la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España).\n"
                        "Método de pago: Transferencia bancaria."
                    )
                elif pay_term == "100_order":
                    ws["B35"] = (
                        "Pago del 100% del valor total de la orden de compra a la recepción de la orden.\n"
                        "Método de pago: Transferencia bancaria"
                    )
                elif pay_term == "90_10":
                    ws["B35"] = (
                        "TERMINOS DE PAGO:\n"
                        "Pago del 90% del Valor total de la orden de compra a la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España) y el 10% restante cuando la documentación final sea aprobada.\n"
                        "Transferencia Bancaria: 60 días desde emisión de factura."
                    )
                elif pay_term == "50_50":
                    ws["B35"] = (
                        "Pago del 50% del valor total de la orden de compra a la recepción de la orden. El 50% restante antes de la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España).\n"
                        "Método de pago: Transferencia bancaria."
                    )
                elif pay_term == "Others":
                    ws["B35"] = "TERMINOS DE PAGO POR DEFINIR"
                    ws["B35"].font = Font(name="Calibri", size=11, bold=True, italic=True, color="FF0000")

                rich_string = CellRichText(
                'Para importes superiores a 30.000,00 €, si es requerido, podremos emitir aval de garantía y estará vigente hasta el final del periodo de garantía indicado.\nEl aval del 10% será emitido con la factura del último suplemento.\n',
                TextBlock(InlineFont(i=True, b=True), 'Por debajo de dicha cantidad, no se emitirán avales.'))
                ws["B31"] = rich_string

                ws["A41"] = (
                    "Si necesita más información relacionada con esta oferta, no dude en ponerse en contacto con:\n"
                    + responsible
                    + "\n"
                    + email
                    + "\n"
                    "Telf.: (+34) 916.582.118"
                )

                for sheet in self.wb_commercial.sheetnames:
                    if sheet not in sheets_confirmed:
                        sheet_to_delete = self.wb_commercial[sheet]
                        self.wb_commercial.remove(sheet_to_delete)

                if int(rev) > 0:
                    sheets_confirmed = ["COVER", "1.2", "1.3", "NOTES"]
                    for sheet in sheets_confirmed:
                        sheet_to_delete = self.wb_commercial[sheet]
                        self.wb_commercial.remove(sheet_to_delete)

                left_text = "Fecha/Date: " + date_offer
                right_text = "Petición nº/Inquiry: " + num_ref

                for sheet in self.wb_commercial.worksheets:
                    sheet.oddFooter.left.text = left_text
                    sheet.oddFooter.right.text = right_text
                    sheet.oddFooter.center.text = "Page &P de &N"

                    sheet.oddFooter.left.size = 9
                    sheet.oddFooter.right.size = 9
                    sheet.oddFooter.center.size = 9
                
                path = self.save_excel_commercial()

                # Creating the technical offer using the commercial one as template
                self.wb_technical = load_workbook(path)

                if int(rev) == 0:
                    sheet_name = "COVER"
                    ws = self.wb_technical[sheet_name]
                    ws["E6"] = offername_technical

                for value_type in df["value_type"].unique():
                    eq_type = (
                        "TW ELEMENTS DATA"
                        if value_type == 1
                        else (
                            "TW+TE ELEMENTS DATA"
                            if value_type == 2
                            else (
                                "TW+BIM ELEMENTS DATA"
                                if value_type == 3
                                else (
                                    "TE ELEMENTS DATA"
                                    if value_type == 4
                                    else (
                                        "BIM ELEMENTS DATA"
                                        if value_type == 5
                                        else (
                                            "TIT ELEMENTS DATA"
                                            if value_type == 6
                                            else (
                                                "SKIN POINT ELEMENTS DATA"
                                                if value_type == 7
                                                else "MULTI-T RO ELEMENTS DATA"
                                            )
                                        )
                                    )
                                )
                            )
                        )
                    )

                    ws = self.wb_technical[eq_type]
                    if int(rev) == 0:
                        ws["G4"] = offername_technical

                    last_row = dict_sheets_data[eq_type][0]
                    num_column_amount = dict_sheets_data[eq_type][1]

                    # self.wb_technical[eq_type].delete_rows(last_row + 5, 20)

                    ws.cell(row=row_amount + 2, column=num_column_amount).value = ""
                    ws.cell(row=row_amount + 4, column=num_column_amount).value = ""
                    ws.cell(row=row_amount + 5, column=num_column_amount).value = ""
                    ws.cell(row=row_amount + 6, column=num_column_amount).value = ""
                    ws.cell(row=row_amount + 8, column=num_column_amount).value = ""

                    ws.cell(row=row_amount + 2, column=num_column_amount - 1).value = ""
                    ws.cell(row=row_amount + 4, column=num_column_amount - 1).value = ""
                    ws.cell(row=row_amount + 5, column=num_column_amount - 1).value = ""
                    ws.cell(row=row_amount + 6, column=num_column_amount - 1).value = ""
                    ws.cell(row=row_amount + 8, column=num_column_amount - 1).value = ""

                # Deleting "Amount" column
                    self.wb_technical[eq_type].delete_cols(num_column_amount, 1)

                # Adjusting the print area
                    new_last_column = num_column_amount
                    last_print_row = 40
                    nuevo_rango_impresion = f'A1:{get_column_letter(new_last_column)}{last_print_row}'
                    self.wb_technical[eq_type].print_area = nuevo_rango_impresion

                    stamp_1 = self.wb_technical[eq_type]._images[1]
                    anchor_actual = stamp_1.anchor

                    from_col, from_col_off = anchor_actual._from.col, anchor_actual._from.colOff
                    from_row, from_row_off = anchor_actual._from.row, anchor_actual._from.rowOff
                    to_col, to_col_off = anchor_actual.to.col, anchor_actual.to.colOff
                    to_row, to_row_off = anchor_actual.to.row, anchor_actual.to.rowOff

                    from_cell = AnchorMarker(col=from_col - 1, colOff=from_col_off, row=from_row, rowOff=from_row_off)
                    to_cell = AnchorMarker(col=to_col - 1, colOff=to_col_off, row=to_row, rowOff=to_row_off)

                    new_anchor = TwoCellAnchor(_from = from_cell, to = to_cell, editAs='absolute')
                    stamp_1.anchor = new_anchor

                    stamp_2 = self.wb_technical[eq_type]._images[2]
                    anchor_actual = stamp_2.anchor

                    from_col, from_col_off = anchor_actual._from.col, anchor_actual._from.colOff
                    from_row, from_row_off = anchor_actual._from.row, anchor_actual._from.rowOff
                    to_col, to_col_off = anchor_actual.to.col, anchor_actual.to.colOff
                    to_row, to_row_off = anchor_actual.to.row, anchor_actual.to.rowOff

                    from_cell = AnchorMarker(col=from_col - 1, colOff=from_col_off, row=from_row, rowOff=from_row_off)
                    to_cell = AnchorMarker(col=to_col - 1, colOff=to_col_off, row=to_row, rowOff=to_row_off)

                    new_anchor = TwoCellAnchor(_from = from_cell, to = to_cell, editAs='absolute')
                    stamp_2.anchor = new_anchor

                if int(rev) == 0:
                    ws = self.wb_technical[self.wb_technical.sheetnames[-2]]
                    ws.cell(row=last_row + 3, column=num_column_amount).value = number_items
                    ws.cell(row=last_row + 3, column=num_column_amount).font = Font(name="Calibri", size=14)

            # Editing sheet NOTES
                if int(rev) == 0:
                    sheet_name = "NOTES"
                    ws = self.wb_technical[sheet_name]

                    rich_string = CellRichText(
                    TextBlock(InlineFont(i=True), 'Estamos ofertando solamente elementos de medida de temperatura, les informamos que en nuestra gama de fabricación con certificación europea, incluye también elementos de caudal e indicadores de nivel de vidrio y magnéticos. (https://www.eipsa.es/productos)\n'),
                    TextBlock(InlineFont(b=True, i=True), 'Los precios ofertados podrían reducirse en caso de compra de toda nuestra gama.'))
                    ws["B6"] = rich_string

                    rich_string = CellRichText(
                    'Plazo de entrega ' + delivery_time + ' semanas desde aprobación de planos y cálculos de la ',
                    TextBlock(InlineFont(b=True, i=True), 'totalidad de los equipos'),
                    TextBlock(InlineFont(i=True),' amparados por el contrato, asi como la documentación crítica (Agosto y las dos últimas semanas de diciembre excluidos).'))
                    ws["B10"] = rich_string

                    rich_string = CellRichText(
                    'Esta oferta ha sido elaborada en base a la información que figura en la hoja de datos correspondiente a cada TAG, ',
                    TextBlock(InlineFont(i=True, u='single'), 'no siendo responsabilidad final de EIPSA la elección del material a utilizar en los equipos ni analizar datos de proceso.\n'),
                    TextBlock(InlineFont(i=True),'La hoja de datos será el único documento técnico/contractual, cualquier otra documentación recibida será considerada como documentación complementaria a efectos informativos.'))
                    ws["B16"] = rich_string

                    if pay_term == "100_delivery":
                        ws["B35"] = (
                            "Pago del 100% del valor total de la orden de compra a la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España).\n"
                            "Método de pago: Transferencia bancaria."
                        )
                    elif pay_term == "100_order":
                        ws["B35"] = (
                            "Pago del 100% del valor total de la orden de compra a la recepción de la orden.\n"
                            "Método de pago: Transferencia bancaria"
                        )
                    elif pay_term == "90_10":
                        ws["B35"] = (
                            "TERMINOS DE PAGO:\n"
                            "Pago del 90% del Valor total de la orden de compra a la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España) y el 10% restante cuando la documentación final sea aprobada.\n"
                            "Transferencia Bancaria: 60 días desde emisión de factura."
                        )
                    elif pay_term == "50_50":
                        ws["B35"] = (
                            "Pago del 50% del valor total de la orden de compra a la recepción de la orden. El 50% restante antes de la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España).\n"
                            "Método de pago: Transferencia bancaria."
                        )
                    elif pay_term == "Others":
                        ws["B35"] = "TERMINOS DE PAGO POR DEFINIR"
                        ws["B35"].font = Font(name="Calibri", size=11, bold=True, italic=True, color="FF0000")

                    rich_string = CellRichText(
                    'Para importes superiores a 30.000,00 €, si es requerido, podremos emitir aval de garantía y estará vigente hasta el final del periodo de garantía indicado.\nEl aval del 10% será emitido con la factura del último suplemento.\n',
                    TextBlock(InlineFont(i=True, b=True), 'Por debajo de dicha cantidad, no se emitirán avales.'))
                    ws["B31"] = rich_string

                    ws["A41"] = (
                        "Si necesita más información relacionada con esta oferta, no dude en ponerse en contacto con:\n"
                        + responsible
                        + "\n"
                        + email
                        + "\n"
                        "Telf.: (+34) 916.582.118"
                    )
                    
                    std = self.wb_technical["1.3"]
                    self.wb_technical.remove(std)

                left_text = "Fecha/Date: " + date_offer
                right_text = "Petición nº/Inquiry: " + num_ref

                for sheet in self.wb_technical.worksheets:
                    sheet.oddFooter.left.text = left_text
                    sheet.oddFooter.right.text = right_text
                    sheet.oddFooter.center.text = "Page &P de &N"

                    sheet.oddFooter.left.size = 9
                    sheet.oddFooter.right.size = 9
                    sheet.oddFooter.center.size = 9
                
                self.save_excel_technical()

                root = Tk()
                root.withdraw()  # Hiding main window Tkinter

                # close communication with the PostgreSQL database server
                # commit the changes
                conn.commit()
        except (Exception, psycopg2.DatabaseError) as error:
            dlg = QtWidgets.QMessageBox()
            new_icon = QtGui.QIcon()
            new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
            dlg.setWindowIcon(new_icon)
            dlg.setWindowTitle("ERP EIPSA")
            dlg.setText("Ha ocurrido el siguiente error:\n"
                        + str(error))
            dlg.setIcon(QtWidgets.QMessageBox.Icon.Critical)
            dlg.exec()
            del dlg, new_icon
        finally:
            if conn is not None:
                conn.close()

    def euros_to_float(self, value):
        """
        Converts a string value representing an amount in euros to a float.
        
        Args:
            value (str): The string representation of an amount in euros, with commas for decimal separation and ' €' for currency indication.
        
        Returns:
            float: The numeric value of the amount in euros.
        """
        value = value.replace(".", "")
        value = value.replace(",", ".")
        value = value[: value.find(" €")]
        return float(value)

    def save_excel_commercial(self):
        """Saves the populated Excel workbook to a specified location.
        Opens a dialog window for the user to select the file path and name.
        """
        # Dialog window to select folder and file name; if path is selected, excel file is saved
        output_path_commercial = asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Archivos de Excel", "*.xlsx")],
            title="Guardar Oferta comercial",
        )
        if output_path_commercial:
            self.wb_commercial.save(output_path_commercial)
            return output_path_commercial

    def save_excel_technical(self):
        """Saves the populated Excel workbook to a specified location.
        Opens a dialog window for the user to select the file path and name.
        """
        output_path_technical = asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Archivos de Excel", "*.xlsx")],
            title="Guardar Oferta técnica",
        )
        if output_path_technical:
            self.wb_technical.save(output_path_technical)

class offer_short_temp_english:
    """
    A class to manage export offer (short format in english) details for temp equipments.
    
    Attributes:
        numoffer (str): Offer number.
        username (str): Name of the user creating the offer.
        rev (str): Revision number of the offer.
        project (str): Name of the project.
        delivery_term (str): Delivery terms for the offer.
        delivery_time (str): Expected delivery time.
        validity (str): Validity period of the offer.
        pay_term (str): Payment terms.
        testinspection (str): Information about testing and inspection.
        revchanges (str): Details of changes made in the revision.
        notes (str): Additional notes, split by line.
    """
    def __init__(self, numoffer, username, rev, project, delivery_term, delivery_time, validity, pay_term, testinspection, revchanges, notes):
        """
        Initializes the offer.

        Args:
            numoffer (str): Offer number.
            username (str): Name of the user creating the offer.
            rev (str): Revision number of the offer.
            project (str): Name of the project.
            delivery_term (str): Delivery terms for the offer.
            delivery_time (str): Expected delivery time.
            validity (str): Validity period of the offer.
            pay_term (str): Payment terms.
            testinspection (str): Information about testing and inspection.
            revchanges (str): Details of changes made in the revision.
            notes (str): Additional notes, split by line.
        """
        notes = notes.split('\n')
        date_offer = date.today().strftime("%d/%m/%Y")
        offername_commercial = numoffer + "-" + "Commercial.Rev" + rev
        offername_technical = numoffer + "-" + "Technical.Rev" + rev

        query_commercial = """
                    SELECT name, surname, email
                    FROM users_data.registration
                    WHERE username = %s
                    """
        query_dataoffer = """
                        SELECT client, num_ref_offer
                        FROM offers
                        WHERE UPPER (num_offer) LIKE UPPER('%%'||%s||'%%')
                        """
        query_tagsdata = """
                        SELECT *
                        FROM tags_data.tags_temp
                        WHERE (
                        UPPER ("num_offer") LIKE UPPER('%%'||%s||'%%')
                        AND
                        "tag_state" NOT IN ('PURCHASED','DELETED')
                        )
                        """
        conn = None
        try:
            # read the connection parameters
            params = config()
            # connect to the PostgreSQL server
            conn = psycopg2.connect(**params)
            cur = conn.cursor()

            cur.execute(query_dataoffer, (numoffer,))
            results_offer = cur.fetchall()
            client = results_offer[0][0]
            num_ref = results_offer[0][1]

            cur.execute(query_commercial, (username,))
            results_commercial = cur.fetchall()
            responsible = results_commercial[0][0] + " " + results_commercial[0][1]
            email = results_commercial[0][2]

            cur.execute(query_tagsdata, (numoffer,))
            data_tags = cur.fetchall()

            if len(data_tags) == 0:
                dlg = QtWidgets.QMessageBox()
                new_icon = QtGui.QIcon()
                new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
                dlg.setWindowIcon(new_icon)
                dlg.setWindowTitle("ERP EIPSA")
                dlg.setText("No hay TAGS importados en la oferta")
                dlg.setIcon(QtWidgets.QMessageBox.Icon.Warning)
                dlg.exec()
                del dlg, new_icon
            else:
                columns = []
                for elt in cur.description:
                    columns.append(elt[0])

                value_type_dict = {
                    "TW": 1,
                    "TW+TE": 2,
                    "TW+TE+TIT": 2,
                    "RETAINING FLANGE":2,
                    "TW+BIM": 3,
                    "TE": 4,
                    "BIM": 5,
                    "TIT": 6,
                    "SKIN+TT": 7,
                    "SKIN POINT": 7,
                    "Multi-T": 8
                }

                df = pd.DataFrame(data=data_tags, columns=columns)
                df["value_type"] = df["item_type"].map(value_type_dict)
                df = df.sort_values(by=["tag", "value_type"])
                df = df.iloc[:, 1:37]
                df["value_type"] = df["item_type"].map(value_type_dict)
                df = df.sort_values(by=["value_type", "tag"])
                df["amount"] = df["amount"].apply(self.euros_to_float)
                total_amount_material = df["amount"].sum()
                df = df.drop([
                        "tag_state",
                        "num_offer",
                        "num_order",
                        "num_po",
                        "position",
                        "subposition",
                        "std_tw",
                        "insulation"
                    ],
                    axis=1,)

                number_items = df.shape[0]
                documentation = number_items * 70

                # Loading Excel Template
                self.wb_commercial = load_workbook(
                    r"\\nas01\DATOS\Comunes\EIPSA-ERP\Plantillas Exportación\PLANTILLA CORTA OFERTA TEMPERATURA - ingles.xlsx"
                )

                # Editing sheet COVER
                sheet_name = "COVER"
                ws = self.wb_commercial[sheet_name]
                ws["E4"] = client
                ws["E6"] = offername_commercial
                ws["E8"] = num_ref
                ws["E10"] = project
                ws["E12"] = date_offer
                ws["E14"] = delivery_term
                ws["E16"] = validity + " days"
                ws["C43"] = responsible
                ws["C45"] = email

                # Editing sheet EQUIPMENT DATA
                sheets_confirmed = ["COVER", "1.2", "1.3", "NOTES"]

                if int(rev) == 0:
                    for item_type in df["item_type"].unique():
                        sheets_confirmed.append(item_type)

                dict_sheets_data = {}

                for value_type in df["value_type"].unique():
                    df_toexport = df[df["value_type"] == value_type]
                    df_toexport.insert(0, "N°", range(1, len(df_toexport) + 1))
                    df_toexport.set_index("N°", inplace=True, drop=False)
                    df_toexport.index.name = None
                    df_toexport = df_toexport.drop(["value_type"], axis=1)

                    eq_type = (
                        "TW ELEMENTS DATA"
                        if value_type == 1
                        else (
                            "TW+TE ELEMENTS DATA"
                            if value_type == 2
                            else (
                                "TW+BIM ELEMENTS DATA"
                                if value_type == 3
                                else (
                                    "TE ELEMENTS DATA"
                                    if value_type == 4
                                    else (
                                        "BIM ELEMENTS DATA"
                                        if value_type == 5
                                        else (
                                            "TIT ELEMENTS DATA"
                                            if value_type == 6
                                            else (
                                                "SKIN POINT ELEMENTS DATA"
                                                if value_type == 7
                                                else "MULTI-T RO ELEMENTS DATA"
                                            )
                                        )
                                    )
                                )
                            )
                        )
                    )

                    if eq_type not in sheets_confirmed:
                        sheets_confirmed.append(eq_type)

                    if eq_type == "TW ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "std_length",
                                "sensor_element",
                                "sheath_stem_material",
                                "sheath_stem_diam",
                                "temp_inf",
                                "temp_sup",
                                "nipple_ext_material",
                                "nipple_ext_length",
                                "head_case_material",
                                "elec_conn_case_diam",
                                "tt_cerblock"],
                            axis=1,)
                    elif eq_type == "TW+TE ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "std_length",
                                "temp_inf",
                                "temp_sup",
                                "nipple_ext_length"],
                            axis=1,)
                    elif eq_type == "TW+BIM ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "std_length",
                                "nipple_ext_length",
                                "tt_cerblock",
                                "puntal",
                                "tube_t"],
                            axis=1,)
                    elif eq_type == "TE ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "tw_type",
                                "size",
                                "rating",
                                "facing",
                                "material_tw",
                                "std_length",
                                "root_diam",
                                "tip_diam",
                                "temp_inf",
                                "temp_sup",
                                "nipple_ext_material",
                                "nipple_ext_length",
                                "head_case_material",
                                "elec_conn_case_diam",
                                "tt_cerblock",
                                "material_flange_lj",
                                "gasket_material",
                                "puntal",
                                "tube_t"],
                            axis=1,)
                    elif eq_type == "BIM ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "tw_type",
                                "size",
                                "rating",
                                "facing",
                                "material_tw",
                                "std_length",
                                "root_diam",
                                "tip_diam",
                                "nipple_ext_length",
                                "tt_cerblock",
                                "material_flange_lj",
                                "gasket_material",
                                "puntal",
                                "tube_t"],
                            axis=1,)
                    elif eq_type == "TIT ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "tw_type",
                                "size",
                                "rating",
                                "facing",
                                "material_tw",
                                "std_length",
                                "root_diam",
                                "tip_diam",
                                "sensor_element",
                                "sheath_stem_material",
                                "sheath_stem_diam",
                                "nipple_ext_material",
                                "nipple_ext_length",
                                "head_case_material",
                                "material_flange_lj",
                                "gasket_material",
                                "puntal",
                                "tube_t"],
                            axis=1,)
                    elif (eq_type == "SKIN POINT ELEMENTS DATA" or eq_type == "SKIN+TT ELEMENTS DATA"):
                        df_toexport = df_toexport.drop([
                                "tw_type",
                                "size",
                                "rating",
                                "facing",
                                "material_tw",
                                "std_length",
                                "root_diam",
                                "tip_diam",
                                "temp_inf",
                                "temp_sup",
                                "material_flange_lj",
                                "gasket_material",
                                "puntal",
                                "tube_t"],
                            axis=1,)
                    elif eq_type == "MULTI-T ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "material_tw",
                                "root_diam",
                                "tip_diam",
                                "temp_inf",
                                "temp_sup",
                                "tt_cerblock",
                                "material_flange_lj",
                                "puntal",
                                "tube_t"],
                            axis=1,)

                    ws = self.wb_commercial[eq_type]
                    ws["G2"] = date_offer
                    ws["G3"] = num_ref
                    ws["G4"] = offername_commercial
                    if revchanges != "":
                        ws["G5"] = rev + " " + revchanges
                        ws["G5"].font = Font(name="Calibri", size=14, bold=True)
                        ws["G5"].fill = PatternFill("solid", fgColor="FFFF00")

                    if int(rev) > 0:
                        for row in ws.iter_rows(min_row=2, max_row=4, min_col=6, max_col=7):
                            for cell in row:
                                cell.value = None
                                cell._style = ws["F1"]._style

                    last_row = ws.max_row

                    num_column_amount = df_toexport.columns.get_loc("amount") + 1

                    for index, row in df_toexport.iterrows():  # Data in desired row
                        for col_num, value in enumerate(row, start=1):
                            cell = ws.cell(row=last_row + 1, column=col_num)
                            cell.value = value
                            if col_num == num_column_amount:
                                cell._style = ws["AG1"]._style
                            else:
                                cell._style = ws["AB1"]._style

                        last_row = ws.max_row

                    ws[f"A{last_row+3}"] = "OFFER VALIDITY: " + validity + " DAYS"
                    ws[f"A{last_row+3}"]._style = ws["AI1"]._style
                    ws[f"A{last_row+4}"] = (
                        "DELIVERY TIME: "
                        + delivery_time
                        + " WEEKS SINCE DRAWING / CALCULATION APPROVAL (AUGUST AND LAST TWO DECEMBER WEEKS EXCLUDED)"
                    )
                    ws[f"A{last_row+4}"]._style = ws["AI1"]._style

                    if notes != "":
                        if isinstance(notes, list):
                            line = last_row + 5
                            for note in notes:
                                ws[f"A{line}"] = note
                                ws[f"A{line}"]._style = ws["AI1"]._style
                                line += 1
                        else:
                            line = last_row + 5
                            ws[f"A{line}"] = notes
                            ws[f"A{line}"]._style = ws["AI1"]._style

                    dict_sheets_data[eq_type] = [last_row, num_column_amount, df_toexport["amount"].sum(), df_toexport.shape[0]]

                ws.cell(row=last_row + 3, column=num_column_amount - 1).value = "QTY. TOTAL"
                ws.cell(row=last_row + 3, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=last_row + 3, column=num_column_amount).value = number_items
                
                row_amount = last_row + 4
                for key, value in dict_sheets_data.items():
                    parts_key = key.split(" ")
                    ws.cell(row=row_amount + 2, column=num_column_amount - 1).value = "TOTAL AMOUNT OF " + parts_key[0] + " " + parts_key[1] + " (QTY: " + str(value[3]) + ")"
                    ws.cell(row=row_amount + 2, column=num_column_amount).value = value[2]
                    ws.cell(row=row_amount + 2, column=num_column_amount - 1)._style = ws["AA1"]._style
                    ws.cell(row=row_amount + 2, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                    ws.cell(row=row_amount + 2, column=num_column_amount)._style = ws["AC1"]._style

                    row_amount += 2

                ws.cell(row=row_amount + 2, column=num_column_amount - 1).value = "TOTAL AMOUNT OF MATERIAL"
                ws.cell(row=row_amount + 2, column=num_column_amount).value = total_amount_material
                ws.cell(row=row_amount + 4, column=num_column_amount - 1).value = "PACKING AND TRANSPORT"
                ws.cell(row=row_amount + 4, column=num_column_amount).value = (f"=MROUND({get_column_letter(num_column_amount)}{row_amount + 2}*0.03,10)" if total_amount_material > 6700 else 200)
                ws.cell(row=row_amount + 5, column=num_column_amount - 1).value = "TESTS & INSPECTION"
                ws.cell(row=row_amount + 5, column=num_column_amount).value = float(testinspection)
                ws.cell(row=row_amount + 6, column=num_column_amount - 1).value = "DOCUMENTATION"
                ws.cell(row=row_amount + 6, column=num_column_amount).value = documentation
                ws.cell(row=row_amount + 8, column=num_column_amount - 1).value = "TOTAL AMOUNT OF BID"
                ws.cell(row=row_amount + 8, column=num_column_amount).value = f"=SUM({get_column_letter(num_column_amount)}{row_amount + 2}:{get_column_letter(num_column_amount)}{row_amount + 6})"

                ws.cell(row=last_row + 3, column=num_column_amount - 1)._style = ws["AA1"]._style
                ws.cell(row=last_row + 3, column=num_column_amount).font = Font(name="Calibri", size=14)
                ws.cell(row=row_amount + 2, column=num_column_amount - 1)._style = ws["AA1"]._style
                ws.cell(row=row_amount + 2, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=row_amount + 2, column=num_column_amount)._style = ws["AC1"]._style
                ws.cell(row=row_amount + 4, column=num_column_amount - 1).font = Font(name="Calibri", size=14)
                ws.cell(row=row_amount + 4, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=row_amount + 4, column=num_column_amount)._style = ws["AC1"]._style
                ws.cell(row=row_amount + 5, column=num_column_amount - 1).font = Font(name="Calibri", size=14)
                ws.cell(row=row_amount + 5, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=row_amount + 5, column=num_column_amount)._style = ws["AC1"]._style
                ws.cell(row=row_amount + 6, column=num_column_amount - 2)._style = ws["AD1"]._style
                ws.cell(row=row_amount + 6, column=num_column_amount - 1)._style = ws["AD1"]._style
                ws.cell(row=row_amount + 6, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=row_amount + 6, column=num_column_amount)._style = ws["AE1"]._style
                ws.cell(row=row_amount + 8, column=num_column_amount - 1)._style = ws["AA1"]._style
                ws.cell(row=row_amount + 8, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=row_amount + 8, column=num_column_amount)._style = ws["AF1"]._style

            # Editing sheet NOTES
                sheet_name = "NOTES"  # Selecting  sheet
                ws = self.wb_commercial[sheet_name]

                rich_string = CellRichText(
                'We are only offering measuring temperature elements. Please be informed that our product range includes flow elements, and glass and magentic level indicators; all with european certification. (https://www.eipsa.es/en/products)\n',
                TextBlock(InlineFont(b=True), 'The prices quoted could be reduced in case of purchasing our full range of products.'))
                ws["B6"] = rich_string

                rich_string = CellRichText(
                TextBlock(InlineFont(i=True), 'Estamos ofertando solamente elementos de medida de temperatura, les informamos que en nuestra gama de fabricación con certificación europea, incluye también elementos de caudal e indicadores de nivel de vidrio y magnéticos. (https://www.eipsa.es/productos)\n'),
                TextBlock(InlineFont(b=True, i=True), 'Los precios ofertados podrían reducirse en caso de compra de toda nuestra gama.'))
                ws["B7"] = rich_string

                rich_string = CellRichText(
                'Delivery time ' + delivery_time + ' weeks since drawing / calculation approval of ',
                TextBlock(InlineFont(b=True), 'all equipment'),
                ' in the contract, as well as critical documentation (August and last two December weeks excluded).')
                ws["B12"] = rich_string

                rich_string = CellRichText(
                'Plazo de entrega ' + delivery_time + ' semanas desde aprobación de planos y cálculos de la ',
                TextBlock(InlineFont(b=True, i=True), 'totalidad de los equipos'),
                TextBlock(InlineFont(i=True),' amparados por el contrato, asi como la documentación crítica (Agosto y las dos últimas semanas de diciembre excluidos).'))
                ws["B13"] = rich_string

                rich_string = CellRichText(
                'Quotation prepared according to the information provided in the datasheet corresponding to each TAG. ',
                TextBlock(InlineFont(u='single'), 'EIPSA does not hold the final responsibility regarding selection of equipment material neither analyze process data.\n'),
                'The datasheet will be considered the only technical/contractual document, any other documentation will be considered as complementary documentation with informative purpose.')
                ws["B21"] = rich_string

                rich_string = CellRichText(
                'Esta oferta ha sido elaborada en base a la información que figura en la hoja de datos correspondiente a cada TAG, ',
                TextBlock(InlineFont(i=True, u='single'), 'no siendo responsabilidad final de EIPSA la elección del material a utilizar en los equipos ni analizar datos de proceso.\n'),
                TextBlock(InlineFont(i=True),'La hoja de datos será el único documento técnico/contractual, cualquier otra documentación recibida será considerada como documentación complementaria a efectos informativos.'))
                ws["B22"] = rich_string

                if pay_term == "100_delivery":
                    ws["B48"] = (
                        "100% of total amount of purchase order upon delivery of material according to Incoterms 2020, FCA (our facilities, Spain).\n"
                        "Payment method: bank transfer"
                    )
                    ws["B49"] = (
                        "Pago del 100% del valor total de la orden de compra a la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España).\n"
                        "Método de pago: Transferencia bancaria."
                    )
                elif pay_term == "100_order":
                    ws["B48"] = (
                        "100 % of the total amount of purchase order upon receipt of purchase order.\n"
                        "Payment method: bank transfer"
                    )
                    ws["B49"] = (
                        "Pago del 100% del valor total de la orden de compra a la recepción de la orden.\n"
                        "Método de pago: Transferencia bancaria"
                    )
                elif pay_term == "90_10":
                    ws["B48"] = (
                        "PAYMENT TERMS:\n"
                        "90 % of the total amount of PO upon delivery of material according to Incoterms 2020, FCA (our facilities, Spain) and 10% when final documentation is approved. \n"
                        "Bank Transfer: 60 days since invoice issue date."
                    )
                    ws["B49"] = (
                        "TERMINOS DE PAGO:\n"
                        "Pago del 90% del Valor total de la orden de compra a la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España) y el 10% restante cuando la documentación final sea aprobada.\n"
                        "Transferencia Bancaria: 60 días desde emisión de factura."
                    )
                elif pay_term == "50_50":
                    ws["B48"] = (
                        "50 % of the total amount of purchase order upon receipt of purchase order. Remaining 50% before material be delivered according to Incoterms 2020, FCA (our facilities, Spain).\n"
                        "Payment method: bank transfer."
                    )
                    ws["B49"] = (
                        "Pago del 50% del valor total de la orden de compra a la recepción de la orden. El 50% restante antes de la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España).\n"
                        "Método de pago: Transferencia bancaria."
                    )
                elif pay_term == "Others":
                    ws["B48"] = "PAYMENT TERMS TO BE DEFINED"
                    ws["B48"].font = Font(name="Calibri", size=11, bold=True, color="FF0000")
                    ws["B49"] = "TERMINOS DE PAGO POR DEFINIR"
                    ws["B49"].font = Font(name="Calibri", size=11, bold=True, italic=True, color="FF0000")

                rich_string = CellRichText(
                'For amounts greater than 30,000.00 € we can issue a warranty bond (if required) valid until the end of the indicated warranty period.\nBond warranty of 10% will be issued with the invoice of the last supplement.\n',
                TextBlock(InlineFont(b=True), 'For lower amounts no warranty bond is issued.'))
                ws["B42"] = rich_string

                rich_string = CellRichText(
                'Para importes superiores a 30.000,00 €, si es requerido, podremos emitir aval de garantía y estará vigente hasta el final del periodo de garantía indicado.\nEl aval del 10% será emitido con la factura del último suplemento.\n',
                TextBlock(InlineFont(i=True, b=True), 'Por debajo de dicha cantidad, no se emitirán avales.'))
                ws["B43"] = rich_string

                ws["A55"] = (
                    "If you require further information related with this offer, please do not hesitate to contact:\n"
                    + responsible
                    + "\n"
                    + email
                    + "\n"
                    "Telf.: (+34) 916.582.118"
                )

                for sheet in self.wb_commercial.sheetnames:
                    if sheet not in sheets_confirmed:
                        sheet_to_delete = self.wb_commercial[sheet]
                        self.wb_commercial.remove(sheet_to_delete)

                if int(rev) > 0:
                    sheets_confirmed = ["COVER", "1.2", "1.3", "NOTES"]
                    for sheet in sheets_confirmed:
                        sheet_to_delete = self.wb_commercial[sheet]
                        self.wb_commercial.remove(sheet_to_delete)

                left_text = "Fecha/Date: " + date_offer
                right_text = "Petición nº/Inquiry: " + num_ref

                for sheet in self.wb_commercial.worksheets:
                    sheet.oddFooter.left.text = left_text
                    sheet.oddFooter.right.text = right_text
                    sheet.oddFooter.center.text = "Page &P de &N"

                    sheet.oddFooter.left.size = 9
                    sheet.oddFooter.right.size = 9
                    sheet.oddFooter.center.size = 9
                
                path = self.save_excel_commercial()

                # Creating the technical offer using the commercial one as template
                self.wb_technical = load_workbook(path)

                if int(rev) == 0:
                    sheet_name = "COVER"
                    ws = self.wb_technical[sheet_name]
                    ws["E6"] = offername_technical

                for value_type in df["value_type"].unique():
                    eq_type = (
                        "TW ELEMENTS DATA"
                        if value_type == 1
                        else (
                            "TW+TE ELEMENTS DATA"
                            if value_type == 2
                            else (
                                "TW+BIM ELEMENTS DATA"
                                if value_type == 3
                                else (
                                    "TE ELEMENTS DATA"
                                    if value_type == 4
                                    else (
                                        "BIM ELEMENTS DATA"
                                        if value_type == 5
                                        else (
                                            "TIT ELEMENTS DATA"
                                            if value_type == 6
                                            else (
                                                "SKIN POINT ELEMENTS DATA"
                                                if value_type == 7
                                                else "MULTI-T RO ELEMENTS DATA"
                                            )
                                        )
                                    )
                                )
                            )
                        )
                    )

                    ws = self.wb_technical[eq_type]
                    if int(rev) == 0:
                        ws["G4"] = offername_technical

                    last_row = dict_sheets_data[eq_type][0]
                    num_column_amount = dict_sheets_data[eq_type][1]

                    # self.wb_technical[eq_type].delete_rows(last_row + 5, 20)

                    ws.cell(row=row_amount + 2, column=num_column_amount).value = ""
                    ws.cell(row=row_amount + 4, column=num_column_amount).value = ""
                    ws.cell(row=row_amount + 5, column=num_column_amount).value = ""
                    ws.cell(row=row_amount + 6, column=num_column_amount).value = ""
                    ws.cell(row=row_amount + 8, column=num_column_amount).value = ""

                    ws.cell(row=row_amount + 2, column=num_column_amount - 1).value = ""
                    ws.cell(row=row_amount + 4, column=num_column_amount - 1).value = ""
                    ws.cell(row=row_amount + 5, column=num_column_amount - 1).value = ""
                    ws.cell(row=row_amount + 6, column=num_column_amount - 1).value = ""
                    ws.cell(row=row_amount + 8, column=num_column_amount - 1).value = ""

                # Deleting "Amount" column
                    self.wb_technical[eq_type].delete_cols(num_column_amount, 1)

                # Adjusting the print area
                    new_last_column = num_column_amount
                    last_print_row = 40
                    nuevo_rango_impresion = f'A1:{get_column_letter(new_last_column)}{last_print_row}'
                    self.wb_technical[eq_type].print_area = nuevo_rango_impresion

                    stamp_1 = self.wb_technical[eq_type]._images[1]
                    anchor_actual = stamp_1.anchor

                    from_col, from_col_off = anchor_actual._from.col, anchor_actual._from.colOff
                    from_row, from_row_off = anchor_actual._from.row, anchor_actual._from.rowOff
                    to_col, to_col_off = anchor_actual.to.col, anchor_actual.to.colOff
                    to_row, to_row_off = anchor_actual.to.row, anchor_actual.to.rowOff

                    from_cell = AnchorMarker(col=from_col - 1, colOff=from_col_off, row=from_row, rowOff=from_row_off)
                    to_cell = AnchorMarker(col=to_col - 1, colOff=to_col_off, row=to_row, rowOff=to_row_off)

                    new_anchor = TwoCellAnchor(_from = from_cell, to = to_cell, editAs='absolute')
                    stamp_1.anchor = new_anchor

                    stamp_2 = self.wb_technical[eq_type]._images[2]
                    anchor_actual = stamp_2.anchor

                    from_col, from_col_off = anchor_actual._from.col, anchor_actual._from.colOff
                    from_row, from_row_off = anchor_actual._from.row, anchor_actual._from.rowOff
                    to_col, to_col_off = anchor_actual.to.col, anchor_actual.to.colOff
                    to_row, to_row_off = anchor_actual.to.row, anchor_actual.to.rowOff

                    from_cell = AnchorMarker(col=from_col - 1, colOff=from_col_off, row=from_row, rowOff=from_row_off)
                    to_cell = AnchorMarker(col=to_col - 1, colOff=to_col_off, row=to_row, rowOff=to_row_off)

                    new_anchor = TwoCellAnchor(_from = from_cell, to = to_cell, editAs='absolute')
                    stamp_2.anchor = new_anchor

                if int(rev) == 0:
                    ws = self.wb_technical[self.wb_technical.sheetnames[-2]]
                    ws.cell(row=last_row + 3, column=num_column_amount).value = number_items
                    ws.cell(row=last_row + 3, column=num_column_amount).font = Font(name="Calibri", size=14)

            # Editing sheet NOTES
                if int(rev) == 0:
                    sheet_name = "NOTES"
                    ws = self.wb_technical[sheet_name]

                    rich_string = CellRichText(
                    'We are only offering measuring temperature elements. Please be informed that our product range includes flow elements, and glass and magentic level indicators; all with european certification. (https://www.eipsa.es/en/products)\n',
                    TextBlock(InlineFont(b=True), 'The prices quoted could be reduced in case of purchasing our full range of products.'))
                    ws["B6"] = rich_string

                    rich_string = CellRichText(
                    TextBlock(InlineFont(i=True), 'Estamos ofertando solamente elementos de medida de temperatura, les informamos que en nuestra gama de fabricación con certificación europea, incluye también elementos de caudal e indicadores de nivel de vidrio y magnéticos. (https://www.eipsa.es/productos)\n'),
                    TextBlock(InlineFont(b=True, i=True), 'Los precios ofertados podrían reducirse en caso de compra de toda nuestra gama.'))
                    ws["B7"] = rich_string

                    rich_string = CellRichText(
                    'Delivery time ' + delivery_time + ' weeks since drawing / calculation approval of ',
                    TextBlock(InlineFont(b=True), 'all equipment'),
                    ' in the contract, as well as critical documentation (August and last two December weeks excluded).')
                    ws["B12"] = rich_string

                    rich_string = CellRichText(
                    'Plazo de entrega ' + delivery_time + ' semanas desde aprobación de planos y cálculos de la ',
                    TextBlock(InlineFont(b=True, i=True), 'totalidad de los equipos'),
                    TextBlock(InlineFont(i=True),' amparados por el contrato, asi como la documentación crítica (Agosto y las dos últimas semanas de diciembre excluidos).'))
                    ws["B13"] = rich_string

                    rich_string = CellRichText(
                    'Quotation prepared according to the information provided in the datasheet corresponding to each TAG. ',
                    TextBlock(InlineFont(u='single'), 'EIPSA does not hold the final responsibility regarding selection of equipment material neither analyze process data.\n'),
                    'The datasheet will be considered the only technical/contractual document, any other documentation will be considered as complementary documentation with informative purpose.')
                    ws["B21"] = rich_string

                    rich_string = CellRichText(
                    'Esta oferta ha sido elaborada en base a la información que figura en la hoja de datos correspondiente a cada TAG, ',
                    TextBlock(InlineFont(i=True, u='single'), 'no siendo responsabilidad final de EIPSA la elección del material a utilizar en los equipos ni analizar datos de proceso.\n'),
                    TextBlock(InlineFont(i=True),'La hoja de datos será el único documento técnico/contractual, cualquier otra documentación recibida será considerada como documentación complementaria a efectos informativos.'))
                    ws["B22"] = rich_string

                    if pay_term == "100_delivery":
                        ws["B48"] = (
                            "100% of total amount of purchase order upon delivery of material according to Incoterms 2020, FCA (our facilities, Spain).\n"
                            "Payment method: bank transfer"
                        )
                        ws["B49"] = (
                            "Pago del 100% del valor total de la orden de compra a la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España).\n"
                            "Método de pago: Transferencia bancaria."
                        )
                    elif pay_term == "100_order":
                        ws["B48"] = (
                            "100 % of the total amount of purchase order upon receipt of purchase order.\n"
                            "Payment method: bank transfer"
                        )
                        ws["B49"] = (
                            "Pago del 100% del valor total de la orden de compra a la recepción de la orden.\n"
                            "Método de pago: Transferencia bancaria"
                        )
                    elif pay_term == "90_10":
                        ws["B48"] = (
                            "PAYMENT TERMS:\n"
                            "90 % of the total amount of PO upon delivery of material according to Incoterms 2020, FCA (our facilities, Spain) and 10% when final documentation is approved. \n"
                            "Bank Transfer: 60 days since invoice issue date."
                        )
                        ws["B49"] = (
                            "TERMINOS DE PAGO:\n"
                            "Pago del 90% del Valor total de la orden de compra a la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España) y el 10% restante cuando la documentación final sea aprobada.\n"
                            "Transferencia Bancaria: 60 días desde emisión de factura."
                        )
                    elif pay_term == "50_50":
                        ws["B48"] = (
                            "50 % of the total amount of purchase order upon receipt of purchase order. Remaining 50% before material be delivered according to Incoterms 2020, FCA (our facilities, Spain).\n"
                            "Payment method: bank transfer."
                        )
                        ws["B49"] = (
                            "Pago del 50% del valor total de la orden de compra a la recepción de la orden. El 50% restante antes de la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España).\n"
                            "Método de pago: Transferencia bancaria."
                        )
                    elif pay_term == "Others":
                        ws["B48"] = "PAYMENT TERMS TO BE DEFINED"
                        ws["B48"].font = Font(name="Calibri", size=11, bold=True, color="FF0000")
                        ws["B49"] = "TERMINOS DE PAGO POR DEFINIR"
                        ws["B49"].font = Font(name="Calibri", size=11, bold=True, italic=True, color="FF0000")

                    rich_string = CellRichText(
                    'For amounts greater than 30,000.00 € we can issue a warranty bond (if required) valid until the end of the indicated warranty period.\nBond warranty of 10% will be issued with the invoice of the last supplement.\n',
                    TextBlock(InlineFont(b=True), 'For lower amounts no warranty bond is issued.'))
                    ws["B42"] = rich_string

                    rich_string = CellRichText(
                    'Para importes superiores a 30.000,00 €, si es requerido, podremos emitir aval de garantía y estará vigente hasta el final del periodo de garantía indicado.\nEl aval del 10% será emitido con la factura del último suplemento.\n',
                    TextBlock(InlineFont(i=True, b=True), 'Por debajo de dicha cantidad, no se emitirán avales.'))
                    ws["B43"] = rich_string

                    ws["A55"] = (
                        "If you require further information related with this offer, please do not hesitate to contact:\n"
                        + responsible
                        + "\n"
                        + email
                        + "\n"
                        "Telf.: (+34) 916.582.118"
                    )
                    
                    std = self.wb_technical["1.3"]
                    self.wb_technical.remove(std)

                left_text = "Fecha/Date: " + date_offer
                right_text = "Petición nº/Inquiry: " + num_ref

                for sheet in self.wb_technical.worksheets:
                    sheet.oddFooter.left.text = left_text
                    sheet.oddFooter.right.text = right_text
                    sheet.oddFooter.center.text = "Page &P de &N"

                    sheet.oddFooter.left.size = 9
                    sheet.oddFooter.right.size = 9
                    sheet.oddFooter.center.size = 9

                self.save_excel_technical()

                root = Tk()
                root.withdraw()  # Hiding main window Tkinter

                # close communication with the PostgreSQL database server
                # commit the changes
                conn.commit()
        except (Exception, psycopg2.DatabaseError) as error:
            dlg = QtWidgets.QMessageBox()
            new_icon = QtGui.QIcon()
            new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
            dlg.setWindowIcon(new_icon)
            dlg.setWindowTitle("ERP EIPSA")
            dlg.setText("Ha ocurrido el siguiente error:\n"
                        + str(error))
            dlg.setIcon(QtWidgets.QMessageBox.Icon.Critical)
            dlg.exec()
            del dlg, new_icon
        finally:
            if conn is not None:
                conn.close()

    def euros_to_float(self, value):
        """
        Converts a string value representing an amount in euros to a float.
        
        Args:
            value (str): The string representation of an amount in euros, with commas for decimal separation and ' €' for currency indication.
        
        Returns:
            float: The numeric value of the amount in euros.
        """
        value = value.replace(".", "")
        value = value.replace(",", ".")
        value = value[: value.find(" €")]
        return float(value)

    def save_excel_commercial(self):
        """Saves the populated Excel workbook to a specified location.
        Opens a dialog window for the user to select the file path and name.
        """
        # Dialog window to select folder and file name; if path is selected, excel file is saved
        output_path_commercial = asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Archivos de Excel", "*.xlsx")],
            title="Guardar Oferta comercial",
        )
        if output_path_commercial:
            self.wb_commercial.save(output_path_commercial)
            return output_path_commercial

    def save_excel_technical(self):
        """Saves the populated Excel workbook to a specified location.
        Opens a dialog window for the user to select the file path and name.
        """
        output_path_technical = asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Archivos de Excel", "*.xlsx")],
            title="Guardar Oferta técnica",
        )
        if output_path_technical:
            self.wb_technical.save(output_path_technical)

class offer_level:
    """
    A class to manage export offer details for level equipments.
    
    Attributes:
        numoffer (str): Offer number.
        username (str): Name of the user creating the offer.
        rev (str): Revision number of the offer.
        project (str): Name of the project.
        delivery_term (str): Delivery terms for the offer.
        delivery_time (str): Expected delivery time.
        validity (str): Validity period of the offer.
        pay_term (str): Payment terms.
        testinspection (str): Information about testing and inspection.
        revchanges (str): Details of changes made in the revision.
        notes (str): Additional notes, split by line.
    """
    def __init__(self, numoffer, username, rev, project, delivery_term, delivery_time, validity, pay_term, testinspection, revchanges, notes,):
        """
        Initializes the offer.

        Args:
            numoffer (str): Offer number.
            username (str): Name of the user creating the offer.
            rev (str): Revision number of the offer.
            project (str): Name of the project.
            delivery_term (str): Delivery terms for the offer.
            delivery_time (str): Expected delivery time.
            validity (str): Validity period of the offer.
            pay_term (str): Payment terms.
            testinspection (str): Information about testing and inspection.
            revchanges (str): Details of changes made in the revision.
            notes (str): Additional notes, split by line.
        """
        notes = notes.split('\n')
        date_offer = date.today().strftime("%d/%m/%Y")
        offername_commercial = numoffer + "-" + "Commercial.Rev" + rev
        offername_technical = numoffer + "-" + "Technical.Rev" + rev

        query_commercial = """
                    SELECT name, surname, email
                    FROM users_data.registration
                    WHERE username = %s
                    """
        query_dataoffer = """
                        SELECT client, num_ref_offer
                        FROM offers
                        WHERE UPPER (num_offer) LIKE UPPER('%%'||%s||'%%')
                        """
        query_tagsdata = """
                        SELECT *
                        FROM tags_data.tags_level
                        WHERE (
                        UPPER ("num_offer") LIKE UPPER('%%'||%s||'%%')
                        AND
                        "tag_state" NOT IN ('PURCHASED','DELETED')
                        )
                        """
        conn = None
        try:
            # read the connection parameters
            params = config()
            # connect to the PostgreSQL server
            conn = psycopg2.connect(**params)
            cur = conn.cursor()

            cur.execute(query_dataoffer, (numoffer,))
            results_offer = cur.fetchall()
            client = results_offer[0][0]
            num_ref = results_offer[0][1]

            cur.execute(query_commercial, (username,))
            results_commercial = cur.fetchall()
            responsible = results_commercial[0][0] + " " + results_commercial[0][1]
            email = results_commercial[0][2]

            cur.execute(query_tagsdata, (numoffer,))
            data_tags = cur.fetchall()

            if len(data_tags) == 0:
                dlg = QtWidgets.QMessageBox()
                new_icon = QtGui.QIcon()
                new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
                dlg.setWindowIcon(new_icon)
                dlg.setWindowTitle("ERP EIPSA")
                dlg.setText("No hay TAGS importados en la oferta")
                dlg.setIcon(QtWidgets.QMessageBox.Icon.Warning)
                dlg.exec()
                del dlg, new_icon
            else:
                columns = []
                for elt in cur.description:
                    columns.append(elt[0])

                value_type_dict = {"Transparent": 1, "Reflex": 1, "Magnetic": 2}

                df = pd.DataFrame(data=data_tags, columns=columns)
                df["value_type"] = df["item_type"].map(value_type_dict)
                df = df.sort_values(by=["tag", "value_type"])
                df = df.iloc[:, 1:38]
                df["value_type"] = df["item_type"].map(value_type_dict)
                df = df.sort_values(by=["value_type", "tag"])
                df["amount"] = df["amount"].apply(self.euros_to_float)
                total_amount_material = df["amount"].sum()
                df = df.drop([
                        "tag_state",
                        "num_offer",
                        "num_order",
                        "num_po",
                        "position",
                        "subposition",
                        "proc_conn_type",
                        "flags",
                        "flange_type",
                        "nipple_hex",
                        "nipple_tub"
                    ],
                    axis=1,)

                number_items = df.shape[0]
                documentation = number_items * 70

                # Loading Excel Template
                self.wb_commercial = load_workbook(
                    r"\\nas01\DATOS\Comunes\EIPSA-ERP\Plantillas Exportación\PLANTILLA OFERTA NIVEL.xlsx"
                )

                # Editing sheet COVER
                sheet_name = "COVER"
                ws = self.wb_commercial[sheet_name]
                ws["E4"] = client
                ws["E6"] = offername_commercial
                ws["E8"] = num_ref
                ws["E10"] = project
                ws["E12"] = date_offer
                ws["E14"] = delivery_term
                ws["E16"] = validity + " days"
                ws["C43"] = responsible
                ws["C45"] = email

                # Editing sheet EQUIPMENT DATA
                sheets_confirmed = ["COVER", "1.2", "1.3", "NOTES"]

                if int(rev) == 0:
                    for item_type in df["item_type"].unique():
                        sheets_confirmed.append(item_type)

                dict_sheets_data = {}

                for value_type in df["value_type"].unique():
                    df_toexport = df[df["value_type"] == value_type]
                    df_toexport.insert(0, "N°", range(1, len(df_toexport) + 1))
                    df_toexport.set_index("N°", inplace=True, drop=False)
                    df_toexport.index.name = None
                    df_toexport = df_toexport.drop(["value_type"], axis=1)

                    eq_type = (
                        "LEVEL GAUGES ELEMENTS DATA"
                        if value_type == 1
                        else "MAGNETIC ELEMENTS DATA"
                        )

                    if eq_type not in sheets_confirmed:
                        sheets_confirmed.append(eq_type)

                    if eq_type == "LEVEL GAUGES ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "dv_size",
                                "dv_rating",
                                "dv_facing",
                                "float_material"],
                            axis=1,)
                    elif eq_type == "MAGNETIC ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "dv_conn",
                                "item_type",
                                "valve_type",
                                "case_cover_material",
                                "illuminator",
                                "ip_code"],
                            axis=1,)

                    ws = self.wb_commercial[eq_type]
                    ws["G2"] = date_offer
                    ws["G3"] = num_ref
                    ws["G4"] = offername_commercial
                    if revchanges != "":
                        ws["G5"] = rev + " " + revchanges
                        ws["G5"].font = Font(name="Calibri", size=14, bold=True)
                        ws["G5"].fill = PatternFill("solid", fgColor="FFFF00")

                    if int(rev) > 0:
                        for row in ws.iter_rows(min_row=2, max_row=4, min_col=6, max_col=7):
                            for cell in row:
                                cell.value = None
                                cell._style = ws["F1"]._style

                    last_row = ws.max_row

                    num_column_amount = df_toexport.columns.get_loc("amount") + 1

                    for index, row in df_toexport.iterrows():  # Data in desired row
                        for col_num, value in enumerate(row, start=1):
                            cell = ws.cell(row=last_row + 1, column=col_num)
                            cell.value = value
                            if col_num == num_column_amount:
                                cell._style = ws["AG1"]._style
                            else:
                                cell._style = ws["AB1"]._style

                        last_row = ws.max_row

                    ws[f"A{last_row+3}"] = "OFFER VALIDITY: " + validity + " DAYS"
                    ws[f"A{last_row+3}"]._style = ws["AI1"]._style
                    ws[f"A{last_row+4}"] = (
                        "DELIVERY TIME: "
                        + delivery_time
                        + " WEEKS SINCE DRAWING APPROVAL (AUGUST AND LAST TWO DECEMBER WEEKS EXCLUDED)"
                    )
                    ws[f"A{last_row+4}"]._style = ws["AI1"]._style

                    if notes != "":
                        if isinstance(notes, list):
                            line = last_row + 5
                            for note in notes:
                                ws[f"A{line}"] = note
                                ws[f"A{line}"]._style = ws["AI1"]._style
                                line += 1
                        else:
                            line = last_row + 5
                            ws[f"A{line}"] = notes
                            ws[f"A{line}"]._style = ws["AI1"]._style

                    dict_sheets_data[eq_type] = [last_row, num_column_amount, df_toexport["amount"].sum(), df_toexport.shape[0]]

                ws.cell(row=last_row + 3, column=num_column_amount - 1).value = "QTY. TOTAL"
                ws.cell(row=last_row + 3, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=last_row + 3, column=num_column_amount).value = number_items
                
                row_amount = last_row + 4
                for key, value in dict_sheets_data.items():
                    parts_key = key.split(" ")
                    ws.cell(row=row_amount + 2, column=num_column_amount - 1).value = "TOTAL AMOUNT OF " + parts_key[0] + " " + parts_key[1] + " (QTY: " + str(value[3]) + ")"
                    ws.cell(row=row_amount + 2, column=num_column_amount).value = value[2]
                    ws.cell(row=row_amount + 2, column=num_column_amount - 1)._style = ws["AA1"]._style
                    ws.cell(row=row_amount + 2, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                    ws.cell(row=row_amount + 2, column=num_column_amount)._style = ws["AC1"]._style

                    row_amount += 2

                ws.cell(row=row_amount + 2, column=num_column_amount - 1).value = "TOTAL AMOUNT OF MATERIAL"
                ws.cell(row=row_amount + 2, column=num_column_amount).value = total_amount_material
                ws.cell(row=row_amount + 4, column=num_column_amount - 1).value = "PACKING AND TRANSPORT"
                ws.cell(row=row_amount + 4, column=num_column_amount).value = (f"=MROUND({get_column_letter(num_column_amount)}{row_amount + 2}*0.03,10)" if total_amount_material > 6700 else 200)
                ws.cell(row=row_amount + 5, column=num_column_amount - 1).value = "TESTS & INSPECTION"
                ws.cell(row=row_amount + 5, column=num_column_amount).value = float(testinspection)
                ws.cell(row=row_amount + 6, column=num_column_amount - 1).value = "DOCUMENTATION"
                ws.cell(row=row_amount + 6, column=num_column_amount).value = documentation
                ws.cell(row=row_amount + 8, column=num_column_amount - 1).value = "TOTAL AMOUNT OF BID"
                ws.cell(row=row_amount + 8, column=num_column_amount).value = f"=SUM({get_column_letter(num_column_amount)}{row_amount + 2}:{get_column_letter(num_column_amount)}{row_amount + 6})"

                ws.cell(row=last_row + 3, column=num_column_amount - 1)._style = ws["AA1"]._style
                ws.cell(row=last_row + 3, column=num_column_amount).font = Font(name="Calibri", size=14)
                ws.cell(row=row_amount + 2, column=num_column_amount - 1)._style = ws["AA1"]._style
                ws.cell(row=row_amount + 2, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=row_amount + 2, column=num_column_amount)._style = ws["AC1"]._style
                ws.cell(row=row_amount + 4, column=num_column_amount - 1).font = Font(name="Calibri", size=14)
                ws.cell(row=row_amount + 4, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=row_amount + 4, column=num_column_amount)._style = ws["AC1"]._style
                ws.cell(row=row_amount + 5, column=num_column_amount - 1).font = Font(name="Calibri", size=14)
                ws.cell(row=row_amount + 5, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=row_amount + 5, column=num_column_amount)._style = ws["AC1"]._style
                ws.cell(row=row_amount + 6, column=num_column_amount - 2)._style = ws["AD1"]._style
                ws.cell(row=row_amount + 6, column=num_column_amount - 1)._style = ws["AD1"]._style
                ws.cell(row=row_amount + 6, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=row_amount + 6, column=num_column_amount)._style = ws["AE1"]._style
                ws.cell(row=row_amount + 8, column=num_column_amount - 1)._style = ws["AA1"]._style
                ws.cell(row=row_amount + 8, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=row_amount + 8, column=num_column_amount)._style = ws["AF1"]._style

            # Editing sheet NOTES
                sheet_name = "NOTES"  # Selecting  sheet
                ws = self.wb_commercial[sheet_name]

                rich_string = CellRichText(
                'We are only offering measuring glass and magnetic level indicators. Please be informed that our product range includes flow elements and temperature elements; all with european certification. (https://www.eipsa.es/en/products)\n',
                TextBlock(InlineFont(b=True), 'The prices quoted could be reduced in case of purchasing our full range of products.'))
                ws["B6"] = rich_string

                rich_string = CellRichText(
                TextBlock(InlineFont(i=True), 'Estamos ofertando solamente indicadores de nivel de vidrio y magnéticos, les informamos que en nuestra gama de fabricación con certificación europea, incluye también elementos de caudal y elementos de medida de temperatura. (https://www.eipsa.es/productos)\n'),
                TextBlock(InlineFont(b=True, i=True), 'Los precios ofertados podrían reducirse en caso de compra de toda nuestra gama.'))
                ws["B7"] = rich_string

                rich_string = CellRichText(
                'Delivery time ' + delivery_time + ' weeks since drawing approval of ',
                TextBlock(InlineFont(b=True), 'all equipment'),
                ' in the contract, as well as critical documentation (August and last two December weeks excluded).')
                ws["B12"] = rich_string

                rich_string = CellRichText(
                'Plazo de entrega ' + delivery_time + ' semanas desde aprobación de planos de la ',
                TextBlock(InlineFont(b=True, i=True), 'totalidad de los equipos'),
                TextBlock(InlineFont(i=True),' amparados por el contrato, asi como la documentación crítica (Agosto y las dos últimas semanas de diciembre excluidos).'))
                ws["B13"] = rich_string

                rich_string = CellRichText(
                'Quotation prepared according to the information provided in the datasheet corresponding to each TAG. ',
                TextBlock(InlineFont(u='single'), 'EIPSA does not hold the final responsibility regarding selection of equipment material neither analyze process data.\n'),
                'The datasheet will be considered the only technical/contractual document, any other documentation will be considered as complementary documentation with informative purpose.')
                ws["B21"] = rich_string

                rich_string = CellRichText(
                'Esta oferta ha sido elaborada en base a la información que figura en la hoja de datos correspondiente a cada TAG, ',
                TextBlock(InlineFont(i=True, u='single'), 'no siendo responsabilidad final de EIPSA la elección del material a utilizar en los equipos ni analizar datos de proceso.\n'),
                TextBlock(InlineFont(i=True),'La hoja de datos será el único documento técnico/contractual, cualquier otra documentación recibida será considerada como documentación complementaria a efectos informativos.'))
                ws["B22"] = rich_string

                if pay_term == "100_delivery":
                    ws["B42"] = (
                        "100% of total amount of purchase order upon delivery of material according to Incoterms 2020, FCA (our facilities, Spain).\n"
                        "Payment method: bank transfer"
                    )
                    ws["B43"] = (
                        "Pago del 100% del valor total de la orden de compra a la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España).\n"
                        "Método de pago: Transferencia bancaria."
                    )
                elif pay_term == "100_order":
                    ws["B42"] = (
                        "100 % of the total amount of purchase order upon receipt of purchase order.\n"
                        "Payment method: bank transfer"
                    )
                    ws["B43"] = (
                        "Pago del 100% del valor total de la orden de compra a la recepción de la orden.\n"
                        "Método de pago: Transferencia bancaria"
                    )
                elif pay_term == "90_10":
                    ws["B42"] = (
                        "PAYMENT TERMS:\n"
                        "90 % of the total amount of PO upon delivery of material according to Incoterms 2020, FCA (our facilities, Spain) and 10% when final documentation is approved. \n"
                        "Bank Transfer: 60 days since invoice issue date."
                    )
                    ws["B43"] = (
                        "TERMINOS DE PAGO:\n"
                        "Pago del 90% del Valor total de la orden de compra a la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España) y el 10% restante cuando la documentación final sea aprobada.\n"
                        "Transferencia Bancaria: 60 días desde emisión de factura."
                    )
                elif pay_term == "50_50":
                    ws["B42"] = (
                        "50 % of the total amount of purchase order upon receipt of purchase order. Remaining 50% before material be delivered according to Incoterms 2020, FCA (our facilities, Spain).\n"
                        "Payment method: bank transfer."
                    )
                    ws["B43"] = (
                        "Pago del 50% del valor total de la orden de compra a la recepción de la orden. El 50% restante antes de la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España).\n"
                        "Método de pago: Transferencia bancaria."
                    )
                elif pay_term == "Others":
                    ws["B42"] = "PAYMENT TERMS TO BE DEFINED"
                    ws["B42"].font = Font(name="Calibri", size=11, bold=True, color="FF0000")
                    ws["B43"] = "TERMINOS DE PAGO POR DEFINIR"
                    ws["B43"].font = Font(name="Calibri", size=11, bold=True, italic=True, color="FF0000")

                rich_string = CellRichText(
                'For amounts greater than 30,000.00 € we can issue a warranty bond (if required) valid until the end of the indicated warranty period.\nBond warranty of 10% will be issued with the invoice of the last supplement.\n',
                TextBlock(InlineFont(b=True), 'For lower amounts no warranty bond is issued.'))
                ws["B45"] = rich_string

                rich_string = CellRichText(
                'Para importes superiores a 30.000,00 €, si es requerido, podremos emitir aval de garantía y estará vigente hasta el final del periodo de garantía indicado.\nEl aval del 10% será emitido con la factura del último suplemento.\n',
                TextBlock(InlineFont(i=True, b=True), 'Por debajo de dicha cantidad, no se emitirán avales.'))
                ws["B46"] = rich_string

                ws["A58"] = (
                    "If you require further information related with this offer, please do not hesitate to contact:\n"
                    + responsible
                    + "\n"
                    + email
                    + "\n"
                    "Telf.: (+34) 916.582.118"
                )

                for sheet in self.wb_commercial.sheetnames:
                    if sheet not in sheets_confirmed:
                        sheet_to_delete = self.wb_commercial[sheet]
                        self.wb_commercial.remove(sheet_to_delete)

                if int(rev) > 0:
                    sheets_confirmed = ["COVER", "1.2", "1.3", "NOTES"]
                    for sheet in sheets_confirmed:
                        sheet_to_delete = self.wb_commercial[sheet]
                        self.wb_commercial.remove(sheet_to_delete)

                left_text = "Fecha/Date: " + date_offer
                right_text = "Petición nº/Inquiry: " + num_ref

                for sheet in self.wb_commercial.worksheets:
                    sheet.oddFooter.left.text = left_text
                    sheet.oddFooter.right.text = right_text
                    sheet.oddFooter.center.text = "Page &P de &N"

                    sheet.oddFooter.left.size = 9
                    sheet.oddFooter.right.size = 9
                    sheet.oddFooter.center.size = 9
                
                path = self.save_excel_commercial()

                # Creating the technical offer using the commercial one as template
                self.wb_technical = load_workbook(path)

                if int(rev) == 0:
                    sheet_name = "COVER"
                    ws = self.wb_technical[sheet_name]
                    ws["E6"] = offername_technical

                for value_type in df["value_type"].unique():
                    eq_type = (
                        "LEVEL GAUGES ELEMENTS DATA"
                        if value_type == 1
                        else "MAGNETIC ELEMENTS DATA"
                        )

                    ws = self.wb_technical[eq_type]
                    if int(rev) == 0:
                        ws["G4"] = offername_technical

                    last_row = dict_sheets_data[eq_type][0]
                    num_column_amount = dict_sheets_data[eq_type][1]

                    # self.wb_technical[eq_type].delete_rows(last_row + 5, 20)

                    ws.cell(row=row_amount + 2, column=num_column_amount).value = ""
                    ws.cell(row=row_amount + 4, column=num_column_amount).value = ""
                    ws.cell(row=row_amount + 5, column=num_column_amount).value = ""
                    ws.cell(row=row_amount + 6, column=num_column_amount).value = ""
                    ws.cell(row=row_amount + 8, column=num_column_amount).value = ""

                    ws.cell(row=row_amount + 2, column=num_column_amount - 1).value = ""
                    ws.cell(row=row_amount + 4, column=num_column_amount - 1).value = ""
                    ws.cell(row=row_amount + 5, column=num_column_amount - 1).value = ""
                    ws.cell(row=row_amount + 6, column=num_column_amount - 1).value = ""
                    ws.cell(row=row_amount + 8, column=num_column_amount - 1).value = ""

                # Deleting "Amount" column
                    self.wb_technical[eq_type].delete_cols(num_column_amount, 1)

                # Adjusting the print area
                    new_last_column = num_column_amount
                    last_print_row = 40
                    nuevo_rango_impresion = f'A1:{get_column_letter(new_last_column)}{last_print_row}'
                    self.wb_technical[eq_type].print_area = nuevo_rango_impresion

                    stamp_1 = self.wb_technical[eq_type]._images[1]
                    anchor_actual = stamp_1.anchor

                    from_col, from_col_off = anchor_actual._from.col, anchor_actual._from.colOff
                    from_row, from_row_off = anchor_actual._from.row, anchor_actual._from.rowOff
                    to_col, to_col_off = anchor_actual.to.col, anchor_actual.to.colOff
                    to_row, to_row_off = anchor_actual.to.row, anchor_actual.to.rowOff

                    from_cell = AnchorMarker(col=from_col - 1, colOff=from_col_off, row=from_row, rowOff=from_row_off)
                    to_cell = AnchorMarker(col=to_col - 1, colOff=to_col_off, row=to_row, rowOff=to_row_off)

                    new_anchor = TwoCellAnchor(_from = from_cell, to = to_cell, editAs='absolute')
                    stamp_1.anchor = new_anchor

                    stamp_2 = self.wb_technical[eq_type]._images[2]
                    anchor_actual = stamp_2.anchor

                    from_col, from_col_off = anchor_actual._from.col, anchor_actual._from.colOff
                    from_row, from_row_off = anchor_actual._from.row, anchor_actual._from.rowOff
                    to_col, to_col_off = anchor_actual.to.col, anchor_actual.to.colOff
                    to_row, to_row_off = anchor_actual.to.row, anchor_actual.to.rowOff

                    from_cell = AnchorMarker(col=from_col - 1, colOff=from_col_off, row=from_row, rowOff=from_row_off)
                    to_cell = AnchorMarker(col=to_col - 1, colOff=to_col_off, row=to_row, rowOff=to_row_off)

                    new_anchor = TwoCellAnchor(_from = from_cell, to = to_cell, editAs='absolute')
                    stamp_2.anchor = new_anchor

                if int(rev) == 0:
                    ws = self.wb_technical[self.wb_technical.sheetnames[-2]]
                    ws.cell(row=last_row + 3, column=num_column_amount).value = number_items
                    ws.cell(row=last_row + 3, column=num_column_amount).font = Font(name="Calibri", size=14)

            # Editing sheet NOTES
                if int(rev) == 0:
                    sheet_name = "NOTES"
                    ws = self.wb_technical[sheet_name]

                    rich_string = CellRichText(
                    'We are only offering measuring glass and magnetic level indicators. Please be informed that our product range includes flow elements and temperature elements; all with european certification. (https://www.eipsa.es/en/products)\n',
                    TextBlock(InlineFont(b=True), 'The prices quoted could be reduced in case of purchasing our full range of products.'))
                    ws["B6"] = rich_string

                    rich_string = CellRichText(
                    TextBlock(InlineFont(i=True), 'Estamos ofertando solamente indicadores de nivel de vidrio y magnéticos, les informamos que en nuestra gama de fabricación con certificación europea, incluye también elementos de caudal y elementos de medida de temperatura. (https://www.eipsa.es/productos)\n'),
                    TextBlock(InlineFont(b=True, i=True), 'Los precios ofertados podrían reducirse en caso de compra de toda nuestra gama.'))
                    ws["B7"] = rich_string

                    rich_string = CellRichText(
                    'Delivery time ' + delivery_time + ' weeks since drawing approval of ',
                    TextBlock(InlineFont(b=True), 'all equipment'),
                    ' in the contract, as well as critical documentation (August and last two December weeks excluded).')
                    ws["B12"] = rich_string

                    rich_string = CellRichText(
                    'Plazo de entrega ' + delivery_time + ' semanas desde aprobación de planos de la ',
                    TextBlock(InlineFont(b=True, i=True), 'totalidad de los equipos'),
                    TextBlock(InlineFont(i=True),' amparados por el contrato, asi como la documentación crítica (Agosto y las dos últimas semanas de diciembre excluidos).'))
                    ws["B13"] = rich_string

                    rich_string = CellRichText(
                    'Quotation prepared according to the information provided in the datasheet corresponding to each TAG. ',
                    TextBlock(InlineFont(u='single'), 'EIPSA does not hold the final responsibility regarding selection of equipment material neither analyze process data.\n'),
                    'The datasheet will be considered the only technical/contractual document, any other documentation will be considered as complementary documentation with informative purpose.')
                    ws["B21"] = rich_string

                    rich_string = CellRichText(
                    'Esta oferta ha sido elaborada en base a la información que figura en la hoja de datos correspondiente a cada TAG, ',
                    TextBlock(InlineFont(i=True, u='single'), 'no siendo responsabilidad final de EIPSA la elección del material a utilizar en los equipos ni analizar datos de proceso.\n'),
                    TextBlock(InlineFont(i=True),'La hoja de datos será el único documento técnico/contractual, cualquier otra documentación recibida será considerada como documentación complementaria a efectos informativos.'))
                    ws["B22"] = rich_string

                    if pay_term == "100_delivery":
                        ws["B42"] = (
                            "100% of total amount of purchase order upon delivery of material according to Incoterms 2020, FCA (our facilities, Spain).\n"
                            "Payment method: bank transfer"
                        )
                        ws["B43"] = (
                            "Pago del 100% del valor total de la orden de compra a la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España).\n"
                            "Método de pago: Transferencia bancaria."
                        )
                    elif pay_term == "100_order":
                        ws["B42"] = (
                            "100 % of the total amount of purchase order upon receipt of purchase order.\n"
                            "Payment method: bank transfer"
                        )
                        ws["B43"] = (
                            "Pago del 100% del valor total de la orden de compra a la recepción de la orden.\n"
                            "Método de pago: Transferencia bancaria"
                        )
                    elif pay_term == "90_10":
                        ws["B42"] = (
                            "PAYMENT TERMS:\n"
                            "90 % of the total amount of PO upon delivery of material according to Incoterms 2020, FCA (our facilities, Spain) and 10% when final documentation is approved. \n"
                            "Bank Transfer: 60 days since invoice issue date."
                        )
                        ws["B43"] = (
                            "TERMINOS DE PAGO:\n"
                            "Pago del 90% del Valor total de la orden de compra a la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España) y el 10% restante cuando la documentación final sea aprobada.\n"
                            "Transferencia Bancaria: 60 días desde emisión de factura."
                        )
                    elif pay_term == "50_50":
                        ws["B42"] = (
                            "50 % of the total amount of purchase order upon receipt of purchase order. Remaining 50% before material be delivered according to Incoterms 2020, FCA (our facilities, Spain).\n"
                            "Payment method: bank transfer."
                        )
                        ws["B43"] = (
                            "Pago del 50% del valor total de la orden de compra a la recepción de la orden. El 50% restante antes de la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España).\n"
                            "Método de pago: Transferencia bancaria."
                        )
                    elif pay_term == "Others":
                        ws["B42"] = "PAYMENT TERMS TO BE DEFINED"
                        ws["B43"].font = Font(name="Calibri", size=11, bold=True, color="FF0000")
                        ws["B42"] = "TERMINOS DE PAGO POR DEFINIR"
                        ws["B43"].font = Font(name="Calibri", size=11, bold=True, italic=True, color="FF0000")

                    rich_string = CellRichText(
                    'For amounts greater than 30,000.00 € we can issue a warranty bond (if required) valid until the end of the indicated warranty period.\nBond warranty of 10% will be issued with the invoice of the last supplement.\n',
                    TextBlock(InlineFont(b=True), 'For lower amounts no warranty bond is issued.'))
                    ws["B45"] = rich_string

                    rich_string = CellRichText(
                    'Para importes superiores a 30.000,00 €, si es requerido, podremos emitir aval de garantía y estará vigente hasta el final del periodo de garantía indicado.\nEl aval del 10% será emitido con la factura del último suplemento.\n',
                    TextBlock(InlineFont(i=True, b=True), 'Por debajo de dicha cantidad, no se emitirán avales.'))
                    ws["B46"] = rich_string

                    ws["A58"] = (
                        "If you require further information related with this offer, please do not hesitate to contact:\n"
                        + responsible
                        + "\n"
                        + email
                        + "\n"
                        "Telf.: (+34) 916.582.118"
                    )

                    std = self.wb_technical["1.3"]
                    self.wb_technical.remove(std)

                left_text = "Fecha/Date: " + date_offer
                right_text = "Petición nº/Inquiry: " + num_ref

                for sheet in self.wb_technical.worksheets:
                    sheet.oddFooter.left.text = left_text
                    sheet.oddFooter.right.text = right_text
                    sheet.oddFooter.center.text = "Page &P de &N"

                    sheet.oddFooter.left.size = 9
                    sheet.oddFooter.right.size = 9
                    sheet.oddFooter.center.size = 9

                self.save_excel_technical()

                root = Tk()
                root.withdraw()  # Hiding main window Tkinter

                # close communication with the PostgreSQL database server
                # commit the changes
                conn.commit()
        except (Exception, psycopg2.DatabaseError) as error:
            dlg = QtWidgets.QMessageBox()
            new_icon = QtGui.QIcon()
            new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
            dlg.setWindowIcon(new_icon)
            dlg.setWindowTitle("ERP EIPSA")
            dlg.setText("Ha ocurrido el siguiente error:\n"
                        + str(error))
            dlg.setIcon(QtWidgets.QMessageBox.Icon.Critical)
            dlg.exec()
            del dlg, new_icon
        finally:
            if conn is not None:
                conn.close()

    def euros_to_float(self, value):
        """
        Converts a string value representing an amount in euros to a float.
        
        Args:
            value (str): The string representation of an amount in euros, with commas for decimal separation and ' €' for currency indication.
        
        Returns:
            float: The numeric value of the amount in euros.
        """
        value = value.replace(".", "")
        value = value.replace(",", ".")
        value = value[: value.find(" €")]
        return float(value)

    def save_excel_commercial(self):
        """Saves the populated Excel workbook to a specified location.
        Opens a dialog window for the user to select the file path and name.
        """
        # Dialog window to select folder and file name; if path is selected, excel file is saved
        output_path_commercial = asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Archivos de Excel", "*.xlsx")],
            title="Guardar Oferta comercial",
        )
        if output_path_commercial:
            self.wb_commercial.save(output_path_commercial)
            return output_path_commercial

    def save_excel_technical(self):
        """Saves the populated Excel workbook to a specified location.
        Opens a dialog window for the user to select the file path and name.
        """
        output_path_technical = asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Archivos de Excel", "*.xlsx")],
            title="Guardar Oferta técnica",
        )
        if output_path_technical:
            self.wb_technical.save(output_path_technical)

class offer_short_level_spanish:
    """
    A class to manage export offer (short format in spanish) details for level equipments.
    
    Attributes:
        numoffer (str): Offer number.
        username (str): Name of the user creating the offer.
        rev (str): Revision number of the offer.
        project (str): Name of the project.
        delivery_term (str): Delivery terms for the offer.
        delivery_time (str): Expected delivery time.
        validity (str): Validity period of the offer.
        pay_term (str): Payment terms.
        testinspection (str): Information about testing and inspection.
        revchanges (str): Details of changes made in the revision.
        notes (str): Additional notes, split by line.
    """
    def __init__(self, numoffer, username, rev, project, delivery_term, delivery_time, validity, pay_term, testinspection, revchanges, notes,):
        """
        Initializes the offer.

        Args:
            numoffer (str): Offer number.
            username (str): Name of the user creating the offer.
            rev (str): Revision number of the offer.
            project (str): Name of the project.
            delivery_term (str): Delivery terms for the offer.
            delivery_time (str): Expected delivery time.
            validity (str): Validity period of the offer.
            pay_term (str): Payment terms.
            testinspection (str): Information about testing and inspection.
            revchanges (str): Details of changes made in the revision.
            notes (str): Additional notes, split by line.
        """
        notes = notes.split('\n')
        date_offer = date.today().strftime("%d/%m/%Y")
        offername_commercial = numoffer + "-" + "Commercial.Rev" + rev
        offername_technical = numoffer + "-" + "Technical.Rev" + rev

        query_commercial = """
                    SELECT name, surname, email
                    FROM users_data.registration
                    WHERE username = %s
                    """
        query_dataoffer = """
                        SELECT client, num_ref_offer
                        FROM offers
                        WHERE UPPER (num_offer) LIKE UPPER('%%'||%s||'%%')
                        """
        query_tagsdata = """
                        SELECT *
                        FROM tags_data.tags_level
                        WHERE (
                        UPPER ("num_offer") LIKE UPPER('%%'||%s||'%%')
                        AND
                        "tag_state" NOT IN ('PURCHASED','DELETED')
                        )
                        """
        conn = None
        try:
            # read the connection parameters
            params = config()
            # connect to the PostgreSQL server
            conn = psycopg2.connect(**params)
            cur = conn.cursor()

            cur.execute(query_dataoffer, (numoffer,))
            results_offer = cur.fetchall()
            client = results_offer[0][0]
            num_ref = results_offer[0][1]

            cur.execute(query_commercial, (username,))
            results_commercial = cur.fetchall()
            responsible = results_commercial[0][0] + " " + results_commercial[0][1]
            email = results_commercial[0][2]

            cur.execute(query_tagsdata, (numoffer,))
            data_tags = cur.fetchall()

            if len(data_tags) == 0:
                dlg = QtWidgets.QMessageBox()
                new_icon = QtGui.QIcon()
                new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
                dlg.setWindowIcon(new_icon)
                dlg.setWindowTitle("ERP EIPSA")
                dlg.setText("No hay TAGS importados en la oferta")
                dlg.setIcon(QtWidgets.QMessageBox.Icon.Warning)
                dlg.exec()
                del dlg, new_icon
            else:
                columns = []
                for elt in cur.description:
                    columns.append(elt[0])

                value_type_dict = {"Transparent": 1, "Reflex": 1, "Magnetic": 2}

                df = pd.DataFrame(data=data_tags, columns=columns)
                df["value_type"] = df["item_type"].map(value_type_dict)
                df = df.sort_values(by=["tag", "value_type"])
                df = df.iloc[:, 1:38]
                df["value_type"] = df["item_type"].map(value_type_dict)
                df = df.sort_values(by=["value_type", "tag"])
                df["amount"] = df["amount"].apply(self.euros_to_float)
                total_amount_material = df["amount"].sum()
                df = df.drop([
                        "tag_state",
                        "num_offer",
                        "num_order",
                        "num_po",
                        "position",
                        "subposition",
                        "proc_conn_type",
                        "flags",
                        "flange_type",
                        "nipple_hex",
                        "nipple_tub"
                    ],
                    axis=1,)

                number_items = df.shape[0]
                documentation = number_items * 70

                # Loading Excel Template
                self.wb_commercial = load_workbook(
                    r"\\nas01\DATOS\Comunes\EIPSA-ERP\Plantillas Exportación\PLANTILLA CORTA OFERTA NIVEL.xlsx"
                )

                # Editing sheet COVER
                sheet_name = "COVER"
                ws = self.wb_commercial[sheet_name]
                ws["E4"] = client
                ws["E6"] = offername_commercial
                ws["E8"] = num_ref
                ws["E10"] = project
                ws["E12"] = date_offer
                ws["E14"] = delivery_term
                ws["E16"] = validity + " days"
                ws["C43"] = responsible
                ws["C45"] = email

                # Editing sheet EQUIPMENT DATA
                sheets_confirmed = ["COVER", "1.2", "1.3", "NOTES"]

                if int(rev) == 0:
                    for item_type in df["item_type"].unique():
                        sheets_confirmed.append(item_type)

                dict_sheets_data = {}

                for value_type in df["value_type"].unique():
                    df_toexport = df[df["value_type"] == value_type]
                    df_toexport.insert(0, "N°", range(1, len(df_toexport) + 1))
                    df_toexport.set_index("N°", inplace=True, drop=False)
                    df_toexport.index.name = None
                    df_toexport = df_toexport.drop(["value_type"], axis=1)

                    eq_type = (
                        "LEVEL GAUGES ELEMENTS DATA"
                        if value_type == 1
                        else "MAGNETIC ELEMENTS DATA"
                        )

                    if eq_type not in sheets_confirmed:
                        sheets_confirmed.append(eq_type)

                    if eq_type == "LEVEL GAUGES ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "dv_size",
                                "dv_rating",
                                "dv_facing",
                                "float_material"],
                            axis=1,)
                    elif eq_type == "MAGNETIC ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "dv_conn",
                                "item_type",
                                "valve_type",
                                "case_cover_material",
                                "illuminator",
                                "ip_code"],
                            axis=1,)

                    ws = self.wb_commercial[eq_type]
                    ws["G2"] = date_offer
                    ws["G3"] = num_ref
                    ws["G4"] = offername_commercial
                    if revchanges != "":
                        ws["G5"] = rev + " " + revchanges
                        ws["G5"].font = Font(name="Calibri", size=14, bold=True)
                        ws["G5"].fill = PatternFill("solid", fgColor="FFFF00")

                    if int(rev) > 0:
                        for row in ws.iter_rows(min_row=2, max_row=4, min_col=6, max_col=7):
                            for cell in row:
                                cell.value = None
                                cell._style = ws["F1"]._style

                    last_row = ws.max_row

                    num_column_amount = df_toexport.columns.get_loc("amount") + 1

                    for index, row in df_toexport.iterrows():  # Data in desired row
                        for col_num, value in enumerate(row, start=1):
                            cell = ws.cell(row=last_row + 1, column=col_num)
                            cell.value = value
                            if col_num == num_column_amount:
                                cell._style = ws["AG1"]._style
                            else:
                                cell._style = ws["AB1"]._style

                        last_row = ws.max_row

                    ws[f"A{last_row+3}"] = "VALIDEZ DE LA OFERTA: " + validity + " DÍAS"
                    ws[f"A{last_row+3}"]._style = ws["AI1"]._style
                    ws[f"A{last_row+4}"] = (
                        "PLAZO DE ENTREGA: "
                        + delivery_time
                        + " SEMANAS DESDE APROBACIÓN DE PLANOS (AGOSTO Y ÚLTIMAS DOS SEMANAS DE DICIEMBRE EXCLUIDOS)"
                    )
                    ws[f"A{last_row+4}"]._style = ws["AI1"]._style

                    if notes != "":
                        if isinstance(notes, list):
                            line = last_row + 5
                            for note in notes:
                                ws[f"A{line}"] = note
                                ws[f"A{line}"]._style = ws["AI1"]._style
                                line += 1
                        else:
                            line = last_row + 5
                            ws[f"A{line}"] = notes
                            ws[f"A{line}"]._style = ws["AI1"]._style

                    dict_sheets_data[eq_type] = [last_row, num_column_amount, df_toexport["amount"].sum(), df_toexport.shape[0]]

                ws.cell(row=last_row + 3, column=num_column_amount - 1).value = "CANTIDAD TOTAL"
                ws.cell(row=last_row + 3, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=last_row + 3, column=num_column_amount).value = number_items
                
                row_amount = last_row + 4
                for key, value in dict_sheets_data.items():
                    parts_key = key.split(" ")
                    ws.cell(row=row_amount + 2, column=num_column_amount - 1).value = "IMPORTE TOTAL " + parts_key[0] + " " + parts_key[1] + " (CANTIDAD: " + str(value[3]) + ")"
                    ws.cell(row=row_amount + 2, column=num_column_amount).value = value[2]
                    ws.cell(row=row_amount + 2, column=num_column_amount - 1)._style = ws["AA1"]._style
                    ws.cell(row=row_amount + 2, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                    ws.cell(row=row_amount + 2, column=num_column_amount)._style = ws["AC1"]._style

                    row_amount += 2

                ws.cell(row=row_amount + 2, column=num_column_amount - 1).value = "IMPORTE TOTAL DEL MATERIAL"
                ws.cell(row=row_amount + 2, column=num_column_amount).value = total_amount_material
                ws.cell(row=row_amount + 4, column=num_column_amount - 1).value = "PACKING Y TRANSPORTE"
                ws.cell(row=row_amount + 4, column=num_column_amount).value = (f"=MROUND({get_column_letter(num_column_amount)}{row_amount + 2}*0.03,10)" if total_amount_material > 6700 else 200)
                ws.cell(row=row_amount + 5, column=num_column_amount - 1).value = "PRUEBAS E INSPECCIÓN"
                ws.cell(row=row_amount + 5, column=num_column_amount).value = float(testinspection)
                ws.cell(row=row_amount + 6, column=num_column_amount - 1).value = "DOCUMENTACIÓN"
                ws.cell(row=row_amount + 6, column=num_column_amount).value = documentation
                ws.cell(row=row_amount + 8, column=num_column_amount - 1).value = "IMPORTE TOTAL DE LA OFERTA"
                ws.cell(row=row_amount + 8, column=num_column_amount).value = f"=SUM({get_column_letter(num_column_amount)}{row_amount + 2}:{get_column_letter(num_column_amount)}{row_amount + 6})"

                ws.cell(row=last_row + 3, column=num_column_amount - 1)._style = ws["AA1"]._style
                ws.cell(row=last_row + 3, column=num_column_amount).font = Font(name="Calibri", size=14)
                ws.cell(row=row_amount + 2, column=num_column_amount - 1)._style = ws["AA1"]._style
                ws.cell(row=row_amount + 2, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=row_amount + 2, column=num_column_amount)._style = ws["AC1"]._style
                ws.cell(row=row_amount + 4, column=num_column_amount - 1).font = Font(name="Calibri", size=14)
                ws.cell(row=row_amount + 4, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=row_amount + 4, column=num_column_amount)._style = ws["AC1"]._style
                ws.cell(row=row_amount + 5, column=num_column_amount - 1).font = Font(name="Calibri", size=14)
                ws.cell(row=row_amount + 5, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=row_amount + 5, column=num_column_amount)._style = ws["AC1"]._style
                ws.cell(row=row_amount + 6, column=num_column_amount - 2)._style = ws["AD1"]._style
                ws.cell(row=row_amount + 6, column=num_column_amount - 1)._style = ws["AD1"]._style
                ws.cell(row=row_amount + 6, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=row_amount + 6, column=num_column_amount)._style = ws["AE1"]._style
                ws.cell(row=row_amount + 8, column=num_column_amount - 1)._style = ws["AA1"]._style
                ws.cell(row=row_amount + 8, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=row_amount + 8, column=num_column_amount)._style = ws["AF1"]._style

            # Editing sheet NOTES
                sheet_name = "NOTES"  # Selecting  sheet
                ws = self.wb_commercial[sheet_name]

                rich_string = CellRichText(
                TextBlock(InlineFont(i=True), 'Estamos ofertando solamente indicadores de nivel de vidrio y magnéticos, les informamos que en nuestra gama de fabricación con certificación europea, incluye también elementos de caudal y elementos de medida de temperatura. (https://www.eipsa.es/productos)\n'),
                TextBlock(InlineFont(b=True, i=True), 'Los precios ofertados podrían reducirse en caso de compra de toda nuestra gama.'))
                ws["B6"] = rich_string

                rich_string = CellRichText(
                'Plazo de entrega ' + delivery_time + ' semanas desde aprobación de planos de la ',
                TextBlock(InlineFont(b=True, i=True), 'totalidad de los equipos'),
                TextBlock(InlineFont(i=True),' amparados por el contrato, asi como la documentación crítica (Agosto y las dos últimas semanas de diciembre excluidos).'))
                ws["B10"] = rich_string

                rich_string = CellRichText(
                'Esta oferta ha sido elaborada en base a la información que figura en la hoja de datos correspondiente a cada TAG, ',
                TextBlock(InlineFont(i=True, u='single'), 'no siendo responsabilidad final de EIPSA la elección del material a utilizar en los equipos ni analizar datos de proceso.\n'),
                TextBlock(InlineFont(i=True),'La hoja de datos será el único documento técnico/contractual, cualquier otra documentación recibida será considerada como documentación complementaria a efectos informativos.'))
                ws["B16"] = rich_string

                if pay_term == "100_delivery":
                    ws["B29"] = (
                        "Pago del 100% del valor total de la orden de compra a la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España).\n"
                        "Método de pago: Transferencia bancaria."
                    )
                elif pay_term == "100_order":
                    ws["B29"] = (
                        "Pago del 100% del valor total de la orden de compra a la recepción de la orden.\n"
                        "Método de pago: Transferencia bancaria"
                    )
                elif pay_term == "90_10":
                    ws["B29"] = (
                        "TERMINOS DE PAGO:\n"
                        "Pago del 90% del Valor total de la orden de compra a la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España) y el 10% restante cuando la documentación final sea aprobada.\n"
                        "Transferencia Bancaria: 60 días desde emisión de factura."
                    )
                elif pay_term == "50_50":
                    ws["B29"] = (
                        "Pago del 50% del valor total de la orden de compra a la recepción de la orden. El 50% restante antes de la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España).\n"
                        "Método de pago: Transferencia bancaria."
                    )
                elif pay_term == "Others":
                    ws["B29"] = "TERMINOS DE PAGO POR DEFINIR"
                    ws["B29"].font = Font(name="Calibri", size=11, bold=True, italic=True, color="FF0000")

                rich_string = CellRichText(
                'Para importes superiores a 30.000,00 €, si es requerido, podremos emitir aval de garantía y estará vigente hasta el final del periodo de garantía indicado.\nEl aval del 10% será emitido con la factura del último suplemento.\n',
                TextBlock(InlineFont(i=True, b=True), 'Por debajo de dicha cantidad, no se emitirán avales.'))
                ws["B27"] = rich_string

                ws["A37"] = (
                    "Si necesita más información relacionada con esta oferta, no dude en ponerse en contacto con:\n"
                    + responsible
                    + "\n"
                    + email
                    + "\n"
                    "Telf.: (+34) 916.582.118"
                )

                for sheet in self.wb_commercial.sheetnames:
                    if sheet not in sheets_confirmed:
                        sheet_to_delete = self.wb_commercial[sheet]
                        self.wb_commercial.remove(sheet_to_delete)

                if int(rev) > 0:
                    sheets_confirmed = ["COVER", "1.2", "1.3", "NOTES"]
                    for sheet in sheets_confirmed:
                        sheet_to_delete = self.wb_commercial[sheet]
                        self.wb_commercial.remove(sheet_to_delete)

                left_text = "Fecha/Date: " + date_offer
                right_text = "Petición nº/Inquiry: " + num_ref

                for sheet in self.wb_commercial.worksheets:
                    sheet.oddFooter.left.text = left_text
                    sheet.oddFooter.right.text = right_text
                    sheet.oddFooter.center.text = "Page &P de &N"

                    sheet.oddFooter.left.size = 9
                    sheet.oddFooter.right.size = 9
                    sheet.oddFooter.center.size = 9
                
                path = self.save_excel_commercial()

                # Creating the technical offer using the commercial one as template
                self.wb_technical = load_workbook(path)

                if int(rev) == 0:
                    sheet_name = "COVER"
                    ws = self.wb_technical[sheet_name]
                    ws["E6"] = offername_technical

                for value_type in df["value_type"].unique():
                    eq_type = (
                        "LEVEL GAUGES ELEMENTS DATA"
                        if value_type == 1
                        else "MAGNETIC ELEMENTS DATA"
                        )

                    ws = self.wb_technical[eq_type]
                    if int(rev) == 0:
                        ws["G4"] = offername_technical

                    last_row = dict_sheets_data[eq_type][0]
                    num_column_amount = dict_sheets_data[eq_type][1]

                    # self.wb_technical[eq_type].delete_rows(last_row + 5, 20)

                    ws.cell(row=row_amount + 2, column=num_column_amount).value = ""
                    ws.cell(row=row_amount + 4, column=num_column_amount).value = ""
                    ws.cell(row=row_amount + 5, column=num_column_amount).value = ""
                    ws.cell(row=row_amount + 6, column=num_column_amount).value = ""
                    ws.cell(row=row_amount + 8, column=num_column_amount).value = ""

                    ws.cell(row=row_amount + 2, column=num_column_amount - 1).value = ""
                    ws.cell(row=row_amount + 4, column=num_column_amount - 1).value = ""
                    ws.cell(row=row_amount + 5, column=num_column_amount - 1).value = ""
                    ws.cell(row=row_amount + 6, column=num_column_amount - 1).value = ""
                    ws.cell(row=row_amount + 8, column=num_column_amount - 1).value = ""

                # Deleting "Amount" column
                    self.wb_technical[eq_type].delete_cols(num_column_amount, 1)

                # Adjusting the print area
                    new_last_column = num_column_amount
                    last_print_row = 40
                    nuevo_rango_impresion = f'A1:{get_column_letter(new_last_column)}{last_print_row}'
                    self.wb_technical[eq_type].print_area = nuevo_rango_impresion

                    stamp_1 = self.wb_technical[eq_type]._images[1]
                    anchor_actual = stamp_1.anchor

                    from_col, from_col_off = anchor_actual._from.col, anchor_actual._from.colOff
                    from_row, from_row_off = anchor_actual._from.row, anchor_actual._from.rowOff
                    to_col, to_col_off = anchor_actual.to.col, anchor_actual.to.colOff
                    to_row, to_row_off = anchor_actual.to.row, anchor_actual.to.rowOff

                    from_cell = AnchorMarker(col=from_col - 1, colOff=from_col_off, row=from_row, rowOff=from_row_off)
                    to_cell = AnchorMarker(col=to_col - 1, colOff=to_col_off, row=to_row, rowOff=to_row_off)

                    new_anchor = TwoCellAnchor(_from = from_cell, to = to_cell, editAs='absolute')
                    stamp_1.anchor = new_anchor

                    stamp_2 = self.wb_technical[eq_type]._images[2]
                    anchor_actual = stamp_2.anchor

                    from_col, from_col_off = anchor_actual._from.col, anchor_actual._from.colOff
                    from_row, from_row_off = anchor_actual._from.row, anchor_actual._from.rowOff
                    to_col, to_col_off = anchor_actual.to.col, anchor_actual.to.colOff
                    to_row, to_row_off = anchor_actual.to.row, anchor_actual.to.rowOff

                    from_cell = AnchorMarker(col=from_col - 1, colOff=from_col_off, row=from_row, rowOff=from_row_off)
                    to_cell = AnchorMarker(col=to_col - 1, colOff=to_col_off, row=to_row, rowOff=to_row_off)

                    new_anchor = TwoCellAnchor(_from = from_cell, to = to_cell, editAs='absolute')
                    stamp_2.anchor = new_anchor

                if int(rev) == 0:
                    ws = self.wb_technical[self.wb_technical.sheetnames[-2]]
                    ws.cell(row=last_row + 3, column=num_column_amount).value = number_items
                    ws.cell(row=last_row + 3, column=num_column_amount).font = Font(name="Calibri", size=14)

            # Editing sheet NOTES
                if int(rev) == 0:
                    sheet_name = "NOTES"
                    ws = self.wb_technical[sheet_name]

                    rich_string = CellRichText(
                    TextBlock(InlineFont(i=True), 'Estamos ofertando solamente indicadores de nivel de vidrio y magnéticos, les informamos que en nuestra gama de fabricación con certificación europea, incluye también elementos de caudal y elementos de medida de temperatura. (https://www.eipsa.es/productos)\n'),
                    TextBlock(InlineFont(b=True, i=True), 'Los precios ofertados podrían reducirse en caso de compra de toda nuestra gama.'))
                    ws["B6"] = rich_string

                    rich_string = CellRichText(
                    'Plazo de entrega ' + delivery_time + ' semanas desde aprobación de planos de la ',
                    TextBlock(InlineFont(b=True, i=True), 'totalidad de los equipos'),
                    TextBlock(InlineFont(i=True),' amparados por el contrato, asi como la documentación crítica (Agosto y las dos últimas semanas de diciembre excluidos).'))
                    ws["B10"] = rich_string

                    rich_string = CellRichText(
                    'Esta oferta ha sido elaborada en base a la información que figura en la hoja de datos correspondiente a cada TAG, ',
                    TextBlock(InlineFont(i=True, u='single'), 'no siendo responsabilidad final de EIPSA la elección del material a utilizar en los equipos ni analizar datos de proceso.\n'),
                    TextBlock(InlineFont(i=True),'La hoja de datos será el único documento técnico/contractual, cualquier otra documentación recibida será considerada como documentación complementaria a efectos informativos.'))
                    ws["B16"] = rich_string

                    if pay_term == "100_delivery":
                        ws["B29"] = (
                            "Pago del 100% del valor total de la orden de compra a la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España).\n"
                            "Método de pago: Transferencia bancaria."
                        )
                    elif pay_term == "100_order":
                        ws["B29"] = (
                            "Pago del 100% del valor total de la orden de compra a la recepción de la orden.\n"
                            "Método de pago: Transferencia bancaria"
                        )
                    elif pay_term == "90_10":
                        ws["B29"] = (
                            "TERMINOS DE PAGO:\n"
                            "Pago del 90% del Valor total de la orden de compra a la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España) y el 10% restante cuando la documentación final sea aprobada.\n"
                            "Transferencia Bancaria: 60 días desde emisión de factura."
                        )
                    elif pay_term == "50_50":
                        ws["B29"] = (
                            "Pago del 50% del valor total de la orden de compra a la recepción de la orden. El 50% restante antes de la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España).\n"
                            "Método de pago: Transferencia bancaria."
                        )
                    elif pay_term == "Others":
                        ws["B29"] = "TERMINOS DE PAGO POR DEFINIR"
                        ws["B29"].font = Font(name="Calibri", size=11, bold=True, italic=True, color="FF0000")

                    rich_string = CellRichText(
                    'Para importes superiores a 30.000,00 €, si es requerido, podremos emitir aval de garantía y estará vigente hasta el final del periodo de garantía indicado.\nEl aval del 10% será emitido con la factura del último suplemento.\n',
                    TextBlock(InlineFont(i=True, b=True), 'Por debajo de dicha cantidad, no se emitirán avales.'))
                    ws["B27"] = rich_string

                    ws["A37"] = (
                        "Si necesita más información relacionada con esta oferta, no dude en ponerse en contacto con:\n"
                        + responsible
                        + "\n"
                        + email
                        + "\n"
                        "Telf.: (+34) 916.582.118"
                    )

                    std = self.wb_technical["1.3"]
                    self.wb_technical.remove(std)

                left_text = "Fecha/Date: " + date_offer
                right_text = "Petición nº/Inquiry: " + num_ref

                for sheet in self.wb_technical.worksheets:
                    sheet.oddFooter.left.text = left_text
                    sheet.oddFooter.right.text = right_text
                    sheet.oddFooter.center.text = "Page &P de &N"

                    sheet.oddFooter.left.size = 9
                    sheet.oddFooter.right.size = 9
                    sheet.oddFooter.center.size = 9

                self.save_excel_technical()

                root = Tk()
                root.withdraw()  # Hiding main window Tkinter

                # close communication with the PostgreSQL database server
                # commit the changes
                conn.commit()
        except (Exception, psycopg2.DatabaseError) as error:
            dlg = QtWidgets.QMessageBox()
            new_icon = QtGui.QIcon()
            new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
            dlg.setWindowIcon(new_icon)
            dlg.setWindowTitle("ERP EIPSA")
            dlg.setText("Ha ocurrido el siguiente error:\n"
                        + str(error))
            dlg.setIcon(QtWidgets.QMessageBox.Icon.Critical)
            dlg.exec()
            del dlg, new_icon
        finally:
            if conn is not None:
                conn.close()

    def euros_to_float(self, value):
        """
        Converts a string value representing an amount in euros to a float.
        
        Args:
            value (str): The string representation of an amount in euros, with commas for decimal separation and ' €' for currency indication.
        
        Returns:
            float: The numeric value of the amount in euros.
        """
        value = value.replace(".", "")
        value = value.replace(",", ".")
        value = value[: value.find(" €")]
        return float(value)

    def save_excel_commercial(self):
        """Saves the populated Excel workbook to a specified location.
        Opens a dialog window for the user to select the file path and name.
        """
        # Dialog window to select folder and file name; if path is selected, excel file is saved
        output_path_commercial = asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Archivos de Excel", "*.xlsx")],
            title="Guardar Oferta comercial",
        )
        if output_path_commercial:
            self.wb_commercial.save(output_path_commercial)
            return output_path_commercial

    def save_excel_technical(self):
        """Saves the populated Excel workbook to a specified location.
        Opens a dialog window for the user to select the file path and name.
        """
        output_path_technical = asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Archivos de Excel", "*.xlsx")],
            title="Guardar Oferta técnica",
        )
        if output_path_technical:
            self.wb_technical.save(output_path_technical)

class offer_short_level_english:
    """
    A class to manage export offer (short format in english) details for level equipments.
    
    Attributes:
        numoffer (str): Offer number.
        username (str): Name of the user creating the offer.
        rev (str): Revision number of the offer.
        project (str): Name of the project.
        delivery_term (str): Delivery terms for the offer.
        delivery_time (str): Expected delivery time.
        validity (str): Validity period of the offer.
        pay_term (str): Payment terms.
        testinspection (str): Information about testing and inspection.
        revchanges (str): Details of changes made in the revision.
        notes (str): Additional notes, split by line.
    """
    def __init__(self, numoffer, username, rev, project, delivery_term, delivery_time, validity, pay_term, testinspection, revchanges, notes,):
        """
        Initializes the offer.

        Args:
            numoffer (str): Offer number.
            username (str): Name of the user creating the offer.
            rev (str): Revision number of the offer.
            project (str): Name of the project.
            delivery_term (str): Delivery terms for the offer.
            delivery_time (str): Expected delivery time.
            validity (str): Validity period of the offer.
            pay_term (str): Payment terms.
            testinspection (str): Information about testing and inspection.
            revchanges (str): Details of changes made in the revision.
            notes (str): Additional notes, split by line.
        """
        notes = notes.split('\n')
        date_offer = date.today().strftime("%d/%m/%Y")
        offername_commercial = numoffer + "-" + "Commercial.Rev" + rev
        offername_technical = numoffer + "-" + "Technical.Rev" + rev

        query_commercial = """
                    SELECT name, surname, email
                    FROM users_data.registration
                    WHERE username = %s
                    """
        query_dataoffer = """
                        SELECT client, num_ref_offer
                        FROM offers
                        WHERE UPPER (num_offer) LIKE UPPER('%%'||%s||'%%')
                        """
        query_tagsdata = """
                        SELECT *
                        FROM tags_data.tags_level
                        WHERE (
                        UPPER ("num_offer") LIKE UPPER('%%'||%s||'%%')
                        AND
                        "tag_state" NOT IN ('PURCHASED','DELETED')
                        )
                        """
        conn = None
        try:
            # read the connection parameters
            params = config()
            # connect to the PostgreSQL server
            conn = psycopg2.connect(**params)
            cur = conn.cursor()

            cur.execute(query_dataoffer, (numoffer,))
            results_offer = cur.fetchall()
            client = results_offer[0][0]
            num_ref = results_offer[0][1]

            cur.execute(query_commercial, (username,))
            results_commercial = cur.fetchall()
            responsible = results_commercial[0][0] + " " + results_commercial[0][1]
            email = results_commercial[0][2]

            cur.execute(query_tagsdata, (numoffer,))
            data_tags = cur.fetchall()

            if len(data_tags) == 0:
                dlg = QtWidgets.QMessageBox()
                new_icon = QtGui.QIcon()
                new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
                dlg.setWindowIcon(new_icon)
                dlg.setWindowTitle("ERP EIPSA")
                dlg.setText("No hay TAGS importados en la oferta")
                dlg.setIcon(QtWidgets.QMessageBox.Icon.Warning)
                dlg.exec()
                del dlg, new_icon
            else:
                columns = []
                for elt in cur.description:
                    columns.append(elt[0])

                value_type_dict = {"Transparent": 1, "Reflex": 1, "Magnetic": 2}

                df = pd.DataFrame(data=data_tags, columns=columns)
                df["value_type"] = df["item_type"].map(value_type_dict)
                df = df.sort_values(by=["tag", "value_type"])
                df = df.iloc[:, 1:38]
                df["value_type"] = df["item_type"].map(value_type_dict)
                df = df.sort_values(by=["value_type", "tag"])
                df["amount"] = df["amount"].apply(self.euros_to_float)
                total_amount_material = df["amount"].sum()
                df = df.drop([
                        "tag_state",
                        "num_offer",
                        "num_order",
                        "num_po",
                        "position",
                        "subposition",
                        "proc_conn_type",
                        "flags",
                        "flange_type",
                        "nipple_hex",
                        "nipple_tub"
                    ],
                    axis=1,)

                number_items = df.shape[0]
                documentation = number_items * 70

                # Loading Excel Template
                self.wb_commercial = load_workbook(
                    r"\\nas01\DATOS\Comunes\EIPSA-ERP\Plantillas Exportación\PLANTILLA CORTA OFERTA NIVEL - ingles.xlsx"
                )

                # Editing sheet COVER
                sheet_name = "COVER"
                ws = self.wb_commercial[sheet_name]
                ws["E4"] = client
                ws["E6"] = offername_commercial
                ws["E8"] = num_ref
                ws["E10"] = project
                ws["E12"] = date_offer
                ws["E14"] = delivery_term
                ws["E16"] = validity + " days"
                ws["C43"] = responsible
                ws["C45"] = email

                # Editing sheet EQUIPMENT DATA
                sheets_confirmed = ["COVER", "1.2", "1.3", "NOTES"]

                if int(rev) == 0:
                    for item_type in df["item_type"].unique():
                        sheets_confirmed.append(item_type)

                dict_sheets_data = {}

                for value_type in df["value_type"].unique():
                    df_toexport = df[df["value_type"] == value_type]
                    df_toexport.insert(0, "N°", range(1, len(df_toexport) + 1))
                    df_toexport.set_index("N°", inplace=True, drop=False)
                    df_toexport.index.name = None
                    df_toexport = df_toexport.drop(["value_type"], axis=1)

                    eq_type = (
                        "LEVEL GAUGES ELEMENTS DATA"
                        if value_type == 1
                        else "MAGNETIC ELEMENTS DATA"
                        )

                    if eq_type not in sheets_confirmed:
                        sheets_confirmed.append(eq_type)

                    if eq_type == "LEVEL GAUGES ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "dv_size",
                                "dv_rating",
                                "dv_facing",
                                "float_material"],
                            axis=1,)
                    elif eq_type == "MAGNETIC ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "dv_conn",
                                "item_type",
                                "valve_type",
                                "case_cover_material",
                                "illuminator",
                                "ip_code"],
                            axis=1,)

                    ws = self.wb_commercial[eq_type]
                    ws["G2"] = date_offer
                    ws["G3"] = num_ref
                    ws["G4"] = offername_commercial
                    if revchanges != "":
                        ws["G5"] = rev + " " + revchanges
                        ws["G5"].font = Font(name="Calibri", size=14, bold=True)
                        ws["G5"].fill = PatternFill("solid", fgColor="FFFF00")

                    if int(rev) > 0:
                        for row in ws.iter_rows(min_row=2, max_row=4, min_col=6, max_col=7):
                            for cell in row:
                                cell.value = None
                                cell._style = ws["F1"]._style

                    last_row = ws.max_row

                    num_column_amount = df_toexport.columns.get_loc("amount") + 1

                    for index, row in df_toexport.iterrows():  # Data in desired row
                        for col_num, value in enumerate(row, start=1):
                            cell = ws.cell(row=last_row + 1, column=col_num)
                            cell.value = value
                            if col_num == num_column_amount:
                                cell._style = ws["AG1"]._style
                            else:
                                cell._style = ws["AB1"]._style

                        last_row = ws.max_row

                    ws[f"A{last_row+3}"] = "OFFER VALIDITY: " + validity + " DAYS"
                    ws[f"A{last_row+3}"]._style = ws["AI1"]._style
                    ws[f"A{last_row+4}"] = (
                        "DELIVERY TIME: "
                        + delivery_time
                        + " WEEKS SINCE DRAWING APPROVAL (AUGUST AND LAST TWO DECEMBER WEEKS EXCLUDED)"
                    )
                    ws[f"A{last_row+4}"]._style = ws["AI1"]._style

                    if notes != "":
                        if isinstance(notes, list):
                            line = last_row + 5
                            for note in notes:
                                ws[f"A{line}"] = note
                                ws[f"A{line}"]._style = ws["AI1"]._style
                                line += 1
                        else:
                            line = last_row + 5
                            ws[f"A{line}"] = notes
                            ws[f"A{line}"]._style = ws["AI1"]._style

                    dict_sheets_data[eq_type] = [last_row, num_column_amount, df_toexport["amount"].sum(), df_toexport.shape[0]]

                ws.cell(row=last_row + 3, column=num_column_amount - 1).value = "QTY. TOTAL"
                ws.cell(row=last_row + 3, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=last_row + 3, column=num_column_amount).value = number_items
                
                row_amount = last_row + 4
                for key, value in dict_sheets_data.items():
                    parts_key = key.split(" ")
                    ws.cell(row=row_amount + 2, column=num_column_amount - 1).value = "TOTAL AMOUNT OF " + parts_key[0] + " " + parts_key[1] + " (QTY: " + str(value[3]) + ")"
                    ws.cell(row=row_amount + 2, column=num_column_amount).value = value[2]
                    ws.cell(row=row_amount + 2, column=num_column_amount - 1)._style = ws["AA1"]._style
                    ws.cell(row=row_amount + 2, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                    ws.cell(row=row_amount + 2, column=num_column_amount)._style = ws["AC1"]._style

                    row_amount += 2

                ws.cell(row=row_amount + 2, column=num_column_amount - 1).value = "TOTAL AMOUNT OF MATERIAL"
                ws.cell(row=row_amount + 2, column=num_column_amount).value = total_amount_material
                ws.cell(row=row_amount + 4, column=num_column_amount - 1).value = "PACKING AND TRANSPORT"
                ws.cell(row=row_amount + 4, column=num_column_amount).value = (f"=MROUND({get_column_letter(num_column_amount)}{row_amount + 2}*0.03,10)" if total_amount_material > 6700 else 200)
                ws.cell(row=row_amount + 5, column=num_column_amount - 1).value = "TESTS & INSPECTION"
                ws.cell(row=row_amount + 5, column=num_column_amount).value = float(testinspection)
                ws.cell(row=row_amount + 6, column=num_column_amount - 1).value = "DOCUMENTATION"
                ws.cell(row=row_amount + 6, column=num_column_amount).value = documentation
                ws.cell(row=row_amount + 8, column=num_column_amount - 1).value = "TOTAL AMOUNT OF BID"
                ws.cell(row=row_amount + 8, column=num_column_amount).value = f"=SUM({get_column_letter(num_column_amount)}{row_amount + 2}:{get_column_letter(num_column_amount)}{row_amount + 6})"

                ws.cell(row=last_row + 3, column=num_column_amount - 1)._style = ws["AA1"]._style
                ws.cell(row=last_row + 3, column=num_column_amount).font = Font(name="Calibri", size=14)
                ws.cell(row=row_amount + 2, column=num_column_amount - 1)._style = ws["AA1"]._style
                ws.cell(row=row_amount + 2, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=row_amount + 2, column=num_column_amount)._style = ws["AC1"]._style
                ws.cell(row=row_amount + 4, column=num_column_amount - 1).font = Font(name="Calibri", size=14)
                ws.cell(row=row_amount + 4, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=row_amount + 4, column=num_column_amount)._style = ws["AC1"]._style
                ws.cell(row=row_amount + 5, column=num_column_amount - 1).font = Font(name="Calibri", size=14)
                ws.cell(row=row_amount + 5, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=row_amount + 5, column=num_column_amount)._style = ws["AC1"]._style
                ws.cell(row=row_amount + 6, column=num_column_amount - 2)._style = ws["AD1"]._style
                ws.cell(row=row_amount + 6, column=num_column_amount - 1)._style = ws["AD1"]._style
                ws.cell(row=row_amount + 6, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=row_amount + 6, column=num_column_amount)._style = ws["AE1"]._style
                ws.cell(row=row_amount + 8, column=num_column_amount - 1)._style = ws["AA1"]._style
                ws.cell(row=row_amount + 8, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=row_amount + 8, column=num_column_amount)._style = ws["AF1"]._style

            # Editing sheet NOTES
                sheet_name = "NOTES"  # Selecting  sheet
                ws = self.wb_commercial[sheet_name]

                rich_string = CellRichText(
                'We are only offering measuring glass and magnetic level indicators. Please be informed that our product range includes flow elements and temperature elements; all with european certification. (https://www.eipsa.es/en/products)\n',
                TextBlock(InlineFont(b=True), 'The prices quoted could be reduced in case of purchasing our full range of products.'))
                ws["B6"] = rich_string

                rich_string = CellRichText(
                TextBlock(InlineFont(i=True), 'Estamos ofertando solamente indicadores de nivel de vidrio y magnéticos, les informamos que en nuestra gama de fabricación con certificación europea, incluye también elementos de caudal y elementos de medida de temperatura. (https://www.eipsa.es/productos)\n'),
                TextBlock(InlineFont(b=True, i=True), 'Los precios ofertados podrían reducirse en caso de compra de toda nuestra gama.'))
                ws["B7"] = rich_string

                rich_string = CellRichText(
                'Delivery time ' + delivery_time + ' weeks since drawing approval of ',
                TextBlock(InlineFont(b=True), 'all equipment'),
                ' in the contract, as well as critical documentation (August and last two December weeks excluded).')
                ws["B12"] = rich_string

                rich_string = CellRichText(
                'Plazo de entrega ' + delivery_time + ' semanas desde aprobación de planos de la ',
                TextBlock(InlineFont(b=True, i=True), 'totalidad de los equipos'),
                TextBlock(InlineFont(i=True),' amparados por el contrato, asi como la documentación crítica (Agosto y las dos últimas semanas de diciembre excluidos).'))
                ws["B13"] = rich_string

                rich_string = CellRichText(
                'Quotation prepared according to the information provided in the datasheet corresponding to each TAG. ',
                TextBlock(InlineFont(u='single'), 'EIPSA does not hold the final responsibility regarding selection of equipment material neither analyze process data.\n'),
                'The datasheet will be considered the only technical/contractual document, any other documentation will be considered as complementary documentation with informative purpose.')
                ws["B21"] = rich_string

                rich_string = CellRichText(
                'Esta oferta ha sido elaborada en base a la información que figura en la hoja de datos correspondiente a cada TAG, ',
                TextBlock(InlineFont(i=True, u='single'), 'no siendo responsabilidad final de EIPSA la elección del material a utilizar en los equipos ni analizar datos de proceso.\n'),
                TextBlock(InlineFont(i=True),'La hoja de datos será el único documento técnico/contractual, cualquier otra documentación recibida será considerada como documentación complementaria a efectos informativos.'))
                ws["B22"] = rich_string

                if pay_term == "100_delivery":
                    ws["B41"] = (
                        "100% of total amount of purchase order upon delivery of material according to Incoterms 2020, FCA (our facilities, Spain).\n"
                        "Payment method: bank transfer"
                    )
                    ws["B42"] = (
                        "Pago del 100% del valor total de la orden de compra a la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España).\n"
                        "Método de pago: Transferencia bancaria."
                    )
                elif pay_term == "100_order":
                    ws["B41"] = (
                        "100 % of the total amount of purchase order upon receipt of purchase order.\n"
                        "Payment method: bank transfer"
                    )
                    ws["B42"] = (
                        "Pago del 100% del valor total de la orden de compra a la recepción de la orden.\n"
                        "Método de pago: Transferencia bancaria"
                    )
                elif pay_term == "90_10":
                    ws["B41"] = (
                        "PAYMENT TERMS:\n"
                        "90 % of the total amount of PO upon delivery of material according to Incoterms 2020, FCA (our facilities, Spain) and 10% when final documentation is approved. \n"
                        "Bank Transfer: 60 days since invoice issue date."
                    )
                    ws["B42"] = (
                        "TERMINOS DE PAGO:\n"
                        "Pago del 90% del Valor total de la orden de compra a la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España) y el 10% restante cuando la documentación final sea aprobada.\n"
                        "Transferencia Bancaria: 60 días desde emisión de factura."
                    )
                elif pay_term == "50_50":
                    ws["B41"] = (
                        "50 % of the total amount of purchase order upon receipt of purchase order. Remaining 50% before material be delivered according to Incoterms 2020, FCA (our facilities, Spain).\n"
                        "Payment method: bank transfer."
                    )
                    ws["B42"] = (
                        "Pago del 50% del valor total de la orden de compra a la recepción de la orden. El 50% restante antes de la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España).\n"
                        "Método de pago: Transferencia bancaria."
                    )
                elif pay_term == "Others":
                    ws["B41"] = "PAYMENT TERMS TO BE DEFINED"
                    ws["B41"].font = Font(name="Calibri", size=11, bold=True, color="FF0000")
                    ws["B42"] = "TERMINOS DE PAGO POR DEFINIR"
                    ws["B42"].font = Font(name="Calibri", size=11, bold=True, italic=True, color="FF0000")

                rich_string = CellRichText(
                'For amounts greater than 30,000.00 € we can issue a warranty bond (if required) valid until the end of the indicated warranty period.\nBond warranty of 10% will be issued with the invoice of the last supplement.\n',
                TextBlock(InlineFont(b=True), 'For lower amounts no warranty bond is issued.'))
                ws["B38"] = rich_string

                rich_string = CellRichText(
                'Para importes superiores a 30.000,00 €, si es requerido, podremos emitir aval de garantía y estará vigente hasta el final del periodo de garantía indicado.\nEl aval del 10% será emitido con la factura del último suplemento.\n',
                TextBlock(InlineFont(i=True, b=True), 'Por debajo de dicha cantidad, no se emitirán avales.'))
                ws["B39"] = rich_string

                ws["A51"] = (
                    "If you require further information related with this offer, please do not hesitate to contact:\n"
                    + responsible
                    + "\n"
                    + email
                    + "\n"
                    "Telf.: (+34) 916.582.118"
                )

                for sheet in self.wb_commercial.sheetnames:
                    if sheet not in sheets_confirmed:
                        sheet_to_delete = self.wb_commercial[sheet]
                        self.wb_commercial.remove(sheet_to_delete)

                if int(rev) > 0:
                    sheets_confirmed = ["COVER", "1.2", "1.3", "NOTES"]
                    for sheet in sheets_confirmed:
                        sheet_to_delete = self.wb_commercial[sheet]
                        self.wb_commercial.remove(sheet_to_delete)

                left_text = "Fecha/Date: " + date_offer
                right_text = "Petición nº/Inquiry: " + num_ref

                for sheet in self.wb_commercial.worksheets:
                    sheet.oddFooter.left.text = left_text
                    sheet.oddFooter.right.text = right_text
                    sheet.oddFooter.center.text = "Page &P de &N"

                    sheet.oddFooter.left.size = 9
                    sheet.oddFooter.right.size = 9
                    sheet.oddFooter.center.size = 9
                
                path = self.save_excel_commercial()

                # Creating the technical offer using the commercial one as template
                self.wb_technical = load_workbook(path)

                if int(rev) == 0:
                    sheet_name = "COVER"
                    ws = self.wb_technical[sheet_name]
                    ws["E6"] = offername_technical

                for value_type in df["value_type"].unique():
                    eq_type = (
                        "LEVEL GAUGES ELEMENTS DATA"
                        if value_type == 1
                        else "MAGNETIC ELEMENTS DATA"
                        )

                    ws = self.wb_technical[eq_type]
                    if int(rev) == 0:
                        ws["G4"] = offername_technical

                    last_row = dict_sheets_data[eq_type][0]
                    num_column_amount = dict_sheets_data[eq_type][1]

                    # self.wb_technical[eq_type].delete_rows(last_row + 5, 20)

                    ws.cell(row=row_amount + 2, column=num_column_amount).value = ""
                    ws.cell(row=row_amount + 4, column=num_column_amount).value = ""
                    ws.cell(row=row_amount + 5, column=num_column_amount).value = ""
                    ws.cell(row=row_amount + 6, column=num_column_amount).value = ""
                    ws.cell(row=row_amount + 8, column=num_column_amount).value = ""

                    ws.cell(row=row_amount + 2, column=num_column_amount - 1).value = ""
                    ws.cell(row=row_amount + 4, column=num_column_amount - 1).value = ""
                    ws.cell(row=row_amount + 5, column=num_column_amount - 1).value = ""
                    ws.cell(row=row_amount + 6, column=num_column_amount - 1).value = ""
                    ws.cell(row=row_amount + 8, column=num_column_amount - 1).value = ""

                # Deleting "Amount" column
                    self.wb_technical[eq_type].delete_cols(num_column_amount, 1)

                # Adjusting the print area
                    new_last_column = num_column_amount
                    last_print_row = 40
                    nuevo_rango_impresion = f'A1:{get_column_letter(new_last_column)}{last_print_row}'
                    self.wb_technical[eq_type].print_area = nuevo_rango_impresion

                    stamp_1 = self.wb_technical[eq_type]._images[1]
                    anchor_actual = stamp_1.anchor

                    from_col, from_col_off = anchor_actual._from.col, anchor_actual._from.colOff
                    from_row, from_row_off = anchor_actual._from.row, anchor_actual._from.rowOff
                    to_col, to_col_off = anchor_actual.to.col, anchor_actual.to.colOff
                    to_row, to_row_off = anchor_actual.to.row, anchor_actual.to.rowOff

                    from_cell = AnchorMarker(col=from_col - 1, colOff=from_col_off, row=from_row, rowOff=from_row_off)
                    to_cell = AnchorMarker(col=to_col - 1, colOff=to_col_off, row=to_row, rowOff=to_row_off)

                    new_anchor = TwoCellAnchor(_from = from_cell, to = to_cell, editAs='absolute')
                    stamp_1.anchor = new_anchor

                    stamp_2 = self.wb_technical[eq_type]._images[2]
                    anchor_actual = stamp_2.anchor

                    from_col, from_col_off = anchor_actual._from.col, anchor_actual._from.colOff
                    from_row, from_row_off = anchor_actual._from.row, anchor_actual._from.rowOff
                    to_col, to_col_off = anchor_actual.to.col, anchor_actual.to.colOff
                    to_row, to_row_off = anchor_actual.to.row, anchor_actual.to.rowOff

                    from_cell = AnchorMarker(col=from_col - 1, colOff=from_col_off, row=from_row, rowOff=from_row_off)
                    to_cell = AnchorMarker(col=to_col - 1, colOff=to_col_off, row=to_row, rowOff=to_row_off)

                    new_anchor = TwoCellAnchor(_from = from_cell, to = to_cell, editAs='absolute')
                    stamp_2.anchor = new_anchor

                if int(rev) == 0:
                    ws = self.wb_technical[self.wb_technical.sheetnames[-2]]
                    ws.cell(row=last_row + 3, column=num_column_amount).value = number_items
                    ws.cell(row=last_row + 3, column=num_column_amount).font = Font(name="Calibri", size=14)

            # Editing sheet NOTES
                if int(rev) == 0:
                    sheet_name = "NOTES"
                    ws = self.wb_technical[sheet_name]

                    rich_string = CellRichText(
                    'We are only offering measuring glass and magnetic level indicators. Please be informed that our product range includes flow elements and temperature elements; all with european certification. (https://www.eipsa.es/en/products)\n',
                    TextBlock(InlineFont(b=True), 'The prices quoted could be reduced in case of purchasing our full range of products.'))
                    ws["B6"] = rich_string

                    rich_string = CellRichText(
                    TextBlock(InlineFont(i=True), 'Estamos ofertando solamente indicadores de nivel de vidrio y magnéticos, les informamos que en nuestra gama de fabricación con certificación europea, incluye también elementos de caudal y elementos de medida de temperatura. (https://www.eipsa.es/productos)\n'),
                    TextBlock(InlineFont(b=True, i=True), 'Los precios ofertados podrían reducirse en caso de compra de toda nuestra gama.'))
                    ws["B7"] = rich_string

                    rich_string = CellRichText(
                    'Delivery time ' + delivery_time + ' weeks since drawing / calculation approval of ',
                    TextBlock(InlineFont(b=True), 'all equipment'),
                    ' in the contract, as well as critical documentation (August and last two December weeks excluded).')
                    ws["B12"] = rich_string

                    rich_string = CellRichText(
                    'Plazo de entrega ' + delivery_time + ' semanas desde aprobación de planos y cálculos de la ',
                    TextBlock(InlineFont(b=True, i=True), 'totalidad de los equipos'),
                    TextBlock(InlineFont(i=True),' amparados por el contrato, asi como la documentación crítica (Agosto y las dos últimas semanas de diciembre excluidos).'))
                    ws["B13"] = rich_string

                    rich_string = CellRichText(
                    'Quotation prepared according to the information provided in the datasheet corresponding to each TAG. ',
                    TextBlock(InlineFont(u='single'), 'EIPSA does not hold the final responsibility regarding selection of equipment material neither analyze process data.\n'),
                    'The datasheet will be considered the only technical/contractual document, any other documentation will be considered as complementary documentation with informative purpose.')
                    ws["B21"] = rich_string

                    rich_string = CellRichText(
                    'Esta oferta ha sido elaborada en base a la información que figura en la hoja de datos correspondiente a cada TAG, ',
                    TextBlock(InlineFont(i=True, u='single'), 'no siendo responsabilidad final de EIPSA la elección del material a utilizar en los equipos ni analizar datos de proceso.\n'),
                    TextBlock(InlineFont(i=True),'La hoja de datos será el único documento técnico/contractual, cualquier otra documentación recibida será considerada como documentación complementaria a efectos informativos.'))
                    ws["B22"] = rich_string

                    if pay_term == "100_delivery":
                        ws["B41"] = (
                            "100% of total amount of purchase order upon delivery of material according to Incoterms 2020, FCA (our facilities, Spain).\n"
                            "Payment method: bank transfer"
                        )
                        ws["B42"] = (
                            "Pago del 100% del valor total de la orden de compra a la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España).\n"
                            "Método de pago: Transferencia bancaria."
                        )
                    elif pay_term == "100_order":
                        ws["B41"] = (
                            "100 % of the total amount of purchase order upon receipt of purchase order.\n"
                            "Payment method: bank transfer"
                        )
                        ws["B42"] = (
                            "Pago del 100% del valor total de la orden de compra a la recepción de la orden.\n"
                            "Método de pago: Transferencia bancaria"
                        )
                    elif pay_term == "90_10":
                        ws["B41"] = (
                            "PAYMENT TERMS:\n"
                            "90 % of the total amount of PO upon delivery of material according to Incoterms 2020, FCA (our facilities, Spain) and 10% when final documentation is approved. \n"
                            "Bank Transfer: 60 days since invoice issue date."
                        )
                        ws["B42"] = (
                            "TERMINOS DE PAGO:\n"
                            "Pago del 90% del Valor total de la orden de compra a la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España) y el 10% restante cuando la documentación final sea aprobada.\n"
                            "Transferencia Bancaria: 60 días desde emisión de factura."
                        )
                    elif pay_term == "50_50":
                        ws["B41"] = (
                            "50 % of the total amount of purchase order upon receipt of purchase order. Remaining 50% before material be delivered according to Incoterms 2020, FCA (our facilities, Spain).\n"
                            "Payment method: bank transfer."
                        )
                        ws["B42"] = (
                            "Pago del 50% del valor total de la orden de compra a la recepción de la orden. El 50% restante antes de la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España).\n"
                            "Método de pago: Transferencia bancaria."
                        )
                    elif pay_term == "Others":
                        ws["B41"] = "PAYMENT TERMS TO BE DEFINED"
                        ws["B41"].font = Font(name="Calibri", size=11, bold=True, color="FF0000")
                        ws["B42"] = "TERMINOS DE PAGO POR DEFINIR"
                        ws["B42"].font = Font(name="Calibri", size=11, bold=True, italic=True, color="FF0000")

                    rich_string = CellRichText(
                    'For amounts greater than 30,000.00 € we can issue a warranty bond (if required) valid until the end of the indicated warranty period.\nBond warranty of 10% will be issued with the invoice of the last supplement.\n',
                    TextBlock(InlineFont(b=True), 'For lower amounts no warranty bond is issued.'))
                    ws["B38"] = rich_string

                    rich_string = CellRichText(
                    'Para importes superiores a 30.000,00 €, si es requerido, podremos emitir aval de garantía y estará vigente hasta el final del periodo de garantía indicado.\nEl aval del 10% será emitido con la factura del último suplemento.\n',
                    TextBlock(InlineFont(i=True, b=True), 'Por debajo de dicha cantidad, no se emitirán avales.'))
                    ws["B39"] = rich_string

                    ws["A51"] = (
                        "If you require further information related with this offer, please do not hesitate to contact:\n"
                        + responsible
                        + "\n"
                        + email
                        + "\n"
                        "Telf.: (+34) 916.582.118"
                    )

                    std = self.wb_technical["1.3"]
                    self.wb_technical.remove(std)

                left_text = "Fecha/Date: " + date_offer
                right_text = "Petición nº/Inquiry: " + num_ref

                for sheet in self.wb_technical.worksheets:
                    sheet.oddFooter.left.text = left_text
                    sheet.oddFooter.right.text = right_text
                    sheet.oddFooter.center.text = "Page &P de &N"

                    sheet.oddFooter.left.size = 9
                    sheet.oddFooter.right.size = 9
                    sheet.oddFooter.center.size = 9

                self.save_excel_technical()

                root = Tk()
                root.withdraw()  # Hiding main window Tkinter

                # close communication with the PostgreSQL database server
                # commit the changes
                conn.commit()
        except (Exception, psycopg2.DatabaseError) as error:
            dlg = QtWidgets.QMessageBox()
            new_icon = QtGui.QIcon()
            new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
            dlg.setWindowIcon(new_icon)
            dlg.setWindowTitle("ERP EIPSA")
            dlg.setText("Ha ocurrido el siguiente error:\n"
                        + str(error))
            dlg.setIcon(QtWidgets.QMessageBox.Icon.Critical)
            dlg.exec()
            del dlg, new_icon
        finally:
            if conn is not None:
                conn.close()

    def euros_to_float(self, value):
        """
        Converts a string value representing an amount in euros to a float.
        
        Args:
            value (str): The string representation of an amount in euros, with commas for decimal separation and ' €' for currency indication.
        
        Returns:
            float: The numeric value of the amount in euros.
        """
        """
        Converts a string value representing an amount in euros to a float.
        
        Args:
            value (str): The string representation of an amount in euros, with commas for decimal separation and ' €' for currency indication.
        
        Returns:
            float: The numeric value of the amount in euros.
        """
        value = value.replace(".", "")
        value = value.replace(",", ".")
        value = value[: value.find(" €")]
        return float(value)

    def save_excel_commercial(self):
        """Saves the populated Excel workbook to a specified location.
        Opens a dialog window for the user to select the file path and name.
        """
        # Dialog window to select folder and file name; if path is selected, excel file is saved
        output_path_commercial = asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Archivos de Excel", "*.xlsx")],
            title="Guardar Oferta comercial",
        )
        if output_path_commercial:
            self.wb_commercial.save(output_path_commercial)
            return output_path_commercial

    def save_excel_technical(self):
        """Saves the populated Excel workbook to a specified location.
        Opens a dialog window for the user to select the file path and name.
        """
        output_path_technical = asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Archivos de Excel", "*.xlsx")],
            title="Guardar Oferta técnica",
        )
        if output_path_technical:
            self.wb_technical.save(output_path_technical)

class offer_flow_temp:
    """
    A class to manage export offer details for flow equipments.
    
    Attributes:
        numoffer (str): Offer number.
        username (str): Name of the user creating the offer.
        rev (str): Revision number of the offer.
        project (str): Name of the project.
        delivery_term (str): Delivery terms for the offer.
        delivery_time (str): Expected delivery time.
        validity (str): Validity period of the offer.
        pay_term (str): Payment terms.
        testinspection (str): Information about testing and inspection.
        revchanges (str): Details of changes made in the revision.
        notes (str): Additional notes, split by line.
    """
    def __init__(self, numoffer, username, rev, project, delivery_term, delivery_time, validity, pay_term, testinspection, revchanges, notes):
        """
        Initializes the offer.

        Args:
            numoffer (str): Offer number.
            username (str): Name of the user creating the offer.
            rev (str): Revision number of the offer.
            project (str): Name of the project.
            delivery_term (str): Delivery terms for the offer.
            delivery_time (str): Expected delivery time.
            validity (str): Validity period of the offer.
            pay_term (str): Payment terms.
            testinspection (str): Information about testing and inspection.
            revchanges (str): Details of changes made in the revision.
            notes (str): Additional notes, split by line.
        """
        notes = notes.split('\n')
        
        date_offer = date.today().strftime("%d/%m/%Y")
        offername_commercial = numoffer + "-" + "Commercial.Rev" + rev
        offername_technical = numoffer + "-" + "Technical.Rev" + rev

        query_commercial = """
                    SELECT name, surname, email
                    FROM users_data.registration
                    WHERE username = %s
                    """
        query_dataoffer = """
                        SELECT client, num_ref_offer
                        FROM offers
                        WHERE UPPER (num_offer) LIKE UPPER('%%'||%s||'%%')
                        """
        query_tagsdata_flow = """
                        SELECT *
                        FROM tags_data.tags_flow
                        WHERE (
                        UPPER ("num_offer") LIKE UPPER('%%'||%s||'%%')
                        AND
                        "tag_state" NOT IN ('PURCHASED','DELETED')
                        )
                        """
        query_tagsdata_temp = """
                        SELECT *
                        FROM tags_data.tags_temp
                        WHERE (
                        UPPER ("num_offer") LIKE UPPER('%%'||%s||'%%')
                        AND
                        "tag_state" NOT IN ('PURCHASED','DELETED')
                        )
                        """
        conn = None
        try:
            # read the connection parameters
            params = config()
            # connect to the PostgreSQL server
            conn = psycopg2.connect(**params)
            cur = conn.cursor()

            cur.execute(query_dataoffer, (numoffer,))
            results_offer = cur.fetchall()
            client = results_offer[0][0]
            num_ref = results_offer[0][1]

            cur.execute(query_commercial, (username,))
            results_commercial = cur.fetchall()
            responsible = results_commercial[0][0] + " " + results_commercial[0][1]
            email = results_commercial[0][2]

            cur.execute(query_tagsdata_flow, (numoffer,))
            data_tags_flow = cur.fetchall()

            columns_flow = []
            for elt in cur.description:
                columns_flow.append(elt[0])

            cur.execute(query_tagsdata_temp, (numoffer,))
            data_tags_temp = cur.fetchall()

            columns_temp = []
            for elt in cur.description:
                columns_temp.append(elt[0])

            if len(data_tags_flow) == 0 or len(data_tags_temp) == 0:
                dlg = QtWidgets.QMessageBox()
                new_icon = QtGui.QIcon()
                new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
                dlg.setWindowIcon(new_icon)
                dlg.setWindowTitle("ERP EIPSA")
                dlg.setText("No hay TAGS importados en la oferta")
                dlg.setIcon(QtWidgets.QMessageBox.Icon.Warning)
                dlg.exec()
                del dlg, new_icon
            else:
                value_type_dict_flow = {
                    "A. Chamber": 1,
                    "C. RING": 1,
                    "F": 1,
                    "F+C.RING": 1,
                    "F+P": 1,
                    "IFO": 1,
                    "M.RUN": 1,
                    "P": 1,
                    "NOZZLE BF": 2,
                    "NOZZLE BW": 2,
                    "NOZZLE F": 2,
                    "PTC-6": 2,
                    "VFM": 3,
                    "VFW": 3,
                    "VWM": 3,
                    "VWW": 3,
                    "WEDGE": 4,
                    "PITOT": 5,
                    "RO": 6,
                    "MULTISTAGE RO": 7,
                }

                # Setting the dataframe with the equipment data
                df_flow = pd.DataFrame(data=data_tags_flow, columns=columns_flow)
                df_flow = df_flow.iloc[:, 1:32]
                df_flow["value_type"] = df_flow["item_type"].map(value_type_dict_flow)
                df_flow = df_flow.sort_values(by=["value_type", "tag"])
                df_flow["amount"] = df_flow["amount"].apply(self.euros_to_float)
                total_amount_material = df_flow["amount"].sum()
                df_flow = df_flow.drop([
                        "tag_state",
                        "num_offer",
                        "num_order",
                        "num_po",
                        "position",
                        "subposition",
                        "flange_type",
                        "plate_std",
                        "pipe_spec",
                        "aprox_weight"
                    ],
                    axis=1,)
                
                value_type_dict_temp = {
                    "TW": 1,
                    "TW+TE": 2,
                    "TW+TE+TIT": 2,
                    "RETAINING FLANGE":2,
                    "TW+BIM": 3,
                    "TE": 4,
                    "BIM": 5,
                    "TIT": 6,
                    "SKIN+TT": 7,
                    "SKIN POINT": 7,
                    "Multi-T": 8
                }

                df_temp = pd.DataFrame(data=data_tags_temp, columns=columns_temp)
                df_temp["value_type"] = df_temp["item_type"].map(value_type_dict_temp)
                df_temp = df_temp.sort_values(by=["tag", "value_type"])
                df_temp = df_temp.iloc[:, 1:37]
                df_temp["value_type"] = df_temp["item_type"].map(value_type_dict_temp)
                df_temp = df_temp.sort_values(by=["value_type", "tag"])
                df_temp["amount"] = df_temp["amount"].apply(self.euros_to_float)
                total_amount_material = df_temp["amount"].sum()
                df_temp = df_temp.drop([
                        "tag_state",
                        "num_offer",
                        "num_order",
                        "num_po",
                        "position",
                        "subposition",
                        "std_tw",
                        "insulation"
                    ],
                    axis=1,)

                number_items = df_flow.shape[0] + df_temp.shape[0]
                documentation = number_items * 70

                # Loading Excel Template
                self.wb_commercial = load_workbook(
                    r"\\nas01\DATOS\Comunes\EIPSA-ERP\Plantillas Exportación\PLANTILLA OFERTA CAUDAL-TEMPERATURA.xlsx"
                )

                # Editing sheet COVER
                sheet_name = "COVER"
                ws = self.wb_commercial[sheet_name]
                ws["E4"] = client
                ws["E6"] = offername_commercial
                ws["E8"] = num_ref
                ws["E10"] = project
                ws["E12"] = date_offer
                ws["E14"] = delivery_term
                ws["E16"] = validity + " days"
                ws["C43"] = responsible
                ws["C45"] = email

                # Editing sheet EQUIPMENT DATA
                sheets_confirmed = ["COVER", "1.2", "1.3", "NOTES"]

                for item_type in df_flow["item_type"].unique():
                    sheets_confirmed.append(item_type)

                for item_type in df_temp["item_type"].unique():
                    sheets_confirmed.append(item_type)

                dict_sheets_data = {}

                for value_type in df_flow["value_type"].unique():
                    df_toexport = df_flow[df_flow["value_type"] == value_type]
                    df_toexport.insert(0, "N°", range(1, len(df_toexport) + 1))
                    df_toexport.set_index("N°", inplace=True, drop=False)
                    df_toexport.index.name = None
                    df_toexport = df_toexport.drop(["value_type"], axis=1)

                    eq_type = (
                        "FLOW ELEMENTS DATA"
                        if value_type == 1
                        else (
                            "NOZZLE ELEMENTS DATA"
                            if value_type == 2
                            else (
                                "VENTURI ELEMENTS DATA"
                                if value_type == 3
                                else (
                                    "WEDGE ELEMENTS DATA"
                                    if value_type == 4
                                    else (
                                        "PITOT ELEMENTS DATA"
                                        if value_type == 5
                                        else (
                                            "RO ELEMENTS DATA"
                                            if value_type == 6
                                            else "MULTISTAGE RO ELEMENTS DATA"
                                        )
                                    )
                                )
                            )
                        )
                    )

                    if eq_type not in sheets_confirmed:
                        sheets_confirmed.append(eq_type)

                    if eq_type == "FLOW ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "tube_material",
                                "plate_type",
                                "valve_conn",
                                "valve_material_body",
                                "stages_number",
                                "aprox_length"],
                            axis=1,)
                    elif eq_type == "NOZZLE ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "plate_type",
                                "plate_thk",
                                "gasket_material",
                                "valve_conn",
                                "valve_material_body",
                                "stages_number"],
                            axis=1,)
                    elif (eq_type == "VENTURI ELEMENTS DATA" or eq_type == "WEDGE ELEMENTS DATA"):
                        df_toexport = df_toexport.drop([
                                "plate_type",
                                "plate_thk",
                                "gasket_material",
                                "bolts_nuts_material",
                                "valve_conn",
                                "valve_material_body",
                                "stages_number"],
                            axis=1,)
                    elif eq_type == "PITOT ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "flange_material",
                                "tube_material",
                                "tapping_num_size",
                                "plate_type",
                                "plate_thk",
                                "gasket_material",
                                "stages_number",
                                "aprox_length"],
                            axis=1,)
                    elif eq_type == "RO ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "flange_material",
                                "tube_material",
                                "tapping_num_size",
                                "gasket_material",
                                "bolts_nuts_material",
                                "valve_conn",
                                "valve_material_body",
                                "stages_number",
                                "aprox_length"],
                            axis=1,)
                    elif eq_type == "MULTISTAGE RO ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "tapping_num_size",
                                "gasket_material",
                                "bolts_nuts_material",
                                "valve_conn",
                                "valve_material_body"],
                            axis=1,)

                    ws = self.wb_commercial[eq_type]
                    ws["G2"] = date_offer
                    ws["G3"] = num_ref
                    ws["G4"] = offername_commercial
                    if revchanges != "":
                        ws["G5"] = rev + " " + revchanges
                        ws["G5"].font = Font(name="Calibri", size=14, bold=True)
                        ws["G5"].fill = PatternFill("solid", fgColor="FFFF00")

                    if int(rev) > 0:
                        for row in ws.iter_rows(min_row=2, max_row=4, min_col=6, max_col=7):
                            for cell in row:
                                cell.value = None
                                cell._style = ws["F1"]._style

                    last_row = ws.max_row

                    # for col_num, col_name in enumerate(df_toexport.columns, start=1):
                    #     cell = ws.cell(row=last_row + 1, column=col_num)
                    #     cell.value = col_name
                    #     cell._style = ws["Y1"]._style

                    # last_row = ws.max_row

                    num_column_amount = df_toexport.columns.get_loc("amount") + 1

                    for index, row in df_toexport.iterrows():  # Data in desired row
                        for col_num, value in enumerate(row, start=1):
                            cell = ws.cell(row=last_row + 1, column=col_num)
                            cell.value = value
                            if col_num == num_column_amount:
                                cell._style = ws["X1"]._style
                            else:
                                cell._style = ws["S1"]._style

                        last_row = ws.max_row

                    if eq_type == "VENTURI ELEMENTS DATA":
                        ws[f"A{last_row+3}"] = "PRICES INCLUDE MACHINED INTEGRAL CENTRE SECTION AND ALL STRUCTURAL WELDS 100% RADIOGRAPHED"
                        ws[f"A{last_row+3}"]._style = ws["Z2"]._style
                    ws[f"A{last_row+4}"] = "OFFER VALIDITY: " + validity + " DAYS"
                    ws[f"A{last_row+4}"]._style = ws["Z1"]._style
                    ws[f"A{last_row+5}"] = (
                        "DELIVERY TIME: "
                        + delivery_time
                        + " WEEKS SINCE DRAWING / CALCULATION APPROVAL (AUGUST AND LAST TWO DECEMBER WEEKS EXCLUDED)"
                    )
                    ws[f"A{last_row+5}"]._style = ws["Z1"]._style

                    if notes != "":
                        if isinstance(notes, list):
                            line = last_row + 6
                            for note in notes:
                                ws[f"A{line}"] = note
                                ws[f"A{line}"]._style = ws["Z1"]._style
                                line += 1
                        else:
                            line = last_row + 6
                            ws[f"A{line}"] = notes
                            ws[f"A{line}"]._style = ws["Z1"]._style


                    dict_sheets_data[eq_type] = [last_row, num_column_amount, df_toexport["amount"].sum(), df_toexport.shape[0]]

                for value_type in df_temp["value_type"].unique():
                    df_toexport = df_temp[df_temp["value_type"] == value_type]
                    df_toexport.insert(0, "N°", range(1, len(df_toexport) + 1))
                    df_toexport.set_index("N°", inplace=True, drop=False)
                    df_toexport.index.name = None
                    df_toexport = df_toexport.drop(["value_type"], axis=1)

                    eq_type = (
                        "TW ELEMENTS DATA"
                        if value_type == 1
                        else (
                            "TW+TE ELEMENTS DATA"
                            if value_type == 2
                            else (
                                "TW+BIM ELEMENTS DATA"
                                if value_type == 3
                                else (
                                    "TE ELEMENTS DATA"
                                    if value_type == 4
                                    else (
                                        "BIM ELEMENTS DATA"
                                        if value_type == 5
                                        else (
                                            "TIT ELEMENTS DATA"
                                            if value_type == 6
                                            else (
                                                "SKIN POINT ELEMENTS DATA"
                                                if value_type == 7
                                                else "MULTI-T ELEMENTS DATA"
                                            )
                                        )
                                    )
                                )
                            )
                        )
                    )

                    if eq_type not in sheets_confirmed:
                        sheets_confirmed.append(eq_type)

                    if eq_type == "TW ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "std_length",
                                "sensor_element",
                                "sheath_stem_material",
                                "sheath_stem_diam",
                                "temp_inf",
                                "temp_sup",
                                "nipple_ext_material",
                                "nipple_ext_length",
                                "head_case_material",
                                "elec_conn_case_diam",
                                "tt_cerblock"],
                            axis=1,)
                    elif eq_type == "TW+TE ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "std_length",
                                "temp_inf",
                                "temp_sup",
                                "nipple_ext_length"],
                            axis=1,)
                    elif eq_type == "TW+BIM ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "std_length",
                                "nipple_ext_length",
                                "tt_cerblock",
                                "puntal",
                                "tube_t"],
                            axis=1,)
                    elif eq_type == "TE ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "tw_type",
                                "size",
                                "rating",
                                "facing",
                                "material_tw",
                                "std_length",
                                "root_diam",
                                "tip_diam",
                                "temp_inf",
                                "temp_sup",
                                "nipple_ext_material",
                                "nipple_ext_length",
                                "head_case_material",
                                "elec_conn_case_diam",
                                "tt_cerblock",
                                "material_flange_lj",
                                "gasket_material",
                                "puntal",
                                "tube_t"],
                            axis=1,)
                    elif eq_type == "BIM ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "tw_type",
                                "size",
                                "rating",
                                "facing",
                                "material_tw",
                                "std_length",
                                "root_diam",
                                "tip_diam",
                                "nipple_ext_length",
                                "tt_cerblock",
                                "material_flange_lj",
                                "gasket_material",
                                "puntal",
                                "tube_t"],
                            axis=1,)
                    elif eq_type == "TIT ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "tw_type",
                                "size",
                                "rating",
                                "facing",
                                "material_tw",
                                "std_length",
                                "ins_length",
                                "root_diam",
                                "tip_diam",
                                "sensor_element",
                                "sheath_stem_material",
                                "sheath_stem_diam",
                                "nipple_ext_material",
                                "nipple_ext_length",
                                "head_case_material",
                                "material_flange_lj",
                                "gasket_material",
                                "puntal",
                                "tube_t"],
                            axis=1,)
                    elif (eq_type == "SKIN POINT ELEMENTS DATA" or eq_type == "SKIN+TT ELEMENTS DATA"):
                        df_toexport = df_toexport.drop([
                                "tw_type",
                                "size",
                                "rating",
                                "facing",
                                "material_tw",
                                "std_length",
                                "root_diam",
                                "tip_diam",
                                "temp_inf",
                                "temp_sup",
                                "material_flange_lj",
                                "gasket_material",
                                "puntal",
                                "tube_t"],
                            axis=1,)
                    elif eq_type == "MULTI-T ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "material_tw",
                                "root_diam",
                                "tip_diam",
                                "temp_inf",
                                "temp_sup",
                                "tt_cerblock",
                                "material_flange_lj",
                                "puntal",
                                "tube_t"],
                            axis=1,)

                    ws = self.wb_commercial[eq_type]
                    ws["G2"] = date_offer
                    ws["G3"] = num_ref
                    ws["G4"] = offername_commercial
                    if revchanges != "":
                        ws["G5"] = rev + " " + revchanges
                        ws["G5"].font = Font(name="Calibri", size=14, bold=True)
                        ws["G5"].fill = PatternFill("solid", fgColor="FFFF00")

                    if int(rev) > 0:
                        for row in ws.iter_rows(min_row=2, max_row=4, min_col=6, max_col=7):
                            for cell in row:
                                cell.value = None
                                cell._style = ws["F1"]._style

                    last_row = ws.max_row

                    num_column_amount = df_toexport.columns.get_loc("amount") + 1

                    for index, row in df_toexport.iterrows():  # Data in desired row
                        for col_num, value in enumerate(row, start=1):
                            cell = ws.cell(row=last_row + 1, column=col_num)
                            cell.value = value
                            if col_num == num_column_amount:
                                cell._style = ws["AG1"]._style
                            else:
                                cell._style = ws["AB1"]._style

                        last_row = ws.max_row

                    ws[f"A{last_row+3}"] = "OFFER VALIDITY: " + validity + " DAYS"
                    ws[f"A{last_row+3}"]._style = ws["AI1"]._style
                    ws[f"A{last_row+4}"] = (
                        "DELIVERY TIME: "
                        + delivery_time
                        + " WEEKS SINCE DRAWING / CALCULATION APPROVAL (AUGUST AND LAST TWO DECEMBER WEEKS EXCLUDED)"
                    )
                    ws[f"A{last_row+4}"]._style = ws["AI1"]._style

                    if notes != "":
                        if isinstance(notes, list):
                            line = last_row + 5
                            for note in notes:
                                ws[f"A{line}"] = note
                                ws[f"A{line}"]._style = ws["AI1"]._style
                                line += 1
                        else:
                            line = last_row + 5
                            ws[f"A{line}"] = notes
                            ws[f"A{line}"]._style = ws["AI1"]._style
                    dict_sheets_data[eq_type] = [last_row, num_column_amount, df_toexport["amount"].sum(), df_toexport.shape[0]]

                ws.cell(row=last_row + 3, column=num_column_amount - 1).value = "QTY. TOTAL"
                ws.cell(row=last_row + 3, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=last_row + 3, column=num_column_amount).value = number_items
                
                row_amount = last_row + 4
                for key, value in dict_sheets_data.items():
                    parts_key = key.split(" ")
                    ws.cell(row=row_amount + 2, column=num_column_amount - 1).value = "TOTAL AMOUNT OF " + parts_key[0] + " " + parts_key[1] + " (QTY: " + str(value[3]) + ")"
                    ws.cell(row=row_amount + 2, column=num_column_amount).value = value[2]
                    ws.cell(row=row_amount + 2, column=num_column_amount - 1)._style = ws["R1"]._style
                    ws.cell(row=row_amount + 2, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                    ws.cell(row=row_amount + 2, column=num_column_amount)._style = ws["T1"]._style

                    row_amount += 2

                ws.cell(row=row_amount + 2, column=num_column_amount - 1).value = "TOTAL AMOUNT OF MATERIAL"
                ws.cell(row=row_amount + 2, column=num_column_amount).value = total_amount_material
                ws.cell(row=row_amount + 4, column=num_column_amount - 1).value = "PACKING AND TRANSPORT"
                ws.cell(row=row_amount + 4, column=num_column_amount).value = (f"=MROUND({get_column_letter(num_column_amount)}{row_amount + 2}*0.03,10)" if total_amount_material > 6700 else 200 )
                ws.cell(row=row_amount + 5, column=num_column_amount - 1).value = "TESTS & INSPECTION"
                ws.cell(row=row_amount + 5, column=num_column_amount).value = float(testinspection)
                ws.cell(row=row_amount + 6, column=num_column_amount - 1).value = "DOCUMENTATION"
                ws.cell(row=row_amount + 6, column=num_column_amount).value = documentation
                ws.cell(row=row_amount + 8, column=num_column_amount - 1).value = "TOTAL AMOUNT OF BID"
                ws.cell(row=row_amount + 8, column=num_column_amount).value = f"=SUM({get_column_letter(num_column_amount)}{row_amount + 2}:{get_column_letter(num_column_amount)}{row_amount + 6})"

                ws.cell(row=last_row + 3, column=num_column_amount - 1)._style = ws["R1"]._style
                ws.cell(row=last_row + 3, column=num_column_amount).font = Font(name="Calibri", size=14)
                ws.cell(row=row_amount + 2, column=num_column_amount - 1)._style = ws["R1"]._style
                ws.cell(row=row_amount + 2, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=row_amount + 2, column=num_column_amount)._style = ws["T1"]._style
                ws.cell(row=row_amount + 4, column=num_column_amount - 1).font = Font(name="Calibri", size=14)
                ws.cell(row=row_amount + 4, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=row_amount + 4, column=num_column_amount)._style = ws["T1"]._style
                ws.cell(row=row_amount + 5, column=num_column_amount - 1).font = Font(name="Calibri", size=14)
                ws.cell(row=row_amount + 5, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=row_amount + 5, column=num_column_amount)._style = ws["T1"]._style
                ws.cell(row=row_amount + 6, column=num_column_amount - 2)._style = ws["U1"]._style
                ws.cell(row=row_amount + 6, column=num_column_amount - 1)._style = ws["U1"]._style
                ws.cell(row=row_amount + 6, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=row_amount + 6, column=num_column_amount)._style = ws["V1"]._style
                ws.cell(row=row_amount + 8, column=num_column_amount - 1)._style = ws["R1"]._style
                ws.cell(row=row_amount + 8, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=row_amount + 8, column=num_column_amount)._style = ws["W1"]._style

            # Editing sheet NOTES
                sheet_name = "NOTES"  # Selecting  sheet
                ws = self.wb_commercial[sheet_name]

                rich_string = CellRichText(
                'We are only offering measuring flow elements. Please be informed that our product range includes temperature elements, and glass and magnetic level indicators; all with european certification. (https://www.eipsa.es/en/products)\n',
                TextBlock(InlineFont(b=True), 'The prices quoted could be reduced in case of purchasing our full range of products.'))
                ws["B6"] = rich_string

                rich_string = CellRichText(
                TextBlock(InlineFont(i=True), 'Estamos ofertando solamente elementos de medida de caudal, les informamos que en nuestra gama de fabricación con certificación europea, incluye también elementos de temperatura e indicadores de nivel de vidrio y magnéticos. (https://www.eipsa.es/productos)\n'),
                TextBlock(InlineFont(b=True, i=True), 'Los precios ofertados podrían reducirse en caso de compra de toda nuestra gama.'))
                ws["B7"] = rich_string

                rich_string = CellRichText(
                'Delivery time ' + delivery_time + ' weeks since drawing / calculation approval of ',
                TextBlock(InlineFont(b=True), 'all equipment'),
                ' in the contract, as well as critical documentation (August and last two December weeks excluded).')
                ws["B12"] = rich_string

                rich_string = CellRichText(
                'Plazo de entrega ' + delivery_time + ' semanas desde aprobación de planos y cálculos de la ',
                TextBlock(InlineFont(b=True, i=True), 'totalidad de los equipos'),
                TextBlock(InlineFont(i=True),' amparados por el contrato, asi como la documentación crítica (Agosto y las dos últimas semanas de diciembre excluidos).'))
                ws["B13"] = rich_string

                rich_string = CellRichText(
                'Quotation prepared according to the information provided in the datasheet corresponding to each TAG. ',
                TextBlock(InlineFont(u='single'), 'EIPSA does not hold the final responsibility regarding selection of equipment material neither analyze process data.\n'),
                'The datasheet will be considered the only technical/contractual document, any other documentation will be considered as complementary documentation with informative purpose.')
                ws["B21"] = rich_string

                rich_string = CellRichText(
                'Esta oferta ha sido elaborada en base a la información que figura en la hoja de datos correspondiente a cada TAG, ',
                TextBlock(InlineFont(i=True, u='single'), 'no siendo responsabilidad final de EIPSA la elección del material a utilizar en los equipos ni analizar datos de proceso.\n'),
                TextBlock(InlineFont(i=True),'La hoja de datos será el único documento técnico/contractual, cualquier otra documentación recibida será considerada como documentación complementaria a efectos informativos.'))
                ws["B22"] = rich_string

                if pay_term == "100_delivery":
                    ws["B45"] = (
                        "100% of total amount of purchase order upon delivery of material according to Incoterms 2020, FCA (our facilities, Spain).\n"
                        "Payment method: bank transfer"
                    )
                    ws["B46"] = (
                        "Pago del 100% del valor total de la orden de compra a la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España).\n"
                        "Método de pago: Transferencia bancaria."
                    )
                elif pay_term == "100_order":
                    ws["B45"] = (
                        "100 % of the total amount of purchase order upon receipt of purchase order.\n"
                        "Payment method: bank transfer"
                    )
                    ws["B46"] = (
                        "Pago del 100% del valor total de la orden de compra a la recepción de la orden.\n"
                        "Método de pago: Transferencia bancaria"
                    )
                elif pay_term == "90_10":
                    ws["B45"] = (
                        "PAYMENT TERMS:\n"
                        "90 % of the total amount of PO upon delivery of material according to Incoterms 2020, FCA (our facilities, Spain) and 10% when final documentation is approved. \n"
                        "Bank Transfer: 60 days since invoice issue date."
                    )
                    ws["B46"] = (
                        "TERMINOS DE PAGO:\n"
                        "Pago del 90% del Valor total de la orden de compra a la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España) y el 10% restante cuando la documentación final sea aprobada.\n"
                        "Transferencia Bancaria: 60 días desde emisión de factura."
                    )
                elif pay_term == "50_50":
                    ws["B45"] = (
                        "50 % of the total amount of purchase order upon receipt of purchase order. Remaining 50% before material be delivered according to Incoterms 2020, FCA (our facilities, Spain).\n"
                        "Payment method: bank transfer."
                    )
                    ws["B46"] = (
                        "Pago del 50% del valor total de la orden de compra a la recepción de la orden. El 50% restante antes de la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España).\n"
                        "Método de pago: Transferencia bancaria."
                    )
                elif pay_term == "Others":
                    ws["B45"] = "PAYMENT TERMS TO BE DEFINED"
                    ws["B45"].font = Font(name="Calibri", size=11, bold=True, color="FF0000")
                    ws["B46"] = "TERMINOS DE PAGO POR DEFINIR"
                    ws["B46"].font = Font(name="Calibri", size=11, bold=True, italic=True, color="FF0000")

                rich_string = CellRichText(
                'For amounts greater than 30,000.00 € we can issue a warranty bond (if required) valid until the end of the indicated warranty period.\nBond warranty of 10% will be issued with the invoice of the last supplement.\n',
                TextBlock(InlineFont(b=True), 'For lower amounts no warranty bond is issued.'))
                ws["B48"] = rich_string

                rich_string = CellRichText(
                'Para importes superiores a 30.000,00 €, si es requerido, podremos emitir aval de garantía y estará vigente hasta el final del periodo de garantía indicado.\nEl aval del 10% será emitido con la factura del último suplemento.\n',
                TextBlock(InlineFont(i=True, b=True), 'Por debajo de dicha cantidad, no se emitirán avales.'))
                ws["B49"] = rich_string

                ws["A60"] = (
                    "If you require further information related with this offer, please do not hesitate to contact:\n"
                    + responsible
                    + "\n"
                    + email
                    + "\n"
                    "Telf.: (+34) 916.582.118"
                )

                for sheet in self.wb_commercial.sheetnames:
                    if sheet not in sheets_confirmed:
                        sheet_to_delete = self.wb_commercial[sheet]
                        self.wb_commercial.remove(sheet_to_delete)

                if int(rev) > 0:
                    sheets_confirmed = ["COVER", "1.2", "1.3", "NOTES"]
                    for sheet in sheets_confirmed:
                        sheet_to_delete = self.wb_commercial[sheet]
                        self.wb_commercial.remove(sheet_to_delete)

                left_text = "Fecha/Date: " + date_offer
                right_text = "Petición nº/Inquiry: " + num_ref

                for sheet in self.wb_commercial.worksheets:
                    sheet.oddFooter.left.text = left_text
                    sheet.oddFooter.right.text = right_text
                    sheet.oddFooter.center.text = "Page &P de &N"

                    sheet.oddFooter.left.size = 9
                    sheet.oddFooter.right.size = 9
                    sheet.oddFooter.center.size = 9
                
                path = self.save_excel_commercial()

                # Creating the technical offer using the commercial one as template
                self.wb_technical = load_workbook(path)

                if int(rev) == 0:
                    sheet_name = "COVER"
                    ws = self.wb_technical[sheet_name]
                    ws["E6"] = offername_technical

                for value_type in df_flow["value_type"].unique():
                    eq_type = (
                        "FLOW ELEMENTS DATA"
                        if value_type == 1
                        else (
                            "NOZZLE ELEMENTS DATA"
                            if value_type == 2
                            else (
                                "VENTURI ELEMENTS DATA"
                                if value_type == 3
                                else (
                                    "WEDGE ELEMENTS DATA"
                                    if value_type == 4
                                    else (
                                        "PITOT ELEMENTS DATA"
                                        if value_type == 5
                                        else (
                                            "RO ELEMENTS DATA"
                                            if value_type == 6
                                            else "MULTISTAGE RO ELEMENTS DATA"
                                        )
                                    )
                                )
                            )
                        )
                    )

                    ws = self.wb_technical[eq_type]
                    if int(rev) == 0:
                        ws["G4"] = offername_technical

                    last_row = dict_sheets_data[eq_type][0]
                    num_column_amount = dict_sheets_data[eq_type][1]

                    # self.wb_technical[eq_type].delete_rows(last_row + 8, 20)
                    ws[f"M{last_row+5}"] = ""
                    ws[f"N{last_row+5}"] = ""

                    ws.cell(row=row_amount + 2, column=num_column_amount).value = ""
                    ws.cell(row=row_amount + 4, column=num_column_amount).value = ""
                    ws.cell(row=row_amount + 5, column=num_column_amount).value = ""
                    ws.cell(row=row_amount + 6, column=num_column_amount).value = ""
                    ws.cell(row=row_amount + 8, column=num_column_amount).value = ""

                    ws.cell(row=row_amount + 2, column=num_column_amount - 1).value = ""
                    ws.cell(row=row_amount + 4, column=num_column_amount - 1).value = ""
                    ws.cell(row=row_amount + 5, column=num_column_amount - 1).value = ""
                    ws.cell(row=row_amount + 6, column=num_column_amount - 1).value = ""
                    ws.cell(row=row_amount + 8, column=num_column_amount - 1).value = ""

                # Deleting "Amount" column
                    self.wb_technical[eq_type].delete_cols(num_column_amount, 1)

                # Adjusting the print area
                    new_last_column = num_column_amount
                    last_print_row = 40
                    nuevo_rango_impresion = f'A1:{get_column_letter(new_last_column)}{last_print_row}'
                    self.wb_technical[eq_type].print_area = nuevo_rango_impresion

                    stamp_1 = self.wb_technical[eq_type]._images[1]
                    anchor_actual = stamp_1.anchor

                    from_col, from_col_off = anchor_actual._from.col, anchor_actual._from.colOff
                    from_row, from_row_off = anchor_actual._from.row, anchor_actual._from.rowOff
                    to_col, to_col_off = anchor_actual.to.col, anchor_actual.to.colOff
                    to_row, to_row_off = anchor_actual.to.row, anchor_actual.to.rowOff

                    from_cell = AnchorMarker(col=from_col - 1, colOff=from_col_off, row=from_row, rowOff=from_row_off)
                    to_cell = AnchorMarker(col=to_col - 1, colOff=to_col_off, row=to_row, rowOff=to_row_off)

                    new_anchor = TwoCellAnchor(_from = from_cell, to = to_cell, editAs='absolute')
                    stamp_1.anchor = new_anchor

                    stamp_2 = self.wb_technical[eq_type]._images[2]
                    anchor_actual = stamp_2.anchor

                    from_col, from_col_off = anchor_actual._from.col, anchor_actual._from.colOff
                    from_row, from_row_off = anchor_actual._from.row, anchor_actual._from.rowOff
                    to_col, to_col_off = anchor_actual.to.col, anchor_actual.to.colOff
                    to_row, to_row_off = anchor_actual.to.row, anchor_actual.to.rowOff

                    from_cell = AnchorMarker(col=from_col - 1, colOff=from_col_off, row=from_row, rowOff=from_row_off)
                    to_cell = AnchorMarker(col=to_col - 1, colOff=to_col_off, row=to_row, rowOff=to_row_off)

                    new_anchor = TwoCellAnchor(_from = from_cell, to = to_cell, editAs='absolute')
                    stamp_2.anchor = new_anchor

                for value_type in df_temp["value_type"].unique():
                    eq_type = (
                        "TW ELEMENTS DATA"
                        if value_type == 1
                        else (
                            "TW+TE ELEMENTS DATA"
                            if value_type == 2
                            else (
                                "TW+BIM ELEMENTS DATA"
                                if value_type == 3
                                else (
                                    "TE ELEMENTS DATA"
                                    if value_type == 4
                                    else (
                                        "BIM ELEMENTS DATA"
                                        if value_type == 5
                                        else (
                                            "TIT ELEMENTS DATA"
                                            if value_type == 6
                                            else (
                                                "SKIN POINT ELEMENTS DATA"
                                                if value_type == 7
                                                else "MULTI-T ELEMENTS DATA"
                                            )
                                        )
                                    )
                                )
                            )
                        )
                    )

                    ws = self.wb_technical[eq_type]
                    if int(rev) == 0:
                        ws["G4"] = offername_technical

                    last_row = dict_sheets_data[eq_type][0]
                    num_column_amount = dict_sheets_data[eq_type][1]

                    # self.wb_technical[eq_type].delete_rows(last_row + 5, 20)

                    ws.cell(row=row_amount + 2, column=num_column_amount).value = ""
                    ws.cell(row=row_amount + 4, column=num_column_amount).value = ""
                    ws.cell(row=row_amount + 5, column=num_column_amount).value = ""
                    ws.cell(row=row_amount + 6, column=num_column_amount).value = ""
                    ws.cell(row=row_amount + 8, column=num_column_amount).value = ""

                    ws.cell(row=row_amount + 2, column=num_column_amount - 1).value = ""
                    ws.cell(row=row_amount + 4, column=num_column_amount - 1).value = ""
                    ws.cell(row=row_amount + 5, column=num_column_amount - 1).value = ""
                    ws.cell(row=row_amount + 6, column=num_column_amount - 1).value = ""
                    ws.cell(row=row_amount + 8, column=num_column_amount - 1).value = ""

                # Deleting "Amount" column
                    self.wb_technical[eq_type].delete_cols(num_column_amount, 1)

                # Adjusting the print area
                    new_last_column = num_column_amount
                    last_print_row = 40
                    nuevo_rango_impresion = f'A1:{get_column_letter(new_last_column)}{last_print_row}'
                    self.wb_technical[eq_type].print_area = nuevo_rango_impresion

                    stamp_1 = self.wb_technical[eq_type]._images[1]
                    anchor_actual = stamp_1.anchor

                    from_col, from_col_off = anchor_actual._from.col, anchor_actual._from.colOff
                    from_row, from_row_off = anchor_actual._from.row, anchor_actual._from.rowOff
                    to_col, to_col_off = anchor_actual.to.col, anchor_actual.to.colOff
                    to_row, to_row_off = anchor_actual.to.row, anchor_actual.to.rowOff

                    from_cell = AnchorMarker(col=from_col - 1, colOff=from_col_off, row=from_row, rowOff=from_row_off)
                    to_cell = AnchorMarker(col=to_col - 1, colOff=to_col_off, row=to_row, rowOff=to_row_off)

                    new_anchor = TwoCellAnchor(_from = from_cell, to = to_cell, editAs='absolute')
                    stamp_1.anchor = new_anchor

                    stamp_2 = self.wb_technical[eq_type]._images[2]
                    anchor_actual = stamp_2.anchor

                    from_col, from_col_off = anchor_actual._from.col, anchor_actual._from.colOff
                    from_row, from_row_off = anchor_actual._from.row, anchor_actual._from.rowOff
                    to_col, to_col_off = anchor_actual.to.col, anchor_actual.to.colOff
                    to_row, to_row_off = anchor_actual.to.row, anchor_actual.to.rowOff

                    from_cell = AnchorMarker(col=from_col - 1, colOff=from_col_off, row=from_row, rowOff=from_row_off)
                    to_cell = AnchorMarker(col=to_col - 1, colOff=to_col_off, row=to_row, rowOff=to_row_off)

                    new_anchor = TwoCellAnchor(_from = from_cell, to = to_cell, editAs='absolute')
                    stamp_2.anchor = new_anchor

                if int(rev) == 0:
                    ws = self.wb_technical[self.wb_technical.sheetnames[-2]]
                    ws.cell(row=last_row + 3, column=num_column_amount).value = number_items
                    ws.cell(row=last_row + 3, column=num_column_amount).font = Font(name="Calibri", size=14)

            # Editing sheet NOTES
                if int(rev) == 0:
                    sheet_name = "NOTES"  # Selecting  sheet
                    ws = self.wb_technical[sheet_name]

                    rich_string = CellRichText(
                    'We are only offering measuring flow elements. Please be informed that our product range includes temperature elements, and glass and magnetic level indicators; all with european certification. (https://www.eipsa.es/en/products)\n',
                    TextBlock(InlineFont(b=True), 'The prices quoted could be reduced in case of purchasing our full range of products.'))
                    ws["B6"] = rich_string

                    rich_string = CellRichText(
                    TextBlock(InlineFont(i=True), 'Estamos ofertando solamente elementos de medida de caudal, les informamos que en nuestra gama de fabricación con certificación europea, incluye también elementos de temperatura e indicadores de nivel de vidrio y magnéticos. (https://www.eipsa.es/productos)\n'),
                    TextBlock(InlineFont(b=True, i=True), 'Los precios ofertados podrían reducirse en caso de compra de toda nuestra gama.'))
                    ws["B7"] = rich_string

                    rich_string = CellRichText(
                    'Delivery time ' + delivery_time + ' weeks since drawing / calculation approval of ',
                    TextBlock(InlineFont(b=True), 'all equipment'),
                    ' in the contract, as well as critical documentation (August and last two December weeks excluded).')
                    ws["B12"] = rich_string

                    rich_string = CellRichText(
                    'Plazo de entrega ' + delivery_time + ' semanas desde aprobación de planos y cálculos de la ',
                    TextBlock(InlineFont(b=True, i=True), 'totalidad de los equipos'),
                    TextBlock(InlineFont(i=True),' amparados por el contrato, asi como la documentación crítica (Agosto y las dos últimas semanas de diciembre excluidos).'))
                    ws["B13"] = rich_string

                    rich_string = CellRichText(
                    'Quotation prepared according to the information provided in the datasheet corresponding to each TAG. ',
                    TextBlock(InlineFont(u='single'), 'EIPSA does not hold the final responsibility regarding selection of equipment material neither analyze process data.\n'),
                    'The datasheet will be considered the only technical/contractual document, any other documentation will be considered as complementary documentation with informative purpose.')
                    ws["B21"] = rich_string

                    rich_string = CellRichText(
                    'Esta oferta ha sido elaborada en base a la información que figura en la hoja de datos correspondiente a cada TAG, ',
                    TextBlock(InlineFont(i=True, u='single'), 'no siendo responsabilidad final de EIPSA la elección del material a utilizar en los equipos ni analizar datos de proceso.\n'),
                    TextBlock(InlineFont(i=True),'La hoja de datos será el único documento técnico/contractual, cualquier otra documentación recibida será considerada como documentación complementaria a efectos informativos.'))
                    ws["B22"] = rich_string

                    if pay_term == "100_delivery":
                        ws["B45"] = (
                            "100% of total amount of purchase order upon delivery of material according to Incoterms 2020, FCA (our facilities, Spain).\n"
                            "Payment method: bank transfer"
                        )
                        ws["B46"] = (
                            "Pago del 100% del valor total de la orden de compra a la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España).\n"
                            "Método de pago: Transferencia bancaria."
                        )
                    elif pay_term == "100_order":
                        ws["B45"] = (
                            "100 % of the total amount of purchase order upon receipt of purchase order.\n"
                            "Payment method: bank transfer"
                        )
                        ws["B46"] = (
                            "Pago del 100% del valor total de la orden de compra a la recepción de la orden.\n"
                            "Método de pago: Transferencia bancaria"
                        )
                    elif pay_term == "90_10":
                        ws["B45"] = (
                            "PAYMENT TERMS:\n"
                            "90 % of the total amount of PO upon delivery of material according to Incoterms 2020, FCA (our facilities, Spain) and 10% when final documentation is approved. \n"
                            "Bank Transfer: 60 days since invoice issue date."
                        )
                        ws["B46"] = (
                            "TERMINOS DE PAGO:\n"
                            "Pago del 90% del Valor total de la orden de compra a la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España) y el 10% restante cuando la documentación final sea aprobada.\n"
                            "Transferencia Bancaria: 60 días desde emisión de factura."
                        )
                    elif pay_term == "50_50":
                        ws["B45"] = (
                            "50 % of the total amount of purchase order upon receipt of purchase order. Remaining 50% before material be delivered according to Incoterms 2020, FCA (our facilities, Spain).\n"
                            "Payment method: bank transfer."
                        )
                        ws["B46"] = (
                            "Pago del 50% del valor total de la orden de compra a la recepción de la orden. El 50% restante antes de la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España).\n"
                            "Método de pago: Transferencia bancaria."
                        )
                    elif pay_term == "Others":
                        ws["B45"] = "PAYMENT TERMS TO BE DEFINED"
                        ws["B45"].font = Font(name="Calibri", size=11, bold=True, color="FF0000")
                        ws["B46"] = "TERMINOS DE PAGO POR DEFINIR"
                        ws["B46"].font = Font(name="Calibri", size=11, bold=True, italic=True, color="FF0000")

                    rich_string = CellRichText(
                    'For amounts greater than 30,000.00 € we can issue a warranty bond (if required) valid until the end of the indicated warranty period.\nBond warranty of 10% will be issued with the invoice of the last supplement.\n',
                    TextBlock(InlineFont(b=True), 'For lower amounts no warranty bond is issued.'))
                    ws["B48"] = rich_string

                    rich_string = CellRichText(
                    'Para importes superiores a 30.000,00 €, si es requerido, podremos emitir aval de garantía y estará vigente hasta el final del periodo de garantía indicado.\nEl aval del 10% será emitido con la factura del último suplemento.\n',
                    TextBlock(InlineFont(i=True, b=True), 'Por debajo de dicha cantidad, no se emitirán avales.'))
                    ws["B49"] = rich_string

                    ws["A60"] = (
                        "If you require further information related with this offer, please do not hesitate to contact:\n"
                        + responsible
                        + "\n"
                        + email
                        + "\n"
                        "Telf.: (+34) 916.582.118"
                    )
                    
                    std = self.wb_technical["1.3"]
                    self.wb_technical.remove(std)

                left_text = "Fecha/Date: " + date_offer
                right_text = "Petición nº/Inquiry: " + num_ref

                for sheet in self.wb_technical.worksheets:
                    sheet.oddFooter.left.text = left_text
                    sheet.oddFooter.right.text = right_text
                    sheet.oddFooter.center.text = "Page &P de &N"

                    sheet.oddFooter.left.size = 9
                    sheet.oddFooter.right.size = 9
                    sheet.oddFooter.center.size = 9

                self.save_excel_technical()

                root = Tk()
                root.withdraw()  # Hiding main window Tkinter

                # close communication with the PostgreSQL database server
                # commit the changes
                conn.commit()
        except (Exception, psycopg2.DatabaseError) as error:
            dlg = QtWidgets.QMessageBox()
            new_icon = QtGui.QIcon()
            new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
            dlg.setWindowIcon(new_icon)
            dlg.setWindowTitle("ERP EIPSA")
            dlg.setText("Ha ocurrido el siguiente error:\n"
                        + str(error))
            dlg.setIcon(QtWidgets.QMessageBox.Icon.Critical)
            dlg.exec()
            del dlg, new_icon
        finally:
            if conn is not None:
                conn.close()

    def euros_to_float(self, value):
        """
        Converts a string value representing an amount in euros to a float.
        
        Args:
            value (str): The string representation of an amount in euros, with commas for decimal separation and ' €' for currency indication.
        
        Returns:
            float: The numeric value of the amount in euros.
        """
        value = value.replace(".", "")
        value = value.replace(",", ".")
        value = value[: value.find(" €")]
        return float(value)

    def save_excel_commercial(self):
        """Saves the populated Excel workbook to a specified location.
        Opens a dialog window for the user to select the file path and name.
        """
        # Dialog window to select folder and file name; if path is selected, excel file is saved
        output_path_commercial = asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Archivos de Excel", "*.xlsx")],
            title="Guardar Oferta comercial",
        )
        if output_path_commercial:
            self.wb_commercial.save(output_path_commercial)
            return output_path_commercial

    def save_excel_technical(self):
        """Saves the populated Excel workbook to a specified location.
        Opens a dialog window for the user to select the file path and name.
        """
        output_path_technical = asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Archivos de Excel", "*.xlsx")],
            title="Guardar Oferta técnica",
        )
        if output_path_technical:
            self.wb_technical.save(output_path_technical)

    def adjust_images(self, sheet):
        """
        Adjusts the width of all images in the provided spreadsheet sheet by decreasing each by 22 units.
        
        Args:
            sheet: The spreadsheet sheet containing images to be adjusted.
        """
        for image in sheet._images:
            image.width -= 22

class offer_flow_temp_level:
    """
    A class to manage export offer details for flow equipments.
    
    Attributes:
        numoffer (str): Offer number.
        username (str): Name of the user creating the offer.
        rev (str): Revision number of the offer.
        project (str): Name of the project.
        delivery_term (str): Delivery terms for the offer.
        delivery_time (str): Expected delivery time.
        validity (str): Validity period of the offer.
        pay_term (str): Payment terms.
        testinspection (str): Information about testing and inspection.
        revchanges (str): Details of changes made in the revision.
        notes (str): Additional notes, split by line.
    """
    def __init__(self, numoffer, username, rev, project, delivery_term, delivery_time, validity, pay_term, testinspection, revchanges, notes):
        """
        Initializes the offer.

        Args:
            numoffer (str): Offer number.
            username (str): Name of the user creating the offer.
            rev (str): Revision number of the offer.
            project (str): Name of the project.
            delivery_term (str): Delivery terms for the offer.
            delivery_time (str): Expected delivery time.
            validity (str): Validity period of the offer.
            pay_term (str): Payment terms.
            testinspection (str): Information about testing and inspection.
            revchanges (str): Details of changes made in the revision.
            notes (str): Additional notes, split by line.
        """
        notes = notes.split('\n')
        
        date_offer = date.today().strftime("%d/%m/%Y")
        offername_commercial = numoffer + "-" + "Commercial.Rev" + rev
        offername_technical = numoffer + "-" + "Technical.Rev" + rev

        query_commercial = """
                    SELECT name, surname, email
                    FROM users_data.registration
                    WHERE username = %s
                    """
        query_dataoffer = """
                        SELECT client, num_ref_offer
                        FROM offers
                        WHERE UPPER (num_offer) LIKE UPPER('%%'||%s||'%%')
                        """
        query_tagsdata_flow = """
                        SELECT *
                        FROM tags_data.tags_flow
                        WHERE (
                        UPPER ("num_offer") LIKE UPPER('%%'||%s||'%%')
                        AND
                        "tag_state" NOT IN ('PURCHASED','DELETED')
                        )
                        """
        query_tagsdata_temp = """
                        SELECT *
                        FROM tags_data.tags_temp
                        WHERE (
                        UPPER ("num_offer") LIKE UPPER('%%'||%s||'%%')
                        AND
                        "tag_state" NOT IN ('PURCHASED','DELETED')
                        )
                        """
        query_tagsdata_level = """
                        SELECT *
                        FROM tags_data.tags_level
                        WHERE (
                        UPPER ("num_offer") LIKE UPPER('%%'||%s||'%%')
                        AND
                        "tag_state" NOT IN ('PURCHASED','DELETED')
                        )
                        """
        conn = None
        try:
            # read the connection parameters
            params = config()
            # connect to the PostgreSQL server
            conn = psycopg2.connect(**params)
            cur = conn.cursor()

            cur.execute(query_dataoffer, (numoffer,))
            results_offer = cur.fetchall()
            client = results_offer[0][0]
            num_ref = results_offer[0][1]

            cur.execute(query_commercial, (username,))
            results_commercial = cur.fetchall()
            responsible = results_commercial[0][0] + " " + results_commercial[0][1]
            email = results_commercial[0][2]

            cur.execute(query_tagsdata_flow, (numoffer,))
            data_tags_flow = cur.fetchall()

            columns_flow = []
            for elt in cur.description:
                columns_flow.append(elt[0])

            cur.execute(query_tagsdata_temp, (numoffer,))
            data_tags_temp = cur.fetchall()

            columns_temp = []
            for elt in cur.description:
                columns_temp.append(elt[0])

            cur.execute(query_tagsdata_level, (numoffer,))
            data_tags_level = cur.fetchall()

            columns_level = []
            for elt in cur.description:
                columns_level.append(elt[0])

            if len(data_tags_flow) == 0 or len(data_tags_temp) == 0 or len(data_tags_level) == 0:
                dlg = QtWidgets.QMessageBox()
                new_icon = QtGui.QIcon()
                new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
                dlg.setWindowIcon(new_icon)
                dlg.setWindowTitle("ERP EIPSA")
                dlg.setText("No hay TAGS importados en la oferta")
                dlg.setIcon(QtWidgets.QMessageBox.Icon.Warning)
                dlg.exec()
                del dlg, new_icon
            else:
                value_type_dict_flow = {
                    "A. Chamber": 1,
                    "C. RING": 1,
                    "F": 1,
                    "F+C.RING": 1,
                    "F+P": 1,
                    "IFO": 1,
                    "M.RUN": 1,
                    "P": 1,
                    "NOZZLE BF": 2,
                    "NOZZLE BW": 2,
                    "NOZZLE F": 2,
                    "PTC-6": 2,
                    "VFM": 3,
                    "VFW": 3,
                    "VWM": 3,
                    "VWW": 3,
                    "WEDGE": 4,
                    "PITOT": 5,
                    "RO": 6,
                    "MULTISTAGE RO": 7,
                }

                # Setting the dataframe with the equipment data
                df_flow = pd.DataFrame(data=data_tags_flow, columns=columns_flow)
                df_flow = df_flow.iloc[:, 1:32]
                df_flow["value_type"] = df_flow["item_type"].map(value_type_dict_flow)
                df_flow = df_flow.sort_values(by=["value_type", "tag"])
                df_flow["amount"] = df_flow["amount"].apply(self.euros_to_float)
                total_amount_material = df_flow["amount"].sum()
                df_flow = df_flow.drop([
                        "tag_state",
                        "num_offer",
                        "num_order",
                        "num_po",
                        "position",
                        "subposition",
                        "flange_type",
                        "plate_std",
                        "pipe_spec",
                        "aprox_weight"
                    ],
                    axis=1,)
                
                value_type_dict_temp = {
                    "TW": 1,
                    "TW+TE": 2,
                    "TW+TE+TIT": 2,
                    "RETAINING FLANGE":2,
                    "TW+BIM": 3,
                    "TE": 4,
                    "BIM": 5,
                    "TIT": 6,
                    "SKIN+TT": 7,
                    "SKIN POINT": 7,
                    "Multi-T": 8
                }

                df_temp = pd.DataFrame(data=data_tags_temp, columns=columns_temp)
                df_temp["value_type"] = df_temp["item_type"].map(value_type_dict_temp)
                df_temp = df_temp.sort_values(by=["tag", "value_type"])
                df_temp = df_temp.iloc[:, 1:37]
                df_temp["value_type"] = df_temp["item_type"].map(value_type_dict_temp)
                df_temp = df_temp.sort_values(by=["value_type", "tag"])
                df_temp["amount"] = df_temp["amount"].apply(self.euros_to_float)
                total_amount_material = df_temp["amount"].sum()
                df_temp = df_temp.drop([
                        "tag_state",
                        "num_offer",
                        "num_order",
                        "num_po",
                        "position",
                        "subposition",
                        "std_tw",
                        "insulation"
                    ],
                    axis=1,)

                value_type_dict = {"Transparent": 1, "Reflex": 1, "Magnetic": 2}

                df_level = pd.DataFrame(data=data_tags_level, columns=columns_level)
                df_level["value_type"] = df_level["item_type"].map(value_type_dict)
                df_level = df_level.sort_values(by=["tag", "value_type"])
                df_level = df_level.iloc[:, 1:38]
                df_level["value_type"] = df_level["item_type"].map(value_type_dict)
                df_level = df_level.sort_values(by=["value_type", "tag"])
                df_level["amount"] = df_level["amount"].apply(self.euros_to_float)
                total_amount_material = df_level["amount"].sum()
                df_level = df_level.drop([
                        "tag_state",
                        "num_offer",
                        "num_order",
                        "num_po",
                        "position",
                        "subposition",
                        "proc_conn_type",
                        "flags",
                        "flange_type",
                        "nipple_hex",
                        "nipple_tub"
                    ],
                    axis=1,)

                number_items = df_flow.shape[0] + df_temp.shape[0] + df_level.shape[0]
                documentation = number_items * 70

                # Loading Excel Template
                self.wb_commercial = load_workbook(
                    r"\\nas01\DATOS\Comunes\EIPSA-ERP\Plantillas Exportación\PLANTILLA OFERTA CAUDAL-TEMPERATURA-NIVEL.xlsx"
                )

                # Editing sheet COVER
                sheet_name = "COVER"
                ws = self.wb_commercial[sheet_name]
                ws["E4"] = client
                ws["E6"] = offername_commercial
                ws["E8"] = num_ref
                ws["E10"] = project
                ws["E12"] = date_offer
                ws["E14"] = delivery_term
                ws["E16"] = validity + " days"
                ws["C43"] = responsible
                ws["C45"] = email

                # Editing sheet EQUIPMENT DATA
                sheets_confirmed = ["COVER", "1.2", "1.3", "NOTES"]

                for item_type in df_flow["item_type"].unique():
                    sheets_confirmed.append(item_type)

                for item_type in df_temp["item_type"].unique():
                    sheets_confirmed.append(item_type)

                for item_type in df_level["item_type"].unique():
                    sheets_confirmed.append(item_type)

                dict_sheets_data = {}

                for value_type in df_flow["value_type"].unique():
                    df_toexport = df_flow[df_flow["value_type"] == value_type]
                    df_toexport.insert(0, "N°", range(1, len(df_toexport) + 1))
                    df_toexport.set_index("N°", inplace=True, drop=False)
                    df_toexport.index.name = None
                    df_toexport = df_toexport.drop(["value_type"], axis=1)

                    eq_type = (
                        "FLOW ELEMENTS DATA"
                        if value_type == 1
                        else (
                            "NOZZLE ELEMENTS DATA"
                            if value_type == 2
                            else (
                                "VENTURI ELEMENTS DATA"
                                if value_type == 3
                                else (
                                    "WEDGE ELEMENTS DATA"
                                    if value_type == 4
                                    else (
                                        "PITOT ELEMENTS DATA"
                                        if value_type == 5
                                        else (
                                            "RO ELEMENTS DATA"
                                            if value_type == 6
                                            else "MULTISTAGE RO ELEMENTS DATA"
                                        )
                                    )
                                )
                            )
                        )
                    )

                    if eq_type not in sheets_confirmed:
                        sheets_confirmed.append(eq_type)

                    if eq_type == "FLOW ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "tube_material",
                                "plate_type",
                                "valve_conn",
                                "valve_material_body",
                                "stages_number",
                                "aprox_length"],
                            axis=1,)
                    elif eq_type == "NOZZLE ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "plate_type",
                                "plate_thk",
                                "gasket_material",
                                "valve_conn",
                                "valve_material_body",
                                "stages_number"],
                            axis=1,)
                    elif (eq_type == "VENTURI ELEMENTS DATA" or eq_type == "WEDGE ELEMENTS DATA"):
                        df_toexport = df_toexport.drop([
                                "plate_type",
                                "plate_thk",
                                "gasket_material",
                                "bolts_nuts_material",
                                "valve_conn",
                                "valve_material_body",
                                "stages_number"],
                            axis=1,)
                    elif eq_type == "PITOT ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "flange_material",
                                "tube_material",
                                "tapping_num_size",
                                "plate_type",
                                "plate_thk",
                                "gasket_material",
                                "stages_number",
                                "aprox_length"],
                            axis=1,)
                    elif eq_type == "RO ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "flange_material",
                                "tube_material",
                                "tapping_num_size",
                                "gasket_material",
                                "bolts_nuts_material",
                                "valve_conn",
                                "valve_material_body",
                                "stages_number",
                                "aprox_length"],
                            axis=1,)
                    elif eq_type == "MULTISTAGE RO ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "tapping_num_size",
                                "gasket_material",
                                "bolts_nuts_material",
                                "valve_conn",
                                "valve_material_body"],
                            axis=1,)

                    ws = self.wb_commercial[eq_type]
                    ws["G2"] = date_offer
                    ws["G3"] = num_ref
                    ws["G4"] = offername_commercial
                    if revchanges != "":
                        ws["G5"] = rev + " " + revchanges
                        ws["G5"].font = Font(name="Calibri", size=14, bold=True)
                        ws["G5"].fill = PatternFill("solid", fgColor="FFFF00")

                    if int(rev) > 0:
                        for row in ws.iter_rows(min_row=2, max_row=4, min_col=6, max_col=7):
                            for cell in row:
                                cell.value = None
                                cell._style = ws["F1"]._style

                    last_row = ws.max_row

                    # for col_num, col_name in enumerate(df_toexport.columns, start=1):
                    #     cell = ws.cell(row=last_row + 1, column=col_num)
                    #     cell.value = col_name
                    #     cell._style = ws["Y1"]._style

                    # last_row = ws.max_row

                    num_column_amount = df_toexport.columns.get_loc("amount") + 1

                    for index, row in df_toexport.iterrows():  # Data in desired row
                        for col_num, value in enumerate(row, start=1):
                            cell = ws.cell(row=last_row + 1, column=col_num)
                            cell.value = value
                            if col_num == num_column_amount:
                                cell._style = ws["X1"]._style
                            else:
                                cell._style = ws["S1"]._style

                        last_row = ws.max_row

                    if eq_type == "VENTURI ELEMENTS DATA":
                        ws[f"A{last_row+3}"] = "PRICES INCLUDE MACHINED INTEGRAL CENTRE SECTION AND ALL STRUCTURAL WELDS 100% RADIOGRAPHED"
                        ws[f"A{last_row+3}"]._style = ws["Z2"]._style
                    ws[f"A{last_row+4}"] = "OFFER VALIDITY: " + validity + " DAYS"
                    ws[f"A{last_row+4}"]._style = ws["Z1"]._style
                    ws[f"A{last_row+5}"] = (
                        "DELIVERY TIME: "
                        + delivery_time
                        + " WEEKS SINCE DRAWING / CALCULATION APPROVAL (AUGUST AND LAST TWO DECEMBER WEEKS EXCLUDED)"
                    )
                    ws[f"A{last_row+5}"]._style = ws["Z1"]._style

                    if notes != "":
                        if isinstance(notes, list):
                            line = last_row + 6
                            for note in notes:
                                ws[f"A{line}"] = note
                                ws[f"A{line}"]._style = ws["Z1"]._style
                                line += 1
                        else:
                            line = last_row + 6
                            ws[f"A{line}"] = notes
                            ws[f"A{line}"]._style = ws["Z1"]._style


                    dict_sheets_data[eq_type] = [last_row, num_column_amount, df_toexport["amount"].sum(), df_toexport.shape[0]]

                for value_type in df_temp["value_type"].unique():
                    df_toexport = df_temp[df_temp["value_type"] == value_type]
                    df_toexport.insert(0, "N°", range(1, len(df_toexport) + 1))
                    df_toexport.set_index("N°", inplace=True, drop=False)
                    df_toexport.index.name = None
                    df_toexport = df_toexport.drop(["value_type"], axis=1)

                    eq_type = (
                        "TW ELEMENTS DATA"
                        if value_type == 1
                        else (
                            "TW+TE ELEMENTS DATA"
                            if value_type == 2
                            else (
                                "TW+BIM ELEMENTS DATA"
                                if value_type == 3
                                else (
                                    "TE ELEMENTS DATA"
                                    if value_type == 4
                                    else (
                                        "BIM ELEMENTS DATA"
                                        if value_type == 5
                                        else (
                                            "TIT ELEMENTS DATA"
                                            if value_type == 6
                                            else (
                                                "SKIN POINT ELEMENTS DATA"
                                                if value_type == 7
                                                else "MULTI-T ELEMENTS DATA"
                                            )
                                        )
                                    )
                                )
                            )
                        )
                    )

                    if eq_type not in sheets_confirmed:
                        sheets_confirmed.append(eq_type)

                    if eq_type == "TW ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "std_length",
                                "sensor_element",
                                "sheath_stem_material",
                                "sheath_stem_diam",
                                "temp_inf",
                                "temp_sup",
                                "nipple_ext_material",
                                "nipple_ext_length",
                                "head_case_material",
                                "elec_conn_case_diam",
                                "tt_cerblock"],
                            axis=1,)
                    elif eq_type == "TW+TE ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "std_length",
                                "temp_inf",
                                "temp_sup",
                                "nipple_ext_length"],
                            axis=1,)
                    elif eq_type == "TW+BIM ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "std_length",
                                "nipple_ext_length",
                                "tt_cerblock",
                                "puntal",
                                "tube_t"],
                            axis=1,)
                    elif eq_type == "TE ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "tw_type",
                                "size",
                                "rating",
                                "facing",
                                "material_tw",
                                "std_length",
                                "root_diam",
                                "tip_diam",
                                "temp_inf",
                                "temp_sup",
                                "nipple_ext_material",
                                "nipple_ext_length",
                                "head_case_material",
                                "elec_conn_case_diam",
                                "tt_cerblock",
                                "material_flange_lj",
                                "gasket_material",
                                "puntal",
                                "tube_t"],
                            axis=1,)
                    elif eq_type == "BIM ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "tw_type",
                                "size",
                                "rating",
                                "facing",
                                "material_tw",
                                "std_length",
                                "root_diam",
                                "tip_diam",
                                "nipple_ext_length",
                                "tt_cerblock",
                                "material_flange_lj",
                                "gasket_material",
                                "puntal",
                                "tube_t"],
                            axis=1,)
                    elif eq_type == "TIT ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "tw_type",
                                "size",
                                "rating",
                                "facing",
                                "material_tw",
                                "std_length",
                                "ins_length",
                                "root_diam",
                                "tip_diam",
                                "sensor_element",
                                "sheath_stem_material",
                                "sheath_stem_diam",
                                "nipple_ext_material",
                                "nipple_ext_length",
                                "head_case_material",
                                "material_flange_lj",
                                "gasket_material",
                                "puntal",
                                "tube_t"],
                            axis=1,)
                    elif (eq_type == "SKIN POINT ELEMENTS DATA" or eq_type == "SKIN+TT ELEMENTS DATA"):
                        df_toexport = df_toexport.drop([
                                "tw_type",
                                "size",
                                "rating",
                                "facing",
                                "material_tw",
                                "std_length",
                                "root_diam",
                                "tip_diam",
                                "temp_inf",
                                "temp_sup",
                                "material_flange_lj",
                                "gasket_material",
                                "puntal",
                                "tube_t"],
                            axis=1,)
                    elif eq_type == "MULTI-T ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "material_tw",
                                "root_diam",
                                "tip_diam",
                                "temp_inf",
                                "temp_sup",
                                "tt_cerblock",
                                "material_flange_lj",
                                "puntal",
                                "tube_t"],
                            axis=1,)

                    ws = self.wb_commercial[eq_type]
                    ws["G2"] = date_offer
                    ws["G3"] = num_ref
                    ws["G4"] = offername_commercial
                    if revchanges != "":
                        ws["G5"] = rev + " " + revchanges
                        ws["G5"].font = Font(name="Calibri", size=14, bold=True)
                        ws["G5"].fill = PatternFill("solid", fgColor="FFFF00")

                    if int(rev) > 0:
                        for row in ws.iter_rows(min_row=2, max_row=4, min_col=6, max_col=7):
                            for cell in row:
                                cell.value = None
                                cell._style = ws["F1"]._style

                    last_row = ws.max_row

                    num_column_amount = df_toexport.columns.get_loc("amount") + 1

                    for index, row in df_toexport.iterrows():  # Data in desired row
                        for col_num, value in enumerate(row, start=1):
                            cell = ws.cell(row=last_row + 1, column=col_num)
                            cell.value = value
                            if col_num == num_column_amount:
                                cell._style = ws["AG1"]._style
                            else:
                                cell._style = ws["AB1"]._style

                        last_row = ws.max_row

                    ws[f"A{last_row+3}"] = "OFFER VALIDITY: " + validity + " DAYS"
                    ws[f"A{last_row+3}"]._style = ws["AI1"]._style
                    ws[f"A{last_row+4}"] = (
                        "DELIVERY TIME: "
                        + delivery_time
                        + " WEEKS SINCE DRAWING / CALCULATION APPROVAL (AUGUST AND LAST TWO DECEMBER WEEKS EXCLUDED)"
                    )
                    ws[f"A{last_row+4}"]._style = ws["AI1"]._style

                    if notes != "":
                        if isinstance(notes, list):
                            line = last_row + 5
                            for note in notes:
                                ws[f"A{line}"] = note
                                ws[f"A{line}"]._style = ws["AI1"]._style
                                line += 1
                        else:
                            line = last_row + 5
                            ws[f"A{line}"] = notes
                            ws[f"A{line}"]._style = ws["AI1"]._style
                    dict_sheets_data[eq_type] = [last_row, num_column_amount, df_toexport["amount"].sum(), df_toexport.shape[0]]

                for value_type in df_level["value_type"].unique():
                    df_toexport = df_level[df_level["value_type"] == value_type]
                    df_toexport.insert(0, "N°", range(1, len(df_toexport) + 1))
                    df_toexport.set_index("N°", inplace=True, drop=False)
                    df_toexport.index.name = None
                    df_toexport = df_toexport.drop(["value_type"], axis=1)

                    eq_type = (
                        "LEVEL GAUGES ELEMENTS DATA"
                        if value_type == 1
                        else "MAGNETIC ELEMENTS DATA"
                        )

                    if eq_type not in sheets_confirmed:
                        sheets_confirmed.append(eq_type)

                    if eq_type == "LEVEL GAUGES ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "dv_size",
                                "dv_rating",
                                "dv_facing",
                                "float_material"],
                            axis=1,)
                    elif eq_type == "MAGNETIC ELEMENTS DATA":
                        df_toexport = df_toexport.drop([
                                "dv_conn",
                                "item_type",
                                "valve_type",
                                "case_cover_material",
                                "illuminator",
                                "ip_code"],
                            axis=1,)

                    ws = self.wb_commercial[eq_type]
                    ws["G2"] = date_offer
                    ws["G3"] = num_ref
                    ws["G4"] = offername_commercial
                    if revchanges != "":
                        ws["G5"] = rev + " " + revchanges
                        ws["G5"].font = Font(name="Calibri", size=14, bold=True)
                        ws["G5"].fill = PatternFill("solid", fgColor="FFFF00")

                    if int(rev) > 0:
                        for row in ws.iter_rows(min_row=2, max_row=4, min_col=6, max_col=7):
                            for cell in row:
                                cell.value = None
                                cell._style = ws["F1"]._style

                    last_row = ws.max_row

                    num_column_amount = df_toexport.columns.get_loc("amount") + 1

                    for index, row in df_toexport.iterrows():  # Data in desired row
                        for col_num, value in enumerate(row, start=1):
                            cell = ws.cell(row=last_row + 1, column=col_num)
                            cell.value = value
                            if col_num == num_column_amount:
                                cell._style = ws["AG1"]._style
                            else:
                                cell._style = ws["AB1"]._style

                        last_row = ws.max_row

                    ws[f"A{last_row+3}"] = "OFFER VALIDITY: " + validity + " DAYS"
                    ws[f"A{last_row+3}"]._style = ws["AI1"]._style
                    ws[f"A{last_row+4}"] = (
                        "DELIVERY TIME: "
                        + delivery_time
                        + " WEEKS SINCE DRAWING APPROVAL (AUGUST AND LAST TWO DECEMBER WEEKS EXCLUDED)"
                    )
                    ws[f"A{last_row+4}"]._style = ws["AI1"]._style

                    if notes != "":
                        if isinstance(notes, list):
                            line = last_row + 5
                            for note in notes:
                                ws[f"A{line}"] = note
                                ws[f"A{line}"]._style = ws["AI1"]._style
                                line += 1
                        else:
                            line = last_row + 5
                            ws[f"A{line}"] = notes
                            ws[f"A{line}"]._style = ws["AI1"]._style

                    dict_sheets_data[eq_type] = [last_row, num_column_amount, df_toexport["amount"].sum(), df_toexport.shape[0]]

                ws.cell(row=last_row + 3, column=num_column_amount - 1).value = "QTY. TOTAL"
                ws.cell(row=last_row + 3, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=last_row + 3, column=num_column_amount).value = number_items
                
                row_amount = last_row + 4
                for key, value in dict_sheets_data.items():
                    parts_key = key.split(" ")
                    ws.cell(row=row_amount + 2, column=num_column_amount - 1).value = "TOTAL AMOUNT OF " + parts_key[0] + " " + parts_key[1] + " (QTY: " + str(value[3]) + ")"
                    ws.cell(row=row_amount + 2, column=num_column_amount).value = value[2]
                    ws.cell(row=row_amount + 2, column=num_column_amount - 1)._style = ws["R1"]._style
                    ws.cell(row=row_amount + 2, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                    ws.cell(row=row_amount + 2, column=num_column_amount)._style = ws["T1"]._style

                    row_amount += 2

                ws.cell(row=row_amount + 2, column=num_column_amount - 1).value = "TOTAL AMOUNT OF MATERIAL"
                ws.cell(row=row_amount + 2, column=num_column_amount).value = total_amount_material
                ws.cell(row=row_amount + 4, column=num_column_amount - 1).value = "PACKING AND TRANSPORT"
                ws.cell(row=row_amount + 4, column=num_column_amount).value = (f"=MROUND({get_column_letter(num_column_amount)}{row_amount + 2}*0.03,10)" if total_amount_material > 6700 else 200 )
                ws.cell(row=row_amount + 5, column=num_column_amount - 1).value = "TESTS & INSPECTION"
                ws.cell(row=row_amount + 5, column=num_column_amount).value = float(testinspection)
                ws.cell(row=row_amount + 6, column=num_column_amount - 1).value = "DOCUMENTATION"
                ws.cell(row=row_amount + 6, column=num_column_amount).value = documentation
                ws.cell(row=row_amount + 8, column=num_column_amount - 1).value = "TOTAL AMOUNT OF BID"
                ws.cell(row=row_amount + 8, column=num_column_amount).value = f"=SUM({get_column_letter(num_column_amount)}{row_amount + 2}:{get_column_letter(num_column_amount)}{row_amount + 6})"

                ws.cell(row=last_row + 3, column=num_column_amount - 1)._style = ws["R1"]._style
                ws.cell(row=last_row + 3, column=num_column_amount).font = Font(name="Calibri", size=14)
                ws.cell(row=row_amount + 2, column=num_column_amount - 1)._style = ws["R1"]._style
                ws.cell(row=row_amount + 2, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=row_amount + 2, column=num_column_amount)._style = ws["T1"]._style
                ws.cell(row=row_amount + 4, column=num_column_amount - 1).font = Font(name="Calibri", size=14)
                ws.cell(row=row_amount + 4, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=row_amount + 4, column=num_column_amount)._style = ws["T1"]._style
                ws.cell(row=row_amount + 5, column=num_column_amount - 1).font = Font(name="Calibri", size=14)
                ws.cell(row=row_amount + 5, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=row_amount + 5, column=num_column_amount)._style = ws["T1"]._style
                ws.cell(row=row_amount + 6, column=num_column_amount - 2)._style = ws["U1"]._style
                ws.cell(row=row_amount + 6, column=num_column_amount - 1)._style = ws["U1"]._style
                ws.cell(row=row_amount + 6, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=row_amount + 6, column=num_column_amount)._style = ws["V1"]._style
                ws.cell(row=row_amount + 8, column=num_column_amount - 1)._style = ws["R1"]._style
                ws.cell(row=row_amount + 8, column=num_column_amount - 1).alignment = Alignment(horizontal='right')
                ws.cell(row=row_amount + 8, column=num_column_amount)._style = ws["W1"]._style

            # Editing sheet NOTES
                sheet_name = "NOTES"  # Selecting  sheet
                ws = self.wb_commercial[sheet_name]

                rich_string = CellRichText(
                'We are only offering measuring flow elements. Please be informed that our product range includes temperature elements, and glass and magnetic level indicators; all with european certification. (https://www.eipsa.es/en/products)\n',
                TextBlock(InlineFont(b=True), 'The prices quoted could be reduced in case of purchasing our full range of products.'))
                ws["B6"] = rich_string

                rich_string = CellRichText(
                TextBlock(InlineFont(i=True), 'Estamos ofertando solamente elementos de medida de caudal, les informamos que en nuestra gama de fabricación con certificación europea, incluye también elementos de temperatura e indicadores de nivel de vidrio y magnéticos. (https://www.eipsa.es/productos)\n'),
                TextBlock(InlineFont(b=True, i=True), 'Los precios ofertados podrían reducirse en caso de compra de toda nuestra gama.'))
                ws["B7"] = rich_string

                rich_string = CellRichText(
                'Delivery time ' + delivery_time + ' weeks since drawing / calculation approval of ',
                TextBlock(InlineFont(b=True), 'all equipment'),
                ' in the contract, as well as critical documentation (August and last two December weeks excluded).')
                ws["B12"] = rich_string

                rich_string = CellRichText(
                'Plazo de entrega ' + delivery_time + ' semanas desde aprobación de planos y cálculos de la ',
                TextBlock(InlineFont(b=True, i=True), 'totalidad de los equipos'),
                TextBlock(InlineFont(i=True),' amparados por el contrato, asi como la documentación crítica (Agosto y las dos últimas semanas de diciembre excluidos).'))
                ws["B13"] = rich_string

                rich_string = CellRichText(
                'Quotation prepared according to the information provided in the datasheet corresponding to each TAG. ',
                TextBlock(InlineFont(u='single'), 'EIPSA does not hold the final responsibility regarding selection of equipment material neither analyze process data.\n'),
                'The datasheet will be considered the only technical/contractual document, any other documentation will be considered as complementary documentation with informative purpose.')
                ws["B21"] = rich_string

                rich_string = CellRichText(
                'Esta oferta ha sido elaborada en base a la información que figura en la hoja de datos correspondiente a cada TAG, ',
                TextBlock(InlineFont(i=True, u='single'), 'no siendo responsabilidad final de EIPSA la elección del material a utilizar en los equipos ni analizar datos de proceso.\n'),
                TextBlock(InlineFont(i=True),'La hoja de datos será el único documento técnico/contractual, cualquier otra documentación recibida será considerada como documentación complementaria a efectos informativos.'))
                ws["B22"] = rich_string

                if pay_term == "100_delivery":
                    ws["B45"] = (
                        "100% of total amount of purchase order upon delivery of material according to Incoterms 2020, FCA (our facilities, Spain).\n"
                        "Payment method: bank transfer"
                    )
                    ws["B46"] = (
                        "Pago del 100% del valor total de la orden de compra a la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España).\n"
                        "Método de pago: Transferencia bancaria."
                    )
                elif pay_term == "100_order":
                    ws["B45"] = (
                        "100 % of the total amount of purchase order upon receipt of purchase order.\n"
                        "Payment method: bank transfer"
                    )
                    ws["B46"] = (
                        "Pago del 100% del valor total de la orden de compra a la recepción de la orden.\n"
                        "Método de pago: Transferencia bancaria"
                    )
                elif pay_term == "90_10":
                    ws["B45"] = (
                        "PAYMENT TERMS:\n"
                        "90 % of the total amount of PO upon delivery of material according to Incoterms 2020, FCA (our facilities, Spain) and 10% when final documentation is approved. \n"
                        "Bank Transfer: 60 days since invoice issue date."
                    )
                    ws["B46"] = (
                        "TERMINOS DE PAGO:\n"
                        "Pago del 90% del Valor total de la orden de compra a la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España) y el 10% restante cuando la documentación final sea aprobada.\n"
                        "Transferencia Bancaria: 60 días desde emisión de factura."
                    )
                elif pay_term == "50_50":
                    ws["B45"] = (
                        "50 % of the total amount of purchase order upon receipt of purchase order. Remaining 50% before material be delivered according to Incoterms 2020, FCA (our facilities, Spain).\n"
                        "Payment method: bank transfer."
                    )
                    ws["B46"] = (
                        "Pago del 50% del valor total de la orden de compra a la recepción de la orden. El 50% restante antes de la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España).\n"
                        "Método de pago: Transferencia bancaria."
                    )
                elif pay_term == "Others":
                    ws["B45"] = "PAYMENT TERMS TO BE DEFINED"
                    ws["B45"].font = Font(name="Calibri", size=11, bold=True, color="FF0000")
                    ws["B46"] = "TERMINOS DE PAGO POR DEFINIR"
                    ws["B46"].font = Font(name="Calibri", size=11, bold=True, italic=True, color="FF0000")

                rich_string = CellRichText(
                'For amounts greater than 30,000.00 € we can issue a warranty bond (if required) valid until the end of the indicated warranty period.\nBond warranty of 10% will be issued with the invoice of the last supplement.\n',
                TextBlock(InlineFont(b=True), 'For lower amounts no warranty bond is issued.'))
                ws["B48"] = rich_string

                rich_string = CellRichText(
                'Para importes superiores a 30.000,00 €, si es requerido, podremos emitir aval de garantía y estará vigente hasta el final del periodo de garantía indicado.\nEl aval del 10% será emitido con la factura del último suplemento.\n',
                TextBlock(InlineFont(i=True, b=True), 'Por debajo de dicha cantidad, no se emitirán avales.'))
                ws["B49"] = rich_string

                ws["A64"] = (
                    "If you require further information related with this offer, please do not hesitate to contact:\n"
                    + responsible
                    + "\n"
                    + email
                    + "\n"
                    "Telf.: (+34) 916.582.118"
                )

                for sheet in self.wb_commercial.sheetnames:
                    if sheet not in sheets_confirmed:
                        sheet_to_delete = self.wb_commercial[sheet]
                        self.wb_commercial.remove(sheet_to_delete)

                if int(rev) > 0:
                    sheets_confirmed = ["COVER", "1.2", "1.3", "NOTES"]
                    for sheet in sheets_confirmed:
                        sheet_to_delete = self.wb_commercial[sheet]
                        self.wb_commercial.remove(sheet_to_delete)

                left_text = "Fecha/Date: " + date_offer
                right_text = "Petición nº/Inquiry: " + num_ref

                for sheet in self.wb_commercial.worksheets:
                    sheet.oddFooter.left.text = left_text
                    sheet.oddFooter.right.text = right_text
                    sheet.oddFooter.center.text = "Page &P de &N"

                    sheet.oddFooter.left.size = 9
                    sheet.oddFooter.right.size = 9
                    sheet.oddFooter.center.size = 9
                
                path = self.save_excel_commercial()

                # Creating the technical offer using the commercial one as template
                self.wb_technical = load_workbook(path)

                if int(rev) == 0:
                    sheet_name = "COVER"
                    ws = self.wb_technical[sheet_name]
                    ws["E6"] = offername_technical

                for value_type in df_flow["value_type"].unique():
                    eq_type = (
                        "FLOW ELEMENTS DATA"
                        if value_type == 1
                        else (
                            "NOZZLE ELEMENTS DATA"
                            if value_type == 2
                            else (
                                "VENTURI ELEMENTS DATA"
                                if value_type == 3
                                else (
                                    "WEDGE ELEMENTS DATA"
                                    if value_type == 4
                                    else (
                                        "PITOT ELEMENTS DATA"
                                        if value_type == 5
                                        else (
                                            "RO ELEMENTS DATA"
                                            if value_type == 6
                                            else "MULTISTAGE RO ELEMENTS DATA"
                                        )
                                    )
                                )
                            )
                        )
                    )

                    ws = self.wb_technical[eq_type]
                    if int(rev) == 0:
                        ws["G4"] = offername_technical

                    last_row = dict_sheets_data[eq_type][0]
                    num_column_amount = dict_sheets_data[eq_type][1]

                    # self.wb_technical[eq_type].delete_rows(last_row + 8, 20)
                    ws[f"M{last_row+5}"] = ""
                    ws[f"N{last_row+5}"] = ""

                    ws.cell(row=row_amount + 2, column=num_column_amount).value = ""
                    ws.cell(row=row_amount + 4, column=num_column_amount).value = ""
                    ws.cell(row=row_amount + 5, column=num_column_amount).value = ""
                    ws.cell(row=row_amount + 6, column=num_column_amount).value = ""
                    ws.cell(row=row_amount + 8, column=num_column_amount).value = ""

                    ws.cell(row=row_amount + 2, column=num_column_amount - 1).value = ""
                    ws.cell(row=row_amount + 4, column=num_column_amount - 1).value = ""
                    ws.cell(row=row_amount + 5, column=num_column_amount - 1).value = ""
                    ws.cell(row=row_amount + 6, column=num_column_amount - 1).value = ""
                    ws.cell(row=row_amount + 8, column=num_column_amount - 1).value = ""

                # Deleting "Amount" column
                    self.wb_technical[eq_type].delete_cols(num_column_amount, 1)

                # Adjusting the print area
                    new_last_column = num_column_amount
                    last_print_row = 40
                    nuevo_rango_impresion = f'A1:{get_column_letter(new_last_column)}{last_print_row}'
                    self.wb_technical[eq_type].print_area = nuevo_rango_impresion

                    stamp_1 = self.wb_technical[eq_type]._images[1]
                    anchor_actual = stamp_1.anchor

                    from_col, from_col_off = anchor_actual._from.col, anchor_actual._from.colOff
                    from_row, from_row_off = anchor_actual._from.row, anchor_actual._from.rowOff
                    to_col, to_col_off = anchor_actual.to.col, anchor_actual.to.colOff
                    to_row, to_row_off = anchor_actual.to.row, anchor_actual.to.rowOff

                    from_cell = AnchorMarker(col=from_col - 1, colOff=from_col_off, row=from_row, rowOff=from_row_off)
                    to_cell = AnchorMarker(col=to_col - 1, colOff=to_col_off, row=to_row, rowOff=to_row_off)

                    new_anchor = TwoCellAnchor(_from = from_cell, to = to_cell, editAs='absolute')
                    stamp_1.anchor = new_anchor

                    stamp_2 = self.wb_technical[eq_type]._images[2]
                    anchor_actual = stamp_2.anchor

                    from_col, from_col_off = anchor_actual._from.col, anchor_actual._from.colOff
                    from_row, from_row_off = anchor_actual._from.row, anchor_actual._from.rowOff
                    to_col, to_col_off = anchor_actual.to.col, anchor_actual.to.colOff
                    to_row, to_row_off = anchor_actual.to.row, anchor_actual.to.rowOff

                    from_cell = AnchorMarker(col=from_col - 1, colOff=from_col_off, row=from_row, rowOff=from_row_off)
                    to_cell = AnchorMarker(col=to_col - 1, colOff=to_col_off, row=to_row, rowOff=to_row_off)

                    new_anchor = TwoCellAnchor(_from = from_cell, to = to_cell, editAs='absolute')
                    stamp_2.anchor = new_anchor

                for value_type in df_temp["value_type"].unique():
                    eq_type = (
                        "TW ELEMENTS DATA"
                        if value_type == 1
                        else (
                            "TW+TE ELEMENTS DATA"
                            if value_type == 2
                            else (
                                "TW+BIM ELEMENTS DATA"
                                if value_type == 3
                                else (
                                    "TE ELEMENTS DATA"
                                    if value_type == 4
                                    else (
                                        "BIM ELEMENTS DATA"
                                        if value_type == 5
                                        else (
                                            "TIT ELEMENTS DATA"
                                            if value_type == 6
                                            else (
                                                "SKIN POINT ELEMENTS DATA"
                                                if value_type == 7
                                                else "MULTI-T ELEMENTS DATA"
                                            )
                                        )
                                    )
                                )
                            )
                        )
                    )

                    ws = self.wb_technical[eq_type]
                    if int(rev) == 0:
                        ws["G4"] = offername_technical

                    last_row = dict_sheets_data[eq_type][0]
                    num_column_amount = dict_sheets_data[eq_type][1]

                    # self.wb_technical[eq_type].delete_rows(last_row + 5, 20)

                    ws.cell(row=row_amount + 2, column=num_column_amount).value = ""
                    ws.cell(row=row_amount + 4, column=num_column_amount).value = ""
                    ws.cell(row=row_amount + 5, column=num_column_amount).value = ""
                    ws.cell(row=row_amount + 6, column=num_column_amount).value = ""
                    ws.cell(row=row_amount + 8, column=num_column_amount).value = ""

                    ws.cell(row=row_amount + 2, column=num_column_amount - 1).value = ""
                    ws.cell(row=row_amount + 4, column=num_column_amount - 1).value = ""
                    ws.cell(row=row_amount + 5, column=num_column_amount - 1).value = ""
                    ws.cell(row=row_amount + 6, column=num_column_amount - 1).value = ""
                    ws.cell(row=row_amount + 8, column=num_column_amount - 1).value = ""

                # Deleting "Amount" column
                    self.wb_technical[eq_type].delete_cols(num_column_amount, 1)

                # Adjusting the print area
                    new_last_column = num_column_amount
                    last_print_row = 40
                    nuevo_rango_impresion = f'A1:{get_column_letter(new_last_column)}{last_print_row}'
                    self.wb_technical[eq_type].print_area = nuevo_rango_impresion

                    stamp_1 = self.wb_technical[eq_type]._images[1]
                    anchor_actual = stamp_1.anchor

                    from_col, from_col_off = anchor_actual._from.col, anchor_actual._from.colOff
                    from_row, from_row_off = anchor_actual._from.row, anchor_actual._from.rowOff
                    to_col, to_col_off = anchor_actual.to.col, anchor_actual.to.colOff
                    to_row, to_row_off = anchor_actual.to.row, anchor_actual.to.rowOff

                    from_cell = AnchorMarker(col=from_col - 1, colOff=from_col_off, row=from_row, rowOff=from_row_off)
                    to_cell = AnchorMarker(col=to_col - 1, colOff=to_col_off, row=to_row, rowOff=to_row_off)

                    new_anchor = TwoCellAnchor(_from = from_cell, to = to_cell, editAs='absolute')
                    stamp_1.anchor = new_anchor

                    stamp_2 = self.wb_technical[eq_type]._images[2]
                    anchor_actual = stamp_2.anchor

                    from_col, from_col_off = anchor_actual._from.col, anchor_actual._from.colOff
                    from_row, from_row_off = anchor_actual._from.row, anchor_actual._from.rowOff
                    to_col, to_col_off = anchor_actual.to.col, anchor_actual.to.colOff
                    to_row, to_row_off = anchor_actual.to.row, anchor_actual.to.rowOff

                    from_cell = AnchorMarker(col=from_col - 1, colOff=from_col_off, row=from_row, rowOff=from_row_off)
                    to_cell = AnchorMarker(col=to_col - 1, colOff=to_col_off, row=to_row, rowOff=to_row_off)

                    new_anchor = TwoCellAnchor(_from = from_cell, to = to_cell, editAs='absolute')
                    stamp_2.anchor = new_anchor

                for value_type in df_level["value_type"].unique():
                    eq_type = (
                        "LEVEL GAUGES ELEMENTS DATA"
                        if value_type == 1
                        else "MAGNETIC ELEMENTS DATA"
                        )

                    ws = self.wb_technical[eq_type]
                    if int(rev) == 0:
                        ws["G4"] = offername_technical

                    last_row = dict_sheets_data[eq_type][0]
                    num_column_amount = dict_sheets_data[eq_type][1]

                    # self.wb_technical[eq_type].delete_rows(last_row + 5, 20)

                    ws.cell(row=row_amount + 2, column=num_column_amount).value = ""
                    ws.cell(row=row_amount + 4, column=num_column_amount).value = ""
                    ws.cell(row=row_amount + 5, column=num_column_amount).value = ""
                    ws.cell(row=row_amount + 6, column=num_column_amount).value = ""
                    ws.cell(row=row_amount + 8, column=num_column_amount).value = ""

                    ws.cell(row=row_amount + 2, column=num_column_amount - 1).value = ""
                    ws.cell(row=row_amount + 4, column=num_column_amount - 1).value = ""
                    ws.cell(row=row_amount + 5, column=num_column_amount - 1).value = ""
                    ws.cell(row=row_amount + 6, column=num_column_amount - 1).value = ""
                    ws.cell(row=row_amount + 8, column=num_column_amount - 1).value = ""

                # Deleting "Amount" column
                    self.wb_technical[eq_type].delete_cols(num_column_amount, 1)

                # Adjusting the print area
                    new_last_column = num_column_amount
                    last_print_row = 40
                    nuevo_rango_impresion = f'A1:{get_column_letter(new_last_column)}{last_print_row}'
                    self.wb_technical[eq_type].print_area = nuevo_rango_impresion

                    stamp_1 = self.wb_technical[eq_type]._images[1]
                    anchor_actual = stamp_1.anchor

                    from_col, from_col_off = anchor_actual._from.col, anchor_actual._from.colOff
                    from_row, from_row_off = anchor_actual._from.row, anchor_actual._from.rowOff
                    to_col, to_col_off = anchor_actual.to.col, anchor_actual.to.colOff
                    to_row, to_row_off = anchor_actual.to.row, anchor_actual.to.rowOff

                    from_cell = AnchorMarker(col=from_col - 1, colOff=from_col_off, row=from_row, rowOff=from_row_off)
                    to_cell = AnchorMarker(col=to_col - 1, colOff=to_col_off, row=to_row, rowOff=to_row_off)

                    new_anchor = TwoCellAnchor(_from = from_cell, to = to_cell, editAs='absolute')
                    stamp_1.anchor = new_anchor

                    stamp_2 = self.wb_technical[eq_type]._images[2]
                    anchor_actual = stamp_2.anchor

                    from_col, from_col_off = anchor_actual._from.col, anchor_actual._from.colOff
                    from_row, from_row_off = anchor_actual._from.row, anchor_actual._from.rowOff
                    to_col, to_col_off = anchor_actual.to.col, anchor_actual.to.colOff
                    to_row, to_row_off = anchor_actual.to.row, anchor_actual.to.rowOff

                    from_cell = AnchorMarker(col=from_col - 1, colOff=from_col_off, row=from_row, rowOff=from_row_off)
                    to_cell = AnchorMarker(col=to_col - 1, colOff=to_col_off, row=to_row, rowOff=to_row_off)

                    new_anchor = TwoCellAnchor(_from = from_cell, to = to_cell, editAs='absolute')
                    stamp_2.anchor = new_anchor

                if int(rev) == 0:
                    ws = self.wb_technical[self.wb_technical.sheetnames[-2]]
                    ws.cell(row=last_row + 3, column=num_column_amount).value = number_items
                    ws.cell(row=last_row + 3, column=num_column_amount).font = Font(name="Calibri", size=14)

            # Editing sheet NOTES
                if int(rev) == 0:
                    sheet_name = "NOTES"  # Selecting  sheet
                    ws = self.wb_technical[sheet_name]

                    rich_string = CellRichText(
                    'We are only offering measuring flow elements. Please be informed that our product range includes temperature elements, and glass and magnetic level indicators; all with european certification. (https://www.eipsa.es/en/products)\n',
                    TextBlock(InlineFont(b=True), 'The prices quoted could be reduced in case of purchasing our full range of products.'))
                    ws["B6"] = rich_string

                    rich_string = CellRichText(
                    TextBlock(InlineFont(i=True), 'Estamos ofertando solamente elementos de medida de caudal, les informamos que en nuestra gama de fabricación con certificación europea, incluye también elementos de temperatura e indicadores de nivel de vidrio y magnéticos. (https://www.eipsa.es/productos)\n'),
                    TextBlock(InlineFont(b=True, i=True), 'Los precios ofertados podrían reducirse en caso de compra de toda nuestra gama.'))
                    ws["B7"] = rich_string

                    rich_string = CellRichText(
                    'Delivery time ' + delivery_time + ' weeks since drawing / calculation approval of ',
                    TextBlock(InlineFont(b=True), 'all equipment'),
                    ' in the contract, as well as critical documentation (August and last two December weeks excluded).')
                    ws["B12"] = rich_string

                    rich_string = CellRichText(
                    'Plazo de entrega ' + delivery_time + ' semanas desde aprobación de planos y cálculos de la ',
                    TextBlock(InlineFont(b=True, i=True), 'totalidad de los equipos'),
                    TextBlock(InlineFont(i=True),' amparados por el contrato, asi como la documentación crítica (Agosto y las dos últimas semanas de diciembre excluidos).'))
                    ws["B13"] = rich_string

                    rich_string = CellRichText(
                    'Quotation prepared according to the information provided in the datasheet corresponding to each TAG. ',
                    TextBlock(InlineFont(u='single'), 'EIPSA does not hold the final responsibility regarding selection of equipment material neither analyze process data.\n'),
                    'The datasheet will be considered the only technical/contractual document, any other documentation will be considered as complementary documentation with informative purpose.')
                    ws["B21"] = rich_string

                    rich_string = CellRichText(
                    'Esta oferta ha sido elaborada en base a la información que figura en la hoja de datos correspondiente a cada TAG, ',
                    TextBlock(InlineFont(i=True, u='single'), 'no siendo responsabilidad final de EIPSA la elección del material a utilizar en los equipos ni analizar datos de proceso.\n'),
                    TextBlock(InlineFont(i=True),'La hoja de datos será el único documento técnico/contractual, cualquier otra documentación recibida será considerada como documentación complementaria a efectos informativos.'))
                    ws["B22"] = rich_string

                    if pay_term == "100_delivery":
                        ws["B45"] = (
                            "100% of total amount of purchase order upon delivery of material according to Incoterms 2020, FCA (our facilities, Spain).\n"
                            "Payment method: bank transfer"
                        )
                        ws["B46"] = (
                            "Pago del 100% del valor total de la orden de compra a la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España).\n"
                            "Método de pago: Transferencia bancaria."
                        )
                    elif pay_term == "100_order":
                        ws["B45"] = (
                            "100 % of the total amount of purchase order upon receipt of purchase order.\n"
                            "Payment method: bank transfer"
                        )
                        ws["B46"] = (
                            "Pago del 100% del valor total de la orden de compra a la recepción de la orden.\n"
                            "Método de pago: Transferencia bancaria"
                        )
                    elif pay_term == "90_10":
                        ws["B45"] = (
                            "PAYMENT TERMS:\n"
                            "90 % of the total amount of PO upon delivery of material according to Incoterms 2020, FCA (our facilities, Spain) and 10% when final documentation is approved. \n"
                            "Bank Transfer: 60 days since invoice issue date."
                        )
                        ws["B46"] = (
                            "TERMINOS DE PAGO:\n"
                            "Pago del 90% del Valor total de la orden de compra a la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España) y el 10% restante cuando la documentación final sea aprobada.\n"
                            "Transferencia Bancaria: 60 días desde emisión de factura."
                        )
                    elif pay_term == "50_50":
                        ws["B45"] = (
                            "50 % of the total amount of purchase order upon receipt of purchase order. Remaining 50% before material be delivered according to Incoterms 2020, FCA (our facilities, Spain).\n"
                            "Payment method: bank transfer."
                        )
                        ws["B46"] = (
                            "Pago del 50% del valor total de la orden de compra a la recepción de la orden. El 50% restante antes de la entrega del material según Incoterm 2020, FCA (nuestras instalaciones, España).\n"
                            "Método de pago: Transferencia bancaria."
                        )
                    elif pay_term == "Others":
                        ws["B45"] = "PAYMENT TERMS TO BE DEFINED"
                        ws["B45"].font = Font(name="Calibri", size=11, bold=True, color="FF0000")
                        ws["B46"] = "TERMINOS DE PAGO POR DEFINIR"
                        ws["B46"].font = Font(name="Calibri", size=11, bold=True, italic=True, color="FF0000")

                    rich_string = CellRichText(
                    'For amounts greater than 30,000.00 € we can issue a warranty bond (if required) valid until the end of the indicated warranty period.\nBond warranty of 10% will be issued with the invoice of the last supplement.\n',
                    TextBlock(InlineFont(b=True), 'For lower amounts no warranty bond is issued.'))
                    ws["B48"] = rich_string

                    rich_string = CellRichText(
                    'Para importes superiores a 30.000,00 €, si es requerido, podremos emitir aval de garantía y estará vigente hasta el final del periodo de garantía indicado.\nEl aval del 10% será emitido con la factura del último suplemento.\n',
                    TextBlock(InlineFont(i=True, b=True), 'Por debajo de dicha cantidad, no se emitirán avales.'))
                    ws["B49"] = rich_string

                    ws["A64"] = (
                        "If you require further information related with this offer, please do not hesitate to contact:\n"
                        + responsible
                        + "\n"
                        + email
                        + "\n"
                        "Telf.: (+34) 916.582.118"
                    )
                    
                    std = self.wb_technical["1.3"]
                    self.wb_technical.remove(std)

                left_text = "Fecha/Date: " + date_offer
                right_text = "Petición nº/Inquiry: " + num_ref

                for sheet in self.wb_technical.worksheets:
                    sheet.oddFooter.left.text = left_text
                    sheet.oddFooter.right.text = right_text
                    sheet.oddFooter.center.text = "Page &P de &N"

                    sheet.oddFooter.left.size = 9
                    sheet.oddFooter.right.size = 9
                    sheet.oddFooter.center.size = 9

                self.save_excel_technical()

                root = Tk()
                root.withdraw()  # Hiding main window Tkinter

                # close communication with the PostgreSQL database server
                # commit the changes
                conn.commit()
        except (Exception, psycopg2.DatabaseError) as error:
            dlg = QtWidgets.QMessageBox()
            new_icon = QtGui.QIcon()
            new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
            dlg.setWindowIcon(new_icon)
            dlg.setWindowTitle("ERP EIPSA")
            dlg.setText("Ha ocurrido el siguiente error:\n"
                        + str(error))
            dlg.setIcon(QtWidgets.QMessageBox.Icon.Critical)
            dlg.exec()
            del dlg, new_icon
        finally:
            if conn is not None:
                conn.close()

    def euros_to_float(self, value):
        """
        Converts a string value representing an amount in euros to a float.
        
        Args:
            value (str): The string representation of an amount in euros, with commas for decimal separation and ' €' for currency indication.
        
        Returns:
            float: The numeric value of the amount in euros.
        """
        value = value.replace(".", "")
        value = value.replace(",", ".")
        value = value[: value.find(" €")]
        return float(value)

    def save_excel_commercial(self):
        """Saves the populated Excel workbook to a specified location.
        Opens a dialog window for the user to select the file path and name.
        """
        # Dialog window to select folder and file name; if path is selected, excel file is saved
        output_path_commercial = asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Archivos de Excel", "*.xlsx")],
            title="Guardar Oferta comercial",
        )
        if output_path_commercial:
            self.wb_commercial.save(output_path_commercial)
            return output_path_commercial

    def save_excel_technical(self):
        """Saves the populated Excel workbook to a specified location.
        Opens a dialog window for the user to select the file path and name.
        """
        output_path_technical = asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Archivos de Excel", "*.xlsx")],
            title="Guardar Oferta técnica",
        )
        if output_path_technical:
            self.wb_technical.save(output_path_technical)

    def adjust_images(self, sheet):
        """
        Adjusts the width of all images in the provided spreadsheet sheet by decreasing each by 22 units.
        
        Args:
            sheet: The spreadsheet sheet containing images to be adjusted.
        """
        for image in sheet._images:
            image.width -= 22


# Templates for commercials
class order_ovr:
    """
    A class to manage order OVR.
    
    Attributes:
        num_order (str): The order number.
        dict_orders (dict): A dictionary to store additional order details.
    """
    def __init__(self, num_order):
        """
        Initializes an order_ovr instance with order number and an empty dictionary for storing orders.
        
        Args:
            num_order (str): The order number.
        """
        self.num_order = num_order
        dict_orders = {}

        query_order_data = ("""
                        SELECT orders."num_order", orders."order_date", orders."num_ref_order"
                        FROM orders
                        WHERE (
                        UPPER(orders."num_order") LIKE UPPER('%%'||%s||'%%')
                        )
                        ORDER BY orders."num_order"
                        """)

        query_flow = ('''
            SELECT tags_data.tags_flow."num_order"
            FROM tags_data.tags_flow
            WHERE UPPER (tags_data.tags_flow."num_order") LIKE UPPER('%%'||%s||'%%')
            ''')
        query_temp = ('''
            SELECT tags_data.tags_temp."num_order"
            FROM tags_data.tags_temp
            WHERE UPPER (tags_data.tags_temp."num_order") LIKE UPPER('%%'||%s||'%%')
            ''')
        query_level = ('''
            SELECT tags_data.tags_level."num_order"
            FROM tags_data.tags_level
            WHERE UPPER (tags_data.tags_level."num_order") LIKE UPPER('%%'||%s||'%%')
            ''')
        query_others = ('''
            SELECT tags_data.tags_others."num_order"
            FROM tags_data.tags_others
            WHERE UPPER (tags_data.tags_others."num_order") LIKE UPPER('%%'||%s||'%%')
            ''')
        conn = None
        try:
        # read the connection parameters
            params = config()
        # connect to the PostgreSQL server
            conn = psycopg2.connect(**params)
            cur = conn.cursor()
        # execution of commands
            cur.execute(query_order_data,(self.num_order,))
            results_orders=cur.fetchall()

            for result in results_orders:
                dict_orders[result[0]] = result[1]

            cur.execute(query_flow,(self.num_order,))
            results_flow=cur.fetchall()
            cur.execute(query_temp,(self.num_order,))
            results_temp=cur.fetchall()
            cur.execute(query_level,(self.num_order,))
            results_level=cur.fetchall()
            cur.execute(query_others,(self.num_order,))
            results_others=cur.fetchall()

            if len(results_flow) != 0:
                self.variable = 'Caudal'
            elif len(results_temp) != 0:
                self.variable = 'Temperatura'
            elif len(results_level) != 0:
                self.variable = 'Nivel'
            elif len(results_others) != 0:
                self.variable = 'Otros'
            else:
                self.variable = ''

        except (Exception, psycopg2.DatabaseError) as error:
            dlg = QtWidgets.QMessageBox()
            new_icon = QtGui.QIcon()
            new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
            dlg.setWindowIcon(new_icon)
            dlg.setWindowTitle("ERP EIPSA")
            dlg.setText("Ha ocurrido el siguiente error:\n"
                        + str(error))
            dlg.setIcon(QtWidgets.QMessageBox.Icon.Critical)
            dlg.exec()
            del dlg, new_icon
        finally:
            if conn is not None:
                conn.close()

            if self.variable == 'Caudal':
                self.table_name = "tags_data.tags_flow"
            elif self.variable == 'Temperatura':
                self.table_name = "tags_data.tags_temp"
            elif self.variable == 'Nivel':
                self.table_name = "tags_data.tags_level"
            elif self.variable == 'Otros':
                self.table_name = "tags_data.tags_others"
            else:
                self.table_name = ''

        commands_tags = f" SELECT num_order, position, subposition, tag, TO_CHAR(contractual_date, 'DD/MM/YYYY'), TO_CHAR(dwg_state_date, 'DD/MM/YYYY'), inspection, TO_CHAR(irc_date, 'DD/MM/YYYY'), TO_CHAR(rn_date, 'DD/MM/YYYY'), TO_CHAR(rn_date + INTERVAL '1 day' * 7, 'DD/MM/YYYY') AS rn_date_plus_7_days FROM {self.table_name} WHERE num_order LIKE UPPER ('%%'||'{self.num_order}'||'%%') ORDER BY num_order"
        self.num_columns = 10
        column_headers = ['Sup.', 'Pos', 'SubPos.', 'TAG', 'PO DELIVERY DATE ', 'DRAWING APPROVAL DATE', 'INSPECTION NUMBER', 'INSPECTION DATE', 'DATE RN', 'NEW DELIVERY DATE']

        try:
        # read the connection parameters
            params = config()
        # connect to the PostgreSQL server
            conn = psycopg2.connect(**params)
            cur = conn.cursor()

            if self.variable != '':
                cur.execute(commands_tags)
                results=cur.fetchall()

        # close communication with the PostgreSQL database server
            cur.close()
        # commit the changes
            conn.commit()

        except (Exception, psycopg2.DatabaseError) as error:
            dlg = QtWidgets.QMessageBox()
            new_icon = QtGui.QIcon()
            new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
            dlg.setWindowIcon(new_icon)
            dlg.setWindowTitle("ERP EIPSA")
            dlg.setText("Ha ocurrido el siguiente error:\n"
                        + str(error))
            dlg.setIcon(QtWidgets.QMessageBox.Icon.Critical)
            dlg.exec()
            del dlg, new_icon
        finally:
            if conn is not None:
                conn.close()

        if self.variable != '' and self.table_name != "tags_data.tags_others":
            data_tags = pd.DataFrame(data=results, columns=column_headers)

            index_tags = data_tags.columns.get_loc('TAG')
            data_tags.insert(index_tags + 1, 'PO DATE',0)

            data_tags['PO DATE'] = data_tags['Sup.'].apply(lambda x: dict_orders[x])
            data_tags['Sup.'] = data_tags['Sup.'].apply(lambda x: x[-1])

            self.wb_ovr = load_workbook(r"\\nas01\DATOS\Comunes\EIPSA-ERP\Plantillas Exportación\PLANTILLA OVR.xlsx")

            sheet_name = "ANNEX I"
            ws = self.wb_ovr[sheet_name]

            if len(dict_orders) > 1:
                ws["A1"] = "ANNEX I - " + results_orders[0][2] + " (" + re.match(r"([A-Z]+-\d+/\d+)", results_orders[0][0]).group(1) + "-S00 to S0" + str(len(dict_orders)) + ")"
            else:
                ws["A1"] = "ANNEX I - " + results_orders[0][2] + " (" + re.match(r"([A-Z]+-\d+/\d+)", results_orders[0][0]).group(1) + "-S00)"

            last_row = 4
            for index, row in data_tags.iterrows():  # Data in desired row
                for col_num, value in enumerate(row, start=1):
                    cell = ws.cell(row=last_row, column=col_num)
                    cell.value = value

                last_row = last_row + 1

            self.save_excel_ovr()

        else:
            dlg_error = QtWidgets.QMessageBox()
            new_icon = QtGui.QIcon()
            new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
            dlg_error.setWindowIcon(new_icon)
            dlg_error.setWindowTitle("Generar OVR")
            dlg_error.setText("No se puede generar el OVR para este pedido")
            dlg_error.setIcon(QtWidgets.QMessageBox.Icon.Warning)
            dlg_error.exec()
            del dlg_error,new_icon

    def save_excel_ovr(self):
        """Saves the populated Excel workbook to a specified location.
        Opens a dialog window for the user to select the file path and name.
        """
        output_path = asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Archivos de Excel", "*.xlsx")],
            title="Guardar OVR",
        )
        if output_path:
            self.wb_ovr.save(output_path)
            return output_path

class doc_situation:
    """
    A class to manage document situation of an order
    
    Attributes:
        num_ref (str): The reference number for the document.
        project (str): The name or identifier of the project.
    """
    def __init__(self, num_ref, project):
        """
        Initializes a doc_situation instance with document reference and project information.
        
        Args:
            num_ref (str): The reference number for the document.
            project (str): The name or identifier of the project.
        """
        self.num_ref = num_ref
        self.project = project

        commands_queryalldoc = ("""
                    SELECT documentation."num_order",orders."num_ref_order",offers."client",TO_CHAR(orders."order_date", 'DD/MM/YYYY'),product_type."variable",
                    documentation."num_doc_client",documentation."num_doc_eipsa",documentation."doc_title",document_type."doc_type",documentation."state",documentation."revision",
                    TO_CHAR(TO_DATE(documentation."state_date", 'DD/MM/YYYY'), 'DD/MM/YYYY'),
                    TO_CHAR(TO_DATE(documentation."date_first_rev", 'DD/MM/YYYY'), 'DD/MM/YYYY'),
                    (TO_DATE(documentation."state_date", 'DD/MM/YYYY') - TO_DATE(documentation."date_first_rev", 'DD/MM/YYYY')) AS difference,
                    CAST(SUBSTRING(offers."delivery_time" FROM POSITION('-' IN offers."delivery_time") + 1) AS INTEGER) AS delivery_weeks,
                    TO_CHAR(
                        (TO_DATE(documentation."state_date", 'DD/MM/YYYY') + INTERVAL '7 days' * 
                        CAST(SUBSTRING(offers."delivery_time" FROM POSITION('-' IN offers."delivery_time") + 1) AS INTEGER)), 
                        'DD/MM/YYYY'
                    ) AS new_date,
                    documentation."tracking"
                    FROM documentation
                    INNER JOIN orders ON (orders."num_order" = documentation."num_order")
                    INNER JOIN offers ON (offers."num_offer" = orders."num_offer")
                    INNER JOIN document_type ON (document_type."id" = documentation."doc_type_id")
                    INNER JOIN product_type ON (product_type."material" = offers."material")
                    WHERE orders."num_ref_order" LIKE (%s||'%%')
                    AND document_type."doc_type" IN ('Cálculo y plano', 'Cálculos', 'Planos', 'Soldadura', 'Pintura')
                    ORDER BY documentation."num_order"
                    """)
        conn = None
        try:
        # read the connection parameters
            params = config()
        # connect to the PostgreSQL server
            conn = psycopg2.connect(**params)
            cur = conn.cursor()
        # execution of commands
            cur.execute(commands_queryalldoc,(self.num_ref,))
            results_orders=cur.fetchall()

        except (Exception, psycopg2.DatabaseError) as error:
            print(error)
            dlg = QtWidgets.QMessageBox()
            new_icon = QtGui.QIcon()
            new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
            dlg.setWindowIcon(new_icon)
            dlg.setWindowTitle("ERP EIPSA")
            dlg.setText("Ha ocurrido el siguiente error:\n"
                        + str(error))
            dlg.setIcon(QtWidgets.QMessageBox.Icon.Critical)
            dlg.exec()
            del dlg, new_icon
        finally:
            if conn is not None:
                conn.close()

        column_headers = ['Nº Pedido', 'Nº PO', 'Cliente', 'Fecha Pedido', 'Material', 'Nº Doc. Cliente', 'Nº Doc. EIPSA', 'Título',
                        'Tipo Doc.', 'Estado', 'Nº Rev.', 'Última Fecha', 'Fecha Rev. 0', 'Días', 'Plazo', 'Fecha Rev.','Seguimiento']

        self.data_docs = pd.DataFrame(data=results_orders, columns=column_headers)

        self.wb_doc = load_workbook(r"\\nas01\DATOS\Comunes\EIPSA-ERP\Plantillas Exportación\PLANTILLA SITUACIÓN DOCS.xlsx")

        sheet_name = "Doc"
        ws = self.wb_doc[sheet_name]

        ws["B1"] = "PROYECTO: " + self.project + " (" + self.num_ref + ")" if self.project is not None else "PROYECTO: SIN PROYECTO (" + self.num_ref + ")"

        ws["Q1"] = date.today().strftime("%d/%m/%y")

        last_row = ws.max_row
        
        for index, row in self.data_docs.iterrows():  # Data in desired row
            for col_num, value in enumerate(row, start=1):
                cell = ws.cell(row=last_row + 1, column=col_num)
                if col_num in [4, 12, 13, 16]:
                    if value is not None:
                        cell.value = datetime.strptime(value, "%d/%m/%Y")
                    else:
                        cell.value = value
                else:
                    cell.value = value

            last_row = ws.max_row

    def save_excel_doc(self):
        """Saves the populated Excel workbook to a specified location.
        Opens a dialog window for the user to select the file path and name.
        """
        # Dialog window to select folder and file name; if path is selected, excel file is saved
        output_path = asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Archivos de Excel", "*.xlsx")],
            title="Guardar Situación Documentos",
        )
        if output_path:
            self.wb_doc.save(output_path)
            wb = load_workbook(output_path)

            # Set date format
            date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
            for col_num in range(17):
                if col_num in [4, 12, 13, 16]:  
                    for row_num in range(3, self.data_docs.shape[0] + 5):
                        cell = wb['Doc'].cell(row=row_num, column=col_num)
                        cell.style = date_style

            for row in wb['Doc'].iter_rows(min_row=3, max_row=wb['Doc'].max_row, min_col=10, max_col=10):
                for cell in row:
                    if cell.value == "Aprobado":
                        cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                    elif cell.value == "Eliminado":
                        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        cell.font = Font(color="FFFFFF", bold=True)

            for row in wb['Doc'].iter_rows(min_row=3, max_row=wb['Doc'].max_row-2, min_col=12, max_col=12):
                for cell in row:
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            for row in wb['Doc'].iter_rows(min_row=3, max_row=wb['Doc'].max_row-2, min_col=14, max_col=14):
                for cell in row:
                    cell.fill = PatternFill(start_color="DA9694", end_color="DA9694", fill_type="solid")

            wb.save(output_path)
            return output_path

class vendor_progress_report:
    """
    A class to manage document situation of an order
    
    Attributes:
        num_ref (str): The reference number for the document.
    """
    def __init__(self, num_ref):
        """
        Initializes a vendor_progress_report instance with project reference.
        
        Args:
            num_ref (str): The reference number for the project.
        """
        self.num_ref = num_ref

        query_flow = ('''
            SELECT tags_data.tags_flow."num_po"
            FROM tags_data.tags_flow
            WHERE UPPER (tags_data.tags_flow."num_po") LIKE UPPER('%%'||%s||'%%')
            ''')
        query_temp = ('''
            SELECT tags_data.tags_temp."num_po"
            FROM tags_data.tags_temp
            WHERE UPPER (tags_data.tags_temp."num_po") LIKE UPPER('%%'||%s||'%%')
            ''')
        query_level = ('''
            SELECT tags_data.tags_level."num_po"
            FROM tags_data.tags_level
            WHERE UPPER (tags_data.tags_level."num_po") LIKE UPPER('%%'||%s||'%%')
            ''')
        query_others = ('''
            SELECT tags_data.tags_others."num_po"
            FROM tags_data.tags_others
            WHERE UPPER (tags_data.tags_others."num_po") LIKE UPPER('%%'||%s||'%%')
            ''')
        conn = None
        try:
        # read the connection parameters
            params = config()
        # connect to the PostgreSQL server
            conn = psycopg2.connect(**params)
            cur = conn.cursor()
        # execution of commands
            cur.execute(query_flow,(num_ref,))
            results_flow=cur.fetchall()
            cur.execute(query_temp,(num_ref,))
            results_temp=cur.fetchall()
            cur.execute(query_level,(num_ref,))
            results_level=cur.fetchall()
            cur.execute(query_others,(num_ref,))
            results_others=cur.fetchall()

            if len(results_flow) != 0 and len(results_temp) != 0:
                self.variable = 'Caudal+Temp'
            elif len(results_flow) != 0 and len(results_level) != 0:
                self.variable = 'Caudal+Nivel'
            elif len(results_flow) != 0:
                self.variable = 'Caudal'
            elif len(results_temp) != 0:
                self.variable = 'Temperatura'
            elif len(results_level) != 0:
                self.variable = 'Nivel'
            elif len(results_others) != 0:
                self.variable = 'Otros'
            else:
                self.variable = ''

        # close communication with the PostgreSQL database server
            cur.close()
        # commit the changes
            conn.commit()
        except (Exception, psycopg2.DatabaseError) as error:
            dlg = QtWidgets.QMessageBox()
            new_icon = QtGui.QIcon()
            new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
            dlg.setWindowIcon(new_icon)
            dlg.setWindowTitle("ERP EIPSA")
            dlg.setText("Ha ocurrido el siguiente error:\n"
                        + str(error))
            dlg.setIcon(QtWidgets.QMessageBox.Icon.Critical)
            dlg.exec()
            del dlg, new_icon
        finally:
            if conn is not None:
                conn.close()

        if self.variable == 'Caudal+Temp':
            self.table_query = "tags_data.tags_flow"
            self.table_query2 = "tags_data.tags_temp"
        elif self.variable =='Caudal+Nivel':
            self.table_query = 'tags_data.tags_flow'
            self.table_query2 = 'tags_data.tags_level'
        elif self.variable == 'Caudal':
            self.table_query = "tags_data.tags_flow"
        elif self.variable == 'Temperatura':
            self.table_query = "tags_data.tags_temp"
        elif self.variable == 'Nivel':
            self.table_query = "tags_data.tags_level"
        elif self.variable == 'Otros':
            self.table_query = "tags_data.tags_others"

        commands_query_data = (
                    f"""SELECT orders."num_ref_order", '' as suppl, tags_data."position", tags_data."subposition", tags_data."tag", '' as empty_column, tags_data."item_type",
                    '' as empty_column2, '' as empty_column3, 'FCA' as incoterm,
                    'Poligono Industrial Igarsa, Naves 4-8, 28860, Paracuellos de Jarama, Spain' as delivery_place, 0 as item_ship_qty,
                    1 as qty_unit, 0 as irc_qty, TO_CHAR(tags_data."irc_date", 'DD/MM/YYYY'), 0 as rn_qty, tags_data."rn_delivery", TO_CHAR(tags_data."rn_date", 'DD/MM/YYYY'),
                    '' as osd_number, TO_CHAR(orders."expected_date", 'DD/MM/YYYY'),
                    CASE 
                        WHEN tags_data."dwg_state_date" IS NULL THEN 'HOLD'
                        ELSE TO_CHAR(tags_data."dwg_state_date" + INTERVAL '7 days' * offers."delivery_time"::integer, 'DD/MM/YYYY')
                    END AS forecast_date,
                    CASE 
                        WHEN tags_data."dwg_state_date" IS NULL THEN NULL
                        ELSE EXTRACT(DAY FROM (tags_data."dwg_state_date" + INTERVAL '7 days' * offers."delivery_time"::integer - orders."expected_date"))
                    END AS deviation_days
                    , 'EIPSA' as manufacturer,
                    'Poligono Industrial Igarsa, Naves 4-8, 28860, Paracuellos de Jarama, Spain' as fab_location, tags_data."inspection", '' as remarks,
                    documentation."state" as drawing_state, documentation."state_date" as drawing_state_date
                    FROM {self.table_query} AS tags_data
                    LEFT JOIN orders ON (tags_data."num_order" = orders."num_order")
                    LEFT JOIN offers ON (orders."num_offer" = offers."num_offer")
                    LEFT JOIN documentation ON (tags_data."dwg_num_doc_eipsa" = documentation."num_doc_eipsa")
                    WHERE tags_data."num_po" LIKE UPPER ('%%'||'{self.num_ref}'||'%%') and tags_data."tag_state" in ('PURCHASED', 'FOR INVOICING')
                    ORDER BY tags_data."tag"
                    """)
        conn = None
        try:
        # read the connection parameters
            params = config()
        # connect to the PostgreSQL server
            conn = psycopg2.connect(**params)
            cur = conn.cursor()
        # execution of commands
            cur.execute(commands_query_data,(self.num_ref,))
            results_progress=cur.fetchall()

        except (Exception, psycopg2.DatabaseError) as error:
            print(error)
            dlg = QtWidgets.QMessageBox()
            new_icon = QtGui.QIcon()
            new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
            dlg.setWindowIcon(new_icon)
            dlg.setWindowTitle("ERP EIPSA")
            dlg.setText("Ha ocurrido el siguiente error:\n"
                        + str(error))
            dlg.setIcon(QtWidgets.QMessageBox.Icon.Critical)
            dlg.exec()
            del dlg, new_icon
        finally:
            if conn is not None:
                conn.close()

        column_headers = ['PO Number', 'Suppl.', 'Pos', 'Sub', 'TAG', '', 'Ident Description',
                        'D1', 'SCH1', 'Incoterm', 'Delivery Place', 'Item Ship. Qty', 'Quantity Unit', 'IRC QTY', 'IRC date',
                        'RN QTY', 'RN Number', 'RN date', 'OSD Number', 'Contractual Delivery Date', 'Forecast Delivery Date', 'DEVIATION (days)',
                        'Manufacturer', 'Fab. Location', 'Final Inspection Date', 'Remarks', 'Drawing State', 'Drawing Date']

        self.data_vpr = pd.DataFrame(data=results_progress, columns=column_headers)


    def save_excel_doc(self):
        """Saves the populated Excel workbook to a specified location.
        Opens a dialog window for the user to select the file path and name.
        """
        file_path = asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])

        if file_path:
            writer = pd.ExcelWriter(file_path, engine='openpyxl')
            self.data_vpr.to_excel(writer, index=False, sheet_name='Sheet1')

            # Set date format
            date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
            for col_num in range(28):
                if col_num in [15, 18, 20, 21, 25, 28]:  
                    for row_num in range(1, self.data_vpr.shape[0] + 5):
                        cell = writer.sheets['Sheet1'].cell(row=row_num, column=col_num)
                        cell.style = date_style

            writer._save()

class spares_two_years:
    """
    A class to manage spares of two years.
    
    Attributes:
        num_order (str): The order number.
        dict_orders (dict): A dictionary to store additional order details.
    """
    def __init__(self, num_order):
        """
        Initializes an spares_two_years instance with order number and an empty dictionary for storing orders.
        
        Args:
            num_order (str): The order number.
        """
        self.num_order = num_order
        dict_orders = {}

        query_order_data = ("""
                        SELECT orders."num_order", orders."order_date", orders."num_ref_order"
                        FROM orders
                        WHERE (
                        UPPER(orders."num_order") LIKE UPPER('%%'||%s||'%%')
                        )
                        ORDER BY orders."num_order"
                        """)

        query_flow = ('''
            SELECT tags_data.tags_flow."num_order"
            FROM tags_data.tags_flow
            WHERE UPPER (tags_data.tags_flow."num_order") LIKE UPPER('%%'||%s||'%%')
            ''')
        query_temp = ('''
            SELECT tags_data.tags_temp."num_order"
            FROM tags_data.tags_temp
            WHERE UPPER (tags_data.tags_temp."num_order") LIKE UPPER('%%'||%s||'%%')
            ''')
        query_level = ('''
            SELECT tags_data.tags_level."num_order"
            FROM tags_data.tags_level
            WHERE UPPER (tags_data.tags_level."num_order") LIKE UPPER('%%'||%s||'%%')
            ''')
        query_others = ('''
            SELECT tags_data.tags_others."num_order"
            FROM tags_data.tags_others
            WHERE UPPER (tags_data.tags_others."num_order") LIKE UPPER('%%'||%s||'%%')
            ''')
        conn = None
        try:
        # read the connection parameters
            params = config()
        # connect to the PostgreSQL server
            conn = psycopg2.connect(**params)
            cur = conn.cursor()
        # execution of commands
            cur.execute(query_order_data,(self.num_order,))
            results_orders=cur.fetchall()

            for result in results_orders:
                dict_orders[result[0]] = result[1]

            cur.execute(query_flow,(self.num_order,))
            results_flow=cur.fetchall()
            cur.execute(query_temp,(self.num_order,))
            results_temp=cur.fetchall()
            cur.execute(query_level,(self.num_order,))
            results_level=cur.fetchall()
            cur.execute(query_others,(self.num_order,))
            results_others=cur.fetchall()

            if len(results_flow) != 0:
                self.variable = 'Caudal'
            elif len(results_temp) != 0:
                self.variable = 'Temperatura'
            elif len(results_level) != 0:
                self.variable = 'Nivel'
            elif len(results_others) != 0:
                self.variable = 'Otros'
            else:
                self.variable = ''

        except (Exception, psycopg2.DatabaseError) as error:
            dlg = QtWidgets.QMessageBox()
            new_icon = QtGui.QIcon()
            new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
            dlg.setWindowIcon(new_icon)
            dlg.setWindowTitle("ERP EIPSA")
            dlg.setText("Ha ocurrido el siguiente error:\n"
                        + str(error))
            dlg.setIcon(QtWidgets.QMessageBox.Icon.Critical)
            dlg.exec()
            del dlg, new_icon
        finally:
            if conn is not None:
                conn.close()

            if self.variable == 'Caudal':
                self.table_name = "tags_data.tags_flow"
            elif self.variable == 'Temperatura':
                self.table_name = "tags_data.tags_temp"
            elif self.variable == 'Nivel':
                self.table_name = "tags_data.tags_level"
            elif self.variable == 'Otros':
                self.table_name = "tags_data.tags_others"
            else:
                self.table_name = ''

        if self.table_name == "tags_data.tags_flow":
            commands_tags = f"""
            SELECT '' AS spare_id, tags.num_order AS model_number, '' AS UNIT,
            tags.tag, tags.line_size || tags.rating || ' Material: ' || tags.gasket_material,
            '' AS group_number, '' AS serial_number,
            tags.dwg_num_doc_eipsa, docs.num_doc_client
            FROM {self.table_name} AS tags
            LEFT JOIN documentation AS docs ON (tags.dwg_num_doc_eipsa = docs.num_doc_eipsa)
            WHERE tags.num_order LIKE UPPER ('%%'||'{self.num_order}'||'%%') AND tags.gasket_material not in ('N/A','OTHERS')
            ORDER BY tags.num_order
            """

        elif self.table_name == "tags_data.tags_temp":
            commands_tags = f"""
            SELECT '' AS spare_id, tags.num_order AS model_number, '' AS UNIT,
            tags.tag, tags.tw_type || ' ' || tags.size || tags.rating || tags.facing || ' Material.' || tags.material_tw ||
            ' - U(mm)=' || tags.ins_length || ' - Rootø(mm)=' || tags.root_diam || ' - Tipø(mm)=' || tags.tip_diam || ' ' || tags.sensor_element,
            '' AS group_number, '' AS serial_number,
            tags.dwg_num_doc_eipsa, docs.num_doc_client
            FROM {self.table_name} AS tags
            LEFT JOIN documentation AS docs ON (tags.dwg_num_doc_eipsa = docs.num_doc_eipsa)
            WHERE tags.num_order LIKE UPPER ('%%'||'{self.num_order}'||'%%')
            ORDER BY tags.num_order
            """

        elif self.table_name == "tags_data.tags_level":
            commands_tags = f"""
            SELECT '' AS spare_id, tags.num_order AS model_number, '' AS UNIT,
            tags.tag, tags.model_num,
            '' AS group_number, '' AS serial_number,
            tags.dwg_num_doc_eipsa, docs.num_doc_client
            FROM {self.table_name} AS tags
            LEFT JOIN documentation AS docs ON (tags.dwg_num_doc_eipsa = docs.num_doc_eipsa)
            WHERE tags.num_order LIKE UPPER ('%%'||'{self.num_order}'||'%%') AND tags.item_type in ('Reflex','Transparent')
            ORDER BY tags.num_order
            """
            

        if self.variable in ['Caudal', 'Temperatura', 'Nivel']:
            try:
            # read the connection parameters
                params = config()
            # connect to the PostgreSQL server
                conn = psycopg2.connect(**params)
                cur = conn.cursor()

                if self.variable != '':
                    cur.execute(commands_tags)
                    results=cur.fetchall()

                    column_headers = ['SPARE ID', 'MODEL NUMBER', 'UNIT', 'TAG', 'DESCRIPTION', 'GROUP', 'SERIAL NUMBER', 'VENDOR DRAWING', 'DRAWING']

            # close communication with the PostgreSQL database server
                cur.close()
            # commit the changes
                conn.commit()

            except (Exception, psycopg2.DatabaseError) as error:
                dlg = QtWidgets.QMessageBox()
                new_icon = QtGui.QIcon()
                new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
                dlg.setWindowIcon(new_icon)
                dlg.setWindowTitle("ERP EIPSA")
                dlg.setText("Ha ocurrido el siguiente error:\n"
                            + str(error))
                dlg.setIcon(QtWidgets.QMessageBox.Icon.Critical)
                dlg.exec()
                del dlg, new_icon
            finally:
                if conn is not None:
                    conn.close()

            data_tags = pd.DataFrame(data=results, columns=column_headers)

        # Identifying the unique descriptions and assigning an ID to each
            df_unique = (
                        data_tags.groupby('DESCRIPTION', as_index=False)
                        .agg({'DESCRIPTION': 'first', 'TAG': 'count'})  # Counting the number of tags for each description
                        .rename(columns={'TAG': 'total_number'})
                        )
            df_unique = df_unique.sort_values(by='DESCRIPTION').reset_index(drop=True)
            df_unique.insert(0, 'desc_id', range(1, len(df_unique) + 1))
            df_unique.insert(1, 'spare_id', "")
            df_unique.insert(2, 'model_number', "")

            if self.variable == 'Nivel':
                df_unique['recommended_pre'] = np.ceil((1 if df_unique['DESCRIPTION'][-1] == 'R' else 2) * int(df_unique['DESCRIPTION'][4]) * df_unique['total_number'] * 0.10).astype(int)
                df_unique['recommended_two'] = np.ceil((1 if df_unique['DESCRIPTION'][-1] == 'R' else 2) * int(df_unique['DESCRIPTION'][4]) * df_unique['total_number'] * 0.20).astype(int)
            else:
                df_unique['recommended_pre'] = np.ceil(df_unique['total_number'] * 0.10).astype(int)
                df_unique['recommended_two'] = np.ceil(df_unique['total_number'] * 0.20).astype(int)
            df_unique['approved'] = None

            df_unique.loc[len(df_unique) + 1] = ["TOTAL", "", "", "Gaskets", "", df_unique['recommended_pre'].sum(), df_unique['recommended_two'].sum(), ""]

            data_tags['GROUP'] = (data_tags['DESCRIPTION'].map(df_unique.set_index('DESCRIPTION')['desc_id']).apply(lambda x: f"{int(x):03d}" if pd.notnull(x) else x))

            data_tags['MODEL NUMBER'] = data_tags.apply(lambda row: 'GAS_SP-' + str(row['MODEL NUMBER'])[:8] + '-' + row['GROUP'], axis=1)
            data_tags['SERIAL NUMBER'] = data_tags.apply(lambda row: 'SN-' + row['MODEL NUMBER'], axis=1)

            self.wb_spares = load_workbook(r"\\nas01\DATOS\Comunes\EIPSA-ERP\Plantillas Exportación\PLANTILLA REPUESTOS.xlsx")

            sheet_name = "APPENDIX 1"
            ws = self.wb_spares[sheet_name]

            last_row = 3
            for index, row in data_tags.iterrows():  # Data in desired row
                for col_num, value in enumerate(row, start=1):
                    cell = ws.cell(row=last_row, column=col_num)
                    cell.value = value

                last_row = last_row + 1

            sheet_name = "APPENDIX 2"
            ws = self.wb_spares[sheet_name]

            last_row = 3
            for index, row in data_tags.iterrows():  # Data in desired row
                for col_num, value in enumerate(row, start=1):
                    cell = ws.cell(row=last_row, column=col_num)
                    cell.value = value

                last_row = last_row + 1

            sheet_name = "SP TABLE"
            ws = self.wb_spares[sheet_name]

            last_row = 3
            for index, row in df_unique.iterrows():  # Data in desired row
                for col_num, value in enumerate(row, start=1):
                    cell = ws.cell(row=last_row, column=col_num)
                    cell.value = value

                last_row = last_row + 1

            self.save_excel_spares()

        else:
            dlg_error = QtWidgets.QMessageBox()
            new_icon = QtGui.QIcon()
            new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
            dlg_error.setWindowIcon(new_icon)
            dlg_error.setWindowTitle("Generar OVR")
            dlg_error.setText("No existen tags de este pedido en el ERP")
            dlg_error.setIcon(QtWidgets.QMessageBox.Icon.Warning)
            dlg_error.exec()
            del dlg_error,new_icon

    def save_excel_spares(self):
        """Saves the populated Excel workbook to a specified location.
        Opens a dialog window for the user to select the file path and name.
        """
        output_path = asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Archivos de Excel", "*.xlsx")],
            title="Guardar Spares",
        )
        if output_path:
            self.wb_spares.save(output_path)
            return output_path

# Templates for technicals
class nuclear_annexes:
    """
    A class to manage nuclear annex details.
    
    Attributes:
        annex (str): The annex identifier.
        numorder (str): The order number associated with the annex.
        ana_code (str): The A.N.A. code.
        ana_order (str): The A.N.A. order.
        line (int): The line number within the annex.
    """
    def __init__(self, annex, numorder, ana_code, ana_order, line):
        """
        Initializes a nuclear_annexes instance with annex details.
        
        Args:
            annex (str): The annex identifier.
            numorder (str): The order number associated with the annex.
            ana_code (str): The A.N.A. code.
            ana_order (str): The A.N.A. order.
            line (int): The line number within the annex.
        """
        self.annex = annex
        self.numorder = numorder
        self.ana_code = ana_code
        self.ana_order = ana_order
        self.line = line

        dict_sensor_types={'1E': 'TIPO E SIMPLE', '1J': 'TIPO J SIMPLE',
        '1K': 'TIPO K SIMPLE', '1N': 'TIPO N SIMPLE',
        '1R': 'TIPO R SIMPLE', '1S': 'TIPO S SIMPLE',
        '1T': 'TIPO T SIMPLE', '2E': 'TIPO E DOBLE',
        '2J': 'TIPO J DOBLE', '2K': 'TIPO K DOBLE',
        '2N': 'TIPO N DOBLE', '2R': 'TIPO R DOBLE',
        '2S': 'TIPO S DOBLE', '2T': 'TIPO T DOBLE',
        '3K': 'TIPO K TRIPLE', '3S': 'TIPO S TRIPLE',
        '1PT100': 'PT100 SIMPLE', '2PT100': 'PT100 DOBLE'}

        commands_calib_data = ("""
                SELECT "tag", "sensor", "master",
                "master_1", "element_1", "error_1", "tolerance_1",
                "master_2", "element_2", "error_2", "tolerance_2",
                "master_3", "element_3", "error_3", "tolerance_3",
                "master_4", "element_4", "error_4", "tolerance_4",
                "notes"
                FROM verification.calibration_thermoelements
                WHERE "num_order" = %s
                """)

        conn = None
        try:
        # read the connection parameters
            params = config()
        # connect to the PostgreSQL server
            conn = psycopg2.connect(**params)
            cur = conn.cursor()
        # execution of commands
            cur.execute(commands_calib_data, (numorder,))
            results = cur.fetchall()

            df = pd.DataFrame(results, columns=["tag", "sensor", "master",
                                    "master_1", "element_1", "error_1", "tolerance_1",
                                    "master_2", "element_2", "error_2", "tolerance_2",
                                    "master_3", "element_3", "error_3", "tolerance_3",
                                    "master_4", "element_4", "error_4", "tolerance_4",
                                    "notes"])

            df.sort_values(by=['tag'])

        # close communication with the PostgreSQL database server
            cur.close()
        # commit the changes
            conn.commit()

        except (Exception, psycopg2.DatabaseError) as error:
            dlg = QtWidgets.QMessageBox()
            new_icon = QtGui.QIcon()
            new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
            dlg.setWindowIcon(new_icon)
            dlg.setWindowTitle("ERP EIPSA")
            dlg.setText("Ha ocurrido el siguiente error:\n"
                        + str(error))
            dlg.setIcon(QtWidgets.QMessageBox.Icon.Critical)
            dlg.exec()
            del dlg, new_icon

        finally:
            if conn is not None:
                conn.close()

        for i in range(1, 5):
            df[f'master_value{i}'] = df.apply(lambda row: self.calculate_master(row[f'master_{i}'], row['master']), axis=1)

        for i in range(1, 5):
            df[f'element_value{i}'] = df.apply(lambda row: self.calculate_element(row[f'element_{i}'], row['sensor']), axis=1)

        df['sensor'] = df['sensor'].map(dict_sensor_types)

        if self.annex == 'Anexo A':
            self.wb = load_workbook(r"\\nas01\DATOS\Comunes\NUCLEAR\FORMATOS ANEXOS\ANEXO A Termopar con union a masa.xlsx")
        elif self.annex == 'Anexo B':
            self.wb = load_workbook(r"\\nas01\DATOS\Comunes\NUCLEAR\FORMATOS ANEXOS\ANEXO B Termopar aislado de masa.xlsx")
        elif self.annex == 'Anexo C':
            self.wb = load_workbook(r"\\nas01\DATOS\Comunes\NUCLEAR\FORMATOS ANEXOS\ANEXO C RTD.xlsx")

        ws_initial = self.wb["Hoja1"]

        ws_initial["C4"] = self.ana_code
        ws_initial["E4"] = self.ana_order
        ws_initial["H4"] = self.line

        df['tag_sliced'] = df['tag'].apply(lambda x: x[:-2])

        unique_values = df['tag_sliced'].unique().tolist()

        for item in unique_values:
            self.wb.copy_worksheet(ws_initial)
            ws_copy = self.wb["Hoja1 Copy"]
            ws_copy.title = item.replace('/', '-')

            new_df = df[df['tag_sliced']==item]

            ws_copy["D6"] = new_df.iloc[0,1]
            ws_copy["H6"] = item

            if self.annex in ['Anexo A', 'Anexo B']:
                ws_copy["G9"] = "A: " + str(new_df.iloc[new_df.index.get_loc(new_df[new_df['tag'] == item + '_A'].index[0]),19]).replace('.',',')[:4] if new_df.shape[0] == 2 else ""
                ws_copy["G10"] = "B: " + str(new_df.iloc[new_df.index.get_loc(new_df[new_df['tag'] == item + '_B'].index[0]),19]).replace('.',',')[:4] if new_df.shape[0] == 2 else str(new_df.iloc[new_df.index.get_loc(new_df[new_df['tag'] == item + '_A'].index[0]),19]).replace('.',',')[:4]

            ws_copy["C23"] = "A: " + str(new_df.iloc[new_df.index.get_loc(new_df[new_df['tag'] == item + '_A'].index[0]),20]).replace('.',',') + " / " +  str(round(new_df.iloc[new_df.index.get_loc(new_df[new_df['tag'] == item + '_A'].index[0]),3],3)).replace('.',',') if new_df.shape[0] == 2 else ""
            ws_copy["C24"] = "B: " + str(new_df.iloc[new_df.index.get_loc(new_df[new_df['tag'] == item + '_B'].index[0]),20]).replace('.',',') + " / " +  str(round(new_df.iloc[new_df.index.get_loc(new_df[new_df['tag'] == item + '_B'].index[0]),3],3)).replace('.',',') if new_df.shape[0] == 2 else str(new_df.iloc[new_df.index.get_loc(new_df[new_df['tag'] == item + '_A'].index[0]),20]).replace('.',',') + " / " +  str(round(new_df.iloc[new_df.index.get_loc(new_df[new_df['tag'] == item + '_A'].index[0]),3],3)).replace('.',',')
            ws_copy["E23"] = "A: " + str(new_df.iloc[new_df.index.get_loc(new_df[new_df['tag'] == item + '_A'].index[0]),24]).replace('.',',') + " / " +  str(round(new_df.iloc[new_df.index.get_loc(new_df[new_df['tag'] == item + '_A'].index[0]),4],3)).replace('.',',') if new_df.shape[0] == 2 else ""
            ws_copy["E24"] = "B: " + str(new_df.iloc[new_df.index.get_loc(new_df[new_df['tag'] == item + '_B'].index[0]),24]).replace('.',',') + " / " +  str(round(new_df.iloc[new_df.index.get_loc(new_df[new_df['tag'] == item + '_B'].index[0]),4],3)).replace('.',',') if new_df.shape[0] == 2 else str(new_df.iloc[new_df.index.get_loc(new_df[new_df['tag'] == item + '_B'].index[0]), 24]).replace('.',',')
            ws_copy["G23"] = ("" if 'TIPO E' in new_df.iloc[0,1] else ("" if 'TIPO J' in new_df.iloc[0,1] else ("" if 'TIPO K' in new_df.iloc[0,1] else str(new_df.iloc[new_df.index.get_loc(new_df[new_df['tag'] == item + '_A'].index[0]),6]))))  if new_df.shape[0] == 2 else ""
            ws_copy["G24"] = ("± 1,7" if 'TIPO E' in new_df.iloc[0,1] else ("± 2,2" if 'TIPO J' in new_df.iloc[0,1] else ("± 2,2" if 'TIPO K' in new_df.iloc[0,1] else str(new_df.iloc[new_df.index.get_loc(new_df[new_df['tag'] == item + '_B'].index[0]),6]))))  if new_df.shape[0] == 2 else ("" if 'TIPO E' in new_df.iloc[0,1] else ("" if 'TIPO J' in new_df.iloc[0,1] else ("" if 'TIPO K' in new_df.iloc[0,1] else str(new_df.iloc[new_df.index.get_loc(new_df[new_df['tag'] == item + '_A'].index[0]),6])))) 

            ws_copy["C25"] = "A: " + str(new_df.iloc[new_df.index.get_loc(new_df[new_df['tag'] == item + '_A'].index[0]),21]).replace('.',',') + " / " +  str(round(new_df.iloc[new_df.index.get_loc(new_df[new_df['tag'] == item + '_A'].index[0]),7],3)).replace('.',',') if new_df.shape[0] == 2 else ""
            ws_copy["C26"] = "B: " + str(new_df.iloc[new_df.index.get_loc(new_df[new_df['tag'] == item + '_B'].index[0]),21]).replace('.',',') + " / " +  str(round(new_df.iloc[new_df.index.get_loc(new_df[new_df['tag'] == item + '_B'].index[0]),7],3)).replace('.',',') if new_df.shape[0] == 2 else str(new_df.iloc[new_df.index.get_loc(new_df[new_df['tag'] == item + '_A'].index[0]),21]).replace('.',',') + " / " +  str(round(new_df.iloc[new_df.index.get_loc(new_df[new_df['tag'] == item + '_A'].index[0]),7],3)).replace('.',',')
            ws_copy["E25"] = "A: " + str(new_df.iloc[new_df.index.get_loc(new_df[new_df['tag'] == item + '_A'].index[0]),25]).replace('.',',') + " / " +  str(round(new_df.iloc[new_df.index.get_loc(new_df[new_df['tag'] == item + '_A'].index[0]),8],3)).replace('.',',') if new_df.shape[0] == 2 else ""
            ws_copy["E26"] = "B: " + str(new_df.iloc[new_df.index.get_loc(new_df[new_df['tag'] == item + '_B'].index[0]),25]).replace('.',',') + " / " +  str(round(new_df.iloc[new_df.index.get_loc(new_df[new_df['tag'] == item + '_B'].index[0]),8],3)).replace('.',',') if new_df.shape[0] == 2 else str(new_df.iloc[new_df.index.get_loc(new_df[new_df['tag'] == item + '_B'].index[0]), 25]).replace('.',',')
            ws_copy["G25"] = ("" if 'TIPO E' in new_df.iloc[0,1] else ("" if 'TIPO J' in new_df.iloc[0,1] else ("" if 'TIPO K' in new_df.iloc[0,1] else str(new_df.iloc[new_df.index.get_loc(new_df[new_df['tag'] == item + '_A'].index[0]),10]))))  if new_df.shape[0] == 2 else ""
            ws_copy["G26"] = ("± 1,7" if 'TIPO E' in new_df.iloc[0,1] else ("± 2,2" if 'TIPO J' in new_df.iloc[0,1] else ("± 2,2" if 'TIPO K' in new_df.iloc[0,1] else str(new_df.iloc[new_df.index.get_loc(new_df[new_df['tag'] == item + '_B'].index[0]),10]))))  if new_df.shape[0] == 2 else ("" if 'TIPO E' in new_df.iloc[0,1] else ("" if 'TIPO J' in new_df.iloc[0,1] else ("" if 'TIPO K' in new_df.iloc[0,1] else str(new_df.iloc[new_df.index.get_loc(new_df[new_df['tag'] == item + '_A'].index[0]),10])))) 

            ws_copy["C27"] = "A: " + str(new_df.iloc[new_df.index.get_loc(new_df[new_df['tag'] == item + '_A'].index[0]),22]).replace('.',',') + " / " +  str(round(new_df.iloc[0,11],3)).replace('.',',') if new_df.shape[0] == 2 else ""
            ws_copy["C28"] = "B: " + str(new_df.iloc[new_df.index.get_loc(new_df[new_df['tag'] == item + '_B'].index[0]),22]).replace('.',',') + " / " +  str(round(new_df.iloc[new_df.index.get_loc(new_df[new_df['tag'] == item + '_B'].index[0]),11],3)).replace('.',',') if new_df.shape[0] == 2 else str(new_df.iloc[new_df.index.get_loc(new_df[new_df['tag'] == item + '_A'].index[0]),22]).replace('.',',') + " / " +  str(round(new_df.iloc[new_df.index.get_loc(new_df[new_df['tag'] == item + '_A'].index[0]),11],3)).replace('.',',')
            ws_copy["E27"] = "A: " + str(new_df.iloc[new_df.index.get_loc(new_df[new_df['tag'] == item + '_A'].index[0]),26]).replace('.',',') + " / " +  str(round(new_df.iloc[0,12],3)).replace('.',',') if new_df.shape[0] == 2 else ""
            ws_copy["E28"] = "B: " + str(new_df.iloc[new_df.index.get_loc(new_df[new_df['tag'] == item + '_B'].index[0]),26]).replace('.',',') + " / " +  str(round(new_df.iloc[new_df.index.get_loc(new_df[new_df['tag'] == item + '_B'].index[0]),12],3)).replace('.',',') if new_df.shape[0] == 2 else str(new_df.iloc[new_df.index.get_loc(new_df[new_df['tag'] == item + '_B'].index[0]), 26]).replace('.',',')
            ws_copy["G27"] = ("" if 'TIPO E' in new_df.iloc[0,1] else ("" if 'TIPO J' in new_df.iloc[0,1] else ("" if 'TIPO K' in new_df.iloc[0,1] else str(new_df.iloc[new_df.index.get_loc(new_df[new_df['tag'] == item + '_A'].index[0]),14]))))  if new_df.shape[0] == 2 else ""
            ws_copy["G28"] = ("± 1,7" if 'TIPO E' in new_df.iloc[0,1] else ("± 2,2" if 'TIPO J' in new_df.iloc[0,1] else ("± 2,2" if 'TIPO K' in new_df.iloc[0,1] else str(new_df.iloc[new_df.index.get_loc(new_df[new_df['tag'] == item + '_B'].index[0]),14]))))  if new_df.shape[0] == 2 else ("" if 'TIPO E' in new_df.iloc[0,1] else ("" if 'TIPO J' in new_df.iloc[0,1] else ("" if 'TIPO K' in new_df.iloc[0,1] else str(new_df.iloc[new_df.index.get_loc(new_df[new_df['tag'] == item + '_A'].index[0]),14])))) 

            ws_copy["B40"] = "PATRÓN " + new_df.iloc[0,2] + ", FLUKE 8842A(030), MEGHOMETRO (022)"

    def save_excel_doc(self):
        """Saves the populated Excel workbook to a specified location.
        Opens a dialog window for the user to select the file path and name.
        """
        #Dialog window to select folder and file name; if path is selected, excel file is saved
        output_path = asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Archivos de Excel", "*.xlsx")],
            title="Guardar Situación Documentos",
        )
        if output_path:
            self.wb.save(output_path)

            del self.wb['Hoja1']

            self.wb.save(output_path)

            return output_path

    def calculate_master(self, temp, master):
        """
        Calculates the standard value based on temperature and the type of master sensor.

        Args:
            temp (float or None): The temperature value to calculate the standard value for. If None, returns 'N/A'.
            master (str): The master sensor type (e.g., 'EIPSA-020', 'EIPSA-TE-01').

        Returns:
            float or str: The calculated standard value, or 'N/A' if the temperature is None.
        """
        if temp is not None:
            if master in ['EIPSA-020', 'EIPSA-TE-01']:
                column_select = 'inta_pt100_values.' + master.replace('-','_')
                commands_intavalues = f"""
                                    SELECT {column_select}
                                    FROM verification.inta_pt100_values
                                    ORDER BY variables
                                    """
                conn = None
                try:
                # read the connection parameters
                    params = config()
                # connect to the PostgreSQL server
                    conn = psycopg2.connect(**params)
                    cur = conn.cursor()
                # execution of commands
                    cur.execute(commands_intavalues)
                    results = cur.fetchall()

                    a_inta = results[0][0]
                    b_inta = results[1][0]
                    c_inta = results[2][0]
                    r_zero = results[3][0]

                # close communication with the PostgreSQL database server
                    cur.close()
                # commit the changes
                    conn.commit()

                except (Exception, psycopg2.DatabaseError) as error:
                    dlg = QtWidgets.QMessageBox()
                    new_icon = QtGui.QIcon()
                    new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
                    dlg.setWindowIcon(new_icon)
                    dlg.setWindowTitle("ERP EIPSA")
                    dlg.setText("Ha ocurrido el siguiente error:\n"
                                + str(error))
                    dlg.setIcon(QtWidgets.QMessageBox.Icon.Critical)
                    dlg.exec()
                    del dlg, new_icon

                finally:
                    if conn is not None:
                        conn.close()

                if temp < 0:
                    final_value = round(r_zero * (1 + a_inta * temp + b_inta * temp**2 + c_inta* (temp - 100) * temp**3), 3)
                else:
                    final_value = round(r_zero * (1 + a_inta * temp + b_inta * temp**2), 3)

            else:
                column_select = 'inta_tc_values.' + master.replace('-','_')
                commands_intavalues = f"""
                                    SELECT {column_select}
                                    FROM verification.inta_tc_values
                                    ORDER BY variables
                                    """
                conn = None
                try:
                # read the connection parameters
                    params = config()
                # connect to the PostgreSQL server
                    conn = psycopg2.connect(**params)
                    cur = conn.cursor()
                # execution of commands
                    cur.execute(commands_intavalues)
                    results = cur.fetchall()

                    a_inta = results[0][0]
                    b_inta = results[1][0]
                    c_inta = results[2][0]

                # close communication with the PostgreSQL database server
                    cur.close()
                # commit the changes
                    conn.commit()

                except (Exception, psycopg2.DatabaseError) as error:
                    dlg = QtWidgets.QMessageBox()
                    new_icon = QtGui.QIcon()
                    new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
                    dlg.setWindowIcon(new_icon)
                    dlg.setWindowTitle("ERP EIPSA")
                    dlg.setText("Ha ocurrido el siguiente error:\n"
                                + str(error))
                    print(error)
                    dlg.setIcon(QtWidgets.QMessageBox.Icon.Critical)
                    dlg.exec()
                    del dlg, new_icon

                finally:
                    if conn is not None:
                        conn.close()

                final_value = round((a_inta + b_inta * temp* + c_inta * temp**2)/1000, 3)

        else:
            final_value = 'N/A'

        return final_value

    def calculate_element(self, temp, sensor):
        """
        Calculates the standard value based on temperature and sensor type.

        Args:
            temp (float or None): The temperature value to calculate the standard value for. If None, returns 'N/A'.
            sensor (str): The type of sensor used for the calculation (e.g., 'PT100', 'K', 'J', etc.).

        Returns:
            float or str: The calculated standard value, or 'N/A' if the temperature is None.
        """
        if temp is not None:
            if 'PT100' in sensor:
                commands_stdvalues = ("""
                                    SELECT values
                                    FROM verification.standard_pt100_values
                                    ORDER BY variables
                                    """)
                conn = None
                try:
                # read the connection parameters
                    params = config()
                # connect to the PostgreSQL server
                    conn = psycopg2.connect(**params)
                    cur = conn.cursor()
                # execution of commands
                    cur.execute(commands_stdvalues)
                    results = cur.fetchall()

                    a_std = results[0][0]
                    b_std = results[1][0]
                    c_std = results[2][0]
                    r_zero = results[3][0]

                # close communication with the PostgreSQL database server
                    cur.close()
                # commit the changes
                    conn.commit()

                except (Exception, psycopg2.DatabaseError) as error:
                    dlg = QtWidgets.QMessageBox()
                    new_icon = QtGui.QIcon()
                    new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
                    dlg.setWindowIcon(new_icon)
                    dlg.setWindowTitle("ERP EIPSA")
                    dlg.setText("Ha ocurrido el siguiente error:\n"
                                + str(error))
                    dlg.setIcon(QtWidgets.QMessageBox.Icon.Critical)
                    dlg.exec()
                    del dlg, new_icon

                finally:
                    if conn is not None:
                        conn.close()

                if temp < 0:
                    final_value = round(r_zero * (1 + a_std * temp + b_std * temp**2 + c_std* (temp - 100) * temp**3), 3)
                else:
                    final_value = round(r_zero * (1 + a_std * temp + b_std * temp**2), 3)

            else:
                if 'B' in sensor:
                    table = 'verification.standard_tc_b_values'
                    column = 'low' if temp <= 630.615 else 'high'
                elif 'C' in sensor:
                    table = 'verification.standard_tc_c_values'
                    column = 'low' if temp <= 630.615 else 'high'
                elif 'E' in sensor:
                    table = 'verification.standard_tc_e_values'
                    column = 'low' if temp <= 0 else 'high'
                elif 'J' in sensor:
                    table = 'verification.standard_tc_j_values'
                    column = 'low' if temp <= 760 else 'high'
                elif 'K' in sensor:
                    table = 'verification.standard_tc_k_values'
                    column = 'low' if temp <= 0 else 'high'
                elif 'N' in sensor:
                    table = 'verification.standard_tc_n_values'
                    column = 'low' if temp <= 0 else 'high'
                elif 'R' in sensor:
                    table = 'verification.standard_tc_r_values'
                    column = 'low' if temp <= 1064.18 else ('medium' if temp <= 1664.5 else 'high')
                elif 'S' in sensor:
                    table = 'verification.standard_tc_s_values'
                    column = 'low' if temp <= 1064.18 else ('medium' if temp <= 1664.5 else 'high')
                elif 'T' in sensor:
                    table = 'verification.standard_tc_t_values'
                    column = 'low' if temp <= 0 else 'high'

                commands_stdvalues = f"""
                                    SELECT {column}
                                    FROM {table}
                                    ORDER BY id
                                    """
                conn = None
                try:
                # read the connection parameters
                    params = config()
                # connect to the PostgreSQL server
                    conn = psycopg2.connect(**params)
                    cur = conn.cursor()
                # execution of commands
                    cur.execute(commands_stdvalues)
                    results = cur.fetchall()

                # close communication with the PostgreSQL database server
                    cur.close()
                # commit the changes
                    conn.commit()

                except (Exception, psycopg2.DatabaseError) as error:
                    dlg = QtWidgets.QMessageBox()
                    new_icon = QtGui.QIcon()
                    new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
                    dlg.setWindowIcon(new_icon)
                    dlg.setWindowTitle("ERP EIPSA")
                    dlg.setText("Ha ocurrido el siguiente error:\n"
                                + str(error))
                    dlg.setIcon(QtWidgets.QMessageBox.Icon.Critical)
                    dlg.exec()
                    del dlg, new_icon

                finally:
                    if conn is not None:
                        conn.close()

                final_value = 0

                if 'K' in sensor:
                    for i in range(len(results)-2):
                        final_value += float(results[i][0]) * float(temp)**i

                    final_value += float(results[11][0]) * exp(float(results[12][0]) * (float(temp) - 126.9686)**2)

                else:
                    for i in range(len(results)):
                        final_value += results[i][0] * temp**i

                final_value = round(final_value, 3)

        else:
            final_value = 'N/A'

        return final_value

class material_order:
    """Class for creating and saving an Excel order for raw materials.

    Attributes:
        wb (Workbook): The loaded Excel workbook.
    """
    def __init__(self, df, num_order, client, variable, num_ot):
        """Initializes the material order by loading a template and filling it with data.

        Args:
            df (DataFrame): DataFrame containing the data to populate the Excel template.
            num_order (str): The order number.
            client (str): The client name.
            variable (str): Additional variable information.
            num_ot (str): The order task number.
        """
        # Loading Excel Template
        self.wb = load_workbook(
            r"\\nas01\DATOS\Comunes\EIPSA-ERP\Plantillas Exportación\Pedido Materia Prima.xlsx"
        )
        sheet_name = "Hoja1"  # Selecting template sheet
        ws = self.wb[sheet_name]
        start_row = 12  # Obtaining last row used
        row_11_style = {}
        for col_num in range(1, 15):
            cell_11 = ws.cell(row=12, column=col_num)
            row_11_style[col_num] = deepcopy(cell_11._style)

        for index, row in df.iterrows():
            for col_num, value in enumerate(row, start=4):
                cell = ws.cell(row=start_row + index, column=col_num)
                cell.value = value
                for num in range(1, 15):
                    cell = ws.cell(row=start_row + index, column=num)
                    cell._style = deepcopy(row_11_style[num])

        # Adding text in cell L4, C5, C6, H1 and H9
        ws["L4"] = num_order
        ws["C5"] = client
        ws["C6"] = variable
        ws["H1"] = int(int(num_ot)+1)
        ws["H9"] = date.today().strftime("%d/%m/%Y")

        root = Tk()
        root.withdraw()  # Hiding main window Tkinter

    def save_excel(self):
        """Saves the populated Excel workbook to a specified location.
        Opens a dialog window for the user to select the file path and name.
        """
        #Dialog window to select folder and file name; if path is selected, excel file is saved
        output_path = asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Archivos de Excel", "*.xlsx")],
            title="Guardar archivo de Excel",
        )
        if output_path:
            self.wb.save(output_path)


# offer_short_flow('O-22/032', 'l.bravo', '0', 'project', 'FCA', '10-12', '30', '90_10', '123', '', '')
# offer_short_temp('O-23/001', 'l.bravo', '0', 'project', 'FCA', '10-12', '30', '90_10', '123', '', '')
# offer_short_level('OE-23/114', 'l.bravo', '0', '-', 'FCA (our facilities on truck)', '12-16', '90', '50_50', '100', '', '')